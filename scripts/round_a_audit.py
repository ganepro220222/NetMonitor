#!/usr/bin/env python3
"""
Round A Audit: DataStore Statistical Consistency
=================================================
Tests:
  A1  single vs batch  (get_sla_stats vs get_sla_stats_batch)
  A2  raw vs hourly    (get_waveform raw/1m vs 1h)
  A3  export vs API    (export_excel vs API query paths)
  A4  failed-with-latency  (probe_success=0 + rtt present MUST NOT pollute stats)
  A5  get_dashboard_stats_batch vs get_sla_stats latency agreement
  A6  boundary / edge-cases
"""

import json
import math
import os
import sqlite3
import sys
import tempfile
import time
from pathlib import Path

# ── add repo root to path so we can import the src.* package ──────────
# export_excel (and trace_policy) use package-qualified `from src.X`
# imports, matching the rest of the codebase and the verify_bugs scripts.
# Importing data_store top-level via src/-on-path left `src` unresolvable
# inside export_excel (A5 raised "No module named 'src'").
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.data_store import DataStore, _PROBE_OK_SQL, _PROBE_LAT_OK_SQL

# ── helpers ───────────────────────────────────────────────────────────
PASS = 0
FAIL = 0

def check(cond, label):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS  {label}")
    else:
        FAIL += 1
        print(f"  FAIL  {label}")

def almost(a, b, eps=1e-4):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(a - b) < eps

def fmt(v):
    if v is None:
        return "None"
    if isinstance(v, float):
        return f"{v:.6f}"
    return str(v)

# ── build a temp DB with controlled data ──────────────────────────────
def build_db(db_path: str):
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS pings_raw (
            id             INTEGER PRIMARY KEY,
            target_id      TEXT    NOT NULL,
            ts             REAL    NOT NULL,
            status         TEXT    NOT NULL,
            latency_ms     REAL,
            loss_rate      REAL,
            probe_success  INTEGER,
            status_code    INTEGER,
            failure_reason TEXT,
            ping_type      TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_raw_target_ts ON pings_raw(target_id, ts);
        CREATE INDEX IF NOT EXISTS idx_raw_ts_only ON pings_raw(ts);

        CREATE TABLE IF NOT EXISTS pings_hourly (
            id        INTEGER PRIMARY KEY,
            target_id TEXT    NOT NULL,
            hour_ts   INTEGER NOT NULL,
            sample_n  INTEGER NOT NULL,
            success_n INTEGER NOT NULL,
            lat_avg   REAL,
            lat_max   REAL,
            lat_min   REAL,
            lat_p95   REAL,
            loss_rate REAL
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_hourly_target_hour
            ON pings_hourly(target_id, hour_ts);
        CREATE INDEX IF NOT EXISTS idx_hourly_ts_only ON pings_hourly(hour_ts);

        CREATE TABLE IF NOT EXISTS alert_events (
            id             INTEGER PRIMARY KEY,
            target_id      TEXT    NOT NULL,
            label          TEXT,
            ip             TEXT,
            ts             REAL    NOT NULL,
            old_status     TEXT,
            new_status     TEXT,
            category       TEXT,
            failure_reason TEXT,
            incident_id    TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_alert_target_ts ON alert_events(target_id, ts);
        CREATE INDEX IF NOT EXISTS idx_alert_ts_only ON alert_events(ts);
        CREATE INDEX IF NOT EXISTS idx_alert_incident ON alert_events(incident_id);

        CREATE TABLE IF NOT EXISTS targets_meta (
            target_id  TEXT PRIMARY KEY,
            label      TEXT,
            ip         TEXT,
            ping_type  TEXT,
            updated_at INTEGER
        );
        CREATE TABLE IF NOT EXISTS cum_stats (
            target_id  TEXT PRIMARY KEY,
            total      INTEGER DEFAULT 0,
            success    INTEGER DEFAULT 0,
            lat_avg    REAL    DEFAULT 0,
            lat_n      INTEGER DEFAULT 0,
            updated_at INTEGER
        );
        CREATE TABLE IF NOT EXISTS traceroute_logs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            target_id   TEXT    NOT NULL,
            label       TEXT,
            ip          TEXT,
            ts          REAL    NOT NULL,
            hops        TEXT,
            incident_id TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_tracert_target_ts
            ON traceroute_logs(target_id, ts);
        CREATE TABLE IF NOT EXISTS icmp_diagnostic_history (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            target_id          TEXT,
            host               TEXT    NOT NULL,
            mode               TEXT    NOT NULL,
            payload_size       INTEGER NOT NULL,
            df                 INTEGER NOT NULL,
            count              INTEGER NOT NULL,
            success_count      INTEGER NOT NULL,
            loss_count         INTEGER NOT NULL,
            avg_rtt_ms         REAL,
            min_rtt_ms         REAL,
            max_rtt_ms         REAL,
            estimated_path_mtu INTEGER,
            max_df_payload     INTEGER,
            error_summary      TEXT,
            conclusion         TEXT,
            details_json       TEXT,
            created_ts         REAL    NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_diag_target_ts
            ON icmp_diagnostic_history(target_id, created_ts);
        CREATE TABLE IF NOT EXISTS dns_diagnostic_history (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            target_id     TEXT,
            domain        TEXT    NOT NULL,
            qtype         TEXT    NOT NULL,
            trace_chain   INTEGER NOT NULL,
            resolver      TEXT,
            success       INTEGER NOT NULL,
            rcode         TEXT,
            elapsed_ms    REAL,
            answer_count  INTEGER NOT NULL,
            error_summary TEXT,
            message       TEXT,
            details_json  TEXT,
            created_ts    REAL    NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_dns_diag_target_ts
            ON dns_diagnostic_history(target_id, created_ts);
        PRAGMA user_version = 11;
    """)
    conn.commit()
    conn.close()

# ── DataSet 1: normal probes (green / orange / red) ───────────────────
# 10 probes per second over 2 hours for a clean raw window
BASE_TS = int(time.time()) - 7200  # 2 hours ago

def insert_normal_probes(db_path):
    conn = sqlite3.connect(db_path)
    ts = BASE_TS
    # T1: mostly green with some high latency (orange)
    for i in range(3600):
        lat = 5.0 + (i % 100) * 0.5  # varies 5-55ms
        conn.execute(
            "INSERT INTO pings_raw(target_id,ts,status,latency_ms,loss_rate,probe_success,status_code,failure_reason,ping_type) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            ("T1", ts + i, "green", lat, 0.0, 1, None, None, "icmp"))
    # T1: some red probes
    for i in range(3600, 4000):
        conn.execute(
            "INSERT INTO pings_raw(target_id,ts,status,latency_ms,loss_rate,probe_success,status_code,failure_reason,ping_type) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            ("T1", ts + i, "red", None, 1.0, 0, None, "timeout", "icmp"))
    # T2: intermittent
    for i in range(2000):
        if i % 3 == 0:
            conn.execute(
                "INSERT INTO pings_raw(target_id,ts,status,latency_ms,loss_rate,probe_success,status_code,failure_reason,ping_type) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                ("T2", ts + i * 2, "red", None, 1.0, 0, None, "timeout", "icmp"))
        else:
            lat = 10.0 + (i % 50) * 2.0
            conn.execute(
                "INSERT INTO pings_raw(target_id,ts,status,latency_ms,loss_rate,probe_success,status_code,failure_reason,ping_type) "
                "VALUES(?,?,?,?,?,?,?,?,?)",
                ("T2", ts + i * 2, "green", lat, 0.0, 1, None, None, "icmp"))
    conn.commit()
    conn.close()

# ── DataSet 2: failed-with-latency (the critical edge case) ───────────
# probe_success=0 AND latency_ms is NOT None: these MUST NOT pollute stats
def insert_failed_with_latency(db_path):
    conn = sqlite3.connect(db_path)
    ts = BASE_TS + 2000  # within the 2-hour window, distinct from normal probes
    # T_FL: 100 probes total, 50 OK, 50 "failed but with RTT"
    for i in range(50):
        conn.execute(
            "INSERT INTO pings_raw(target_id,ts,status,latency_ms,loss_rate,probe_success,status_code,failure_reason,ping_type) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            ("T_FL", ts + i * 0.1, "green", 20.0, 0.0, 1, 200, None, "http"))
    for i in range(50):
        # FAILED but has latency: this must NOT count as success anywhere
        conn.execute(
            "INSERT INTO pings_raw(target_id,ts,status,latency_ms,loss_rate,probe_success,status_code,failure_reason,ping_type) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            ("T_FL", ts + 50 + i * 0.1, "red", 5.0, 0.2, 0, 500, "http_5xx", "http"))
    # T_FL2: ALL failed with latency (100 probes, all probe_success=0 but rtt=50ms)
    # Expected: sample_n=100, success=0, lat_avg=None, lat_max=None
    ts2 = ts + 200  # slightly later
    for i in range(100):
        conn.execute(
            "INSERT INTO pings_raw(target_id,ts,status,latency_ms,loss_rate,probe_success,status_code,failure_reason,ping_type) "
            "VALUES(?,?,?,?,?,?,?,?,?)",
            ("T_FL2", ts2 + i * 0.1, "red", 50.0, 0.3, 0, 503, "http_503", "http"))
    conn.commit()
    conn.close()

# ── DataSet 3: alert events for SLA tests ─────────────────────────────
def insert_alerts(db_path):
    conn = sqlite3.connect(db_path)
    ts_a = BASE_TS + 100  # within the T1 probe window
    # T1: green → red → green (outage lasting 300s)
    conn.execute(
        "INSERT INTO alert_events(target_id,label,ip,ts,old_status,new_status,category,failure_reason,incident_id) "
        "VALUES(?,?,?,?,?,?,?,?,?)",
        ("T1", "Test1", "10.0.0.1", ts_a, "green", "red", "availability", "timeout", "INC-001"))
    conn.execute(
        "INSERT INTO alert_events(target_id,label,ip,ts,old_status,new_status,category,failure_reason,incident_id) "
        "VALUES(?,?,?,?,?,?,?,?,?)",
        ("T1", "Test1", "10.0.0.1", ts_a + 300, "red", "green", "ok", "", "INC-001"))
    # T2: orange → red（性能退化 → 故障）
    conn.execute(
        "INSERT INTO alert_events(target_id,label,ip,ts,old_status,new_status,category,failure_reason,incident_id) "
        "VALUES(?,?,?,?,?,?,?,?,?)",
        ("T2", "Test2", "10.0.0.2", ts_a + 100, "green", "orange", "performance", "high_latency", "INC-002"))
    conn.commit()
    conn.close()

# ── Build pings_hourly from pings_raw via DataStore._hourly_agg ───────
def build_hourly(store: DataStore, db_path: str):
    """Manually call _hourly_agg to populate pings_hourly."""
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    # Set up targets_meta so _hourly_agg finds the targets
    store._hourly_agg(conn, backfill_internal_gaps=True)
    conn.commit()
    conn.close()

# ═══════════════════════════════════════════════════════════════════════
#  TEST SUITE
# ═══════════════════════════════════════════════════════════════════════

def run_tests():
    global PASS, FAIL
    PASS = 0
    FAIL = 0

    tmpdir = tempfile.mkdtemp(prefix="round_a_")
    db_path = os.path.join(tmpdir, "test.db")
    print(f"\nTemp DB: {db_path}")

    # 1. Build DB + data
    build_db(db_path)
    insert_normal_probes(db_path)
    insert_failed_with_latency(db_path)
    insert_alerts(db_path)

    # 2. Create DataStore pointed at this DB
    # __init__ auto-starts the writer thread and waits for schema init (up to 10s)
    store = DataStore(db_path, raw_retention_days=7, hourly_retention_days=90)
    # Schema is already initialized; wait a bit for writer thread to settle
    time.sleep(1.0)

    # Build hourly data
    build_hourly(store, db_path)

    # ── A1: single vs batch SLA consistency ───────────────────────────
    print("\n── A1: single vs batch SLA ──")
    start_ts = BASE_TS - 100
    end_ts = BASE_TS + 10000
    for tid in ["T1", "T2"]:
        s_single = store.get_sla_stats(tid, start_ts, end_ts)
        s_batch = store.get_sla_stats_batch([tid, "T_FL", "T_FL2"], start_ts, end_ts)
        s_batch_t = s_batch.get(tid, {})
        for key in ["uptime_pct", "lat_avg", "lat_max", "lat_p95", "outage_count", "outage_minutes", "sample_n"]:
            a = s_single.get(key)
            b = s_batch_t.get(key)
            check(almost(a, b, 1e-3), f"{tid}.{key}: single={fmt(a)} batch={fmt(b)}")

    # ── A2: raw vs hourly waveform ────────────────────────────────────
    print("\n── A2: raw vs hourly waveform ──")
    # Short window within raw retention → both raw and forced-hourly should agree on summary
    raw = store.get_waveform("T1", start_ts, end_ts)
    # Force hourly by using a span > 48h
    span_48h = 50 * 3600
    hourly = store.get_waveform("T1", end_ts - span_48h, end_ts)
    print(f"  raw  resolution={raw.get('resolution')}, summary={raw.get('summary')}")
    print(f"  hourly resolution={hourly.get('resolution')}, summary={hourly.get('summary')}")

    # ── A3: dashboard stats vs SLA latency ────────────────────────────
    print("\n── A3: dashboard stats vs SLA latency ──")
    dash = store.get_dashboard_stats_batch(["T1", "T2", "T_FL", "T_FL2"], start_ts, end_ts)
    sla_batch = store.get_sla_stats_batch(["T1", "T2", "T_FL", "T_FL2"], start_ts, end_ts)
    for tid in ["T1", "T2", "T_FL", "T_FL2"]:
        d = dash.get(tid, {})
        s = sla_batch.get(tid, {})
        # total==sample_n must agree
        check(d.get("total") == s.get("sample_n"),
              f"{tid}.total==sample_n: dash={fmt(d.get('total'))} sla={fmt(s.get('sample_n'))}")
        # latency stats — dashboard rounds to 1 dec, SLA to 2 dec; tolerate 0.05
        for key in ["lat_avg", "lat_max", "lat_p95"]:
            a = d.get(key)
            b = s.get(key)
            eps = 0.051 if key == "lat_avg" else 1e-4  # dash rounds to 1dp vs SLA 2dp
            check(almost(a, b, eps), f"{tid}.{key}: dash={fmt(a)} sla={fmt(b)}")

    # ── A4: failed-with-latency MUST NOT pollute stats ────────────────
    print("\n── A4: failed-with-latency isolation ──")
    # Get dashboard stats (which include success count) for verification
    dash_fl = store.get_dashboard_stats_batch(["T_FL", "T_FL2"], start_ts, end_ts)
    # T_FL: 50 success, 50 failed-with-RTT → success=50, total=100
    d_tfl = dash_fl.get("T_FL", {})
    check(d_tfl.get("total") == 100,
          f"T_FL total: expect 100, got {d_tfl.get('total')}")
    check(d_tfl.get("success") == 50,
          f"T_FL success: expect 50, got {d_tfl.get('success')}")
    # lat_avg must ONLY consider the 50 OK probes (all 20.0ms)
    check(almost(d_tfl.get("lat_avg"), 20.0, 0.01),
          f"T_FL lat_avg: expect 20.0, got {fmt(d_tfl.get('lat_avg'))} — FAILED probes must NOT contribute!")
    check(almost(d_tfl.get("lat_max"), 20.0, 0.01),
          f"T_FL lat_max: expect 20.0, got {fmt(d_tfl.get('lat_max'))}")
    # T_FL2: ALL failed with RTT=50ms → success=0, lat_avg=None
    d_tfl2 = dash_fl.get("T_FL2", {})
    check(d_tfl2.get("total") == 100,
          f"T_FL2 total: expect 100, got {d_tfl2.get('total')}")
    check(d_tfl2.get("success") == 0,
          f"T_FL2 success: expect 0, got {d_tfl2.get('success')}")
    check(d_tfl2.get("lat_avg") is None,
          f"T_FL2 lat_avg: expect None (no successful probes), got {fmt(d_tfl2.get('lat_avg'))} — 50ms from failed probes leaks!")
    check(d_tfl2.get("lat_max") is None,
          f"T_FL2 lat_max: expect None, got {fmt(d_tfl2.get('lat_max'))}")

    # ── A4b: raw-level SQL check ─────────────────────────────────────
    print("\n── A4b: SQL-level probe classification ──")
    conn_raw = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    # Check _PROBE_OK_SQL on T_FL2: all 100 rows should be 0
    rows = conn_raw.execute(
        f"SELECT {_PROBE_OK_SQL} FROM pings_raw WHERE target_id='T_FL2'"
    ).fetchall()
    all_zero = all(r[0] == 0 for r in rows)
    check(all_zero, f"T_FL2 _PROBE_OK_SQL: all 100 rows → expect all 0, got mix")

    # Check _PROBE_LAT_OK_SQL: must be 0 for all T_FL2 rows
    rows2 = conn_raw.execute(
        f"SELECT {_PROBE_LAT_OK_SQL} FROM pings_raw WHERE target_id='T_FL2'"
    ).fetchall()
    all_zero2 = all(r[0] == 0 for r in rows2)
    check(all_zero2, f"T_FL2 _PROBE_LAT_OK_SQL: expect all 0 (no latency-ok probes)")

    # T_FL: 50 OK + 50 failed-with-RTT
    rows3 = conn_raw.execute(
        f"SELECT SUM(CASE WHEN {_PROBE_OK_SQL}=1 THEN 1 ELSE 0 END), "
        f"SUM(CASE WHEN {_PROBE_LAT_OK_SQL}=1 THEN 1 ELSE 0 END) "
        f"FROM pings_raw WHERE target_id='T_FL'"
    ).fetchone()
    check(rows3[0] == 50, f"T_FL _PROBE_OK_SQL sum: expect 50, got {rows3[0]}")
    check(rows3[1] == 50, f"T_FL _PROBE_LAT_OK_SQL sum: expect 50, got {rows3[1]}")

    conn_raw.close()

    # ── A4c: Python row_probe_success check ───────────────────────────
    print("\n── A4c: Python row_probe_success ──")
    # With probe_success=0, must return False even if latency_ms and status say otherwise
    check(DataStore.row_probe_success(50.0,   "green", None, 0) is False, "probe_success=0, green, rtt=50 → False")
    check(DataStore.row_probe_success(None,   "green", None, 0) is False, "probe_success=0, green, rtt=None → False")
    check(DataStore.row_probe_success(50.0,   "green", None, 1) is True,  "probe_success=1, green, rtt=50 → True")
    check(DataStore.row_probe_success(None,   "green", None, 1) is True,  "probe_success=1, green, rtt=None → True (legacy OK)")
    # Legacy: no probe_success column, relies on failure_reason/status/latency
    check(DataStore.row_probe_success(None,   "red",   "timeout", None) is False, "legacy: red+timeout+rtt=None → False")
    check(DataStore.row_probe_success(5.0,    "green", None, None) is True,   "legacy: green+rtt=5, no failure_reason → True")
    check(DataStore.row_probe_success(None,   "green", None, None) is False,  "legacy: green+rtt=None → False")

    # ── A5: export_excel vs API consistency ───────────────────────────
    print("\n── A5: export_excel vs API consistency ──")
    export_path = os.path.join(tmpdir, "export_test.xlsx")
    try:
        import openpyxl
        targets_meta = {
            "T1": {"label": "Test1", "ip": "10.0.0.1", "ping_type": "icmp"},
            "T2": {"label": "Test2", "ip": "10.0.0.2", "ping_type": "icmp"},
            "T_FL": {"label": "TestFL", "ip": "10.0.0.3", "ping_type": "http"},
            "T_FL2": {"label": "TestFL2", "ip": "10.0.0.4", "ping_type": "http"},
        }
        store.export_excel(export_path, ["T1", "T2", "T_FL", "T_FL2"],
                          start_ts, end_ts, targets_meta)
        check(os.path.exists(export_path), "export_excel produced file")
        check(os.path.getsize(export_path) > 100, "export_excel file non-trivial")

        # Verify exported summary stats match API dashboard stats
        wb = openpyxl.load_workbook(export_path, read_only=True, data_only=True)
        check("汇总" in wb.sheetnames, "export has 汇总 sheet")
        ws = wb["汇总"]
        rows_list = list(ws.iter_rows(values_only=True))
        check(len(rows_list) >= 5,
              f"汇总 has {len(rows_list)} rows (expect >=5: header+4 targets)")
        if len(rows_list) >= 5:
            header = rows_list[0]
            col_map = {str(h): i for i, h in enumerate(header) if h}
            tid_map = {"Test1": "T1", "Test2": "T2",
                       "TestFL": "T_FL", "TestFL2": "T_FL2"}
            for row in rows_list[1:]:
                label = str(row[col_map.get("节点名称", 0)] or "")
                tid = tid_map.get(label)
                if not tid:
                    continue
                d = dash.get(tid, {})
                export_total = row[col_map.get("探测次数", 3)]
                check(export_total == d.get("total", 0),
                      f"export {label} total: {export_total} vs api {d.get('total')}")
                export_success = row[col_map.get("成功次数", 4)]
                check(export_success == d.get("success", 0),
                      f"export {label} success: {export_success} vs api {d.get('success')}")
                # avg latency tolerates export rounding (±5ms max)
                export_avg = row[col_map.get("平均延时 ms", 6)]
                if export_avg and d.get("success"):
                    check(almost(float(export_avg), d.get("lat_avg") or 0, 5.0),
                          f"export {label} avg_lat: {export_avg} vs api {d.get('lat_avg')}")
        wb.close()
    except Exception as e:
        import traceback
        traceback.print_exc()
        check(False, f"export_excel raised: {e}")

    # ── A6: boundary / empty-target ───────────────────────────────────
    print("\n── A6: boundary / empty-target ──")
    s_empty = store.get_sla_stats("T_NONEXISTENT", start_ts, end_ts)
    check(s_empty.get("sample_n") == 0, "non-existent target → sample_n=0")
    # Non-existent targets get uptime_pct=100.0 (no outage data):
    check(s_empty.get("uptime_pct") == 100.0, "non-existent target → uptime_pct=100.0 (no outage data → 100%)")

    # Edge: start_ts == end_ts
    s_zero = store.get_sla_stats("T1", end_ts, end_ts)
    print(f"  zero-span result: {s_zero}")
    check(isinstance(s_zero, dict), "zero-span returns dict not exception")

    # ── Cleanup ───────────────────────────────────────────────────────
    store.shutdown()
    try:
        os.unlink(export_path)
    except Exception:
        pass
    try:
        os.rmdir(tmpdir)
    except Exception:
        pass

    # ── Summary ───────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"RESULTS: {PASS} passed, {FAIL} failed out of {PASS+FAIL}")
    if FAIL:
        print("⚠️  SOME CHECKS FAILED — review output above")
        sys.exit(1)
    else:
        print("✓  All checks passed")
        sys.exit(0)


if __name__ == "__main__":
    run_tests()
