"""Regression checks for bug 26 (stats API + Excel vs raw retention)."""
import os
import sqlite3
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.data_store import DataStore


def _insert_hourly(conn, tid, hour_ts, lat=42.0, n=60):
    conn.execute(
        "INSERT INTO pings_hourly "
        "(target_id,hour_ts,sample_n,success_n,lat_avg,lat_max,"
        " lat_min,lat_p95,loss_rate) VALUES (?,?,?,?,?,?,?,?,?)",
        (tid, hour_ts, n, n, lat, lat, lat, lat, 0.0))


def _insert_raw(conn, tid, ts, lat=10.0):
    conn.execute(
        "INSERT INTO pings_raw (target_id,ts,status,latency_ms,loss_rate,"
        "status_code,failure_reason,ping_type) VALUES (?,?,?,?,?,?,?,?)",
        (tid, ts, "green", lat, 0.0, None, None, "icmp"))


def _stats_for_target(ds, tid, start, end):
    """Mirror /api/stats aggregation for one target."""
    import math as _math
    use_hourly = ds.should_use_hourly(start, end, 86400)
    rows = ds.get_history(tid, start, end, use_hourly=use_hourly)
    if not rows:
        return None
    if use_hourly:
        total = sum(r["sample_n"] for r in rows)
        success = sum(r["success_n"] for r in rows)
        lat_pairs = [(r["lat_avg"], r["success_n"]) for r in rows
                     if r["lat_avg"] is not None and r["success_n"]]
        w_n = sum(n for _v, n in lat_pairs)
        lat_avg = (sum(v * n for v, n in lat_pairs) / w_n if w_n else None)
    else:
        total = len(rows)
        success = sum(1 for r in rows if r["status"] in ("green", "orange"))
        lats = [r["latency_ms"] for r in rows if r["latency_ms"] is not None]
        lat_avg = sum(lats) / len(lats) if lats else None
    return {"total": total, "success": success, "lat_avg": lat_avg,
            "use_hourly": use_hourly}


def test_old_24h_api_stats():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    try:
        ds = DataStore(path, raw_retention_days=7)
        time.sleep(0.3)
        tid = "t"
        now = int(time.time())
        hour_ts = ((now - 30 * 86400) // 3600) * 3600
        start_ts = hour_ts
        end_ts = hour_ts + 24 * 3600
        conn = sqlite3.connect(path)
        _insert_hourly(conn, tid, hour_ts)
        conn.commit()
        conn.close()

        stats = _stats_for_target(ds, tid, start_ts, end_ts)
        ok = (stats is not None and stats["total"] == 60
              and stats["use_hourly"])
        print(f"Bug26 stats: {stats} -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def test_recent_24h_api_stats_raw():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    try:
        ds = DataStore(path, raw_retention_days=7)
        time.sleep(0.3)
        tid = "t"
        now = int(time.time())
        start_ts = now - 24 * 3600
        end_ts = now
        conn = sqlite3.connect(path)
        _insert_raw(conn, tid, now - 3600)
        conn.commit()
        conn.close()

        stats = _stats_for_target(ds, tid, start_ts, end_ts)
        ok = (stats is not None and stats["total"] == 1
              and not stats["use_hourly"])
        print(f"Bug26 recent stats: {stats} -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def test_spanning_cutoff_stats():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    try:
        ds = DataStore(path, raw_retention_days=7)
        time.sleep(0.3)
        tid = "t"
        now = int(time.time())
        old_hour = ((now - 9 * 86400) // 3600) * 3600
        start_ts = old_hour
        end_ts = now - 3600
        conn = sqlite3.connect(path)
        _insert_hourly(conn, tid, old_hour, lat=42.0)
        _insert_raw(conn, tid, now - 7200)
        conn.commit()
        conn.close()

        stats = _stats_for_target(ds, tid, start_ts, end_ts)
        ok = stats is not None and stats["total"] >= 60 and stats["use_hourly"]
        print(f"Bug26 span stats: total={stats['total'] if stats else 0} "
              f"hourly={stats['use_hourly'] if stats else None} -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def test_old_24h_export():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    xlsx = path + ".xlsx"
    try:
        ds = DataStore(path, raw_retention_days=7)
        time.sleep(0.3)
        tid = "t"
        now = int(time.time())
        hour_ts = ((now - 30 * 86400) // 3600) * 3600
        start_ts = hour_ts
        end_ts = hour_ts + 24 * 3600
        conn = sqlite3.connect(path)
        _insert_hourly(conn, tid, hour_ts)
        conn.execute(
            "INSERT INTO targets_meta (target_id,label,ip,ping_type) "
            "VALUES (?,?,?,?)",
            (tid, "node", "1.2.3.4", "icmp"))
        conn.commit()
        conn.close()

        meta = {tid: {"label": "node", "ip": "1.2.3.4", "ping_type": "icmp"}}
        ds.export_excel(xlsx, [tid], start_ts, end_ts, meta)
        import openpyxl
        wb = openpyxl.load_workbook(xlsx)
        row = [c.value for c in wb["汇总"][2]]
        total = row[3]
        ok = total == 60
        print(f"Bug26 export old: summary_row={row} -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
        try:
            os.unlink(xlsx)
        except OSError:
            pass


def test_recent_export_raw_rows():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    xlsx = path + ".xlsx"
    try:
        ds = DataStore(path, raw_retention_days=7)
        time.sleep(0.3)
        tid = "t"
        now = int(time.time())
        start_ts = now - 3600
        end_ts = now
        conn = sqlite3.connect(path)
        _insert_raw(conn, tid, now - 600)
        conn.execute(
            "INSERT INTO targets_meta (target_id,label,ip,ping_type) "
            "VALUES (?,?,?,?)",
            (tid, "node", "1.2.3.4", "icmp"))
        conn.commit()
        conn.close()

        meta = {tid: {"label": "node", "ip": "1.2.3.4", "ping_type": "icmp"}}
        ds.export_excel(xlsx, [tid], start_ts, end_ts, meta)
        import openpyxl
        wb = openpyxl.load_workbook(xlsx)
        node_rows = wb[wb.sheetnames[1]].max_row
        ok = node_rows >= 2
        print(f"Bug26 export recent: node_data_rows={node_rows - 1} -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
        try:
            os.unlink(xlsx)
        except OSError:
            pass


def main():
    results = [
        test_old_24h_api_stats(),
        test_recent_24h_api_stats_raw(),
        test_spanning_cutoff_stats(),
        test_old_24h_export(),
        test_recent_export_raw_rows(),
    ]
    failed = [i for i, ok in enumerate(results) if not ok]
    if failed:
        print("FAILED:", failed)
        sys.exit(1)
    print("All bug 26 checks passed.")


if __name__ == "__main__":
    main()
