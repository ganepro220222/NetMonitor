"""Regression checks for bug 70 (Excel alert timeline same-ts ordering)."""
import os
import sqlite3
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.data_store import DataStore

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_STORE = os.path.join(ROOT, "src", "data_store.py")


def _insert_alerts(conn, tid, ts, pairs):
    for i, (old_s, new_s, iid) in enumerate(pairs, start=1):
        conn.execute(
            "INSERT INTO alert_events "
            "(target_id, label, ip, ts, old_status, new_status, "
            " category, incident_id) "
            "VALUES (?, 'n', '1.2.3.4', ?, ?, ?, 'availability', ?)",
            (tid, ts, old_s, new_s, iid),
        )
    conn.commit()


def _incident_headers(ws):
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell = row[0] if row else None
        if isinstance(cell, str) and "故障事件" in cell:
            out.append(cell)
    return out


def _status_changes(ws):
    """Collect 'old → new' labels from the status-change column (F)."""
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) > 5 and row[5]:
            chg = str(row[5])
            if "→" in chg or "->" in chg:
                out.append(chg)
    return out


def test_excel_same_ts_two_closed_incidents():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    xlsx = path + ".xlsx"
    try:
        ds = DataStore(path)
        ds._schema_ready.wait(timeout=5)
        now = int(time.time())
        ts = float(now)
        start_ts = now - 3600
        end_ts = now + 3600
        conn = sqlite3.connect(path)
        _insert_alerts(conn, "T", ts, [
            ("green", "red",   "INC-A"),
            ("red",   "green", "INC-A"),
            ("green", "red",   "INC-B"),
            ("red",   "green", "INC-B"),
        ])
        conn.execute(
            "INSERT INTO targets_meta (target_id,label,ip,ping_type) "
            "VALUES (?,?,?,?)",
            ("T", "node", "1.2.3.4", "icmp"),
        )
        conn.commit()
        conn.close()

        meta = {"T": {"label": "node", "ip": "1.2.3.4", "ping_type": "icmp"}}
        ds.export_excel(xlsx, ["T"], start_ts, end_ts, meta)
        ds.shutdown()

        import openpyxl
        wb = openpyxl.load_workbook(xlsx)
        ws = wb["告警时间线"]
        headers = _incident_headers(ws)
        changes = _status_changes(ws)

        bad_trunc = any("开始于查询范围外" in h for h in headers)
        bad_open = any("查询结束时仍未恢复" in h for h in headers)
        closed = [h for h in headers if "故障事件 #" in h
                  and "开始于查询范围外" not in h
                  and "查询结束时仍未恢复" not in h]
        ok = (
            not bad_trunc
            and not bad_open
            and len(closed) == 2
            and changes.count("正常 → 严重") >= 2
            and changes.count("严重 → 正常") >= 2
        )
        print(f"Bug70 closed={len(closed)} trunc={bad_trunc} open={bad_open} "
              f"red={changes.count('正常 → 严重')} "
              f"green={changes.count('严重 → 正常')} -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
        try:
            os.unlink(xlsx)
        except OSError:
            pass


def test_get_alert_history_returns_id_ordered():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    try:
        ds = DataStore(path)
        ds._schema_ready.wait(timeout=5)
        ts = 1780506232.0
        conn = sqlite3.connect(path)
        _insert_alerts(conn, "T", ts, [
            ("green", "red",   "INC-A"),
            ("red",   "green", "INC-A"),
            ("green", "red",   "INC-B"),
            ("red",   "green", "INC-B"),
        ])
        conn.close()
        hist = ds.get_alert_history("T", start_ts=0, limit=-1)
        ds.shutdown()
        ids = [e["id"] for e in hist]
        ok = ids == sorted(ids, reverse=True)
        print(f"Bug70 get_alert_history ids={ids} (expect 4..1) -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


def test_source_id_sort():
    with open(DATA_STORE, encoding="utf-8") as f:
        text = f.read()
    hist = text.split("def get_alert_history", 1)[1].split(
        "def export_excel", 1)[0]
    build = text.split("def _build_incidents(events):", 1)[1].split(
        "# ── Render the sheet", 1)[0]
    ok = (
        "SELECT id, target_id" in hist
        and "ORDER BY ts DESC, id DESC" in hist
        and '"id": r[0]' in hist
        and "key=lambda e: (e['ts'], e.get('id', 0))" in build
    )
    print(f"Bug70 source id + (ts,id) sort -> {ok}")
    return ok


def main():
    results = [
        ("excel", test_excel_same_ts_two_closed_incidents()),
        ("history", test_get_alert_history_returns_id_ordered()),
        ("source", test_source_id_sort()),
    ]
    failed = [name for name, ok in results if not ok]
    if failed:
        print("FAILED:", failed)
        sys.exit(1)
    print("All bug 70 checks passed.")


if __name__ == "__main__":
    main()
