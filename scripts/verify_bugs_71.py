"""Regression checks for bug 71 (Excel export timestamps show milliseconds)."""
import os
import re
import sqlite3
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.data_store import DataStore, fmt_ts_ms

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_STORE = os.path.join(ROOT, "src", "data_store.py")


def _insert_raw(conn, tid, ts, lat=12.0):
    conn.execute(
        "INSERT INTO pings_raw "
        "(target_id, ts, status, latency_ms, loss_rate, ping_type) "
        "VALUES (?, ?, 'green', ?, 0, 'icmp')",
        (tid, ts, lat),
    )


def _insert_alert(conn, tid, ts, old_s, new_s):
    conn.execute(
        "INSERT INTO alert_events "
        "(target_id, label, ip, ts, old_status, new_status, category) "
        "VALUES (?, 'n', '1.2.3.4', ?, ?, ?, 'availability')",
        (tid, ts, old_s, new_s),
    )


def test_raw_sheet_ms_timestamps():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    xlsx = path + ".xlsx"
    try:
        base = float(int(time.time()))
        ds = DataStore(path)
        ds._schema_ready.wait(timeout=5)
        conn = sqlite3.connect(path)
        _insert_raw(conn, "T", base + 0.10, 10.0)
        _insert_raw(conn, "T", base + 0.70, 11.0)
        conn.execute(
            "INSERT INTO targets_meta (target_id,label,ip,ping_type) "
            "VALUES (?,?,?,?)",
            ("T", "node", "1.2.3.4", "icmp"),
        )
        conn.commit()
        conn.close()
        start_ts = int(base) - 60
        end_ts = int(base) + 60
        meta = {"T": {"label": "node", "ip": "1.2.3.4", "ping_type": "icmp"}}
        ds.export_excel(xlsx, ["T"], start_ts, end_ts, meta)
        ds.shutdown()

        import openpyxl
        wb = openpyxl.load_workbook(xlsx)
        ws = wb[wb.sheetnames[1]]
        times = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        ok = (
            len(times) == 2
            and times[0] != times[1]
            and re.search(r"\.\d{3}$", str(times[0]))
            and re.search(r"\.\d{3}$", str(times[1]))
            and fmt_ts_ms(base + 0.10) in str(times[0])
            and fmt_ts_ms(base + 0.70) in str(times[1])
        )
        print(f"Bug71 raw times={times} -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
        try:
            os.unlink(xlsx)
        except OSError:
            pass


def test_alert_timeline_ms_timestamps():
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    xlsx = path + ".xlsx"
    try:
        base = float(int(time.time()))
        ts = base + 10.0
        ds = DataStore(path)
        ds._schema_ready.wait(timeout=5)
        conn = sqlite3.connect(path)
        _insert_alert(conn, "T", ts + 0.10, "green", "red")
        _insert_alert(conn, "T", ts + 0.70, "red", "green")
        conn.execute(
            "INSERT INTO targets_meta (target_id,label,ip,ping_type) "
            "VALUES (?,?,?,?)",
            ("T", "node", "1.2.3.4", "icmp"),
        )
        conn.commit()
        conn.close()
        start_ts = int(base)
        end_ts = int(base) + 3600
        meta = {"T": {"label": "node", "ip": "1.2.3.4", "ping_type": "icmp"}}
        ds.export_excel(xlsx, ["T"], start_ts, end_ts, meta)
        ds.shutdown()

        import openpyxl
        wb = openpyxl.load_workbook(xlsx)
        ws = wb["告警时间线"]
        event_times = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Detail rows: col E (idx 4) = 事件时间, col F (idx 5) = 状态变化.
            # A 探测类型 column (col D) was later inserted ahead of these,
            # shifting the time/transition one column right of the old layout.
            if row and row[4] and row[5] and "→" in str(row[5]):
                event_times.append(str(row[4]))
        ok = (
            len(event_times) == 2
            and event_times[0] != event_times[1]
            and all(re.search(r"\.\d{3}$", t) for t in event_times)
        )
        print(f"Bug71 alert times={event_times} -> {ok}")
        return ok
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
        try:
            os.unlink(xlsx)
        except OSError:
            pass


def test_fmt_ts_ms_unit():
    ts = 1780506227.123
    s = fmt_ts_ms(ts)
    ok = s.endswith(".123") and "1780506227" not in s
    print(f"Bug71 fmt_ts_ms({ts})={s!r} -> {ok}")
    return ok


def test_source_uses_fmt_ts_ms():
    with open(DATA_STORE, encoding="utf-8") as f:
        text = f.read()
    export = text.split("def export_excel", 1)[1].split(
        "def get_alert_history", 1)[0]
    ok = (
        "def fmt_ts_ms(ts)" in text
        and "dt   = fmt_ts_ms(r[\"ts\"])" in export
        and "return fmt_ts_ms(ts)" in export
        and 'strftime("%Y-%m-%d %H:%M:%S")' not in export.split(
            "告警时间线", 1)[0]
    )
    print(f"Bug71 source fmt_ts_ms in export -> {ok}")
    return ok


def main():
    results = [
        ("raw_sheet", test_raw_sheet_ms_timestamps()),
        ("alert_sheet", test_alert_timeline_ms_timestamps()),
        ("fmt_unit", test_fmt_ts_ms_unit()),
        ("source", test_source_uses_fmt_ts_ms()),
    ]
    failed = [name for name, ok in results if not ok]
    if failed:
        print("FAILED:", failed)
        sys.exit(1)
    print("All bug 71 checks passed.")


if __name__ == "__main__":
    main()
