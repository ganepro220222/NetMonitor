from __future__ import annotations
"""
web_server.py
=============
Key fixes in this version:

[Traceroute - Critical]
  Previous versions used CREATE_NO_WINDOW + shell=True.
  On Windows this combination breaks pipe inheritance: cmd.exe is created
  without a console, stdout/stderr pipes are not properly inherited, and
  result.stdout comes back as b"". The exception handler then returns the
  single "break" hop that was confusing users.

  Fix: use STARTUPINFO.dwFlags |= STARTF_USESHOWWINDOW + SW_HIDE.
  This hides the console window while keeping pipe inheritance intact.
  Also: full path to tracert.exe, -w 1000 (shorter probe timeout),
  120-second subprocess timeout, improved hop line parser.

[Statistics]
  accumStat() previously counted success only when latency_ms != null.
  For HTTP monitors, a non-matching status code returns latency_ms=null
  even though the server IS responding — causing uptime% to show 0%.
  Fix: count success based on status field (green/orange = up).

[Bar chart]
  HTTP latency (~1000ms) vs ICMP (~5ms) on the same linear axis makes
  ICMP bars invisible. Added a log/linear toggle button.

[Pause/resume]
  pause_target() sets web card to "paused" state (gray, paused icon).
  resume_target() clears the paused state; next SSE update shows real status.
"""

import json
import logging
import os
import platform
import queue
import re
import socket
import subprocess
import threading
import time
from datetime import datetime
from typing import Optional

try:
    from flask import Flask, Response, request
    FLASK_AVAILABLE = True
except ImportError:
    FLASK_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────────
# Traceroute
# ─────────────────────────────────────────────────────────────────────
# Implementation moved to src.traceroute_util so the C-side DNS-diag
# "追踪此 IP" feature can share the exact same parser.  Re-exported here
# under the original name to keep web_server.py call sites unchanged.
from src.traceroute_util import run_traceroute   # noqa: F401


def dns_check(hostname: str) -> dict:
    try:
        results = socket.getaddrinfo(hostname, None)
        ips = list(dict.fromkeys(r[4][0] for r in results))
        return {"success": True, "ips": ips}
    except socket.gaierror as e:
        return {"success": False, "error": str(e), "ips": []}


def tcp_check(ip: str, port: int, timeout: float = 2.5) -> dict:
    """TCP port liveness check used by the traceroute scheduler to
    fill the 'ports' badge alongside each tracert run.

    Uses socket.create_connection() (dual-stack) rather than a raw
    AF_INET socket.  AF_INET refuses colon-style IPv6 addresses with
    'Address family for hostname not supported', which would make every
    IPv6 node's port badge perpetually red regardless of actual state.
    create_connection() raises on failure (collapses the old
    'returns non-zero' + 'throws' paths into one), so success only
    needs the try-branch.
    """
    start = time.time()
    try:
        s = socket.create_connection((ip, port), timeout=timeout)
        elapsed = (time.time() - start) * 1000
        s.close()
        return {"port": port, "success": True,
                "latency_ms": round(elapsed, 1)}
    except Exception as e:
        return {"port": port, "success": False,
                "latency_ms": None, "error": str(e)}


# ─────────────────────────────────────────────────────────────────────
# Background traceroute scheduler
# ─────────────────────────────────────────────────────────────────────

class _TracerouteScheduler:
    DEFAULT_INTERVAL = 300

    def __init__(self, cache: dict, cache_lock: threading.Lock,
                 interval: int = 300, push_fn=None):
        self._cache     = cache
        self._lock      = cache_lock
        self.interval   = interval
        self.max_hops   = 30   # configurable via WebServer.set_traceroute_max_hops()
        self._push_fn   = push_fn
        self._data_store = None   # injected via set_data_store(); persists scheduled runs
        self._prev_route_changed: dict[str, bool] = {}  # Rule B: require 2 consecutive changes
        self._targets: dict[str, str] = {}
        self._targets_lock = threading.Lock()
        self._pending: set = set()
        self._pending_lock = threading.Lock()
        # Per-tid monotonic generation counter.  Bumped whenever the
        # target's identity or history baseline changes (IP edit,
        # remove, or explicit history wipe).  _run_one captures the
        # generation at start and re-checks before writing -- a
        # mismatch means the in-flight tracert belongs to a
        # superseded version of the target and its result must NOT
        # touch the cache / DataStore (otherwise wipe_target_history
        # finishes "cleanly", then an in-flight tracert that was
        # blocked on the OLD ip returns and resurrects the very rows
        # the user just deleted).
        self._gen: dict[str, int] = {}
        self._gen_lock = threading.Lock()
        threading.Thread(target=self._loop, daemon=True, name="tracer-sched").start()

    def bump_generation(self, tid: str) -> None:
        """Invalidate any in-flight traceroute for tid.

        Call this whenever the target's identity changes (IP edit) or
        its history is wiped: the next _run_one() ending will detect
        the version mismatch and discard its result.  Idempotent;
        cheap (one lock acquisition).

        Also drops _prev_route_changed[tid] so the 2-consecutive-runs
        route-change debounce can't carry state across an identity
        change.  Without this, an old target that had armed the
        debounce (one prior route change observed) would, after the
        user repurposed the tid to a NEW target, mark the new
        target's FIRST normal route change as a persistent
        route_changed=True -- corrupting the route-change history
        the moment the new node appears.
        """
        with self._gen_lock:
            self._gen[tid] = self._gen.get(tid, 0) + 1
        # _prev_route_changed lives on the scheduler thread's data
        # path; only the scheduler thread writes it in _run_one, so a
        # dict.pop from another thread races at worst with a dict.get
        # (atomic in CPython) -- the worst-case outcome is the next
        # _run_one observes False either way, which is the desired
        # post-bump state.
        self._prev_route_changed.pop(tid, None)

    def _current_generation(self, tid: str) -> int:
        with self._gen_lock:
            return self._gen.get(tid, 0)

    def set_targets(self, targets: dict):
        """
        targets can be {tid: "ip_str"} (legacy) or
        {tid: {"ip": "...", "probe_ports": [80, 443, ...]}} (preferred).
        """
        with self._targets_lock:
            normalized = {}
            for tid, v in targets.items():
                if isinstance(v, str):
                    normalized[tid] = {"ip": v, "probe_ports": [80, 443]}
                else:
                    normalized[tid] = {"ip": v.get("ip", ""),
                                       "label": v.get("label", tid),
                                       "probe_ports": v.get("probe_ports", [80, 443])}
            self._targets = normalized

    def request_now(self, tid: str):
        with self._pending_lock:
            self._pending.add(tid)

    def _loop(self):
        last_full = time.time()   # defer first full scan
        while True:
            time.sleep(5)

            with self._pending_lock:
                pending = list(self._pending)
                self._pending.clear()
            with self._targets_lock:
                snap = dict(self._targets)

            for tid in pending:
                info = snap.get(tid)
                if info:
                    self._run_one(tid, info)
                    time.sleep(1)

            now = time.time()
            if snap and now - last_full >= self.interval:
                for tid, info in snap.items():
                    self._run_one(tid, info)
                    time.sleep(2)
                    # Allow red-alert tracert to preempt mid full-scan.
                    with self._pending_lock:
                        urgent = list(self._pending)
                        self._pending.clear()
                    for p_tid in urgent:
                        # Look up against the LIVE _targets (not the
                        # full-scan snapshot taken minutes ago): a node
                        # added during this scan that immediately fires a
                        # red alert is not in `snap`, so snap.get() would
                        # return None and silently drop its first urgent
                        # tracert -- the very tracert operators need most.
                        with self._targets_lock:
                            p_info = self._targets.get(p_tid)
                        if p_info:
                            self._run_one(p_tid, p_info)
                            time.sleep(1)
                last_full = time.time()

    def _run_one(self, tid: str, target_info):
        """
        Run traceroute + TCP checks for one target and store in cache.
        target_info: {"ip": "...", "probe_ports": [...]} or plain ip string.
        TCP checks are included so the "ICMP blocked but service up" detection
        works correctly even when the background scheduler updates the cache.
        """
        if isinstance(target_info, str):
            ip, probe_ports = target_info, [80, 443]
        else:
            ip           = target_info.get("ip", "")
            probe_ports  = target_info.get("probe_ports", [80, 443])

        # Capture identity at start so we can detect supersession on
        # return.  bump_generation() called by an IP edit / target
        # removal / history wipe while we're inside run_traceroute()
        # will leave start_gen != current_gen and we'll discard.
        #
        # start_gen is the SCHEDULER's per-tid generation, used for the
        # cache-validity check below.  start_ds_gen is the DATASTORE's
        # per-tid generation, captured separately and forwarded to
        # record_traceroute as captured_generation.  The two counters
        # are NOT synchronised: scheduler bumps on every IP edit (to
        # invalidate the in-flight tracert's cache write); DataStore
        # bumps only on wipe_target_history / on_target_removed (to
        # invalidate the in-flight tracert's DB INSERT).  Passing the
        # scheduler value to DataStore's captured_generation -- as the
        # old code did -- caused EVERY post-IP-edit tracert to be
        # silently dropped by the writer-side guard, because the
        # scheduler had incremented past 0 while DataStore stayed at
        # 0 so the guard's "captured != current" check was
        # permanently true.
        start_gen    = self._current_generation(tid)
        start_ds_gen = (self._data_store.current_target_generation(tid)
                         if self._data_store is not None else None)
        start_ip     = ip

        is_domain = not re.match(r"^\d{1,3}(\.\d{1,3}){3}$", ip)
        dns       = dns_check(ip) if is_domain else None
        resolve   = (dns["ips"][0] if dns and dns.get("success") else ip)
        hops      = run_traceroute(resolve, max_hops=self.max_hops)
        # Always run TCP checks so ICMP-blocked detection remains valid
        tcp_checks = [tcp_check(resolve, p) for p in probe_ports[:6]]
        entry     = {
            "ip": ip, "resolve_ip": resolve, "hops": hops,
            "dns": dns, "tcp_checks": tcp_checks,
            "updated_at": datetime.now().strftime("%H:%M:%S"),
            "updated_ts": time.time(),
        }
        # Validity gate: if the target's identity (IP) changed or the
        # user wiped its history while run_traceroute() was blocked,
        # this result no longer represents the current node.  Drop it
        # entirely -- no cache write, no DataStore record, no SSE
        # broadcast.  Without this gate, wipe_target_history was a
        # promise the scheduler couldn't keep: the user clicks
        # "更换监控对象（清空历史）", the wipe completes, then a
        # tracert that was started against the OLD ip returns and
        # re-inserts the very row the user just deleted.
        #
        # The gen+ip re-check AND the cache write live INSIDE
        # self._lock.  The previous shape (check then drop the lock
        # then re-acquire to write) had a TOCTOU window: an IP edit
        # that ran update_target between the gen check and the cache
        # write would see EMPTY -> pop nothing -> bump generation,
        # while our stale task acquired the lock right after the pop
        # and wrote its old-IP entry; the freshly-bumped generation
        # was observed too late.  update_target /
        # on_target_removed / wipe_target_history all take this same
        # _tracer_lock before clearing the cache, so checking inside
        # it serialises us against them: either we observe the
        # bumped generation and bail, or our write is the one the
        # subsequent cache pop will clean up.
        committed = False
        route_changed = False
        with self._lock:
            if self._current_generation(tid) != start_gen:
                return
            with self._targets_lock:
                cur_target = self._targets.get(tid)
            if cur_target is None or cur_target.get("ip") != start_ip:
                return
            old_entry = self._cache.get(tid)
            self._cache[tid] = entry
            committed = True

            # ── Route-change detection (record-only, no alert) ──
            # Compare new hop-IP sequence against previous cached
            # result.  Two noise-reduction rules to avoid false
            # positives:
            #   A) Ignore single-hop ICMP jitter (* ↔ IP on one
            #      intermediate hop)
            #   B) Require change to persist across two consecutive
            #      runs
            # Kept under the same lock as the cache write so
            # _prev_route_changed can't be cleared by a concurrent
            # bump_generation between our read and our update.
            if old_entry:
                def _stable_ips(hops_list):
                    """Extract IP sequence, treating * as None."""
                    return [h.get("ip") for h in (hops_list or [])]

                new_ips = _stable_ips(hops)
                old_ips = _stable_ips(old_entry.get("hops", []))

                if new_ips and old_ips and new_ips != old_ips:
                    # Rule A: ignore if only one hop differs and it's
                    # a * ↔ IP swap
                    diffs = [(i, o, n) for i, (o, n) in enumerate(zip(old_ips, new_ips))
                             if o != n]
                    # Also catch length differences (new hops added/removed)
                    len_diff = abs(len(new_ips) - len(old_ips))
                    trivial = (len(diffs) <= 1 and len_diff == 0 and
                               all(o is None or n is None for _, o, n in diffs))
                    if not trivial:
                        # Rule B: only flag if the previous run also
                        # showed a change from the one before it
                        # (stored in _prev_route_changed)
                        prev_changed = self._prev_route_changed.get(tid, False)
                        if prev_changed:
                            route_changed = True
                        self._prev_route_changed[tid] = True
                    else:
                        self._prev_route_changed[tid] = False
                else:
                    self._prev_route_changed[tid] = False

            if route_changed:
                entry["route_changed"] = True

        # Persist to traceroute_logs so the Web "历史记录" tab has data.
        # Without this, history only contains red-alert-triggered entries
        # from alert_manager; nodes that never go red never get any history.
        # captured_generation closes the residual race between our cache
        # commit and the actual DB INSERT: if a wipe lands in the queue
        # between our queue.put and the writer thread's execute, the
        # writer drops the row.  See data_store.record_traceroute.
        # IMPORTANT: forward the DATASTORE-side generation snapshot
        # (start_ds_gen), NOT the scheduler-side one; see start_gen
        # comment above.
        if committed and self._data_store:
            try:
                label = target_info.get("label", tid) if isinstance(target_info, dict) else tid
                self._data_store.record_traceroute(
                    tid, label, ip, entry["updated_ts"], hops,
                    route_changed=route_changed,
                    captured_generation=start_ds_gen)
            except Exception:
                pass

        # Broadcast port-status so browsers update their indicators
        # without waiting for the user to open the diagnosis modal.
        # This fires only when the background scheduler finishes a cycle,
        # not on every ping — so it won't flood the SSE stream.
        #
        # Identity re-check before the push: between the lock release
        # above and this point, another thread can legitimately edit
        # the target's IP (or repurpose the tid).  Without this guard
        # the browser would receive a port_status snapshot taken
        # against the OLD ip and bind it to a tid that now points at
        # the NEW node.  Bumped generation OR a different scheduler
        # ip both indicate "this result no longer belongs to the
        # current identity"; in either case drop the push silently.
        # The payload also gains an "ip" field so the browser can
        # cross-check on its side (defence in depth -- the server-side
        # check is the primary guard).
        if committed and self._push_fn and tcp_checks:
            still_current = (self._current_generation(tid) == start_gen)
            if still_current:
                with self._targets_lock:
                    cur_target = self._targets.get(tid)
                if (cur_target is None
                    or cur_target.get("ip") != start_ip):
                    still_current = False
            if still_current:
                try:
                    self._push_fn(json.dumps({
                        "type":       "port_status",
                        "tid":        tid,
                        "ip":         start_ip,
                        "ports":      tcp_checks,
                        "updated_at": entry["updated_at"],
                    }))
                except Exception:
                    pass   # never crash the scheduler thread over a push failure


# ─────────────────────────────────────────────────────────────────────
# Dashboard HTML
# ─────────────────────────────────────────────────────────────────────



# ─────────────────────────────────────────────────────────────────────
# Quick-tracert standalone window HTML
# ─────────────────────────────────────────────────────────────────────
QUICK_TRACERT_HTML = r"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<title>⚡ 快速追踪 — {label}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
  background:#0f1117;color:#e2e8f0;font-size:13px;padding:16px}}
h1{{font-size:15px;font-weight:600;margin-bottom:4px;color:#f1f5f9}}
.meta{{font-size:11px;color:#64748b;margin-bottom:14px}}
#status{{font-size:11px;padding:4px 10px;border-radius:20px;display:inline-block;
  margin-bottom:14px;background:#1e293b;color:#94a3b8}}
#status.running{{background:rgba(59,130,246,.15);color:#60a5fa}}
#status.done{{background:rgba(34,197,94,.12);color:#4ade80}}
#status.error{{background:rgba(239,68,68,.15);color:#f87171}}
table{{width:100%;border-collapse:collapse}}
th{{font-size:10px;font-weight:600;color:#475569;text-align:left;
   padding:4px 8px;border-bottom:1px solid #1e293b;letter-spacing:.5px}}
td{{padding:6px 8px;border-bottom:1px solid #0f1117;vertical-align:middle}}
tr.hop-ok td:first-child{{color:#4ade80}}
tr.hop-timeout td:first-child{{color:#475569}}
tr.hop-target{{background:rgba(59,130,246,.07)}}
tr.hop-target td:first-child{{color:#60a5fa;font-weight:700}}
.rtt{{font-family:monospace;font-size:12px;color:#a3e635}}
.rtt.slow{{color:#fb923c}}
.rtt.nil{{color:#334155}}
.ip-cell{{font-family:monospace;font-size:12px;color:#e2e8f0}}
.ip-cell.nil{{color:#334155}}
.bar{{display:inline-block;height:8px;border-radius:4px;background:#3b82f6;
  min-width:2px;vertical-align:middle;margin-left:8px;opacity:.7}}
.hop-num{{width:32px;text-align:center;font-weight:700;font-size:14px}}
#ctrl{{margin-top:16px;display:flex;gap:8px;align-items:center}}
button{{padding:5px 16px;border:none;border-radius:6px;cursor:pointer;
  font-size:12px;font-weight:600}}
.btn-run{{background:#2563eb;color:#fff}}
.btn-run:hover{{background:#1d4ed8}}
.btn-close{{background:#1e293b;color:#94a3b8}}
.btn-close:hover{{background:#334155}}
.note{{font-size:10px;color:#475569;margin-top:2px}}
</style>
</head>
<body>
<h1>⚡ 快速路由追踪</h1>
<div class="meta">目标：<b>{label}</b> &nbsp;|&nbsp; IP：<b>{ip}</b></div>
<div id="status">等待开始…</div>
<table>
  <thead><tr>
    <th style="width:32px">跳</th>
    <th>IP 地址</th>
    <th style="width:80px">延时</th>
    <th style="width:160px">延时图</th>
  </tr></thead>
  <tbody id="tbody"></tbody>
</table>
<div id="ctrl">
  <button class="btn-run" id="btn" onclick="startTrace()">开始追踪</button>
  <button class="btn-close" onclick="window.close()">关闭窗口</button>
  <span class="note">使用 -d 参数跳过 DNS 反查，速度更快</span>
</div>
<script>
// QS is "tid=..." or "ip=..." — formed server-side from validated input,
// so it's safe to drop straight into the URL.  No client-side encoding
// needed because tid is uuid-like and ip is already validated.
const QS    = "{qs}";
const LABEL = "{label}";
let es = null, maxMs = 1;

function startTrace(){{
  // Reset
  document.getElementById('tbody').innerHTML = '';
  document.getElementById('status').textContent = '追踪中…';
  document.getElementById('status').className = 'running';
  document.getElementById('btn').disabled = true;
  maxMs = 1;
  if(es){{es.close();es=null;}}

  es = new EventSource('/api/traceroute/quick?' + QS);

  es.onmessage = function(e){{
    const msg = JSON.parse(e.data);
    if(msg.type === 'hop'){{
      addHop(msg);
    }} else if(msg.type === 'done'){{
      es.close(); es = null;
      document.getElementById('status').textContent = '追踪完成 ✓';
      document.getElementById('status').className = 'done';
      document.getElementById('btn').disabled = false;
      document.getElementById('btn').textContent = '重新追踪';
    }} else if(msg.type === 'error'){{
      es.close(); es = null;
      document.getElementById('status').textContent = '错误: ' + msg.message;
      document.getElementById('status').className = 'error';
      document.getElementById('btn').disabled = false;
    }}
  }};

  es.onerror = function(){{
    if(es){{es.close();es=null;}}
    document.getElementById('status').textContent = '连接错误，请重试';
    document.getElementById('status').className = 'error';
    document.getElementById('btn').disabled = false;
  }};
}}

function addHop(h){{
  if(h.ms && h.ms > maxMs) maxMs = h.ms;
  const isTarget = h.is_target;
  const cls = isTarget ? 'hop-target' : (h.ip ? 'hop-ok' : 'hop-timeout');
  const ipHtml = h.ip
    ? `<span class="ip-cell">${{h.ip}}</span>`
    : `<span class="ip-cell nil">* * *</span>`;
  const rttCls = !h.ms ? 'nil' : (h.ms > 100 ? 'slow' : '');
  const rttHtml = h.ms
    ? `<span class="rtt ${{rttCls}}">${{h.ms}} ms</span>`
    : `<span class="rtt nil">—</span>`;
  const barW = h.ms ? Math.max(2, Math.round(h.ms / maxMs * 120)) : 0;
  const barHtml = h.ms
    ? `<span class="bar" style="width:${{barW}}px"></span>` : '';
  const tr = document.createElement('tr');
  tr.className = cls;
  tr.innerHTML =
    `<td class="hop-num">${{h.hop}}</td>` +
    `<td>${{ipHtml}}</td>` +
    `<td>${{rttHtml}}</td>` +
    `<td>${{barHtml}}</td>`;
  document.getElementById('tbody').appendChild(tr);
  // Re-scale bars when maxMs grew
  document.querySelectorAll('.bar').forEach(b=>{{
    const ms = parseFloat(b.parentElement.previousElementSibling.textContent);
    if(!isNaN(ms)) b.style.width = Math.max(2,Math.round(ms/maxMs*120))+'px';
  }});
}}

// Auto-start
startTrace();
</script>
</body>
</html>"""

DASHBOARD_HTML = r"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>网络连通性监测</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"
        onerror="console.warn('[netmonitor] Chart.js CDN load failed -- charts disabled, status cards unaffected'); window.__chartUnavailable=true;"></script>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#111318;--surface:#1a1d24;--surface2:#20242d;--card:#1e2229;
  --border:#2a2e3a;--border2:#323644;--text:#dde1ec;--dim:#6b7280;--dim2:#4b5363;
  --green:#22c55e;--orange:#f59e0b;--red:#ef4444;--gray:#9098a8;
  --blue:#3b82f6;--blue2:#60a5fa;--radius:12px;--shadow:0 2px 16px rgba(0,0,0,.4);}
.theme-dark  input[type="date"],.theme-dark  input[type="datetime-local"]{color-scheme:dark}
.theme-light input[type="date"],.theme-light input[type="datetime-local"]{color-scheme:light}
.theme-dark  input[type="date"]::-webkit-calendar-picker-indicator,
.theme-dark  input[type="datetime-local"]::-webkit-calendar-picker-indicator{filter:invert(1) brightness(1.2);cursor:pointer}
@media(prefers-color-scheme:dark){
  .theme-system input[type="date"],.theme-system input[type="datetime-local"]{color-scheme:dark}
  .theme-system input[type="date"]::-webkit-calendar-picker-indicator,
  .theme-system input[type="datetime-local"]::-webkit-calendar-picker-indicator{filter:invert(1) brightness(1.2);cursor:pointer}
}
.theme-light{--bg:#f0f2f5;--surface:#fff;--surface2:#f8f9fc;--card:#fff;
  --border:#e2e5ec;--border2:#d0d5e0;--text:#1a202c;--dim:#6b7280;
  --shadow:0 2px 12px rgba(0,0,0,.08);}
@media(prefers-color-scheme:light){.theme-system{
  --bg:#f0f2f5;--surface:#fff;--surface2:#f8f9fc;--card:#fff;
  --border:#e2e5ec;--border2:#d0d5e0;--text:#1a202c;--dim:#6b7280;
  --shadow:0 2px 12px rgba(0,0,0,.08);}}
body{background:var(--bg);color:var(--text);
  font-family:'Segoe UI','Microsoft YaHei',system-ui,sans-serif;
  min-height:100vh;display:flex;flex-direction:column;}
#alert-banner{background:linear-gradient(90deg,#7f1d1d,#991b1b);
  border-bottom:2px solid #dc2626;padding:10px 20px;
  display:flex;align-items:center;gap:10px;font-size:14px;font-weight:700;}
#alert-banner.hidden{display:none!important}
.dismiss{margin-left:auto;cursor:pointer;font-size:20px;background:none;
  border:none;color:#fff;opacity:.7;padding:0 4px}.dismiss:hover{opacity:1}
header{background:var(--surface);border-bottom:1px solid var(--border);
  padding:0 20px;height:54px;display:flex;align-items:center;gap:12px;
  position:sticky;top:0;z-index:90;box-shadow:var(--shadow);flex-shrink:0}
.logo{font-size:20px}.site-title{font-size:15px;font-weight:700}
.hdr-right{margin-left:auto;display:flex;align-items:center;gap:8px}
#conn-pill{font-size:11px;padding:3px 11px;border-radius:20px;
  background:rgba(34,197,94,.12);border:1px solid #22c55e;color:#4ade80}
#conn-pill.disc{background:rgba(239,68,68,.12);border-color:#ef4444;color:#f87171}
.theme-btns{display:flex;border:1px solid var(--border2);border-radius:6px;position:relative}
.tbtn{background:none;border:none;cursor:pointer;padding:5px 9px;font-size:14px;
  color:var(--dim);transition:.15s;border-right:1px solid var(--border2)}
.tmenu-item{display:block;width:100%;padding:7px 12px;background:none;border:none;border-radius:6px;cursor:pointer;color:var(--text);font-size:13px;text-align:left;white-space:nowrap}
.tmenu-item:hover{background:var(--surface)}
.tbtn:last-child{border-right:none}
.tbtn.active,.tbtn:hover{background:var(--border2);color:var(--text)}
#snd-btn,#sel-btn{background:none;border:1px solid var(--border2);border-radius:6px;
  cursor:pointer;padding:5px 9px;font-size:14px;color:var(--dim);transition:.15s}
#snd-btn.on{color:var(--green);border-color:rgba(34,197,94,.4)}
#snd-btn:hover,#sel-btn:hover{background:var(--border2);color:var(--text)}
.tabs{background:var(--surface2);border-bottom:1px solid var(--border);padding:0 20px;display:flex;flex-shrink:0}
.tab{background:none;border:none;cursor:pointer;padding:11px 18px;font-size:13px;
  color:var(--dim);border-bottom:2px solid transparent;transition:.15s;font-weight:500}
.tab.active{color:var(--text);border-bottom-color:var(--blue);font-weight:600}
.sumbar{background:var(--surface2);border-bottom:1px solid var(--border);
  padding:0 20px;height:40px;display:flex;align-items:center;gap:20px;font-size:13px;flex-shrink:0}
.sum-chip{display:flex;align-items:center;gap:6px;padding:3px 11px;
  border-radius:20px;border:1px solid var(--border2);font-weight:500}
#tab-monitor{flex:1;display:flex;flex-direction:column;overflow:hidden}
#tab-stats{display:none;overflow-y:auto}
.grid-wrap{flex:1;overflow-y:auto;padding:16px 20px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:14px}
.card{background:var(--card);border:1px solid var(--border);border-left:5px solid var(--gray);
  border-radius:var(--radius);padding:18px 20px;
  transition:border-color .3s,box-shadow .3s;box-shadow:var(--shadow);cursor:pointer}
.card:hover{box-shadow:0 6px 24px rgba(0,0,0,.6);border-color:var(--blue2)}
.card.green{border-left-color:var(--green)}.card.orange{border-left-color:var(--orange)}
.card.red{border-left-color:var(--red);border-color:rgba(239,68,68,.3)}
.card.paused{border-left-color:var(--gray);opacity:.7;cursor:default}
@keyframes pulse-red{
  0%,100%{
    border-color:rgba(239,68,68,.35);
    box-shadow:0 0 6px rgba(239,68,68,.15),var(--shadow);}
  50%{
    border-color:rgba(239,68,68,.9);
    box-shadow:0 0 22px rgba(239,68,68,.5),0 0 50px rgba(239,68,68,.15),var(--shadow);}}
/* will-change tells the browser to promote this element to its own GPU
   compositing layer. box-shadow and border-color then animate on the
   compositor without triggering main-thread repaints. background was
   removed from the keyframes because animating it forces full-element
   repaint on every frame — the largest source of frame drops. */

/* ── 科技大屏：毛玻璃卡片 ─────────────────────────────────────────── */
.card{
  backdrop-filter:blur(8px) saturate(150%);
  -webkit-backdrop-filter:blur(8px) saturate(150%);
  transition:box-shadow .3s,border-color .3s,transform .15s;
}
.card:hover{transform:translateY(-2px);}

/* ── 状态发光（呼吸灯）─────────────────────────────────────────────── */
.card.red{
  border-color:rgba(239,68,68,.6);
  box-shadow:0 0 0 1px rgba(239,68,68,.3),
             0 0 18px rgba(239,68,68,.25),
             inset 0 0 20px rgba(239,68,68,.04);
  animation:pulse-red 2s infinite;
  will-change:box-shadow,border-color;
}
.card.orange{
  border-color:rgba(245,158,11,.4);
  box-shadow:0 0 12px rgba(245,158,11,.15);
}
.card.green{
  border-color:rgba(34,197,94,.2);
  box-shadow:0 0 8px rgba(34,197,94,.08);
}

/* ── 水波纹：状态圆点（正常时向外扩散）───────────────────────────────── */
.sdot{position:relative;display:inline-block;}
.card.green .sdot::after{
  content:'';position:absolute;top:50%;left:50%;
  width:12px;height:12px;border-radius:50%;
  background:rgba(34,197,94,.4);
  transform:translate(-50%,-50%) scale(1);
  animation:ripple 2.4s ease-out infinite;
  pointer-events:none;
}
@keyframes ripple{
  0%{transform:translate(-50%,-50%) scale(.6);opacity:.8}
  70%{transform:translate(-50%,-50%) scale(2.4);opacity:0}
  100%{transform:translate(-50%,-50%) scale(.6);opacity:0}
}

/* ── 延时数字过渡动画 ─────────────────────────────────────────────── */
.lat-big{
  transition:color .4s;
  font-variant-numeric:tabular-nums;
}
@keyframes num-pop{
  0%{transform:scale(1.08);opacity:.7}
  100%{transform:scale(1);opacity:1}
}
.lat-big.updated{animation:num-pop .25s ease-out;}

/* ── 通用加载 spinner（波形图等异步面板用） ────────────────────── */
.waveform-spinner{
  position:absolute;inset:0;display:flex;align-items:center;justify-content:center;
  background:rgba(0,0,0,0);pointer-events:none;z-index:5;
}
.waveform-spinner::after{
  content:"";width:28px;height:28px;border-radius:50%;
  border:3px solid var(--border2);border-top-color:var(--blue);
  animation:wf-spin .8s linear infinite;
}
@keyframes wf-spin{ to{ transform:rotate(360deg); } }

/* ── 全屏按钮样式 ─────────────────────────────────────────────────── */
#fullscreen-btn{
  padding:5px 12px;border:none;border-radius:6px;cursor:pointer;
  font-size:12px;font-weight:600;background:var(--border);color:var(--dim);
  transition:.15s;
}
#fullscreen-btn:hover{background:var(--border2);color:var(--text);}

/* ── 科技网格背景纹理 ─────────────────────────────────────────────── */
body.theme-dark{
  background-image:
    linear-gradient(rgba(59,130,246,.03) 1px,transparent 1px),
    linear-gradient(90deg,rgba(59,130,246,.03) 1px,transparent 1px);
  background-size:40px 40px;
}
.card.red{animation:pulse-red 2s infinite;will-change:box-shadow,border-color}
/* ── Port-status badges (Feature A) ───────────────────────────────── */
.bar-label{font-size:11px;font-weight:600;color:var(--text);
  margin-bottom:4px;letter-spacing:.4px;text-transform:uppercase}
.ports-row{display:flex;align-items:center;flex-wrap:wrap;gap:4px;
  margin-top:6px;padding-top:6px;border-top:1px solid var(--border)}
.port-pill{display:inline-flex;align-items:center;gap:3px;font-size:10px;
  padding:2px 7px;border-radius:10px;font-weight:600;cursor:default}
.port-pill.open{background:rgba(34,197,94,.15);color:#16a34a}
.port-pill.closed{background:rgba(239,68,68,.12);color:#dc2626}
.ports-ts{font-size:9px;color:var(--dim);margin-left:auto}
.card-top{display:flex;align-items:flex-start;gap:10px;margin-bottom:14px;min-width:0}
.card-top>div{min-width:0;flex:1;overflow:hidden}
.sdot{font-size:18px;margin-top:4px;flex-shrink:0}
.cname{font-size:15px;font-weight:700;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:100%}
.cip{font-size:12px;color:var(--dim);margin-top:3px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:100%}
.ctype{font-size:10px;padding:1px 7px;border-radius:20px;margin-top:3px;display:inline-block;
  background:var(--border);color:var(--dim)}
.lat-big{font-size:38px;font-weight:800;line-height:1;letter-spacing:-1px}
.lat-u{font-size:18px;font-weight:600;margin-left:2px;opacity:.8}
.lat-sub{font-size:12px;color:var(--dim);margin-top:5px}
.tout{font-size:20px;font-weight:800;color:var(--red)}
.paused-label{font-size:18px;font-weight:700;color:var(--dim)}
.cbadge{font-size:12px;padding:3px 10px;border-radius:20px;font-weight:600;margin-top:8px;display:inline-block}
.b-loss{background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.4);color:#f87171}
.b-ok{background:rgba(34,197,94,.1);border:1px solid rgba(34,197,94,.3);color:#4ade80}
.b-warn{background:rgba(245,158,11,.1);border:1px solid rgba(245,158,11,.3);color:#fbbf24}
.hist-section{flex-shrink:0;background:var(--surface);border-top:1px solid var(--border)}
.hist-toggle{width:100%;background:none;border:none;cursor:pointer;
  display:flex;align-items:center;gap:8px;padding:10px 20px;
  font-size:13px;font-weight:600;color:var(--text);text-align:left}
.hist-toggle:hover{background:var(--surface2)}
.htarrow{transition:transform .2s;font-size:11px;color:var(--dim)}
.htarrow.open{transform:rotate(90deg)}
.hist-body{max-height:200px;overflow-y:auto;transition:max-height .2s ease}
.hist-body.collapsed{max-height:0;overflow:hidden}
.hist-empty{padding:18px 20px;text-align:center;color:var(--dim);font-size:12px}
.hrow{display:flex;align-items:center;padding:7px 20px;gap:8px;font-size:12px;border-bottom:1px solid var(--border)}
.hrow:hover{background:rgba(255,255,255,.02)}
.htime{color:var(--dim);width:54px;flex-shrink:0;font-variant-numeric:tabular-nums}
.hname{font-weight:700;flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.hip{color:var(--dim);font-size:11px;flex-shrink:0;margin-left:4px}
.hbadge{font-size:10px;padding:2px 8px;border-radius:20px;font-weight:700;flex-shrink:0;margin-left:6px}
.stats-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:16px;padding:16px 20px}
.stat-card{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);padding:16px;box-shadow:var(--shadow)}
.stat-title{font-size:13px;font-weight:600;color:var(--text);margin-bottom:12px;display:flex;align-items:center;gap:6px}
.chart-wrap{position:relative;height:180px}
.full-chart{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);padding:16px;margin:0 20px 20px;box-shadow:var(--shadow)}
.full-chart-wrap{position:relative;height:240px}
/* Modals */
.modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:200;display:flex;align-items:center;justify-content:center}
.modal-overlay.hidden{display:none!important}
.modal-box{background:var(--surface);border:1px solid var(--border);border-radius:14px;
  width:min(92vw,680px);max-height:88vh;overflow-y:auto;box-shadow:0 8px 40px rgba(0,0,0,.6);
  display:flex;flex-direction:column;transition:width .18s ease}
/* Phase 3C-1.5 / 3C-3: widen the diagnostic modal while DNS compare
   mode is active so the diff matrix has breathing room.  Three tiers
   driven by the number of CUSTOM nameservers in chip list:
     1 custom (2 cols total)  → 900 px  (3C-1.5 baseline, --dns-compare)
     2 custom (3 cols total)  → 1000 px (--dns-compare-mid)
     3 custom (4 cols total)  → 1100 px (--dns-compare-wide)
   Picking the right class lives in _syncDnsModalWidth (JS).  Width
   transitions are CSS-smoothed so toggling chips does not jank. */
.modal-box.modal-box--dns-compare{width:min(96vw,900px)}
.modal-box.modal-box--dns-compare-mid{width:min(97vw,1000px)}
.modal-box.modal-box--dns-compare-wide{width:min(98vw,1100px)}
.modal-head{padding:16px 20px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px;flex-shrink:0}
.modal-title{font-size:15px;font-weight:700;flex:1}
.modal-close{background:none;border:none;cursor:pointer;font-size:20px;color:var(--dim);padding:0 4px}
.modal-close:hover{color:var(--text)}
.modal-body{padding:16px 20px;flex:1}
.modal-footer{padding:12px 20px;border-top:1px solid var(--border);display:flex;justify-content:space-between;align-items:center;flex-shrink:0}
.btn{padding:7px 18px;border-radius:8px;border:none;cursor:pointer;font-size:13px;font-weight:600;transition:.15s}
.btn-primary{background:var(--blue);color:#fff}.btn-primary:hover{opacity:.85}
.btn-ghost{background:var(--border);color:var(--text)}.btn-ghost:hover{background:var(--border2)}
/* Topology */
.topo-wrap{display:flex;flex-direction:column;gap:0;padding:4px 0}
.topo-node{display:flex;align-items:center;gap:12px;padding:8px 14px;border-radius:10px;
  border:1px solid var(--border);background:var(--card);transition:.15s}
.topo-node.you{background:rgba(59,130,246,.08);border-color:rgba(59,130,246,.3)}
.topo-node.destination{background:rgba(107,114,128,.08)}
.topo-node.ok{background:rgba(34,197,94,.06);border-color:rgba(34,197,94,.25)}
.topo-node.filtered{background:rgba(245,158,11,.06);border-color:rgba(245,158,11,.25)}
.topo-node.break{background:rgba(239,68,68,.08);border-color:rgba(239,68,68,.35)}
.topo-node.after{opacity:.35}
.node-circle{width:38px;height:38px;border-radius:50%;display:flex;align-items:center;
  justify-content:center;font-size:13px;font-weight:700;flex-shrink:0}
.topo-node.you .node-circle{background:rgba(59,130,246,.2);border:2px solid var(--blue);color:var(--blue2)}
.topo-node.ok .node-circle{background:rgba(34,197,94,.2);border:2px solid var(--green);color:var(--green)}
.topo-node.filtered .node-circle{background:rgba(245,158,11,.2);border:2px solid var(--orange);color:var(--orange)}
.topo-node.break .node-circle{background:rgba(239,68,68,.2);border:2px solid var(--red);color:var(--red)}
.topo-node.after .node-circle{background:rgba(107,114,128,.15);border:2px solid var(--gray);color:var(--gray)}
.topo-node.destination .node-circle{background:rgba(107,114,128,.15);border:2px solid var(--blue2);color:var(--blue2)}
.node-info{flex:1;min-width:0}
.node-ip{font-size:12px;font-family:monospace;font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.node-host{font-size:10px;color:var(--dim);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:1px}
.node-rtt{font-size:12px;color:var(--dim);flex-shrink:0;margin-left:auto}
.node-tag{font-size:10px;padding:1px 7px;border-radius:10px;flex-shrink:0;margin-left:6px}
.topo-node.filtered .node-tag{background:rgba(245,158,11,.2);color:var(--orange)}
.topo-node.break .node-tag{background:rgba(239,68,68,.2);color:var(--red)}
.topo-connector{display:flex;flex-direction:column;align-items:center;padding:1px 0;margin-left:30px}
.connector-line{width:2px;height:16px;background:var(--border2)}
.connector-arrow{color:var(--dim2);font-size:10px;margin-top:-2px}
.diag-section{margin-top:14px;padding-top:12px;border-top:1px solid var(--border)}
.diag-title{font-size:12px;font-weight:700;color:var(--dim);margin-bottom:8px;display:flex;align-items:center;gap:6px}
.diag-tabs{display:flex;gap:4px;margin:0 10px}
.diag-tab{padding:3px 12px;border-radius:4px;font-size:12px;border:1px solid var(--border);
  background:transparent;color:var(--dim);cursor:pointer;transition:all .15s}
.diag-tab:hover{background:var(--surface2)}
.diag-tab.active{background:var(--blue);color:#fff;border-color:var(--blue)}
.diag-row{display:flex;align-items:center;gap:8px;font-size:12px;padding:5px 0;border-bottom:1px solid var(--border2)}
.diag-row:last-child{border-bottom:none}
.d-ok{color:var(--green)}.d-fail{color:var(--red)}.d-info{color:var(--blue2)}
/* ICMP Diag tab (Task B) */
.icmp-form{padding:4px 2px}
.icmp-row{display:flex;align-items:center;gap:10px;margin-bottom:10px;flex-wrap:wrap}
.icmp-row label{font-size:12px;color:var(--dim);min-width:72px}
.icmp-input{background:var(--surface2);border:1px solid var(--border2);
  border-radius:5px;padding:4px 8px;color:var(--text);font-size:12px;
  font-family:inherit;width:70px}
.icmp-input:focus{outline:none;border-color:var(--blue)}
.icmp-pill{padding:3px 10px;border-radius:4px;font-size:11px;
  border:1px solid var(--border2);background:transparent;color:var(--dim);
  cursor:pointer;transition:.15s}
.icmp-pill:hover{background:var(--surface2);color:var(--text)}
.icmp-pill.active{background:var(--blue);color:#fff;border-color:var(--blue)}
.icmp-preset{padding:2px 8px;border-radius:4px;font-size:11px;
  border:1px solid var(--border2);background:transparent;color:var(--dim);
  cursor:pointer;transition:.15s;font-family:monospace}
.icmp-preset:hover{background:var(--surface2);color:var(--text);border-color:var(--blue)}
.icmp-switch{width:34px;height:18px;border-radius:9px;background:var(--border2);
  position:relative;cursor:pointer;transition:.15s;border:none;padding:0}
.icmp-switch::after{content:"";position:absolute;width:14px;height:14px;
  border-radius:50%;background:#fff;top:2px;left:2px;transition:.15s}
.icmp-switch.on{background:var(--blue)}
.icmp-switch.on::after{left:18px}
.icmp-warn{font-size:11px;color:var(--orange);padding:6px 10px;
  background:rgba(245,158,11,.08);border:1px solid rgba(245,158,11,.25);
  border-radius:5px;margin:6px 0 10px}
.icmp-actions{display:flex;gap:8px;margin-top:8px;flex-wrap:wrap}
.icmp-progress{background:var(--surface2);border:1px solid var(--border);
  border-radius:6px;padding:10px 12px;margin-top:14px;
  max-height:240px;overflow-y:auto;font-family:monospace;font-size:11px;
  line-height:1.8}
.icmp-progress .row{display:flex;align-items:center;gap:10px;
  padding:2px 0;border-bottom:1px dashed var(--border2)}
.icmp-progress .row:last-child{border-bottom:none}
.icmp-progress .row .lbl{color:var(--dim);min-width:90px}
.icmp-progress .row .ic-ok{color:var(--green);min-width:14px}
.icmp-progress .row .ic-bad{color:var(--red);min-width:14px}
.icmp-conclusion{padding:10px 12px;border-radius:6px;margin-top:12px;
  font-size:12px;line-height:1.8;background:rgba(59,130,246,.06);
  border:1px solid rgba(59,130,246,.25);color:var(--text)}
.icmp-conclusion.ok{background:rgba(34,197,94,.07);border-color:rgba(34,197,94,.3)}
.icmp-conclusion.bad{background:rgba(239,68,68,.08);border-color:rgba(239,68,68,.3)}
.icmp-conclusion.warn{background:rgba(245,158,11,.08);border-color:rgba(245,158,11,.3)}
.icmp-stats{display:flex;flex-wrap:wrap;gap:14px;font-size:11px;
  color:var(--dim);margin-top:6px;font-family:monospace}
.icmp-stats b{color:var(--text);font-family:inherit}
/* DNS Diag tab (Phase 2) — reuses .icmp-form / .icmp-row / .icmp-input /
   .icmp-pill / .icmp-switch / .icmp-actions / .icmp-conclusion above,
   so only the chain-box and answer-table styles are unique here.       */
.dns-chain{font-family:monospace;font-size:11px;background:var(--surface2);
  border:1px solid var(--border);border-radius:6px;padding:8px 12px;
  white-space:pre;line-height:1.6;margin-top:10px;
  max-height:140px;overflow:auto;color:var(--text)}
.dns-summary{font-size:11px;color:var(--dim);font-family:monospace;
  display:flex;flex-wrap:wrap;gap:18px;margin-top:12px}
.dns-summary b{color:var(--text);font-family:inherit}
.dns-ans-tbl{width:100%;border-collapse:collapse;font-size:11px;
  font-family:monospace;margin-top:10px}
.dns-ans-tbl th{font-size:10px;font-weight:600;color:var(--dim);
  text-align:left;padding:5px 8px;border-bottom:1px solid var(--border);
  letter-spacing:.4px}
.dns-ans-tbl td{padding:5px 8px;border-bottom:1px solid var(--border2);
  vertical-align:middle;color:var(--text)}
.dns-ans-tbl tr:last-child td{border-bottom:none}
.dns-ans-tbl .dns-qtype{color:var(--blue2);font-weight:600}
.dns-ans-tbl .dns-ttl{color:var(--dim);text-align:right;width:60px}
.dns-trace-btn{background:transparent;border:1px solid var(--border2);
  color:var(--dim);border-radius:4px;padding:2px 8px;font-size:11px;
  cursor:pointer;transition:.15s;font-family:inherit}
.dns-trace-btn:hover{border-color:var(--blue);color:var(--blue2);
  background:var(--surface2)}
/* Phase 3C-1.5: diff-first compare layout.
   Top  → 单行结论条 (verdict bar)
   Mid  → 全宽差异矩阵 (rows = answer values, cols = resolvers)
   Bot  → 每解析器一个可折叠详情 (CNAME 链 + 完整答案表)
   The matrix lives full-width because side-by-side panels can't fit
   in a 680 px modal — 96vw/900px is reserved via modal-box--dns-compare. */
.dns-cmp-verdict{padding:11px 14px;border-radius:8px;
  font-size:13px;font-weight:600;margin-top:10px;
  display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.dns-cmp-verdict.ok {background:rgba(34,197,94,.13);color:var(--green);
  border:1px solid rgba(34,197,94,.35)}
.dns-cmp-verdict.bad{background:rgba(239,68,68,.13);color:var(--red);
  border:1px solid rgba(239,68,68,.35)}
.dns-cmp-verdict .dns-cmp-mark{font-size:17px;line-height:1}
.dns-cmp-verdict .dns-cmp-elapsed{margin-left:auto;font-size:12px;
  font-weight:500;color:var(--dim);font-family:monospace}

/* Diff matrix container.  Border + radius give the matrix a "card"
   feel that visually anchors it as the primary view. */
.dns-cmp-diff{margin-top:12px;border:1px solid var(--border);
  border-radius:8px;overflow:hidden;background:var(--surface)}
.dns-cmp-diff table{width:100%;border-collapse:collapse;
  table-layout:fixed;font-size:12px}
.dns-cmp-diff thead th{font-size:11px;font-weight:700;text-align:left;
  padding:9px 13px;background:var(--surface2);
  border-bottom:1px solid var(--border);color:var(--text);
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
/* 3C-3: side 0 (system) is the only "blue" column; sides 1..N reuse
   the purple --cus tone — at 3 / 4 columns the eye groups baseline-
   vs-customs which matches the operator's mental model better than
   trying to colour-pick each custom side individually. */
.dns-cmp-diff thead th.col-sys{color:var(--blue2)}
.dns-cmp-diff thead th.col-cus{color:#a78bfa}
.dns-cmp-diff tbody td{padding:9px 13px;vertical-align:top;
  border-bottom:1px solid var(--border2);
  overflow-wrap:anywhere;word-break:normal;line-height:1.5}
.dns-cmp-diff tbody tr:last-child td{border-bottom:none}
.dns-cmp-diff tbody tr:hover td{background:rgba(127,127,127,.05)}
.dns-cmp-diff .dns-cmp-val{font-family:'Menlo','Consolas','Courier New',monospace;
  font-size:12px;color:var(--text)}
.dns-cmp-diff .dns-cmp-qtype{display:inline-block;font-size:10px;
  color:var(--dim);font-family:monospace;margin-left:6px;
  padding:1px 6px;background:var(--border);border-radius:3px;
  vertical-align:1px}
.dns-cmp-diff .cell-ok{color:var(--green);font-weight:600;
  font-family:'Menlo','Consolas','Courier New',monospace;font-size:12px}
.dns-cmp-diff .cell-bad{color:var(--red);font-weight:600;
  font-family:'Menlo','Consolas','Courier New',monospace;font-size:12px}
.dns-cmp-diff .dns-cmp-empty{text-align:center;color:var(--dim);
  font-size:11px;padding:18px 12px;font-style:italic}

/* Collapsible per-resolver detail.  Hidden by default when sides
   agree; auto-expanded when there's something worth inspecting
   (mismatch or differing CNAME chain) — that logic lives in JS. */
.dns-cmp-detail{margin-top:10px;border:1px solid var(--border);
  border-radius:8px;overflow:hidden;background:var(--surface)}
.dns-cmp-detail-hdr{display:flex;align-items:center;gap:10px;
  padding:10px 14px;cursor:pointer;user-select:none;
  background:var(--surface2);font-size:12px;font-weight:600;
  transition:background .14s}
.dns-cmp-detail-hdr:hover{background:var(--bg)}
.dns-cmp-detail-hdr .dns-cmp-chev{display:inline-block;width:12px;
  text-align:center;font-size:10px;color:var(--dim);
  transition:transform .15s}
.dns-cmp-detail-hdr.sys .dns-cmp-resolver{color:var(--blue2)}
.dns-cmp-detail-hdr.cus .dns-cmp-resolver{color:#a78bfa}
.dns-cmp-detail-hdr .dns-cmp-resolver-meta{margin-left:auto;
  font-size:11px;color:var(--dim);font-family:monospace;font-weight:500}
.dns-cmp-detail-body{display:none;padding:12px 14px;
  border-top:1px solid var(--border);background:var(--bg)}
.dns-cmp-detail-body.open{display:block}
.dns-cmp-detail-body .dns-chain{max-height:140px;margin:0 0 10px}
.dns-cmp-detail-body .dns-ans-tbl{margin-top:0}
.dns-cmp-detail-body .dns-cmp-msg{font-size:11px;color:var(--dim);
  margin-top:8px;line-height:1.5}
/* Phase 3C-3 — multi-resolver chip input for compare mode.  Chips
   replace the single-IP text field; preset buttons + an Add button
   feed into the same chip list so operators can build a 1..3 IP
   compare set in a few clicks.  Purple tone matches the verdict
   detail headers so the eye groups chip → "对照解析器" column. */
.dns-cmp-ns-list{display:inline-flex;flex-wrap:wrap;gap:6px;
  align-items:center;vertical-align:middle}
.dns-cmp-ns-chip{display:inline-flex;align-items:center;gap:4px;
  padding:2px 4px 2px 10px;border-radius:14px;
  background:rgba(167,139,250,.15);
  border:1px solid rgba(167,139,250,.35);
  color:#a78bfa;font-size:11px;
  font-family:'Menlo','Consolas','Courier New',monospace;
  line-height:1.6}
.dns-cmp-ns-chip .dns-cmp-ns-x{background:none;border:none;cursor:pointer;
  color:#a78bfa;font-size:15px;padding:0 4px;line-height:1;
  font-family:inherit;transition:.15s}
.dns-cmp-ns-chip .dns-cmp-ns-x:hover{color:#fca5a5}
.dns-cmp-ns-count{font-size:11px;color:var(--dim);margin-left:4px;
  font-family:monospace}
.dns-cmp-ns-count.full{color:var(--orange)}
.dns-cmp-ns-empty{font-size:11px;color:var(--dim);font-style:italic}

/* Phase 3C-4 — authority-trace tree view (rev2: real block layout).
   Top   → 单行结论条 (verdict bar, reusing .dns-cmp-verdict palette so
           operator's eye groups it with compare/single verdicts).
   Mid   → 全宽 monospace 树.  Each hop is its own <div> with a header
           row (zone / dest / rtt) and an indented body containing
           one <div> per detail line (NS list, glue lines, message).
           Critically: NO white-space:pre-wrap, NO literal "\n" text
           nodes — relying purely on block layout means no stray
           blank lines between hops and detail lines wrap cleanly.
   Bot   → 最终答案表 (复用 .dns-ans-tbl) + 结论 box.                */
.dns-auth-tree{font-family:'Menlo','Consolas','Courier New',monospace;
  font-size:11.5px;line-height:1.55;background:var(--surface2);
  border:1px solid var(--border);border-radius:8px;padding:10px 14px;
  margin-top:12px;color:var(--text)}
.dns-auth-tree .dns-auth-hop{padding:5px 0}
.dns-auth-tree .dns-auth-hop + .dns-auth-hop{
  border-top:1px dashed var(--border2);margin-top:2px;padding-top:7px}
.dns-auth-hop-head{display:flex;flex-wrap:wrap;align-items:baseline;
  gap:6px;column-gap:4px}
.dns-auth-prefix{color:var(--dim);flex:0 0 auto}
.dns-auth-zone{font-weight:700;flex:0 0 auto}
.dns-auth-hop.ok   .dns-auth-zone{color:var(--blue2)}
.dns-auth-hop.bad  .dns-auth-zone{color:var(--red)}
.dns-auth-hop.warn .dns-auth-zone{color:var(--orange)}
.dns-auth-dest{color:var(--dim);overflow-wrap:anywhere;word-break:break-word;
  flex:1 1 auto;min-width:0}
.dns-auth-dest b{color:var(--text);font-weight:500}
.dns-auth-rtt{color:var(--dim);font-variant-numeric:tabular-nums;
  white-space:nowrap;flex:0 0 auto;margin-left:auto}
.dns-auth-hop-body{padding-left:20px;margin-top:2px}
.dns-auth-line{display:block;padding:1px 0;color:var(--text);
  overflow-wrap:anywhere;word-break:break-word}
.dns-auth-line .lbl{color:var(--dim);margin-right:6px;
  display:inline-block;min-width:34px}
.dns-auth-line .err{color:var(--red);font-weight:600}
.dns-auth-line .warn{color:var(--orange)}
.dns-auth-line .info{color:var(--dim);font-style:italic}
.dns-auth-line .ns-name{color:var(--text)}
.dns-auth-line .glue-name{color:var(--text)}
.dns-auth-line .glue-arrow{color:var(--dim);margin:0 4px}
.dns-auth-line .glue-ip{color:#86efac}
.dns-auth-summary{font-size:11px;color:var(--dim);font-family:monospace;
  display:flex;flex-wrap:wrap;gap:18px;margin-top:6px}
.dns-auth-summary b{color:var(--text);font-family:inherit}

/* Phase 3C-5 — DNSSEC validation result layout.
   Top   → 4-state verdict bar (.secure / .insecure / .bogus / .indeterminate).
           Colours encode the verdict semantics: green=SECURE (trust
           chain fully verified), grey=INSECURE (zone未签名 — common
           CN case, NOT a failure), red=BOGUS (signature failed —
           actual incident), amber=INDETERMINATE (couldn't tell).
   Mid   → Verification chain TABLE (NOT a tree — chain is linear,
           one row per zone from root to leaf, distinct from the
           trace_auth tree which shows delegation referrals).
   Bot   → Answer table (.dns-ans-tbl reused) + AD/RRSIG metadata
           strip + footer conclusion echo.
   The intent of the visual language is to make the INSECURE case
   read as "normal grey informational" rather than "red error" —
   operator must be able to tell BOGUS apart from INSECURE at a
   glance, which is the whole point of validating client-side. */
.dns-sec-verdict{padding:11px 14px;border-radius:8px;
  font-size:13px;font-weight:600;margin-top:10px;
  display:flex;align-items:center;gap:10px;flex-wrap:wrap}
.dns-sec-verdict.secure{background:rgba(34,197,94,.13);color:var(--green);
  border:1px solid rgba(34,197,94,.35)}
.dns-sec-verdict.insecure{background:rgba(127,127,127,.13);color:var(--dim);
  border:1px solid rgba(127,127,127,.30)}
.dns-sec-verdict.bogus{background:rgba(239,68,68,.13);color:var(--red);
  border:1px solid rgba(239,68,68,.35)}
.dns-sec-verdict.indeterminate{background:rgba(245,158,11,.13);
  color:var(--orange);border:1px solid rgba(245,158,11,.35)}
.dns-sec-verdict .dns-sec-mark{font-size:17px;line-height:1}
.dns-sec-verdict .dns-sec-tag{display:inline-block;padding:1px 8px;
  border-radius:4px;font-size:11px;font-weight:700;letter-spacing:.5px;
  background:rgba(0,0,0,.18);font-family:monospace}
.dns-sec-verdict .dns-sec-elapsed{margin-left:auto;font-size:12px;
  font-weight:500;color:inherit;opacity:.8;font-family:monospace}

/* Verification chain — a TABLE intentionally (not a tree).  Each
   row is one zone; the column layout makes "where did the trust
   break?" answerable at a glance.  Colours echo the verdict
   palette: green check on a fully-verified link, grey ○ on the
   first INSECURE boundary (chain stops there), red ✗ on BOGUS. */
.dns-sec-chain{margin-top:12px;border:1px solid var(--border);
  border-radius:8px;overflow:hidden;background:var(--surface)}
.dns-sec-chain table{width:100%;border-collapse:collapse;font-size:12px}
.dns-sec-chain thead th{font-size:11px;font-weight:700;text-align:left;
  padding:9px 13px;background:var(--surface2);
  border-bottom:1px solid var(--border);color:var(--text)}
.dns-sec-chain tbody td{padding:9px 13px;vertical-align:top;
  border-bottom:1px solid var(--border2);line-height:1.55}
.dns-sec-chain tbody tr:last-child td{border-bottom:none}
.dns-sec-chain .sec-zone{font-family:'Menlo','Consolas','Courier New',monospace;
  font-weight:700;color:var(--text);white-space:nowrap}
.dns-sec-chain .sec-mark{font-size:14px;line-height:1;width:24px;
  text-align:center;font-family:monospace}
.dns-sec-chain .sec-mark-ok{color:var(--green)}
.dns-sec-chain .sec-mark-insecure{color:var(--dim)}
.dns-sec-chain .sec-mark-bad{color:var(--red)}
.dns-sec-chain .sec-mark-pending{color:var(--orange)}
.dns-sec-chain .sec-algo{font-family:monospace;color:var(--dim);
  font-size:11px;white-space:nowrap}
.dns-sec-chain .sec-note{color:var(--text);font-size:11.5px;
  overflow-wrap:anywhere;word-break:normal}
.dns-sec-chain tbody tr.row-insecure td{background:rgba(127,127,127,.06)}
.dns-sec-chain tbody tr.row-bad td{background:rgba(239,68,68,.06)}

/* AD / RRSIG metadata strip — sits between the chain and the answer
   table.  Single-row monospace; clearly informational, NOT part of
   the verdict (especially the AD bit, which we don't trust). */
.dns-sec-meta{font-size:11px;color:var(--dim);font-family:monospace;
  display:flex;flex-wrap:wrap;gap:18px;margin-top:10px;
  padding:8px 12px;background:var(--surface);border:1px solid var(--border);
  border-radius:6px}
.dns-sec-meta b{color:var(--text);font-family:inherit}
.dns-sec-meta .sec-meta-ad-hint{color:var(--dim);font-style:italic;
  font-family:inherit}

.icmp-hist-item{border:1px solid var(--border);border-radius:6px;
  margin-bottom:6px;overflow:hidden}
.icmp-hist-head{display:flex;align-items:center;gap:10px;
  padding:9px 14px;cursor:pointer;user-select:none;background:var(--surface)}
.icmp-hist-head:hover{background:var(--surface2)}
.icmp-hist-body{display:none;background:var(--bg);padding:10px 14px;
  border-top:1px solid var(--border)}
.icmp-hist-mode{font-size:10px;font-weight:600;padding:1px 7px;
  border-radius:3px;background:var(--border);color:var(--text);font-family:monospace}
/* Card selection */
.chip-grid{display:flex;flex-wrap:wrap;gap:8px;padding:4px 0}
.chip{display:flex;align-items:center;gap:6px;padding:6px 12px;border-radius:20px;
  border:1px solid var(--border2);cursor:pointer;font-size:12px;transition:.15s;user-select:none}
.chip:hover{border-color:var(--blue)}
.chip.selected{background:rgba(59,130,246,.15);border-color:var(--blue);color:var(--blue2)}
.chip-dot{width:8px;height:8px;border-radius:50%}
::-webkit-scrollbar{width:5px;height:5px}::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
@media(max-width:600px){.grid{grid-template-columns:1fr}.grid-wrap{padding:10px 12px}
  header,.sumbar,.hist-toggle,.hrow{padding-left:12px;padding-right:12px}.modal-box{width:95vw}}
</style>
</head>
<body class="theme-system" id="body">
<div id="alert-banner" class="hidden">
  <span style="font-size:18px">🔴</span><span>网络告警：</span>
  <span id="alert-text" style="color:#fca5a5"></span>
  <button class="dismiss" onclick="dismissAlert()">×</button>
</div>
<header>
  <span class="logo">🌐</span><span class="site-title">网络连通性监测</span>
  <div class="hdr-right">
    <button id="sel-btn" onclick="openCardSel()" title="自定义显示的卡片">⊞</button>
    <button id="snd-btn" onclick="toggleSound()" title="告警声音">🔔</button>
    <button id="fullscreen-btn" onclick="toggleFullscreen()" title="大屏全屏模式（F11）">⛶</button>
    <div class="theme-btns" style="position:relative">
      <button id="theme-toggle" class="tbtn" title="切换主题"
              onclick="toggleThemeMenu(event)" style="min-width:32px">🌙</button>
      <div id="theme-menu" style="display:none;position:fixed;background:var(--surface2);
          border:1px solid var(--border);border-radius:8px;
          box-shadow:var(--shadow);z-index:9999;padding:4px;min-width:130px">
        <button onclick="setTheme('dark');closeThemeMenu()" class="tmenu-item">🌙 暗色</button>
        <button onclick="setTheme('light');closeThemeMenu()" class="tmenu-item">☀️ 亮色</button>
        <button onclick="setTheme('system');closeThemeMenu()" class="tmenu-item">💻 跟随系统</button>
      </div>
    </div>
    <span id="conn-pill">● 已连接</span>
  </div>
</header>
<div class="tabs">
  <button class="tab active" onclick="switchTab('monitor',this)">📡 实时监控</button>
  <button class="tab" onclick="switchTab('stats',this)">📊 统计分析</button>
  <button class="tab" onclick="switchTab('sla',this)">📋 SLA 报告</button>
</div>
<div class="sumbar">
  <div class="sum-chip" style="border-color:rgba(107,114,128,.3)"><span id="s-total">共 0 个</span></div>
  <div class="sum-chip" style="border-color:rgba(34,197,94,.3);color:var(--green)"><span>●</span><span id="s-green">正常 0</span></div>
  <div class="sum-chip" style="border-color:rgba(245,158,11,.3);color:var(--orange)"><span>●</span><span id="s-orange">警告 0</span></div>
  <div class="sum-chip" style="border-color:rgba(239,68,68,.3);color:var(--red)"><span>●</span><span id="s-red">告警 0</span></div>
  <span id="upd" style="margin-left:auto;font-size:11px;color:var(--dim)">连接中...</span>
</div>
<div id="tab-monitor">
  <div class="grid-wrap"><div class="grid" id="grid"><div style="color:var(--dim);font-size:13px;padding:24px;grid-column:1/-1">加载中...</div></div></div>
  <div class="hist-section">
    <button class="hist-toggle" onclick="toggleHist()">
      <span class="htarrow open" id="htarrow">▶</span><span>📋 告警历史</span>
      <span id="hcount" style="font-size:11px;color:var(--dim);font-weight:400;margin-left:4px"></span>
    </button>
    <div class="hist-body" id="hist-body"><div class="hist-empty">暂无告警记录</div></div>
  </div>
</div>
<div id="tab-stats">

  <!-- 统一时间控制栏 -->
  <div id="time-ctrl" style="display:flex;align-items:center;gap:10px;
       flex-wrap:wrap;padding:12px 20px;border-bottom:1px solid var(--border)">
    <div style="display:flex;gap:6px">
      <button class="tbtn tsel" data-sec="3600"    onclick="setQuick(this)">近1小时</button>
      <button class="tbtn"      data-sec="86400"   onclick="setQuick(this)">近24小时</button>
      <button class="tbtn"      data-sec="604800"  onclick="setQuick(this)">近7天</button>
      <button class="tbtn"      data-sec="2592000" onclick="setQuick(this)">近30天</button>
    </div>
    <span style="color:var(--dim);font-size:11px">或自定义：</span>
    <input type="datetime-local" id="dt-start"
           style="font-size:12px;padding:3px 8px;border-radius:6px;
                  background:var(--border);color:var(--text);border:1px solid var(--border2)">
    <span style="color:var(--dim)">—</span>
    <input type="datetime-local" id="dt-end"
           style="font-size:12px;padding:3px 8px;border-radius:6px;
                  background:var(--border);color:var(--text);border:1px solid var(--border2)">
    <button onclick="applyTimeRange()"
            style="padding:4px 14px;border-radius:6px;border:none;cursor:pointer;
                   font-size:12px;font-weight:600;background:var(--blue);color:#fff">
      🔍 查询并刷新图表
    </button>
    <button onclick="doExportExcel()"
            style="padding:4px 14px;border-radius:6px;border:none;cursor:pointer;
                   font-size:12px;font-weight:600;
                   background:rgba(34,197,94,.15);color:#4ade80">
      📥 导出 Excel
    </button>
    <span id="time-info" style="font-size:10px;color:var(--dim)"></span>
  </div>

  <!-- Loading 遮罩层 -->
  <div id="stats-loading"
       style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.55);
              z-index:9999;align-items:center;justify-content:center">
    <div style="background:var(--card);padding:28px 40px;border-radius:12px;text-align:center">
      <div style="font-size:24px;margin-bottom:10px">⏳</div>
      <div id="loading-msg" style="color:var(--text);font-size:14px;font-weight:500">
        正在处理，请稍候...
      </div>
    </div>
  </div>

  <div class="stats-grid" id="donut-grid"></div>
  <div class="full-chart">
    <div class="stat-title">
      <span>📊 平均延时对比（ms）</span>
      <span style="font-size:11px;color:var(--dim)">各协议独立 Y 轴</span>
    </div>
    <!-- Responsive grid: ICMP full-width on row 1, TCP+HTTP side-by-side on row 2 -->
    <div id="bar-grid" style="display:flex;flex-direction:column;gap:12px">
      <div id="bar-icmp-group" style="display:none">
        <div class="bar-label">ICMP PING</div>
        <div id="bar-icmp-wrap" style="min-height:60px"><canvas id="bar-icmp"></canvas></div>
      </div>
      <div id="bar-row2" style="display:flex;gap:12px">
        <div id="bar-tcp-group" style="display:none;flex:1;min-width:0">
          <div class="bar-label">TCP PING</div>
          <div id="bar-tcp-wrap" style="min-height:60px"><canvas id="bar-tcp"></canvas></div>
        </div>
        <div id="bar-http-group" style="display:none;flex:1;min-width:0">
          <div class="bar-label">HTTP / HTTPS</div>
          <div id="bar-http-wrap" style="min-height:60px"><canvas id="bar-http"></canvas></div>
        </div>
        <div id="bar-dns-group" style="display:none;flex:1;min-width:0">
          <div class="bar-label">DNS 监控</div>
          <div id="bar-dns-wrap" style="min-height:60px"><canvas id="bar-dns"></canvas></div>
        </div>
      </div>
    </div>
    <div id="bar-empty" style="color:var(--dim);font-size:12px;padding:20px 0;text-align:center">
      暂无统计数据
    </div>
  </div>
  <!-- Waveform trend chart section -->
  <div class="full-chart" id="waveform-section">
    <div class="stat-title" style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px">
      <span>📈 延时趋势波形图</span>
      <div style="display:flex;gap:6px;align-items:center">
        <select id="waveform-tid" onchange="loadWaveform()"
                style="font-size:12px;padding:3px 8px;border-radius:6px;
                       background:var(--border);color:var(--text);border:1px solid var(--border2)">
          <option value="">— 选择节点 —</option>
        </select>
        <span id="waveform-resolution" style="font-size:10px;color:var(--dim)"></span>
        <button onclick="loadWaveform()"
                style="padding:3px 10px;border-radius:5px;border:none;cursor:pointer;
                       font-size:11px;background:var(--blue);color:#fff">🔄 刷新</button>
        <button onclick="exportWaveform()"
                style="padding:3px 10px;border-radius:5px;border:none;cursor:pointer;
                       font-size:11px;background:var(--green);color:#fff">⬇ Excel</button>
      </div>
    </div>
    <div id="waveform-summary" style="display:flex;gap:16px;flex-wrap:wrap;
         padding:8px 0 12px;font-size:12px;color:var(--dim)"></div>
    <div style="position:relative;height:220px">
      <canvas id="waveform-chart"></canvas>
      <div id="waveform-empty" style="position:absolute;inset:0;display:flex;
           align-items:center;justify-content:center;color:var(--dim);font-size:13px">
        请选择节点后点击刷新
      </div>
      <div id="waveform-loading" class="waveform-spinner" style="display:none"></div>
    </div>
  </div>
</div>
<!-- Diagnosis modal -->
<div class="modal-overlay hidden" id="diag-overlay" onclick="if(event.target===this)closeDiag()">
  <div class="modal-box">
    <div class="modal-head">
      <span id="diag-title" class="modal-title">诊断</span>
      <div class="diag-tabs">
        <button class="diag-tab active" id="tab-current" onclick="switchDiagTab('current')">路由追踪</button>
        <button class="diag-tab" id="tab-history" onclick="switchDiagTab('history')">路由历史</button>
        <button class="diag-tab" id="tab-icmp" onclick="switchDiagTab('icmp')">ICMP 诊断</button>
        <button class="diag-tab" id="tab-dns" onclick="switchDiagTab('dns')">DNS 诊断</button>
      </div>
      <button class="modal-close" onclick="closeDiag()">×</button>
    </div>
    <div class="modal-body" id="diag-body"></div>
    <div class="modal-footer">
      <span id="diag-updated" style="font-size:11px;color:var(--dim)"></span>
      <div style="display:flex;gap:8px">
        <button class="btn btn-ghost" id="diag-quick-btn" onclick="openQuickTrace()"
                title="在独立窗口中运行快速路由追踪（-d 跳过 DNS 反查）"
                style="color:#ca8a04;border-color:rgba(234,179,8,.4)">⚡ 快速追踪</button>
        <button class="btn btn-ghost" id="diag-refresh" onclick="refreshDiag()">🔄 刷新</button>
        <button class="btn btn-ghost" onclick="closeDiag()">关闭</button>
      </div>
    </div>
  </div>
</div>
<!-- Card selection modal -->
<div class="modal-overlay hidden" id="sel-overlay" onclick="if(event.target===this)closeCardSel()">
  <div class="modal-box">
    <div class="modal-head">
      <span class="modal-title">⊞ 自定义监控面板显示的卡片</span>
      <button class="modal-close" onclick="closeCardSel()">×</button>
    </div>
    <div class="modal-body">
      <div style="font-size:12px;color:var(--dim);margin-bottom:12px">选择要在监控面板中显示的节点。未选中的节点仍在监测，只是不显示在面板上。此设置保存在当前浏览器，不影响其他设备。</div>
      <div class="chip-grid" id="chip-grid"></div>
      <div style="margin-top:14px;display:flex;gap:8px">
        <button class="btn btn-ghost" onclick="selectAllChips()">全选</button>
        <button class="btn btn-ghost" onclick="selectNoneChips()">清除</button>
      </div>
    </div>
    <div class="modal-footer" style="justify-content:flex-end">
      <button class="btn btn-ghost" onclick="closeCardSel()">取消</button>
      <button class="btn btn-primary" onclick="saveCardSel()">保存</button>
    </div>
  </div>
</div>

<div id="tab-sla" style="display:none;padding:12px 20px">
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:14px;flex-wrap:wrap">
    <span style="font-size:13px;font-weight:600;color:var(--text)">统计周期：</span>
    <button class="tbtn active" id="slab-7d"  onclick="setSLAPeriod(7)">近 7 天</button>
    <button class="tbtn" id="slab-30d" onclick="setSLAPeriod(30)">近 30 天</button>
    <button class="tbtn" id="slab-mo"  onclick="setSLACurrentMonth()">本月</button>
    <span style="color:var(--dim);font-size:12px;margin-left:4px">自定义：</span>
    <input type="date" id="sla-s" style="font-size:12px;padding:3px 6px;background:var(--surface2);color:var(--text);border:1px solid var(--border);border-radius:4px">
    <span style="color:var(--dim)">—</span>
    <input type="date" id="sla-e" style="font-size:12px;padding:3px 6px;background:var(--surface2);color:var(--text);border:1px solid var(--border);border-radius:4px">
    <button onclick="loadSLA()"
            style="padding:7px 16px;border:none;border-radius:6px;cursor:pointer;
                   font-size:12px;font-weight:600;background:var(--blue);color:#fff">
      🔍 查询
    </button>
    <button onclick="exportSLAExcel()"
            style="padding:7px 16px;border:none;border-radius:6px;cursor:pointer;
                   font-size:12px;font-weight:600;
                   background:rgba(34,197,94,.15);color:#4ade80">
      📥 导出 Excel
    </button>
    <span id="sla-hint" style="font-size:11px;color:var(--dim);margin-left:6px"></span>
  </div>
  <!-- 数据说明 -->
  <div style="margin:0 0 10px;padding:8px 14px;border-radius:6px;
       background:rgba(59,130,246,.06);border:1px solid rgba(59,130,246,.18);
       font-size:11px;color:var(--dim);line-height:1.8">
    📌 <b style="color:var(--text)">数据说明：</b>
    <b>连通率 / 故障次数 / 累计中断</b> — 基于实时告警事件，精确到秒；
    &nbsp;<b>均延时 / P95 延时 / 最大延时</b> — 基于每小时聚合数据，
    当前不完整小时的数据最长滞后 <b>1 小时</b>才会计入。
  </div>
  <div style="overflow-x:auto">
  </div>

  <table id="sla-table" style="width:100%;border-collapse:collapse;font-size:13px">
    <thead>
      <tr style="background:var(--surface2);color:var(--dim);text-align:left">
        <th style="padding:9px 14px;border-bottom:1px solid var(--border);cursor:pointer" onclick="sortSLA(0)">节点 ↕</th>
        <th style="padding:9px 14px;border-bottom:1px solid var(--border);cursor:pointer;text-align:center" onclick="sortSLA(1)">连通率 ↕</th>
        <th style="padding:9px 14px;border-bottom:1px solid var(--border);cursor:pointer;text-align:right" onclick="sortSLA(2)">均延时 ↕</th>
        <th style="padding:9px 14px;border-bottom:1px solid var(--border);cursor:pointer;text-align:right" onclick="sortSLA(3)">P95 延时 ↕</th>
        <th style="padding:9px 14px;border-bottom:1px solid var(--border);cursor:pointer;text-align:right" onclick="sortSLA(4)">最大延时 ↕</th>
        <th style="padding:9px 14px;border-bottom:1px solid var(--border);cursor:pointer;text-align:center" onclick="sortSLA(5)">故障次数 ↕</th>
        <th style="padding:9px 14px;border-bottom:1px solid var(--border);cursor:pointer;text-align:right" onclick="sortSLA(6)">累计中断 ↕</th>
      </tr>
    </thead>
    <tbody id="sla-tbody">
      <tr><td colspan="7" style="text-align:center;padding:40px;color:var(--dim)">点击「查询」加载 SLA 报告</td></tr>
    </tbody>
  </table>
  </div>
</div>

<script>
const targets={},hist=[],stats={};
const MAX_HIST=60,MAX_LAT=300;
const barCharts={};let histOpen=true;
const portStatus={}; // {tid: {ports:[...], updated_at:''}}
const donutCharts={};
let currentDiagTid=null;
let _diagReqId=0;   // Incremented on each new loadDiag/refreshDiag call.
                     // The fetch callback checks this before rendering to prevent
                     // a stale traceroute response from overwriting a newer request.
let pinnedTargets=JSON.parse(localStorage.getItem('pinnedTargets')||'null');
function isVisible(tid){return pinnedTargets===null||pinnedTargets.includes(tid);}

// ── Alert Audio ────────────────────────────────────────────────────────
// Strategy: play server-provided WAV files via <audio> elements.
// WAV files are served by Flask from the assets/ directory on the C-side
// machine. This is far more reliable than oscillator synthesis because:
//  • Browsers can block or muffle synthesised tones (AudioContext policy)
//  • Real WAV files play through the browser's standard media pipeline
//  • C-side operator can replace the WAVs to customise alert sounds
//
// Audio unlock: browsers require a user gesture before playing audio.
// We create the <audio> elements on first click, then they stay ready.
let soundEnabled=localStorage.getItem('web_sound')!=='false';
let _audioReady=false;
const _audios={};
function _ensureAudio(){
  if(_audioReady)return;
  _audioReady=true;
  ['alarm','warning','recovery'].forEach(k=>{
    const a=new Audio(`/api/sound/alert_${k}.wav`);
    a.preload='auto';
    _audios[k]=a;
  });
}
document.addEventListener('click',_ensureAudio,{once:false,passive:true});
function _playWav(key){
  if(!soundEnabled)return;
  _ensureAudio();
  const a=_audios[key];
  if(!a)return;
  try{a.currentTime=0;a.play().catch(()=>{});}catch{}
}
function playWarning(){_playWav('warning');}
function playAlarm(){_playWav('alarm');}
function playRecovery(){_playWav('recovery');}
function toggleSound(){soundEnabled=!soundEnabled;localStorage.setItem('web_sound',soundEnabled);
  document.getElementById('snd-btn').classList.toggle('on',soundEnabled);}
(()=>document.getElementById('snd-btn').classList.toggle('on',soundEnabled))();

// Theme
let cTheme=localStorage.getItem('theme')||'system';
function toggleFullscreen(){
  if(!document.fullscreenElement){
    document.documentElement.requestFullscreen().catch(()=>{});
  }else{
    document.exitFullscreen();
  }
}
document.addEventListener('fullscreenchange',()=>{
  const btn=document.getElementById('fullscreen-btn');
  if(btn)btn.textContent=document.fullscreenElement?'✕ 退出':'⛶';
});

// 延时数字弹跳动效
function _animateLatUpdate(tid){
  const el=document.querySelector('#lat-'+tid+' .lat-big');
  if(!el)return;
  el.classList.remove('updated');
  void el.offsetWidth;
  el.classList.add('updated');
  setTimeout(()=>el.classList.remove('updated'),300);
}

function applyTheme(t){document.getElementById('body').className='theme-'+t;
  const _icons={dark:'🌙',light:'☀️',system:'💻'};
  const _tog=document.getElementById('theme-toggle');
  if(_tog)_tog.textContent=_icons[t]||'🌙';
  // Only select theme buttons, not other tbtn-class elements (e.g. time range buttons)
  document.querySelectorAll('.theme-btns .tbtn').forEach((b,i)=>
    b.classList.toggle('active',['dark','light','system'][i]===t));
  const dark=t==='dark'||(t==='system'&&matchMedia('(prefers-color-scheme:dark)').matches);
  // typeof Chart guard: the only Chart.js source is a public CDN.
  // When the host is offline / DNS-blocked / behind a strict proxy
  // the script tag silently fails, leaving `Chart` undefined.
  // Touching Chart.defaults here used to throw ReferenceError mid-
  // init, which prevented the SSE connect and card render code from
  // ever running -- so the dashboard appeared completely broken
  // even though node monitoring was unaffected.  Skip the Chart
  // theming when the library is missing; status cards, history
  // panel and SSE bring-up keep working.  Charts are degraded
  // (empty placeholder) but do not block the dashboard.
  if(typeof Chart!=='undefined'){
    Chart.defaults.color=dark?'#9ca3af':'#4b5563';
    Chart.defaults.borderColor=dark?'#2a2e3a':'#e2e5ec';
  }
  // Rebuild donut gradient colors (canvas gradients are theme-dependent)
  Object.entries(donutCharts).forEach(([tid,ch])=>{
    try{
      const ds=ch.data.datasets[0];
      const cv=ch.canvas;
      if(!cv)return;
      const ctx2=cv.getContext('2d');
      const g=ctx2.createLinearGradient(0,0,0,100);
      g.addColorStop(0,'#22d3ee');g.addColorStop(1,'#22c55e');
      ds.backgroundColor=[g,'rgba(239,68,68,.18)'];
      ch.update('none');
    }catch{}
  });
  Object.values(barCharts).forEach(c=>{try{c.update();}catch{}});
  // ── Calendar icon color-scheme (3-layer defence) ──────────────────
  // Layer 1: html root color-scheme → browser renders ALL native controls
  //          (calendar icon, scrollbars, etc.) in the right mode.
  //          This is the most reliable signal Chrome/Edge actually respects.
  document.documentElement.style.colorScheme=dark?'dark':'light';
  // Layer 2: per-input inline color-scheme (redundant but harmless)
  const cs=dark?'dark':'light';
  document.querySelectorAll('input[type="date"],input[type="datetime-local"]').forEach(el=>el.style.colorScheme=cs);
  // Layer 3: filter:invert via dynamic <style> as final fallback
  let calStyle=document.getElementById('_cal_style');
  if(!calStyle){calStyle=document.createElement('style');calStyle.id='_cal_style';document.head.appendChild(calStyle);}
  calStyle.textContent=dark
    ?'input[type="date"]::-webkit-calendar-picker-indicator,input[type="datetime-local"]::-webkit-calendar-picker-indicator{filter:invert(1) brightness(1.2) !important;cursor:pointer}'
    :'input[type="date"]::-webkit-calendar-picker-indicator,input[type="datetime-local"]::-webkit-calendar-picker-indicator{filter:none !important;cursor:pointer}';}
function setTheme(t){cTheme=t;localStorage.setItem('theme',t);applyTheme(t);}
function toggleThemeMenu(e){
  e.stopPropagation();
  const m=document.getElementById('theme-menu');
  if(!m)return;
  if(m.style.display==='none'||!m.style.display){
    const r=e.currentTarget.getBoundingClientRect();
    m.style.top=(r.bottom+6)+'px';
    m.style.right=(window.innerWidth-r.right)+'px';
    m.style.display='block';
  }else{
    m.style.display='none';
  }
}
function closeThemeMenu(){const m=document.getElementById('theme-menu');if(m)m.style.display='none';}
document.addEventListener('click',()=>closeThemeMenu());
applyTheme(cTheme);
matchMedia('(prefers-color-scheme:dark)').addEventListener('change',()=>{if(cTheme==='system')applyTheme('system');});

function switchTab(n,btn){
  document.querySelectorAll('.tab').forEach(b=>b.classList.remove('active'));btn.classList.add('active');
  document.getElementById('tab-monitor').style.display=n==='monitor'?'flex':'none';
  document.getElementById('tab-stats').style.display=n==='stats'?'block':'none';
  document.getElementById('tab-sla').style.display=n==='sla'?'block':'none';
  if(n==='stats')setTimeout(initStatsTab,50);
  if(n==='sla')setTimeout(initSLATab,50);}
function toggleHist(){histOpen=!histOpen;
  document.getElementById('hist-body').classList.toggle('collapsed',!histOpen);
  document.getElementById('htarrow').classList.toggle('open',histOpen);}

const ST={green:{color:'var(--green)',label:'正常',bg:'rgba(34,197,94,.1)',br:'rgba(34,197,94,.3)'},
  orange:{color:'var(--orange)',label:'警告',bg:'rgba(245,158,11,.1)',br:'rgba(245,158,11,.3)'},
  red:{color:'var(--red)',label:'告警',bg:'rgba(239,68,68,.1)',br:'rgba(239,68,68,.3)'},
  gray:{color:'var(--gray)',label:'未知',bg:'rgba(107,114,128,.1)'},
  paused:{color:'var(--dim)',label:'已暂停',bg:'rgba(107,114,128,.1)'}};
const esc=s=>String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#039;');
const st=s=>ST[s]||ST.gray;
function formatFailure(r){
  if(!r)return '请求超时';
  const l=String(r).toLowerCase();
  if(l==='keyword_not_found')return '⚠ 关键字未匹配（内容异常）';
  if(l==='keyword_found')    return '⚠ 关键字命中（排除模式：内容含异常标志）';
  if(l.startsWith('dns_hijack:'))return '⚠ DNS劫持:'+l.split(':')[1];
  if(l==='nxdomain')return '域名不存在';
  if(l==='dns_no_a_record')return '无 A 记录';
  if(l==='dns_no_domain')return '未配置查询域名';
  if(l.includes('dns'))return 'DNS 失败';
  if(l.includes('tls'))return 'TLS 失败';
  if(l.includes('tcp_timeout')||l.includes('recv_timeout'))return '超时';
  if(l.includes('tcp'))return 'TCP 失败';
  if(l.includes('timeout'))return '超时';
  if(l.startsWith('status_'))return `HTTP ${l.split('_',2)[1]}`;
  if(l.includes('refused'))return '连接拒绝';
  return '请求失败';
}

function renderCards(){
  const grid=document.getElementById('grid');
  const items=Object.values(targets).filter(t=>isVisible(t.tid));
  if(!items.length){grid.innerHTML='<div style="color:var(--dim);padding:40px;grid-column:1/-1;text-align:center">暂无显示节点<br><small style="opacity:.6">点击右上角 ⊞ 选择显示哪些节点</small></div>';updateSummary();return;}
  grid.innerHTML=items.map(t=>{
    const s=st(t.status);
    const typeLabel={icmp:'ICMP',tcp:'TCP',http:'HTTP',https:'HTTPS',dns:'DNS'}[t.ping_type||'icmp']||'ICMP';
    const isPaused=t.status==='paused';
    // Failure reason translation for the latency slot when failed
    const failText=t.failure_reason?formatFailure(t.failure_reason):'请求超时';
    const latHtml=isPaused
      ?`<div class="paused-label">⏸ 已暂停</div>`
      :t.latency_ms!=null&&t.status!=='red'
        ?`<div class="lat-big" style="color:${s.color}">${t.latency_ms.toFixed(1)}<span class="lat-u">ms</span></div>`
        :t.latency_ms!=null  // failed but with measured time (e.g. wrong HTTP status)
          ?`<div class="tout" style="font-size:16px">${esc(failText)}<span style="font-size:11px;color:var(--dim);margin-left:6px">${t.latency_ms.toFixed(0)} ms</span></div>`
          :`<div class="tout">${esc(failText)}</div>`;
    // Subline: alert category, then jitter / HTTP status / type-specific hint
    const subParts=[];
    if(!isPaused){
      if(t.status==='orange'&&t.alert_category==='performance')subParts.push('<span style="color:var(--orange);font-weight:600">⚡ 性能告警</span>');
      else if((t.status==='orange'||t.status==='red')&&t.alert_category==='availability')subParts.push('<span style="color:var(--red);font-weight:600">🔴 可用性告警</span>');
      if((t.ping_type==='http'||t.ping_type==='https')&&t.status_code!=null)subParts.push(`HTTP ${t.status_code}`);
      else if((t.ping_type==='http'||t.ping_type==='https')&&!subParts.length)subParts.push('含DNS+TLS');
      if(t.ping_type==='dns'&&t.dns_domain)subParts.push(`查询: ${esc(t.dns_domain)}`);
      if(t.jitter_ms!=null)subParts.push(`抖动 ${t.jitter_ms.toFixed(1)} ms`);
    }
    const jit=isPaused?'':`<div class="lat-sub">${subParts.join('  ·  ')||'抖动 -- ms'}</div>`;
    const loss=!isPaused&&t.loss_rate>0.01?`<div class="cbadge b-loss">丢包 ${(t.loss_rate*100).toFixed(0)}%</div>`:'';
    const ok=!isPaused&&!loss&&t.status==='green'?`<div class="cbadge b-ok">连通</div>`:'';
    const warn=!isPaused&&!loss&&t.status==='orange'&&t.alert_category==='availability'?`<div class="cbadge b-warn">质量劣化</div>`:'';
    // Click binding via data-tid + event delegation (see grid click
    // listener below).  Inline onclick="openDiag('${esc(t.label)}',...)"
    // was unsafe: esc() does HTML-entity escape, but the attribute
    // body is a JS string, and the browser HTML-decodes the
    // attribute BEFORE the JS engine sees it -- so &#039; turns
    // back into ' and a label of "');alert(1);//" escapes the
    // string literal.  Switching to a data-attribute + delegated
    // listener moves the values out of the HTML-attribute -> JS
    // pipeline entirely.
    const clickAttr=isPaused?'':'data-clickable="1"';
    // id attributes on mutable parts enable patchCard() to update in-place
    // without rebuilding the entire grid on every SSE event.
    return`<div class="card ${t.status||'gray'}" id="card-${t.tid}" data-tid="${esc(t.tid)}" ${clickAttr}>
      <div class="card-top">
        <span class="sdot" id="dot-${t.tid}" style="color:${s.color}">●</span>
        <div><div class="cname" title="${esc(t.label)}">${esc(t.label)}</div><div class="cip" title="${esc(t.ip)}">${esc(t.ip)}</div>
          <span class="ctype">${typeLabel}</span></div></div>
      <div id="lat-${t.tid}">${latHtml}</div>
      <div id="sub-${t.tid}">${jit}</div>
      <div id="badges-${t.tid}">${loss}${ok}${warn}</div>
      <div id="ports-${t.tid}">${buildPortsHtml(t)}</div>
    </div>`;}).join('');
  // One delegated click listener per render; idempotent guard via
  // dataset.bound so we don't stack handlers on every SSE update.
  if(!grid.dataset.bound){
    grid.addEventListener('click',ev=>{
      const card=ev.target.closest('[data-tid][data-clickable]');
      if(!card)return;
      const tid=card.dataset.tid;
      const t=targets[tid];
      if(!t||t.status==='paused')return;
      openDiag(t.tid,t.label,t.ip);
    });
    grid.dataset.bound='1';
  }
  updateSummary();}
function buildLatHtml(t,s){
  if(t.status==='paused')return'<div class="paused-label">⏸ 已暂停</div>';
  const failText=t.failure_reason?formatFailure(t.failure_reason):'请求超时';
  if(t.latency_ms!=null&&t.status!=='red')
    return`<div class="lat-big" style="color:${s.color}">${t.latency_ms.toFixed(1)}<span class="lat-u">ms</span></div>`;
  if(t.latency_ms!=null)
    return`<div class="tout" style="font-size:16px">${esc(failText)}<span style="font-size:11px;color:var(--dim);margin-left:6px">${t.latency_ms.toFixed(0)} ms</span></div>`;
  return`<div class="tout">${esc(failText)}</div>`;
}
function buildSubHtml(t){
  if(t.status==='paused')return'';
  const parts=[];
  if(t.status==='orange'&&t.alert_category==='performance')parts.push('<span style="color:var(--orange);font-weight:600">⚡ 性能告警</span>');
  else if((t.status==='orange'||t.status==='red')&&t.alert_category==='availability')parts.push('<span style="color:var(--red);font-weight:600">🔴 可用性告警</span>');
  if((t.ping_type==='http'||t.ping_type==='https')&&t.status_code!=null)parts.push(`HTTP ${t.status_code}`);
  else if((t.ping_type==='http'||t.ping_type==='https')&&!parts.length)parts.push('含DNS+TLS');
      if(t.ping_type==='dns'&&t.dns_domain)parts.push(`查询: ${esc(t.dns_domain)}`);
  if(t.jitter_ms!=null)parts.push(`抖动 ${t.jitter_ms.toFixed(1)} ms`);
    if(t.keyword_ok===false)parts.push('<span style="color:var(--orange)">⚠ 关键字未匹配</span>');
  else if(t.keyword_ok===true&&t.status==='green')parts.push('<span style="color:var(--green)">✓ 关键字</span>');
  return`<div class="lat-sub">${parts.join('  ·  ')||'抖动 -- ms'}</div>`;
}
function buildBadgesHtml(t){
  if(t.status==='paused')return'';
  const loss=t.loss_rate>0.01?`<div class="cbadge b-loss">丢包 ${(t.loss_rate*100).toFixed(0)}%</div>`:'';
  const ok=!loss&&t.status==='green'?'<div class="cbadge b-ok">连通</div>':'';
  const warn=!loss&&t.status==='orange'&&t.alert_category==='availability'?'<div class="cbadge b-warn">质量劣化</div>':'';
    let cert='';
  if(t.cert_days_left!=null){
    if(t.cert_days_left<0)cert='<div class="cbadge b-loss">🔒 证书已过期</div>';
    else if(t.cert_days_left<=15)cert=`<div class="cbadge b-loss">🔒 证书${t.cert_days_left}天</div>`;
    else if(t.cert_days_left<=30)cert=`<div class="cbadge b-warn">🔒 证书${t.cert_days_left}天</div>`;
  }
  return loss+ok+warn+cert;
}
function buildPortsHtml(t){
  const ps=portStatus[t.tid];
  // Only TCP nodes have tcp_probe_ports; hide row if no data yet.
  if(!ps||!ps.ports||!ps.ports.length||t.ping_type!=='tcp')return '';
  // Defence in depth: if the cached snapshot belonged to a different
  // IP than the current target identity (e.g. an SSE port_status that
  // raced with an IP edit slipped past the message-level filter),
  // refuse to render it.  The cache is keyed by tid, so without this
  // check a stale entry from the OLD identity would keep showing on
  // the new card until the next scheduler cycle landed a fresh
  // snapshot.
  if(ps.ip&&t.ip&&ps.ip!==t.ip)return '';
  const pills=ps.ports.map(p=>{
    const cls=p.success?'open':'closed';
    const tip=p.success?`端口 ${p.port} 开放 (${p.latency_ms}ms)`:`端口 ${p.port} 关闭/不可达`;
    return `<span class="port-pill ${cls}" title="${tip}">${p.success?'✓':'✗'} ${p.port}</span>`;
  }).join('');
  return `<div class="ports-row">${pills}<span class="ports-ts">探测 ${ps.updated_at}</span></div>`;
}
// patchCard: update a single card DOM element in-place.
// replaces the old pattern of calling renderCards() (full grid rebuild)
// on every SSE update — which restarted CSS animations and lost in-flight
// click events every second.
function patchCard(t){
  const el=document.getElementById('card-'+t.tid);
  if(!el){renderCards();return;}
  const s=st(t.status);
  const newClass='card '+(t.status||'gray');
  if(el.className!==newClass)el.className=newClass;
  const isPaused=t.status==='paused';
  // Click eligibility is toggled via data-clickable; the delegated
  // listener installed in renderCards() reads this and looks up the
  // current target snapshot in `targets[]`.  This replaces the old
  // inline onclick string -- see the renderCards click-attr comment
  // for the security rationale (HTML-decoded labels could break out
  // of the JS string literal).
  if(isPaused)el.removeAttribute('data-clickable');
  else el.setAttribute('data-clickable','1');
  const dot=document.getElementById('dot-'+t.tid);
  if(dot)dot.style.color=s.color;
  const latEl=document.getElementById('lat-'+t.tid);
  if(latEl){const _nl=buildLatHtml(t,s);
    if(latEl.innerHTML!==_nl){latEl.innerHTML=_nl;_animateLatUpdate(t.tid);}}
  const subEl=document.getElementById('sub-'+t.tid);if(subEl)subEl.innerHTML=buildSubHtml(t);
  const badgeEl=document.getElementById('badges-'+t.tid);if(badgeEl)badgeEl.innerHTML=buildBadgesHtml(t);
  const portsEl=document.getElementById('ports-'+t.tid);if(portsEl)portsEl.innerHTML=buildPortsHtml(t);
}
function updateSummary(){
  const all=Object.values(targets),vis=all.filter(t=>isVisible(t.tid));
  const cnt={green:0,orange:0,red:0};
  vis.forEach(t=>{if(cnt[t.status]!==undefined)cnt[t.status]++;});
  document.getElementById('s-total').textContent=`共 ${vis.length}/${all.length} 个`;
  document.getElementById('s-green').textContent=`正常 ${cnt.green}`;
  document.getElementById('s-orange').textContent=`警告 ${cnt.orange}`;
  document.getElementById('s-red').textContent=`告警 ${cnt.red}`;}

let titleTimer=null;

// ── 浏览器通知（远程告警推送）───────────────────────────────────────
// 其他局域网设备打开页面后，系统告警会通过浏览器系统通知弹窗 + 网页音效推送。
// 使用 Notification API（W3C 标准），Chrome/Edge/Firefox 均支持。
let _notifEnabled = false;

function initNotifications(){
  if(!('Notification' in window))return;
  if(Notification.permission==='granted'){
    _notifEnabled=true;
  }else if(Notification.permission!=='denied'){
    // 在用户首次与页面交互后才弹权限请求（浏览器策略要求）
    document.addEventListener('click',function _ask(){
      document.removeEventListener('click',_ask);
      Notification.requestPermission().then(p=>{
        _notifEnabled=(p==='granted');
      });
    },{once:true});
  }
}

function pushNotification(title, body, icon){
  if(!_notifEnabled) return;
  try{
    const n=new Notification(title,{
      body:body,
      icon:icon||'data:image/svg+xml,<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24"><text y="18" font-size="18">🔴</text></svg>',
      tag:'netmonitor-alert',    // 同一 tag 会覆盖前一条，不堆积
      requireInteraction:false,  // 自动关闭
    });
    setTimeout(()=>n.close(),8000);
  }catch(e){}
}

function checkAlerts(){
  const reds=Object.values(targets).filter(t=>t.status==='red'&&isVisible(t.tid));
  if(reds.length){const names=reds.map(t=>t.label).join('、');
    document.getElementById('alert-banner').classList.remove('hidden');
    document.getElementById('alert-text').textContent=names+' 连接中断！';
    if(!titleTimer){let f=false;titleTimer=setInterval(()=>{document.title=f?`🔴 告警 ${names}`:'网络连通性监测';f=!f;},1000);}
  }else{document.getElementById('alert-banner').classList.add('hidden');
    if(titleTimer){clearInterval(titleTimer);titleTimer=null;document.title='网络连通性监测';}}
}
function dismissAlert(){document.getElementById('alert-banner').classList.add('hidden');
  if(titleTimer){clearInterval(titleTimer);titleTimer=null;document.title='网络连通性监测';}}

function addHist(label,ip,oldSt,newSt,timeStr){
  if((oldSt==='gray'||oldSt==='paused')&&newSt==='green')return;
  if(oldSt===newSt)return;
  hist.unshift({t:timeStr||new Date().toTimeString().slice(0,8),label,ip,oldSt,newSt});
  if(hist.length>MAX_HIST)hist.pop();renderHist();}
function renderHist(){
  const body=document.getElementById('hist-body'),hc=document.getElementById('hcount');
  if(!hist.length){body.innerHTML='<div class="hist-empty">暂无告警记录</div>';hc.textContent='';return;}
  // Count nodes whose LAST event is still orange/red (not yet recovered).
  // Raw event count inflates the number after recovery.
  const lastSt={};
  hist.forEach(h=>{lastSt[h.label+'|'+h.ip]=h.newSt;});
  const active=Object.values(lastSt).filter(s=>s==='red'||s==='orange').length;
  hc.textContent=`（${hist.length}条${active?' · ⚠ '+active+'条未恢复':''}）`;
  body.innerHTML=hist.map(h=>{const os=st(h.oldSt),ns=st(h.newSt);
    return`<div class="hrow"><span class="htime">${h.t}</span>
      <span style="color:${os.color}">●</span><span style="color:var(--dim2)">→</span>
      <span style="color:${ns.color}">●</span>
      <span class="hname">${esc(h.label)}</span><span class="hip">${esc(h.ip)}</span>
      <span class="hbadge" style="color:${ns.color};background:${ns.bg};border:1px solid ${ns.br||ns.color}">${ns.label}</span>
    </div>`;}).join('');}

// Diagnosis
function openDiag(tid,label,ip){
  // Reset any previous ICMP/DNS-tab state so opening diag on a new node
  // never inherits a running poll from the previously diagnosed one.
  stopIcmpPolling();
  _resetIcmpForm();
  stopDnsPolling();
  _resetDnsForm();
  currentDiagTid=tid;
  document.getElementById('diag-title').textContent=`🔍 诊断：${label} (${ip})`;
  document.getElementById('diag-body').innerHTML='<div style="text-align:center;padding:30px;color:var(--dim)">⏳ 正在加载路由诊断数据...</div>';
  document.getElementById('diag-updated').textContent='';
  document.getElementById('diag-overlay').classList.remove('hidden');
  // Reset to current tab on each open
  switchDiagTab('current', /*silent*/true);
  loadDiag(tid,ip);}

let _currentDiagTab = 'current';
function switchDiagTab(tab, silent){
  _currentDiagTab = tab;
  // Invalidate any in-flight request from the previous tab so it can't
  // overwrite the content of the tab we just switched to.
  if(!silent) ++_diagReqId;
  // Stop ICMP / DNS polling whenever we leave their respective tabs.
  // Active task on the server keeps running — closeDiag does not cancel
  // it either, mirroring the C-side "leave running, come back later"
  // semantic.
  if(tab!=='icmp') stopIcmpPolling();
  if(tab!=='dns')  stopDnsPolling();
  document.getElementById('tab-current').classList.toggle('active', tab==='current');
  document.getElementById('tab-history').classList.toggle('active', tab==='history');
  document.getElementById('tab-icmp')   .classList.toggle('active', tab==='icmp');
  document.getElementById('tab-dns')    .classList.toggle('active', tab==='dns');
  // Show/hide footer buttons based on tab — quick-tracert and refresh
  // only apply to the route-trace tab.
  const qBtn = document.getElementById('diag-quick-btn');
  const rBtn = document.getElementById('diag-refresh');
  if(qBtn) qBtn.style.display = tab==='current' ? '' : 'none';
  if(rBtn) rBtn.style.display = tab==='current' ? '' : 'none';
  document.getElementById('diag-updated').textContent = '';
  if(!silent && currentDiagTid){
    if(tab==='history')      loadDiagHistory(currentDiagTid);
    else if(tab==='icmp')    renderIcmpTab();
    else if(tab==='dns')     renderDnsTab();
    else{
      const t=targets[currentDiagTid];
      if(t) loadDiag(currentDiagTid, t.ip);
    }
  }
  // 3C-1.5: when leaving the DNS tab (or entering it while still in
  // single mode) make sure the modal width drops back to its default.
  _syncDnsModalWidth();
}

function loadDiagHistory(tid){
  const myId=++_diagReqId;
  document.getElementById('diag-body').innerHTML=
    '<div style="text-align:center;padding:30px;color:var(--dim)">⏳ 加载历史记录...</div>';
  fetch(`/api/traceroute/history?tid=${encodeURIComponent(tid)}&limit=5`)
    .then(r=>r.json())
    .then(rows=>{
      if(myId!==_diagReqId)return;   // tab was switched again, discard
      renderDiagHistory(rows);})
    .catch(()=>{
      if(myId!==_diagReqId)return;
      document.getElementById('diag-body').innerHTML=
        '<div style="color:var(--red);padding:20px">历史记录加载失败</div>';});}

function _traceStatus(hops){
  if(!hops||!hops.length) return {label:'失败',color:'var(--red)',bg:'rgba(239,68,68,.15)'};
  const valid=hops.filter(h=>h.ip||h.rtt_ms);
  if(!valid.length) return {label:'失败',color:'var(--red)',bg:'rgba(239,68,68,.15)'};
  let tailTimeout=0;
  for(let i=hops.length-1;i>=0;i--){
    if(!hops[i].ip&&!hops[i].rtt_ms) tailTimeout++;
    else break;
  }
  if(tailTimeout>=3) return {label:'超时',color:'var(--orange)',bg:'rgba(245,158,11,.15)'};
  if(tailTimeout===0) return {label:'正常',color:'var(--green)',bg:'rgba(34,197,94,.15)'};
  return {label:'中断',color:'#ca8a04',bg:'rgba(202,138,4,.15)'};}

function renderDiagHistory(rows){
  if(!rows||!rows.length){
    document.getElementById('diag-body').innerHTML=
      '<div style="color:var(--dim);padding:24px;text-align:center">暂无历史路由追踪记录<br>'
      +'<span style="font-size:11px">在红色告警触发时、或按设置的追踪间隔周期自动写入</span></div>';
    return;}
  let html='<div style="padding:6px 2px">';
  rows.forEach((row,idx)=>{
    const dt=new Date(row.ts*1000);
    const timeStr=dt.getFullYear()+'-'
      +String(dt.getMonth()+1).padStart(2,'0')+'-'
      +String(dt.getDate()).padStart(2,'0')+' '
      +String(dt.getHours()).padStart(2,'0')+':'
      +String(dt.getMinutes()).padStart(2,'0')+':'
      +String(dt.getSeconds()).padStart(2,'0');
    const hops=row.hops||[];
    const st=_traceStatus(hops);
    const changed=row.route_changed;
    const rowId='hist-row-'+idx;
    const chevId='hist-chev-'+idx;
    html+=`<div style="border:1px solid ${changed?'rgba(245,158,11,.4)':' var(--border)'};
           border-radius:6px;margin-bottom:6px;overflow:hidden">
      <div style="display:flex;align-items:center;gap:10px;padding:9px 14px;
           background:${changed?'rgba(245,158,11,.05)':'var(--surface)'};cursor:pointer;user-select:none"
           onclick="toggleHistRow('${rowId}','${chevId}')">
        <span style="font-size:12px;color:var(--dim);flex:1;min-width:0;white-space:nowrap">${esc(timeStr)}</span>
        <span style="font-size:11px;color:var(--dim2)">${hops.length} 跳</span>
        ${changed?`<span style="font-size:11px;font-weight:600;color:var(--orange);
               background:rgba(245,158,11,.15);border-radius:3px;padding:1px 7px">⚠ 路径变化</span>`:''}
        <span style="font-size:11px;font-weight:600;color:${st.color};
               background:${st.bg};border-radius:3px;padding:1px 7px">${st.label}</span>
        <span id="${chevId}" style="font-size:12px;color:var(--dim);transition:transform .2s;display:inline-block">›</span>
      </div>
      <div id="${rowId}" style="display:none;background:var(--bg)">`;
    if(!hops.length){
      html+='<div style="padding:12px 14px;color:var(--dim);font-size:12px">无跳点数据</div>';
    } else {
      html+='<table style="width:100%;border-collapse:collapse;font-size:11px;font-family:monospace">';
      // Find last hop that reached an IP (likely destination)
      const lastValid=hops.reduce((acc,h,i)=>(h.ip?i:acc),-1);
      hops.forEach((h,i)=>{
        const isTarget = i===lastValid && h.ip;
        const isTimeout = !h.ip && !h.rtt_ms;
        const rowStyle = isTarget
          ? 'background:rgba(34,197,94,.07);'
          : isTimeout ? 'opacity:.5;' : '';
        const ipText = h.ip
          ? `<span style="${isTarget?'color:var(--green);font-weight:600':''}">
              ${esc(h.ip)}</span>${isTarget?'<span style="color:var(--green);font-size:10px"> ✓</span>':''}`
          : `<span style="color:var(--dim2)">*</span>`;
        const rttText = h.rtt_ms!=null
          ? `<b>${Math.round(h.rtt_ms)}</b><span style="color:var(--dim2)"> ms</span>`
          : `<span style="color:var(--dim2)">timeout</span>`;
        html+=`<tr style="border-bottom:1px solid var(--border);${rowStyle}">
          <td style="padding:4px 14px;color:var(--dim2);width:40px">${h.hop}</td>
          <td style="padding:4px 8px">${ipText}</td>
          <td style="padding:4px 14px;text-align:right">${rttText}</td></tr>`;});
      html+='</table>';}
    html+='</div></div>';});
  html+='</div>';
  document.getElementById('diag-body').innerHTML=html;
  document.getElementById('diag-updated').textContent=`最近 ${rows.length} 条记录`;}

function toggleHistRow(id,chevId){
  const el=document.getElementById(id);
  const chev=document.getElementById(chevId);
  const open=el.style.display==='none';
  el.style.display=open?'block':'none';
  if(chev) chev.style.transform=open?'rotate(90deg)':'rotate(0deg)';}

function closeDiag(){
  // Stop the polling timers first to avoid races with cancel requests.
  // We do NOT cancel a running task on close — that mirrors C-side
  // behaviour where closing the modal leaves the run going so the user
  // can come back and see results. Closing the browser tab is a true
  // abandonment but the server-side TTL + global semaphore handle that.
  stopIcmpPolling();
  stopDnsPolling();
  document.getElementById('diag-overlay').classList.add('hidden');
  currentDiagTid=null;
}
function openQuickTrace(){
  if(!currentDiagTid)return;
  const t=targets[currentDiagTid];if(!t)return;
  const url='/traceroute/window?tid='+encodeURIComponent(currentDiagTid);
  window.open(url,'_blank',
    'width=620,height=700,resizable=yes,scrollbars=yes,menubar=no,toolbar=no,location=no');
}
function _pollDiag(tid, ip, myId, attempts){
  if(myId!==_diagReqId)return;
  fetch(`/api/traceroute?tid=${encodeURIComponent(tid)}`)
    .then(r=>r.json())
    .then(data=>{
      if(myId!==_diagReqId)return;
      if(data.status==='pending'){
        if(attempts>=30){
          document.getElementById('diag-body').innerHTML=
            '<div style="color:var(--red);padding:20px">⚠ 诊断超时，请稍后手动刷新。</div>';
          return;
        }
        const dots='.'.repeat((attempts%3)+1);
        document.getElementById('diag-updated').textContent=
          `正在运行路由追踪${dots}（约 10~30 秒）`;
        setTimeout(()=>_pollDiag(tid,ip,myId,attempts+1),5000);
      } else {
        renderDiag(data,ip);
      }
    })
    .catch(e=>{if(myId!==_diagReqId)return;
      document.getElementById('diag-body').innerHTML=
        `<div style="color:var(--red);padding:20px">诊断失败：${esc(e.message)}</div>`;});}
function refreshDiag(){
  if(!currentDiagTid)return;
  const t=targets[currentDiagTid];if(!t)return;
  const myId=++_diagReqId;
  document.getElementById('diag-body').innerHTML=
    '<div style="text-align:center;padding:30px;color:var(--dim)">🔄 正在重新运行路由追踪...</div>';
  document.getElementById('diag-updated').textContent='等待后台调度...';
  fetch(`/api/traceroute?tid=${encodeURIComponent(currentDiagTid)}&force=1`)
    .then(r=>r.json())
    .then(data=>{
      if(myId!==_diagReqId)return;
      if(data.status==='pending'){
        setTimeout(()=>_pollDiag(currentDiagTid,t.ip,myId,1),5000);
      } else {
        renderDiag(data,t.ip);
      }
    })
    .catch(e=>{if(myId!==_diagReqId)return;
      document.getElementById('diag-body').innerHTML=
        `<div style="color:var(--red);padding:20px">诊断失败：${esc(e.message)}</div>`;});}
function loadDiag(tid,ip){
  const myId=++_diagReqId;
  fetch(`/api/traceroute?tid=${encodeURIComponent(tid)}`)
    .then(r=>r.json())
    .then(data=>{
      if(myId!==_diagReqId)return;
      if(data.status==='pending'){
        document.getElementById('diag-updated').textContent='后台诊断运行中...';
        setTimeout(()=>_pollDiag(tid,ip,myId,1),5000);
      } else {
        renderDiag(data,ip);
      }
    })
    .catch(e=>{if(myId!==_diagReqId)return;
      document.getElementById('diag-body').innerHTML=
        `<div style="color:var(--red);padding:20px">诊断失败：${esc(e.message)}</div>`;});}

function renderDiag(data,ip){
  const hops=data.hops||[];
  const upd=data.updated_at||'';
  if(upd)document.getElementById('diag-updated').textContent=`缓存自 ${upd}`;

  // Show error if execution failed
  if(hops.length===1&&hops[0].error){
    document.getElementById('diag-body').innerHTML=
      `<div style="padding:16px;border-radius:8px;background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3)">
        <div style="color:var(--red);font-weight:700;margin-bottom:8px">⚠ 路由追踪执行失败</div>
        <div style="font-size:12px;color:var(--dim);font-family:monospace;word-break:break-all">${esc(hops[0].error)}</div>
        <div style="font-size:12px;color:var(--dim);margin-top:8px">
          常见原因：tracert.exe 路径问题、权限问题，或路由追踪被防火墙阻止。<br>
          请确认可以在命令提示符中运行：tracert ${esc(data.resolve_ip||ip)}
        </div>
      </div>`;
    return;
  }

  // ICMP-blocked detection
  const hasBreak=hops.some(h=>h.status==='break'||h.status==='after');
  const tcpOk=data.tcp_checks&&data.tcp_checks.some(c=>c.success);
  const icmpBlocked=hasBreak&&tcpOk;

  // Build topology
  let html='<div class="topo-wrap">';
  html+=`<div class="topo-node you"><div class="node-circle">💻</div>
    <div class="node-info"><div class="node-ip">本机</div></div></div>`;

  hops.forEach(h=>{
    const cls=h.status||'ok';
    const addr=h.ip?(h.hostname&&h.hostname!==h.ip?`${h.hostname} [${h.ip}]`:h.ip):'*';
    const rtt=h.rtt_ms?`${h.rtt_ms} ms`:'*';
    const tag=(cls==='filtered')?'<span class="node-tag">ICMP过滤</span>':
               (cls==='break')?'<span class="node-tag">断路点</span>':'';
    html+=`<div class="topo-connector"><div class="connector-line"></div><div class="connector-arrow">▼</div></div>
    <div class="topo-node ${cls}">
      <div class="node-circle">${h.hop}</div>
      <div class="node-info"><div class="node-ip">${esc(addr)}</div></div>
      <div class="node-rtt">${rtt}</div>${tag}
    </div>`;});

  html+=`<div class="topo-connector"><div class="connector-line"></div><div class="connector-arrow">▼</div></div>
  <div class="topo-node destination">
    <div class="node-circle">🎯</div>
    <div class="node-info"><div class="node-ip">目标：${esc(data.resolve_ip||ip)}</div>
      ${data.ip&&data.ip!==data.resolve_ip?`<div class="node-host">原始：${esc(data.ip)}</div>`:''}
    </div>
  </div></div>`;

  // ICMP blocked note
  if(icmpBlocked)html+=`<div style="margin-top:12px;padding:10px 14px;border-radius:8px;
    background:rgba(34,197,94,.08);border:1px solid rgba(34,197,94,.3);font-size:12px">
    <span style="color:var(--green);font-weight:700">✓ 目标过滤了 ICMP，但 TCP 端口检测正常</span><br>
    <span style="color:var(--dim)">traceroute 显示"断路"是因为目标设备屏蔽了 ICMP 探测包，实际服务正常运行。
    这在 DNS 服务器、CDN 节点、部分企业防火墙中非常常见。</span></div>`;

  // HTTP timings breakdown (last successful probe)
  const t=targets[currentDiagTid];
  if(t&&t.timings&&(t.ping_type==='http'||t.ping_type==='https')){
    const tm=t.timings;
      // Certificate expiry notice — use t (already fetched above) not
      // data.target which does not exist in the traceroute API response.
  if(t&&t.cert_days_left!=null){
    const d=t.cert_days_left;
    const color=d<0||d<=15?'var(--red)':d<=30?'var(--orange)':'var(--green)';
    const msg=d<0?'🔒 证书已过期！':d<=30?`🔒 证书即将到期：还剩 <b style="color:${color}">${d} 天</b>，请尽快续签`:`🔒 证书有效期剩余 <b style="color:${color}">${d} 天</b>`;
    const bg=d<=15?'rgba(239,68,68,.08)':d<=30?'rgba(245,158,11,.08)':'rgba(34,197,94,.06)';
    html+=`<div style="padding:8px 12px;border-radius:6px;margin-bottom:8px;font-size:12px;background:${bg}">${msg}</div>`;
  }
html+=`<div class="diag-section"><div class="diag-title">⏱ HTTP 时间分解（上次成功探测）</div>`;
    const rows=[
      ['DNS 解析',         tm.dns_ms,   '主机名查询时间'],
      ['TCP 握手',         tm.tcp_ms,   '三次握手建连'],
      ['TLS 握手',         tm.tls_ms,   'HTTPS 加密协商（仅 HTTPS）'],
      ['TTFB（服务器处理）', tm.ttfb_ms,  '发送请求 → 收到首字节'],
      ['总耗时',           tm.total_ms, ''],
    ];
    rows.forEach(([lbl,v,note])=>{
      if(v==null)return;
      // colour-code: long phases are reddish
      let c='var(--dim)';
      if(v>1000)c='var(--red)';
      else if(v>300)c='var(--orange)';
      else if(v>0)c='var(--green)';
      html+=`<div class="diag-row">
        <span style="min-width:140px">${lbl}</span>
        <span style="color:${c};font-weight:600;font-family:monospace">${v.toFixed(1)} ms</span>
        <span style="color:var(--dim);font-size:10px">${note}</span>
      </div>`;
    });
    // Diagnostic hint based on timings
    if(tm.tcp_ms>300||tm.tls_ms>500)
      html+=`<div style="margin-top:8px;font-size:11px;color:var(--orange)">💡 TCP/TLS 较慢 → 倾向于网络层问题（路由、丢包、距离远）</div>`;
    if(tm.ttfb_ms>1000)
      html+=`<div style="margin-top:8px;font-size:11px;color:var(--orange)">💡 TTFB 较高 → 倾向于服务器处理慢（应用层瓶颈）</div>`;
    html+=`</div>`;
  }

  // DNS
  if(data.dns){const dns=data.dns;
    html+=`<div class="diag-section"><div class="diag-title">🌐 DNS 解析（追踪用）</div>`;
    if(dns.success)html+=`<div class="diag-row"><span class="d-ok">✓</span> 解析成功 → ${esc(dns.ips.join(', '))}</div>`;
    else html+=`<div class="diag-row"><span class="d-fail">✕</span> 解析失败：${esc(dns.error||'')}</div>`;
    html+=`</div>`;}

  // DNS monitor specific section (only for dns ping_type cards)
  {const t=targets[currentDiagTid];
  if(t&&t.ping_type==='dns'){
    html+=`<div class="diag-section"><div class="diag-title">🔍 DNS 监控查询结果</div>`;
    html+=`<div class="diag-row"><span style="min-width:90px;color:var(--dim)">DNS 服务器</span>
      <span style="font-family:monospace" class="d-info">${esc(t.ip)}</span></div>`;
    if(t.dns_domain)
      html+=`<div class="diag-row"><span style="min-width:90px;color:var(--dim)">查询域名</span>
        <span style="font-family:monospace">${esc(t.dns_domain)}</span></div>`;
    const fr=(t.failure_reason||'');
    const fl=fr.toLowerCase();
    if(t.status==='green'||t.status==='orange'){
      html+=`<div class="diag-row"><span class="d-ok">✓</span> 解析正常`;
      if(t.latency_ms!=null)html+=` <span style="color:var(--green);font-family:monospace">${t.latency_ms} ms</span>`;
      html+=`</div>`;
    }else if(fl.startsWith('dns_hijack:')){
      const gotIPs=fr.split(':')[1]||'';
      html+=`<div class="diag-row"><span class="d-fail">✕</span>
        <b style="color:var(--red)">⚠ 检测到 DNS 劫持！</b></div>`;
      html+=`<div class="diag-row"><span style="min-width:90px;color:var(--dim)">实际解析到</span>
        <span style="font-family:monospace;color:var(--red)">${esc(gotIPs)}</span>
        <span style="font-size:10px;color:var(--dim);margin-left:8px">（与预期 IP 不符）</span></div>`;
      html+=`<div style="margin-top:6px;padding:8px 10px;border-radius:6px;font-size:11px;
        background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.25);color:var(--dim)">
        可能原因：ISP 劫持、DNS 投毒、DNS 服务器被篡改，或预期 IP 配置有误。
        建议用可信 DNS（如 8.8.8.8）交叉验证该域名的真实解析结果。
      </div>`;
    }else if(fl==='nxdomain'){
      html+=`<div class="diag-row"><span class="d-fail">✕</span> 域名不存在 (NXDOMAIN) — 该域名在 DNS 中无记录</div>`;
    }else if(fl==='dns_no_a_record'){
      html+=`<div class="diag-row"><span class="d-fail">✕</span> 无 A 记录返回 — DNS 服务器响应但无 IPv4 地址</div>`;
    }else if(fl==='dns_timeout'){
      html+=`<div class="diag-row"><span class="d-fail">✕</span> DNS 查询超时 — DNS 服务器未在规定时间内响应</div>`;
    }else if(fl==='dns_no_domain'){
      html+=`<div class="diag-row"><span style="color:var(--orange)">⚠</span> 未配置查询域名 — 请在节点设置中填写 dns_domain</div>`;
    }else if(fr){
      html+=`<div class="diag-row"><span class="d-fail">✕</span> ${esc(formatFailure(fr))}</div>`;
    }
    html+=`</div>`;
  }}
  if(data.tcp_checks&&data.tcp_checks.length){
    html+=`<div class="diag-section"><div class="diag-title">🔌 TCP 端口检测</div>`;
    data.tcp_checks.forEach(c=>{
      const icon=c.success?'<span class="d-ok">✓</span>':'<span class="d-fail">✕</span>';
      const lat=c.latency_ms?` ${c.latency_ms}ms`:'';
      html+=`<div class="diag-row">${icon} 端口 ${c.port} ${c.success?'开放'+lat:'关闭/不可达'}</div>`;});
    html+=`</div>`;}

  html+=`<div style="margin-top:14px;font-size:11px;color:var(--dim);line-height:2">
    <span style="color:var(--green)">●</span> 正常 &nbsp;
    <span style="color:var(--orange)">●</span> ICMP过滤（路由正常）&nbsp;
    <span style="color:var(--red)">●</span> 断路点
  </div>`;
  document.getElementById('diag-body').innerHTML=html;}

// ICMP Advanced Diagnostic Tab (Task B)
// =====================================
// Form state survives tab switches inside the same modal open. On modal
// open we reset to defaults via _resetIcmpForm() so the previous node's
// settings don't bleed into the new node.
const _icmpForm = {
  mode:         'single',
  payload_size: 32,
  df:           false,
  count:        4,
  timeout_s:    2.0,
  interval_s:   1.0,
};
let _icmpTaskId    = null;
let _icmpPollTimer = null;
// Human-readable error code map — mirrors C-side _ERROR_LABELS.
const _ICMP_ERR_LABEL = {
  ok:                  '成功',
  timeout:             '超时',
  frag_needed:         '需要分片（DF 阻止）',
  host_unreachable:    '目标不可达',
  network_unreachable: '网络不可达',
  permission_denied:   '权限不足',
  param_invalid:       '参数无效',
  tool_unavailable:    'ping 命令不可用',
  unknown_error:       '未知错误',
};
const _ICMP_PAYLOAD_PRESETS = [0,32,256,512,1200,1400,1472];
const _ICMP_MODE_LABEL = {single:'单次诊断', stability:'稳定性测试', pmtu:'PMTU 探测'};

function _resetIcmpForm(){
  _icmpForm.mode='single'; _icmpForm.payload_size=32;
  _icmpForm.df=false; _icmpForm.count=4;
  _icmpForm.timeout_s=2.0; _icmpForm.interval_s=1.0;
  _icmpTaskId=null;
}

function _icmpHost(){
  const t = currentDiagTid ? targets[currentDiagTid] : null;
  return t ? (t.ip||'') : '';
}

// Build and inject the ICMP tab DOM into diag-body. Called whenever the
// user switches to the ICMP tab — re-render is cheap and lets us refresh
// the displayed host if it changes (it does not today, but keeps the
// invariant simple).
function renderIcmpTab(){
  const host=_icmpHost();
  const m=_icmpForm.mode;
  const isPmtu=(m==='pmtu');
  const isStab=(m==='stability');
  const presetBtns=_ICMP_PAYLOAD_PRESETS.map(v=>
    `<button class="icmp-preset" onclick="setIcmpPayload(${v})">${v}</button>`).join('');
  let html=`
  <div class="icmp-form">
    <div class="icmp-row">
      <label>目标</label>
      <span style="font-family:monospace;color:var(--text)">${esc(host||'(无)')}</span>
    </div>
    <div class="icmp-row">
      <label>模式</label>
      <button class="icmp-pill ${m==='single'?'active':''}" onclick="setIcmpMode('single')">单次诊断</button>
      <button class="icmp-pill ${m==='stability'?'active':''}" onclick="setIcmpMode('stability')">稳定性测试</button>
      <button class="icmp-pill ${m==='pmtu'?'active':''}" onclick="setIcmpMode('pmtu')">PMTU 探测</button>
    </div>`;
  if(!isPmtu){
    html+=`
    <div class="icmp-row" id="icmp-row-payload">
      <label>载荷 (B)</label>
      <input class="icmp-input" id="icmp-payload" type="number" min="0" max="9000"
             value="${_icmpForm.payload_size}" oninput="onIcmpPayloadInput()">
      <span style="font-size:11px;color:var(--dim);margin-right:4px">预设:</span>
      ${presetBtns}
    </div>
    <div class="icmp-row" id="icmp-row-df">
      <label>禁止分片</label>
      <button class="icmp-switch ${_icmpForm.df?'on':''}" id="icmp-df-sw"
              onclick="toggleIcmpDf()" title="DF: 路径上需要分片时由设备返回错误"></button>
      <span style="font-size:11px;color:var(--dim)">DF</span>
      <label style="margin-left:16px">次数</label>
      <input class="icmp-input" id="icmp-count" type="number" min="1" max="50"
             value="${_icmpForm.count}" oninput="_icmpForm.count=parseInt(this.value)||1">
      <label style="margin-left:16px">超时(秒)</label>
      <input class="icmp-input" id="icmp-timeout" type="number" min="0.2" max="10" step="0.1"
             value="${_icmpForm.timeout_s}" oninput="_icmpForm.timeout_s=parseFloat(this.value)||2.0">
      ${isStab?`<label style="margin-left:16px">间隔(秒)</label>
      <input class="icmp-input" id="icmp-interval" type="number" min="0.2" max="5" step="0.1"
             value="${_icmpForm.interval_s}" oninput="_icmpForm.interval_s=parseFloat(this.value)||1.0">`:''}
    </div>
    ${(_icmpForm.df||_icmpForm.payload_size>1000)?`<div class="icmp-warn">⚠ 大包/DF 为高级测试，请按需手动执行。</div>`:''}`;
  } else {
    html+=`
    <div class="icmp-row">
      <label>超时(秒)</label>
      <input class="icmp-input" id="icmp-timeout" type="number" min="0.2" max="10" step="0.1"
             value="${_icmpForm.timeout_s}" oninput="_icmpForm.timeout_s=parseFloat(this.value)||2.0">
      <span style="font-size:11px;color:var(--dim);margin-left:8px">
        PMTU 模式将通过二分搜索自动测试不同 payload (DF=on)
      </span>
    </div>`;
  }
  const running=(_icmpTaskId!==null);
  html+=`
    <div class="icmp-actions">
      <button class="btn btn-primary" id="icmp-start" onclick="startIcmpDiag()"
              ${running?'disabled style="opacity:.5;cursor:not-allowed"':''}>▶ 开始诊断</button>
      <button class="btn btn-ghost" id="icmp-stop" onclick="stopIcmpDiag()"
              ${running?'':'disabled style="opacity:.5;cursor:not-allowed"'}>■ 停止</button>
      <button class="btn btn-ghost" onclick="loadIcmpHistory()">📜 历史</button>
    </div>
  </div>
  <div id="icmp-output"></div>`;
  document.getElementById('diag-body').innerHTML=html;
  // Re-attach polling renderer immediately if a task is in flight (e.g.
  // user switched away and back during a run). The polling timer itself
  // is still alive; we only need to re-render whatever state it last saw.
  if(running){ pollIcmpStatus(); }
}

function setIcmpMode(m){ _icmpForm.mode=m; renderIcmpTab(); }
function setIcmpPayload(v){
  _icmpForm.payload_size=Math.max(0, Math.min(9000, v|0));
  // Only re-render if the warning toggles (crossing the 1000B threshold)
  renderIcmpTab();
}
function onIcmpPayloadInput(){
  const el=document.getElementById('icmp-payload'); if(!el) return;
  const v=parseInt(el.value); if(!isNaN(v)) _icmpForm.payload_size=Math.max(0,Math.min(9000,v));
  // Avoid full re-render on every keystroke — only update warning
  const warns=document.querySelectorAll('.icmp-warn');
  const need=(_icmpForm.df||_icmpForm.payload_size>1000);
  if(need&&!warns.length) renderIcmpTab();
  else if(!need&&warns.length) renderIcmpTab();
}
function toggleIcmpDf(){
  _icmpForm.df=!_icmpForm.df;
  const sw=document.getElementById('icmp-df-sw');
  if(sw) sw.classList.toggle('on', _icmpForm.df);
  // Re-evaluate warning visibility
  const warns=document.querySelectorAll('.icmp-warn');
  const need=(_icmpForm.df||_icmpForm.payload_size>1000);
  if(need!==warns.length>0) renderIcmpTab();
}

// Start / stop / poll
// ---------------------------------------------------------------------
function startIcmpDiag(){
  if(!currentDiagTid){
    _showIcmpError('未选中节点'); return;
  }
  const host=_icmpHost();
  if(!host){ _showIcmpError('节点没有有效的目标地址'); return; }
  // Defensive: stop any leftover poll from a previous task before we
  // fire a new one. Server will 409 if we're somehow already running.
  stopIcmpPolling();
  const body={
    target_id:    currentDiagTid,
    host:         host,
    mode:         _icmpForm.mode,
    payload_size: _icmpForm.payload_size,
    df:           _icmpForm.df,
    count:        _icmpForm.count,
    timeout_s:    _icmpForm.timeout_s,
    interval_s:   _icmpForm.interval_s,
  };
  document.getElementById('icmp-output').innerHTML=
    '<div style="padding:10px;color:var(--dim);font-size:12px">⏳ 正在启动诊断...</div>';
  fetch('/api/icmp_diag/start',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify(body),
  }).then(async r=>{
    const data=await r.json().catch(()=>({}));
    if(r.status===200 && data.task_id){
      _icmpTaskId=data.task_id;
      // Re-render so Start/Stop button states flip.
      renderIcmpTab();
      // Kick off polling: first call right now (instant feedback) then
      // every 1s. The renderer cancels the timer when the task hits a
      // terminal state (done / error / cancelled).
      pollIcmpStatus();
      _icmpPollTimer=setInterval(pollIcmpStatus, 1000);
    } else {
      const msg=data.error||`HTTP ${r.status}`;
      _showIcmpError(msg);
    }
  }).catch(e=>_showIcmpError('请求失败: '+(e&&e.message||e)));
}

function stopIcmpDiag(){
  if(!_icmpTaskId) return;
  const tid=_icmpTaskId;
  fetch('/api/icmp_diag/cancel',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({task_id:tid}),
  }).catch(()=>{});
  // Let the next poll see status=cancelled and render the final state.
}

function stopIcmpPolling(){
  if(_icmpPollTimer){ clearInterval(_icmpPollTimer); _icmpPollTimer=null; }
}

function pollIcmpStatus(){
  if(!_icmpTaskId) return;
  const tid=_icmpTaskId;
  fetch('/api/icmp_diag/status?task_id='+encodeURIComponent(tid))
    .then(r=>r.json())
    .then(data=>{
      // Race: user clicked stop / opened a new task / closed modal
      // between our request and response — discard stale data.
      if(tid!==_icmpTaskId) return;
      renderIcmpProgress(data);
      if(['done','error','cancelled'].includes(data.status)){
        stopIcmpPolling();
        // Cache the final output HTML, re-render the form to flip
        // button states, then restore the output block.
        const out=document.getElementById('icmp-output');
        const cached = out ? out.innerHTML : '';
        _icmpTaskId=null;
        renderIcmpTab();
        const newOut=document.getElementById('icmp-output');
        if(newOut && cached) newOut.innerHTML=cached;
      }
    })
    .catch(()=>{ /* transient network blip — keep polling */ });
}

function _showIcmpError(msg){
  const out=document.getElementById('icmp-output');
  if(!out) return;
  out.innerHTML=
    `<div class="icmp-conclusion bad" style="margin-top:10px">⚠ ${esc(msg)}</div>`;
}

// Progress / conclusion rendering
// ---------------------------------------------------------------------
function _icmpErrLabel(code){ return _ICMP_ERR_LABEL[code] || code; }

function renderIcmpProgress(data){
  const out=document.getElementById('icmp-output');
  if(!out) return;
  const mode=data.mode||'single';
  const prog=data.progress||{};
  let html='';
  if(mode==='pmtu'){
    const steps=prog.pmtu_steps||[];
    const cnt=steps.length;
    html+=`<div style="margin-top:12px;font-size:12px;color:var(--dim);
           display:flex;justify-content:space-between">
      <span>已测试 ${cnt} 个 payload</span>
      <span>${data.status==='running'?'探测中...':''}</span>
    </div>`;
    html+='<div class="icmp-progress">';
    steps.forEach((s,i)=>{
      const ic=s.success?'<span class="ic-ok">✓</span>':'<span class="ic-bad">✗</span>';
      const lbl=s.success?'通过':_icmpErrLabel(s.error_code);
      html+=`<div class="row">
        <span class="lbl">测试 #${i+1}</span>
        ${ic}
        <span>payload = <b>${s.payload}</b> B</span>
        <span style="color:var(--dim);margin-left:auto">${esc(lbl)}</span>
      </div>`;
    });
    if(!steps.length){
      html+='<div style="padding:10px;color:var(--dim2);text-align:center">等待首次探测...</div>';
    }
    html+='</div>';
  } else {
    const samples=prog.samples||[];
    const completed=prog.completed||0;
    const total=prog.total||0;
    html+=`<div style="margin-top:12px;font-size:12px;color:var(--dim);
           display:flex;justify-content:space-between">
      <span>进度 ${completed} / ${total||'?'}</span>
      <span>${data.status==='running'?'探测中...':''}</span>
    </div>`;
    html+='<div class="icmp-progress">';
    samples.forEach(s=>{
      const ic=s.success?'<span class="ic-ok">✓</span>':'<span class="ic-bad">✗</span>';
      const rtt=s.success && s.rtt_ms!=null
        ? `<b>${s.rtt_ms}</b> ms`
        : `<span style="color:var(--dim)">${esc(_icmpErrLabel(s.error_code))}</span>`;
      html+=`<div class="row">
        <span class="lbl">探测 ${s.seq} / ${total||'?'}</span>
        ${ic}
        <span>${rtt}</span>
      </div>`;
    });
    if(!samples.length){
      html+='<div style="padding:10px;color:var(--dim2);text-align:center">等待首次探测...</div>';
    }
    html+='</div>';
  }
  if(data.status==='done' && data.result){
    html+=_renderIcmpConclusion(data.result);
  } else if(data.status==='error'){
    html+=`<div class="icmp-conclusion bad">⚠ 诊断失败：${esc(data.error||'未知错误')}</div>`;
  } else if(data.status==='cancelled'){
    html+=`<div class="icmp-conclusion warn">■ 诊断已被取消</div>`;
  }
  out.innerHTML=html;
}

function _renderIcmpConclusion(r){
  // Visual flavour from success ratio: green / orange / red.
  const total=r.count||0, ok=r.success_count||0;
  let cls='ok';
  if(total===0 || ok===0) cls='bad';
  else if(ok<total)       cls='warn';
  let stats='';
  if(r.mode==='pmtu'){
    const mtu=r.estimated_path_mtu;
    const mp=r.max_df_payload;
    stats=`<div class="icmp-stats">
      <span>最大成功 payload: <b>${mp!=null?mp+' B':'—'}</b></span>
      <span>估算路径 MTU: <b>${mtu!=null?mtu+' B':'—'}</b></span>
      <span>探测次数: <b>${total}</b></span>
    </div>`;
  } else {
    const avg=r.avg_rtt_ms!=null?r.avg_rtt_ms+' ms':'—';
    const mn=r.min_rtt_ms!=null?r.min_rtt_ms+' ms':'—';
    const mx=r.max_rtt_ms!=null?r.max_rtt_ms+' ms':'—';
    const loss=total?Math.round((total-ok)/total*100):0;
    stats=`<div class="icmp-stats">
      <span>成功: <b>${ok}/${total}</b></span>
      <span>丢包率: <b>${loss}%</b></span>
      <span>平均: <b>${avg}</b></span>
      <span>最小: <b>${mn}</b></span>
      <span>最大: <b>${mx}</b></span>
    </div>`;
  }
  return `<div class="icmp-conclusion ${cls}">
    <div style="font-weight:600;margin-bottom:4px">📋 结论</div>
    <div>${esc(r.conclusion||'')}</div>
    ${stats}
  </div>`;
}

function loadIcmpHistory(){
  if(!currentDiagTid){ _showIcmpError('未选中节点'); return; }
  const out=document.getElementById('icmp-output');
  if(out) out.innerHTML=
    '<div style="padding:14px;color:var(--dim);text-align:center">⏳ 加载历史...</div>';
  fetch('/api/icmp_diag/history?tid='+encodeURIComponent(currentDiagTid)+'&limit=50')
    .then(r=>r.json())
    .then(rows=>renderIcmpHistory(rows||[]))
    .catch(e=>{
      if(out) out.innerHTML=
        `<div class="icmp-conclusion bad">历史加载失败：${esc(e&&e.message||e)}</div>`;
    });
}

function renderIcmpHistory(rows){
  const out=document.getElementById('icmp-output');
  if(!out) return;
  if(!rows.length){
    out.innerHTML=
      '<div style="padding:20px;color:var(--dim);text-align:center;font-size:12px">'+
      '暂无 ICMP 诊断历史记录<br><span style="font-size:11px">'+
      '运行一次诊断即可在此查看历史</span></div>';
    return;
  }
  let html=`<div style="margin-top:10px;font-size:12px;color:var(--dim)">
    最近 ${rows.length} 条诊断记录（含 C 端与 Web 端）
  </div>`;
  rows.forEach((row,idx)=>{ html += _renderIcmpHistRow(row, idx); });
  out.innerHTML=html;
}

function _renderIcmpHistRow(row, idx){
  const dt=new Date((row.created_ts||0)*1000);
  const timeStr=dt.getFullYear()+'-'
    +String(dt.getMonth()+1).padStart(2,'0')+'-'
    +String(dt.getDate()).padStart(2,'0')+' '
    +String(dt.getHours()).padStart(2,'0')+':'
    +String(dt.getMinutes()).padStart(2,'0')+':'
    +String(dt.getSeconds()).padStart(2,'0');
  const mode=_ICMP_MODE_LABEL[row.mode]||row.mode;
  const total=row.count||0, ok=row.success_count||0;
  const lossPct=total?Math.round((total-ok)/total*100):0;
  let stCol='var(--green)', stBg='rgba(34,197,94,.15)', stTxt='正常';
  if(total===0||ok===0){ stCol='var(--red)'; stBg='rgba(239,68,68,.15)'; stTxt='失败'; }
  else if(ok<total){ stCol='var(--orange)'; stBg='rgba(245,158,11,.15)'; stTxt='不稳定'; }
  const bodyId='icmp-hist-body-'+idx;
  const chevId='icmp-hist-chev-'+idx;
  const dfStr=row.df?' DF':'';
  const pStr=row.mode==='pmtu'?'':` payload=${row.payload_size}B${dfStr}`;
  // See _renderDnsHistRow for the layout rationale — same pattern:
  // natural-width timestamp + flex-grow status pill (via margin-left:
  // auto) keeps the right group right-aligned without ever forcing a
  // text overflow into a neighbouring chip.
  return `<div class="icmp-hist-item">
    <div class="icmp-hist-head" onclick="toggleIcmpHist('${bodyId}','${chevId}')">
      <span style="font-size:12px;color:var(--dim);white-space:nowrap;
            flex:0 0 auto">${esc(timeStr)}</span>
      <span class="icmp-hist-mode">${esc(mode)}</span>
      <span style="font-size:11px;color:var(--dim);
            overflow:hidden;text-overflow:ellipsis;white-space:nowrap;
            flex:1 1 auto;min-width:0">${esc(pStr)}</span>
      <span style="font-size:11px;font-weight:600;color:${stCol};
            background:${stBg};border-radius:3px;padding:1px 7px;
            margin-left:auto;flex:0 0 auto;white-space:nowrap">${stTxt}</span>
      <span style="font-size:11px;color:var(--dim);font-family:monospace;
            flex:0 0 auto;white-space:nowrap">${ok}/${total} · ${lossPct}%</span>
      <span id="${chevId}" style="font-size:12px;color:var(--dim);
            transition:transform .2s;display:inline-block;flex:0 0 auto">›</span>
    </div>
    <div id="${bodyId}" class="icmp-hist-body">${_renderIcmpHistDetail(row)}</div>
  </div>`;
}

function _renderIcmpHistDetail(row){
  let html='';
  if(row.conclusion){
    html+=`<div style="font-size:12px;color:var(--text);line-height:1.7;margin-bottom:8px">
      <b style="color:var(--dim);font-weight:600">结论：</b>${esc(row.conclusion)}
    </div>`;
  }
  if(row.mode==='pmtu'){
    html+=`<div class="icmp-stats">
      <span>最大成功 payload: <b>${row.max_df_payload!=null?row.max_df_payload+' B':'—'}</b></span>
      <span>估算路径 MTU: <b>${row.estimated_path_mtu!=null?row.estimated_path_mtu+' B':'—'}</b></span>
      <span>探测次数: <b>${row.count||0}</b></span>
    </div>`;
  } else {
    const total=row.count||0, ok=row.success_count||0;
    const lossPct=total?Math.round((total-ok)/total*100):0;
    html+=`<div class="icmp-stats">
      <span>成功: <b>${ok}/${total}</b></span>
      <span>丢包率: <b>${lossPct}%</b></span>
      <span>平均: <b>${row.avg_rtt_ms!=null?row.avg_rtt_ms+' ms':'—'}</b></span>
      <span>最小: <b>${row.min_rtt_ms!=null?row.min_rtt_ms+' ms':'—'}</b></span>
      <span>最大: <b>${row.max_rtt_ms!=null?row.max_rtt_ms+' ms':'—'}</b></span>
    </div>`;
  }
  const samples=(row.details&&row.details.samples)||[];
  if(samples.length){
    html+='<div class="icmp-progress" style="max-height:180px;margin-top:10px">';
    samples.forEach(s=>{
      const ok2=s.ok!==undefined?s.ok:s.success;
      const rtt=s.rtt!==undefined?s.rtt:s.rtt_ms;
      const err=s.err||s.error_code||'';
      const ic=ok2?'<span class="ic-ok">✓</span>':'<span class="ic-bad">✗</span>';
      const txt=ok2 && rtt!=null
        ? `<b>${rtt}</b> ms`
        : `<span style="color:var(--dim)">${esc(_icmpErrLabel(err))}</span>`;
      html+=`<div class="row">
        <span class="lbl">探测 ${s.seq}</span>${ic}<span>${txt}</span>
      </div>`;
    });
    html+='</div>';
  }
  return html;
}

function toggleIcmpHist(bodyId,chevId){
  const el=document.getElementById(bodyId);
  const chev=document.getElementById(chevId);
  if(!el) return;
  const open=(el.style.display!=='block');
  el.style.display=open?'block':'none';
  if(chev) chev.style.transform=open?'rotate(90deg)':'rotate(0deg)';
}

// (icmp history block end)

// DNS Advanced Diagnostic Tab (Phase 2)
// =====================================
// Mirrors the ICMP tab in shape (form → start → polled status → render)
// but with three deliberate simplifications:
//   • No incremental progress — DNS is a single dnspython call; we poll
//     /status at 500ms (slightly faster than ICMP's 1s) and render the
//     full result the moment status flips to done|cancelled|error.
//   • No history button — DNS history table is Phase 3.
//   • Mode is just qtype (A/AAAA/CNAME/MX/NS) instead of single/stab/pmtu.
const _dnsForm = {
  domain:        '',
  qtype:         'A',
  timeout_s:     2,
  trace_chain:   true,
  // Phase 3B: optional custom resolver — only consulted in single mode.
  //   resolver_mode = 'system' → ignore nameserver (always send null)
  //   resolver_mode = 'custom' → send nameserver verbatim after strip
  resolver_mode: 'system',
  nameserver:    '',
  // Phase 3C-1: top-level query mode.  When set to 'compare' the
  // resolver-mode segmented control is hidden; the chip list below
  // becomes the canonical input.  The operator's last-typed values
  // are preserved across mode flips so toggling does not destroy state.
  query_mode:    'single',
  // Phase 3C-3: multi-resolver chip list for compare mode.  At least 1,
  // at most _DNS_COMPARE_MAX (3) custom IPs.  Order matters — it
  // determines diff-matrix column order.  De-dup / IP-literal validation
  // happens in addDnsCompareNs before we ever push.
  compare_nameservers: [],
};
// Phase 3C-3: keep this constant in sync with
// src.dns_diag._MAX_COMPARE_CUSTOM_NS.  A mismatch would let the UI
// queue 4+ chips that the server then rejects with a less helpful
// error — keep them aligned in one obvious place.
const _DNS_COMPARE_MAX = 3;
let _dnsTaskId    = null;
let _dnsPollTimer = null;
// Phase 3C-2: TXT / SOA added.  Order matches src.dns_diag._VALID_QTYPES;
// the pill list is rendered in source order so the most common types
// (A/AAAA) stay leftmost and TXT/SOA sit after the structural types.
const _DNS_QTYPES = ['A','AAAA','CNAME','MX','NS','TXT','SOA'];
// Phase 3B: keep this list in sync with src/ui/dns_diag_window.py's
// _DNS_PRESETS so the C / Web tabs offer the same one-click pills.
const _DNS_PRESETS = ['8.8.8.8','1.1.1.1','114.114.114.114'];
// Mirror src.dns_diag.ERROR_LABELS (Chinese labels).  Used to translate
// the rcode / step.status codes that come back from /api/dns_diag/status.
const _DNS_ERR_LABEL = {
  ok:                   '成功',
  timeout:              '超时',
  nxdomain:             '域名不存在',
  no_answer:            '无匹配记录',
  servfail:             '解析器 SERVFAIL',
  refused:              '解析器拒绝',
  resolver_unreachable: '解析器不可达',
  invalid_domain:       '域名格式错误',
  cancelled:            '已取消',
  cname_loop:           'CNAME 链回环',
  cname_too_deep:       'CNAME 链过深',
  unknown_error:        '未知错误',
};

function _resetDnsForm(){
  _dnsForm.domain        = '';
  _dnsForm.qtype         = 'A';
  _dnsForm.timeout_s     = 2;
  _dnsForm.trace_chain   = true;
  _dnsForm.resolver_mode = 'system';
  _dnsForm.nameserver    = '';
  _dnsForm.query_mode    = 'single';
  _dnsForm.compare_nameservers = [];
  _dnsTaskId = null;
}

// Predfill rule mirrors C-side main_window._open_dns_diag: for a DNS
// monitor card, prefer its configured query domain; otherwise fall back
// to the target's host/IP field.
function _dnsPrefill(){
  const t = currentDiagTid ? targets[currentDiagTid] : null;
  if(!t) return '';
  if(t.ping_type === 'dns')
    return t.dns_domain || t.ip || '';
  return t.ip || '';
}

function renderDnsTab(){
  // Lazy prefill: only on first render after switching to this tab, so
  // the operator's own edits aren't wiped on re-renders triggered by
  // qtype changes or run state flipping.
  if(!_dnsForm.domain) _dnsForm.domain = _dnsPrefill();
  const isAAAA  = (_dnsForm.qtype === 'A' || _dnsForm.qtype === 'AAAA');
  const running = (_dnsTaskId !== null);
  const qtypeBtns = _DNS_QTYPES.map(q =>
    `<button class="icmp-pill ${_dnsForm.qtype===q?'active':''}"
             onclick="setDnsQtype('${q}')">${q}</button>`).join('');
  // Phase 3C-4 / 3C-5: also disable the trace checkbox in trace_auth
  // AND dnssec modes — CNAME chasing is performed by the engine in
  // both modes (one-step follow for trace_auth, internal walk for
  // dnssec), so a separate per-query toggle here is misleading.  We
  // disable rather than hide so the operator sees the control is
  // intentional, just gated by current mode.
  const traceLocked   = (!isAAAA)
                         || (_dnsForm.query_mode === 'trace_auth')
                         || (_dnsForm.query_mode === 'dnssec');
  const traceDisabled = traceLocked ? 'disabled' : '';
  const traceStyle    = traceLocked ? 'opacity:.4;cursor:not-allowed' : '';
  // Phase 3B + 3C-3: derive resolver-row layout from both query_mode
  // and resolver_mode.
  //   • single mode → classic 3B layout: system / custom segmented
  //     control + single IP entry + 3 preset pills.
  //   • compare mode → 3C-3 chip-list layout: chips of added IPs +
  //     entry + Add button + counter + preset pills (each preset
  //     adds to the chip list rather than overwriting a single field).
  //   • dnssec mode → re-uses the single-mode layout verbatim — the
  //     resolver is just a TRANSPORT in this mode, and operators
  //     overwhelmingly want to swap to 1.1.1.1 / 8.8.8.8 to bypass
  //     ISP recursors that strip RRSIGs.
  const isCompare    = (_dnsForm.query_mode === 'compare');
  // Phase 3C-4 — third mode.  Trace_auth hides BOTH the single-mode
  // resolver pills AND the compare chip list because the engine
  // walks delegation from the IPv4 root hints and ignores any
  // operator-supplied resolver IP.  We also disable the CNAME-chain
  // checkbox since trace_auth handles CNAME chasing internally (one-
  // step follow at end of chain — see run_dns_auth_trace).
  const isTraceAuth  = (_dnsForm.query_mode === 'trace_auth');
  // Phase 3C-5 — fourth mode.  Dnssec re-uses the single-mode
  // resolver controls (the resolver is the transport for fetching
  // DNSKEY / DS / target rrsets) but with DO=1 + CD=1 baked into
  // the engine.  The CNAME-chain checkbox is force-off; the engine
  // chases CNAMEs internally to identify the leaf zone.
  const isDnssec     = (_dnsForm.query_mode === 'dnssec');
  const isCustomNs   = (_dnsForm.resolver_mode === 'custom');
  const resolverLabel = isCompare ? '对照解析器' : '解析器';

  // Build the compare-mode resolver row.  Kept as a separate variable
  // so the markup template stays linear and we don't nest 4 deep
  // ternaries inside the template literal.
  let resolverRowHtml;
  if(isTraceAuth){
    // Trace_auth: no resolver row.  Show a brief explanatory note in
    // its place so the operator understands why the IP/chip controls
    // disappeared and what the engine is about to do.
    resolverRowHtml = `
    <div class="icmp-row" style="align-items:flex-start">
      <label>解析器</label>
      <div style="flex:1;font-size:11px;color:var(--dim);line-height:1.55;
            background:rgba(96,165,250,.06);
            border:1px solid rgba(96,165,250,.25);
            border-radius:6px;padding:8px 10px">
        权威追踪从 IPv4 根服务器迭代查询 delegation 链至权威 NS，
        <b style="color:var(--text)">不经过系统递归解析器</b>。
        需本机可访问公网 <b style="font-family:monospace">UDP/53</b>；
        专网封外网时第一跳即超时。
      </div>
    </div>`;
  } else if(isCompare){
    const chips = (_dnsForm.compare_nameservers || []).map((ip, idx) =>
      `<span class="dns-cmp-ns-chip" title="${esc(ip)}">
         <span>${esc(ip)}</span>
         <button class="dns-cmp-ns-x"
                 onclick="removeDnsCompareNs(${idx})"
                 title="移除">×</button>
       </span>`).join('');
    const n = (_dnsForm.compare_nameservers || []).length;
    const full = (n >= _DNS_COMPARE_MAX);
    const countCls = full ? 'dns-cmp-ns-count full' : 'dns-cmp-ns-count';
    const addDisabled = full ? 'disabled style="opacity:.45;cursor:not-allowed"' : '';
    // Preset buttons in compare mode push to the chip list.  When the
    // list is full (3 IPs already) we visually disable so the operator
    // sees the cap without typing.
    const presetCmpPills = _DNS_PRESETS.map(ip => {
      const already = (_dnsForm.compare_nameservers || []).indexOf(ip) >= 0;
      const cls = already ? 'icmp-pill active' : 'icmp-pill';
      const dis = (full && !already) ? 'disabled style="opacity:.45;cursor:not-allowed"' : '';
      return `<button class="${cls}" ${dis}
                onclick="addDnsCompareNs('${esc(ip)}')"
                title="${already ? '已添加' : '添加 ' + esc(ip)}">${esc(ip)}</button>`;
    }).join('');
    const chipsArea = n
      ? `<span class="dns-cmp-ns-list" id="dns-cmp-ns-list">${chips}</span>`
      : `<span class="dns-cmp-ns-empty">（请先添加至少 1 个对照解析器）</span>`;
    resolverRowHtml = `
    <div class="icmp-row">
      <label>${resolverLabel}</label>
      ${chipsArea}
      <input class="icmp-input" id="dns-ns" type="text" style="width:140px;margin-left:8px"
             placeholder="例如 8.8.8.8" ${addDisabled}
             onkeydown="if(event.key==='Enter'){event.preventDefault();addDnsCompareNsFromInput();}">
      <button class="icmp-pill" ${addDisabled}
              onclick="addDnsCompareNsFromInput()"
              title="将输入框的 IP 加入对照列表">+ 添加</button>
      <span class="${countCls}">${n}/${_DNS_COMPARE_MAX}</span>
      <span style="margin-left:12px;display:inline-flex;gap:6px">${presetCmpPills}</span>
    </div>`;
  } else {
    const nsEntryAttrs = isCustomNs ? '' : 'disabled style="opacity:.45;cursor:not-allowed"';
    const nsPlaceholder = '例如 8.8.8.8';
    const presetPills  = _DNS_PRESETS.map(ip =>
      `<button class="icmp-pill" onclick="setDnsPreset('${esc(ip)}')"
               title="使用 ${esc(ip)} 作为解析器">${esc(ip)}</button>`).join('');
    const resolverModePills =
      `<button class="icmp-pill ${_dnsForm.resolver_mode==='system'?'active':''}"
               onclick="setDnsResolverMode('system')">系统默认</button>
       <button class="icmp-pill ${_dnsForm.resolver_mode==='custom'?'active':''}"
               onclick="setDnsResolverMode('custom')">自定义</button>`;
    resolverRowHtml = `
    <div class="icmp-row">
      <label>${resolverLabel}</label>
      ${resolverModePills}
      <input class="icmp-input" id="dns-ns" type="text" style="width:160px;margin-left:8px"
             placeholder="${nsPlaceholder}" ${nsEntryAttrs}
             value="${esc(_dnsForm.nameserver)}"
             oninput="_dnsForm.nameserver=this.value">
      <span style="margin-left:10px;display:inline-flex;gap:6px">${presetPills}</span>
    </div>`;
  }
  let html = `
  <div class="icmp-form">
    <div class="icmp-row">
      <label>域名</label>
      <input class="icmp-input" id="dns-domain" type="text" style="width:280px"
             value="${esc(_dnsForm.domain)}"
             oninput="_dnsForm.domain=this.value">
      <span style="font-size:11px;color:var(--dim);margin-left:8px">
        手动触发的 DNS 查询，不影响监控主链路
      </span>
    </div>
    <div class="icmp-row">
      <label>类型</label>
      ${qtypeBtns}
      <label style="margin-left:18px">超时(秒)</label>
      <input class="icmp-input" id="dns-timeout" type="number"
             min="0.2" max="10" step="0.1"
             value="${_dnsForm.timeout_s}"
             oninput="_dnsForm.timeout_s=parseFloat(this.value)||2.0">
      <button class="icmp-switch ${_dnsForm.trace_chain && isAAAA?'on':''}"
              id="dns-trace-sw" ${traceDisabled}
              style="margin-left:16px;${traceStyle}"
              onclick="toggleDnsTrace()"
              title="A/AAAA 查询时展开 CNAME 链至终端 IP"></button>
      <span style="font-size:11px;color:var(--dim)">展开 CNAME 链</span>
    </div>
    <div class="icmp-row">
      <label>查询模式</label>
      <button class="icmp-pill ${(_dnsForm.query_mode==='single')?'active':''}"
              onclick="setDnsQueryMode('single')"
              title="只查一次（系统默认或指定解析器）">单次查询</button>
      <button class="icmp-pill ${isCompare?'active':''}"
              onclick="setDnsQueryMode('compare')"
              title="同一域名顺序查询系统默认与最多 3 个对照解析器">多解析器对比</button>
      <button class="icmp-pill ${isTraceAuth?'active':''}"
              onclick="setDnsQueryMode('trace_auth')"
              title="从 IPv4 根服务器迭代追踪 delegation 链至权威 NS（RD=0）">权威追踪</button>
      <button class="icmp-pill ${isDnssec?'active':''}"
              onclick="setDnsQueryMode('dnssec')"
              title="本机独立验证 DNSSEC 签名链（不信任递归解析器的 AD 位）">DNSSEC 验证</button>
    </div>
    ${resolverRowHtml}
    <div class="icmp-warn" style="color:var(--dim);background:transparent;
         border-color:var(--border2)">
      ${isTraceAuth
        ? '权威追踪从根开始逐级 referral 至权威 NS，最后再查目标 qtype；'
          + '若答案为 CNAME（A / AAAA），将自动对 CNAME 目标再做一次完整追踪。'
          + '<br>本模式 <b>不</b> 使用系统或自定义解析器，亦 <b>不</b> 验证 DNSSEC。'
        : (isDnssec
          ? '本机独立验证 DNSSEC 信任链（root → TLD → leaf）；'
            + 'DO=1 + CD=1 抓取 DNSKEY / DS / RRSIG 后在客户端完成签名校验，'
            + '<b>不</b> 信任解析器的 AD 位。'
            + '<br>四态结果：<b style="color:var(--green)">SECURE</b>（已签名且通过）/'
            + '<b style="color:var(--dim)">INSECURE</b>（zone 未签名）/'
            + '<b style="color:var(--red)">BOGUS</b>（签名错误）/'
            + '<b style="color:var(--orange)">INDETERMINATE</b>（未能判定）。'
            + '<br>建议选用 1.1.1.1 / 8.8.8.8 等公共解析器作为传输，以避免运营商递归剥离 RRSIG。'
          : (isAAAA
              ? 'A / AAAA 查询会按所选项自动展开 CNAME 链至终端 IP。'
              : `已选 ${_dnsForm.qtype}：仅查询该类型，不自动跟随 CNAME。`
                + '如需 IP 解析请改用 A / AAAA。'))}
      ${isCompare
        ? '<br>对比按"系统默认 → 自定义"顺序依次查询；仅比较答案值集合，CNAME 链 / TTL 差异作为附加提示。'
        : ''}
    </div>
    <div class="icmp-actions">
      <button class="btn btn-primary" id="dns-start" onclick="startDnsDiag()"
              ${running?'disabled style="opacity:.5;cursor:not-allowed"':''}>▶ 开始查询</button>
      <button class="btn btn-ghost" id="dns-stop" onclick="stopDnsDiag()"
              ${running?'':'disabled style="opacity:.5;cursor:not-allowed"'}>■ 停止</button>
      <button class="btn btn-ghost" onclick="loadDnsHistory()"
              ${running?'disabled style="opacity:.5;cursor:not-allowed"':''}>📜 历史</button>
    </div>
  </div>
  <div id="dns-output"></div>`;
  document.getElementById('diag-body').innerHTML = html;
  // Phase 3C-1.5 — keep the modal width in sync with the active mode.
  // Done here (rather than only in setDnsQueryMode) because renderDnsTab
  // is also called on tab re-entry from history / icmp, where the user
  // may have flipped between modes without re-syncing the modal class.
  _syncDnsModalWidth();
  // If we switched away and back during a run, re-attach to the polling
  // state by triggering one immediate status fetch.  The interval timer
  // is still alive (we only stop it on terminal status).
  if(running){ pollDnsStatus(); }
}

function setDnsQtype(q){
  _dnsForm.qtype = q;
  // CNAME-chain checkbox is only meaningful for A/AAAA — silently force
  // it off when switching to MX/NS/CNAME so the UI state matches what
  // the server will actually do (the trace_chain field is ignored for
  // non-A/AAAA on the server side, but keeping them in sync prevents
  // operator confusion).
  if(q!=='A' && q!=='AAAA') _dnsForm.trace_chain = false;
  renderDnsTab();
}
function toggleDnsTrace(){
  if(_dnsForm.qtype!=='A' && _dnsForm.qtype!=='AAAA') return;
  _dnsForm.trace_chain = !_dnsForm.trace_chain;
  const sw = document.getElementById('dns-trace-sw');
  if(sw) sw.classList.toggle('on', _dnsForm.trace_chain);
}

// Phase 3B helpers — resolver mode + preset selection.  Both end with a
// full renderDnsTab() because (a) the active-pill highlighting requires
// re-rendering the buttons and (b) the entry's disabled state is baked
// into the markup, so a class toggle alone isn't enough.  Operator's
// current entry value is preserved across re-renders via _dnsForm.nameserver.
function setDnsResolverMode(mode){
  if(mode !== 'system' && mode !== 'custom') return;
  // Snapshot current entry value before re-render so a user who typed
  // an IP then clicked 系统默认 keeps it visible (just inert) — same UX
  // as the C-side window.
  const cur = document.getElementById('dns-ns');
  if(cur) _dnsForm.nameserver = cur.value || '';
  _dnsForm.resolver_mode = mode;
  renderDnsTab();
}

// Phase 3C-1: flip the top-level 查询模式.  We re-render so the
// resolver-row label (解析器 ↔ 对照解析器) and the optional
// system/custom segmented control update in lockstep with the
// hint text.  The user's typed nameserver and resolver_mode values
// are preserved across the flip — switching back to single after
// running a compare keeps the previously-chosen single-mode resolver
// (e.g. 1.1.1.1) in place, which matches the C-side behaviour.
//
// Phase 3C-3: only snapshot the entry to _dnsForm.nameserver while
// LEAVING single mode.  In compare mode the entry is a transient
// chip-staging field; preserving its (usually empty) content into
// _dnsForm.nameserver would silently wipe the operator's single-mode
// resolver choice the moment they flip to compare.
function setDnsQueryMode(mode){
  // Phase 3C-4 / 3C-5 — accept 'trace_auth' and 'dnssec'.  Like compare,
  // we snapshot the single-mode entry on the way OUT of single AND on
  // the way OUT of dnssec (which also uses the dns-ns input) so the
  // operator's previously-typed IP survives a round-trip through
  // dnssec → compare → single without being silently wiped.  This is
  // the same defensive snapshot pattern as 3C-3's single ↔ compare
  // flow.
  if(mode !== 'single' && mode !== 'compare'
      && mode !== 'trace_auth' && mode !== 'dnssec') return;
  if(_dnsForm.query_mode === 'single' || _dnsForm.query_mode === 'dnssec'){
    const cur = document.getElementById('dns-ns');
    if(cur) _dnsForm.nameserver = cur.value || '';
  }
  _dnsForm.query_mode = mode;
  // renderDnsTab() rebuilds the inner body and will call
  // _syncDnsModalWidth() itself, but invoking it BEFORE the rerender
  // makes the modal's CSS width transition kick off at the same time
  // as the segment pill highlight — without this the modal width
  // visibly lags the button highlight by ~50 ms.
  _syncDnsModalWidth();
  renderDnsTab();
}
function setDnsPreset(ip){
  // Preset clicks always mean "I want this resolver" — flipping mode
  // is implicit, otherwise the populated entry would silently no-op
  // under 系统默认.
  _dnsForm.nameserver    = ip;
  _dnsForm.resolver_mode = 'custom';
  renderDnsTab();
}

// ════════════════════════════════════════════════════════════════════
// Phase 3C-3 — multi-resolver compare chip-list helpers.
//
// `_dnsForm.compare_nameservers` is the source of truth.  Mutations
// here all end with a full renderDnsTab() so the chip strip, counter,
// preset highlights AND modal width recompute together — partial DOM
// patches drift out of sync fast (the modal width transition is on a
// CSS class set in JS).
// ════════════════════════════════════════════════════════════════════

// IPv4 + bracketed/plain IPv6 literal check.  Mirrors src.dns_diag.
// _validate_nameserver — duplicated here so the operator gets immediate
// feedback without a round-trip; the server validator is still the
// final word on shape correctness.
function _isValidNsLiteral(s){
  if(typeof s !== 'string') return false;
  s = s.trim();
  if(!s || s.length > 45) return false;
  if(s.indexOf('%') >= 0) return false;   // zone IDs not supported
  // IPv4
  if(/^(\d{1,3})(\.\d{1,3}){3}$/.test(s)){
    return s.split('.').every(p => {
      const n = parseInt(p, 10);
      return n >= 0 && n <= 255 && String(n) === p;
    });
  }
  // IPv6 — accept the most common forms via a lax pattern; the server
  // still does an authoritative inet_pton check.
  if(/^[0-9a-fA-F:]+$/.test(s) && s.indexOf(':') >= 0) return true;
  return false;
}

function addDnsCompareNs(ip){
  ip = (ip || '').trim();
  if(!ip){ _showDnsError('请输入有效的 IP'); return; }
  if(!_isValidNsLiteral(ip)){
    _showDnsError('解析器须为合法的 IPv4 / IPv6 地址'); return;
  }
  const list = _dnsForm.compare_nameservers || (_dnsForm.compare_nameservers = []);
  if(list.indexOf(ip) >= 0){
    _showDnsError('该解析器已添加'); return;
  }
  if(list.length >= _DNS_COMPARE_MAX){
    _showDnsError(`最多支持 ${_DNS_COMPARE_MAX} 个对照解析器`); return;
  }
  list.push(ip);
  _clearDnsError();
  renderDnsTab();
}

function addDnsCompareNsFromInput(){
  const el = document.getElementById('dns-ns');
  const raw = el ? el.value : '';
  // Snapshot then clear input so the operator can chain-add quickly;
  // a focus restore at the end keeps the keyboard flow tight.
  if(el) el.value = '';
  addDnsCompareNs(raw);
  // renderDnsTab() rebuilt the element — re-grab and re-focus.
  setTimeout(() => {
    const fresh = document.getElementById('dns-ns');
    if(fresh) try { fresh.focus(); } catch(_e){}
  }, 0);
}

function removeDnsCompareNs(idx){
  const list = _dnsForm.compare_nameservers || [];
  if(idx < 0 || idx >= list.length) return;
  list.splice(idx, 1);
  _clearDnsError();
  renderDnsTab();
}

// Best-effort: clear a stale "请输入" / "已添加" error after the
// operator fixes the offending input.  No-op if there's no output yet
// (e.g. first chip add on an empty form).
function _clearDnsError(){
  const out = document.getElementById('dns-output');
  if(out && out.innerHTML.indexOf('icmp-conclusion bad') >= 0){
    out.innerHTML = '';
  }
}

function startDnsDiag(){
  if(!currentDiagTid){ _showDnsError('未选中节点'); return; }
  const domain = (document.getElementById('dns-domain')?.value || '').trim();
  if(!domain){ _showDnsError('请输入要查询的域名'); return; }
  _dnsForm.domain = domain;
  const isCompare   = (_dnsForm.query_mode === 'compare');
  const isTraceAuth = (_dnsForm.query_mode === 'trace_auth');
  const isDnssec    = (_dnsForm.query_mode === 'dnssec');
  // Phase 3C-3: compare mode uses the chip-list (nameservers[]) rather
  // than the single-IP entry.  If the operator typed something into the
  // text field but didn't click "+ 添加", grab it as a courtesy so a
  // half-finished compare doesn't fail validation with "请添加至少 1 个"
  // when there's a perfectly good IP sitting right there.
  //
  // Phase 3C-5 dnssec mode reuses the single-mode entry handling — the
  // resolver IS used (as a DO=1 + CD=1 transport for fetching DNSKEY /
  // DS / target rrsets), with the operator's choice of system default
  // vs. custom IP carrying the same meaning as single mode.
  let nsBody  = null;
  let nssBody = null;
  if(isTraceAuth){
    // Phase 3C-4: trace_auth iterates from root and ignores any
    // operator-supplied resolver entirely.  Make that explicit by
    // wiping both nameserver slots before they hit the wire — even
    // if the form happens to still hold values from a prior mode.
    nsBody  = null;
    nssBody = null;
  } else if(isDnssec){
    // Phase 3C-5: dnssec — mirror single-mode resolver handling.
    // System default → null (engine uses system resolver list);
    // custom → operator-typed IP literal.  trace_chain is ignored
    // by the engine; we still null nssBody to avoid leaking a
    // compare-leftover chip list into a dnssec request body.
    nssBody = null;
    if(_dnsForm.resolver_mode === 'custom'){
      const cur = document.getElementById('dns-ns');
      const raw = (cur ? cur.value : _dnsForm.nameserver) || '';
      const ns  = raw.trim();
      if(!ns){ _showDnsError('请输入解析器 IP'); return; }
      _dnsForm.nameserver = ns;
      nsBody = ns;
    }
  } else if(isCompare){
    const cur = document.getElementById('dns-ns');
    const pendingRaw = cur ? (cur.value || '').trim() : '';
    if(pendingRaw){
      // Validate inline and auto-add — saves the operator one click.
      if(_isValidNsLiteral(pendingRaw) &&
         (_dnsForm.compare_nameservers || []).indexOf(pendingRaw) < 0 &&
         (_dnsForm.compare_nameservers || []).length < _DNS_COMPARE_MAX){
        _dnsForm.compare_nameservers.push(pendingRaw);
        if(cur) cur.value = '';
      }
    }
    const list = (_dnsForm.compare_nameservers || []).slice();
    if(!list.length){
      _showDnsError('请至少添加 1 个对照解析器 IP');
      renderDnsTab();   // refresh chip strip in case we auto-added then bailed
      return;
    }
    nssBody = list;
  } else if(_dnsForm.resolver_mode === 'custom'){
    const cur = document.getElementById('dns-ns');
    const raw = (cur ? cur.value : _dnsForm.nameserver) || '';
    const ns  = raw.trim();
    if(!ns){ _showDnsError('请输入解析器 IP'); return; }
    _dnsForm.nameserver = ns;
    nsBody = ns;
  }
  // Defensive: stop any leftover poll before kicking off a new run.
  stopDnsPolling();
  const body = {
    target_id:   currentDiagTid,
    domain:      domain,
    qtype:       _dnsForm.qtype,
    timeout_s:   _dnsForm.timeout_s,
    // Phase 3C-4 / 3C-5: trace_auth and dnssec both handle CNAME
    // chasing internally (trace_auth via end-of-chain follow,
    // dnssec via _follow_cname_for_dnssec) — force-off so the
    // backend doesn't double-walk.
    trace_chain: (isTraceAuth || isDnssec) ? false : !!_dnsForm.trace_chain,
    nameserver:  nsBody,    // single + dnssec only; null in compare/auth
    nameservers: nssBody,   // 3C-3 array; null in single/auth/dnssec
    mode:        isDnssec    ? 'dnssec'
                : isTraceAuth ? 'trace_auth'
                : isCompare   ? 'compare'
                              : 'single',
    auth_max_hops: isTraceAuth ? 12 : undefined,
  };
  document.getElementById('dns-output').innerHTML =
    '<div style="padding:10px;color:var(--dim);font-size:12px">⏳ 正在解析...</div>';
  fetch('/api/dns_diag/start',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify(body),
  }).then(async r=>{
    const data = await r.json().catch(()=>({}));
    if(r.status===200 && data.task_id){
      _dnsTaskId = data.task_id;
      renderDnsTab();   // flip Start/Stop button states
      // DNS queries are sub-second typically, so poll faster than ICMP.
      pollDnsStatus();
      _dnsPollTimer = setInterval(pollDnsStatus, 500);
    } else {
      _showDnsError(data.error || `HTTP ${r.status}`);
    }
  }).catch(e=>_showDnsError('请求失败: ' + (e&&e.message||e)));
}

function stopDnsDiag(){
  if(!_dnsTaskId) return;
  const tid = _dnsTaskId;
  fetch('/api/dns_diag/cancel',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({task_id: tid}),
  }).catch(()=>{});
  // Next poll sees status=cancelled and renders the terminal state.
}

function stopDnsPolling(){
  if(_dnsPollTimer){ clearInterval(_dnsPollTimer); _dnsPollTimer = null; }
}

function pollDnsStatus(){
  if(!_dnsTaskId) return;
  const tid = _dnsTaskId;
  fetch('/api/dns_diag/status?task_id=' + encodeURIComponent(tid))
    .then(r=>r.json())
    .then(data=>{
      // Race: user stopped / opened a new query between request and
      // response — discard stale data.
      if(tid !== _dnsTaskId) return;
      renderDnsProgress(data);
      if(['done','error','cancelled'].includes(data.status)){
        stopDnsPolling();
        // Same cached-restore dance as ICMP: re-render the form to flip
        // button states, then put the rendered output back.
        const out = document.getElementById('dns-output');
        const cached = out ? out.innerHTML : '';
        _dnsTaskId = null;
        renderDnsTab();
        const newOut = document.getElementById('dns-output');
        if(newOut && cached) newOut.innerHTML = cached;
      }
    })
    .catch(()=>{ /* transient blip — keep polling */ });
}

function _showDnsError(msg){
  const out = document.getElementById('dns-output');
  if(!out) return;
  out.innerHTML =
    `<div class="icmp-conclusion bad" style="margin-top:10px">⚠ ${esc(msg)}</div>`;
}

function _dnsErrLabel(code){ return _DNS_ERR_LABEL[code] || code || ''; }

// Open quick-tracert in a new window for a resolved A/AAAA IP.  Uses
// the same /traceroute/window endpoint as the route-trace tab, but with
// ?ip= instead of ?tid= — the server validates the IP literal strictly
// before it ever reaches the tracert subprocess.
function traceDnsIp(ip){
  if(!ip) return;
  const url = '/traceroute/window?ip=' + encodeURIComponent(ip);
  window.open(url, '_blank',
    'width=620,height=700,resizable=yes,scrollbars=yes,menubar=no,toolbar=no,location=no');
}

function renderDnsProgress(data){
  const out = document.getElementById('dns-output');
  if(!out) return;
  // While running we don't have a partial result to render, just show
  // a one-line placeholder.  DNS round-trips are typically <100ms so
  // this state is briefly visible at most.
  if(data.status === 'running'){
    // Phase 3C-1 / 3C-3: compare mode runs N queries back-to-back, so
    // the placeholder text tells the operator we're explicitly
    // comparing ("正在对比解析…") instead of the generic single-query
    // spinner.  /status surfaces mode + nameservers[] for the running
    // task; fall back to the legacy `nameserver` summary string for
    // any racy pre-3C-3 in-flight response.
    const isCompare   = (data.mode === 'compare');
    const isTraceAuth = (data.mode === 'trace_auth');
    const isDnssec    = (data.mode === 'dnssec');
    let inner;
    if(isTraceAuth){
      // Trace_auth walks delegation from root — no resolver IP makes
      // sense in the placeholder, just announce the target and the
      // fact that we're descending the chain.
      inner = `⏳ 正在从根服务器追踪 ${esc(data.domain || '')}`
            + ` (${esc(data.qtype || '')}) 的权威 NS ...`;
    } else if(isDnssec){
      // Phase 3C-5: DNSSEC walk fetches DNSKEY + DS + leaf rrset for
      // each zone in the delegation, plus the embedded anchor check.
      // Mention the resolver only when explicit — system default is
      // implied otherwise.
      const ns = (data.nameserver || '').trim();
      inner = `🔐 正在验证 ${esc(data.domain || '')}`
            + ` (${esc(data.qtype || '')}) 的 DNSSEC 签名链`
            + (ns ? ` （经由 ${esc(ns)}）` : '')
            + ' ...';
    } else if(isCompare){
      let nsDesc = '';
      const nss = Array.isArray(data.nameservers) ? data.nameservers : [];
      if(nss.length === 1)      nsDesc = nss[0];
      else if(nss.length >= 2)  nsDesc = nss[0] + ', +' + (nss.length - 1);
      else                      nsDesc = data.nameserver || '对照解析器';
      inner = `⏳ 正在对比解析（系统 → ${esc(nsDesc)}）...`;
    } else {
      inner = `⏳ 正在解析 ${esc(data.domain||'')} (${esc(data.qtype||'')})...`;
    }
    out.innerHTML =
      '<div style="padding:12px;color:var(--dim);font-size:12px;text-align:center">'
      + inner + '</div>';
    return;
  }
  if(data.status === 'cancelled'){
    out.innerHTML =
      `<div class="icmp-conclusion warn" style="margin-top:10px">■ 查询已取消</div>`;
    return;
  }
  if(data.status === 'error'){
    out.innerHTML =
      `<div class="icmp-conclusion bad" style="margin-top:10px">`
      + `⚠ 查询失败：${esc(data.error || '未知错误')}</div>`;
    return;
  }
  // status === 'done' — dispatch on result.mode (Phase 3C-1).  The
  // helper picks the single vs compare renderer; pre-3C-1 results
  // without a mode key fall through to single (legacy contract).
  out.innerHTML = _renderDnsRun(data.result || {});
}

// Shared dispatcher — pick single vs compare vs trace_auth renderer
// based on result.mode.  Used by both the live /status polling path
// and the history detail / replay paths so all three render identically.
function _renderDnsRun(r){
  if(r && r.mode === 'compare')    return _renderDnsCompareBlock(r);
  if(r && r.mode === 'trace_auth') return _renderDnsAuthBlock(r);
  // Phase 3C-5: dnssec is the fourth mode.  Its renderer is distinct
  // from single mode because the verdict bar carries four states (not
  // just success/failure) and the verification chain layout has zero
  // overlap with the CNAME-chain pre tag used by single mode.
  if(r && r.mode === 'dnssec')     return _renderDnsSecBlock(r);
  return _renderDnsResultBlock(r);
}

// Shared render: takes a result-shaped object (the dict produced by
// src.dns_diag.dns_result_to_dict — same shape returned by /status and
// stored in dns_diagnostic_history.details_json) and returns the inner
// HTML for the result panel.  Pure function: no DOM lookups, no side
// effects.  Callers control where to insert the returned string.
function _renderDnsResultBlock(r){
  const ok  = !!r.success;
  const cls = ok ? 'ok' : 'bad';
  let html = `<div class="dns-summary">
    <span>解析器: <b>${esc(r.resolver || '系统默认')}</b></span>
    <span>RCODE: <b>${esc(r.rcode || '—')}</b></span>
    <span>耗时: <b>${r.elapsed_ms!=null ? Number(r.elapsed_ms).toFixed(1)+' ms' : '—'}</b></span>
  </div>`;
  // CNAME chain (monospace tree).  For non-A/AAAA we don't try to draw
  // a chain — the answer table itself is the full result.
  const chain = r.cname_chain || [];
  if(chain.length){
    let chainText = '';
    chain.forEach((name, i) => {
      if(i === 0) chainText += name + '\n';
      else        chainText += '  '.repeat(i) + '└─ CNAME ' + name + '\n';
    });
    if(r.qtype === 'A' || r.qtype === 'AAAA'){
      const indent = '  '.repeat(chain.length);
      (r.answers || []).forEach(a => {
        chainText += indent + '└─ ' + a.qtype + ' ' + a.value + '\n';
      });
    }
    html += `<div class="dns-chain">${esc(chainText)}</div>`;
  }
  // Answer table — empty answers also render the header so the operator
  // sees explicit "(无返回记录)" instead of a missing element.
  const answers = r.answers || [];
  html += '<table class="dns-ans-tbl"><thead><tr>'
        + '<th>名称</th><th style="width:60px">类型</th><th>值</th>'
        + '<th style="width:60px;text-align:right">TTL</th>'
        + '<th style="width:100px">操作</th></tr></thead><tbody>';
  if(!answers.length){
    html += '<tr><td colspan="5" style="text-align:center;color:var(--dim);'
          + 'padding:14px">(无返回记录)</td></tr>';
  } else {
    answers.forEach(a => {
      const ttl = (a.ttl!=null) ? a.ttl : '—';
      const action = (a.qtype === 'A' || a.qtype === 'AAAA')
        ? `<button class="dns-trace-btn"
                   onclick="traceDnsIp('${esc(a.value)}')">追踪此 IP</button>`
        : '';
      html += `<tr>
        <td>${esc(a.name)}</td>
        <td class="dns-qtype">${esc(a.qtype)}</td>
        <td style="overflow-wrap:anywhere;word-break:normal;max-width:520px">${esc(a.value)}</td>
        <td class="dns-ttl">${ttl}</td>
        <td>${action}</td>
      </tr>`;
    });
  }
  html += '</tbody></table>';
  // Conclusion box
  let conclusion = r.message || '';
  if(!conclusion && !ok) conclusion = '查询失败';
  if(!conclusion &&  ok) conclusion = `解析成功，共 ${answers.length} 条 ${r.qtype} 记录`;
  html += `<div class="icmp-conclusion ${cls}" style="margin-top:12px">
    <div style="font-weight:600;margin-bottom:4px">📋 结论</div>
    <div>${esc(conclusion)}</div>
  </div>`;
  return html;
}

// ════════════════════════════════════════════════════════════════════
// Phase 3C-1.5 — diff-first compare renderer.
//
// Inputs come from src.dns_diag.dns_compare_to_dict.  Layout:
//   ① verdict bar      —— 一致 / 不一致 + 结论 + 总耗时
//   ② diff matrix      —— 全宽：值 | 系统 | 对照 (cell = ✓ TTL n / ✗ 缺失)
//   ③ 折叠详情 × 2     —— 每个解析器：RCODE / CNAME 链 / 完整答案表 /
//                         每条 A/AAAA 行带「追踪此 IP」按钮
//
// 当两侧答案完全一致且 CNAME 链相同时，两个折叠面板默认收起；
// 任一差异（answer set 不同 / 链不同 / 任一侧失败）则默认展开，
// 让运维员看见现场快照。
// ════════════════════════════════════════════════════════════════════

// Toggle the diagnostic modal width based on the current DNS query mode.
// Called from renderDnsTab(), switchDiagTab(), setDnsQueryMode().
//
// Phase 3C-3: three tiers driven by the chip count.  Two columns
// (2 sides) fits at 900 px; three at 1000; four at 1100.  The class
// is mutually-exclusive — only one width tier active at a time, so
// flipping chip count cleanly transitions modal width via CSS.
function _syncDnsModalWidth(){
  const box = document.querySelector('#diag-overlay .modal-box');
  if(!box) return;
  const inDns = (_currentDiagTab === 'dns' && _dnsForm);
  const isCompare    = inDns && _dnsForm.query_mode === 'compare';
  // Phase 3C-4: trace_auth lives in tier-1 (900 px) — the same width
  // the verdict + diff matrix use for 2-side compare.  We do NOT
  // promote to tier-2/3 because the tree grows vertically, not
  // horizontally, so a wider modal just adds blank gutter on the
  // right.  Keeping it at 900 also matches the C-side window's
  // 880×720 footprint so the two UIs feel proportional.
  const isTraceAuth  = inDns && _dnsForm.query_mode === 'trace_auth';
  // Phase 3C-5: dnssec lives in tier-1 (900 px) like trace_auth.  The
  // verification chain TABLE is intentionally compact (one row per
  // zone, never more than ~6 rows for any real-world domain), so a
  // tier-2 promotion would just add blank gutter without surfacing
  // additional information.  Keeping dnssec and trace_auth at the
  // same width tier also avoids a visible modal-resize judder when
  // operators flip between the two modes back-to-back.
  const isDnssec     = inDns && _dnsForm.query_mode === 'dnssec';
  let tier = 0;   // 0 = base 680, 1 = 900, 2 = 1000, 3 = 1100
  if(isCompare){
    const n = (_dnsForm.compare_nameservers || []).length;
    if(n <= 1) tier = 1;
    else if(n === 2) tier = 2;
    else tier = 3;
  } else if(isTraceAuth || isDnssec){
    tier = 1;
  }
  box.classList.toggle('modal-box--dns-compare',      tier === 1);
  box.classList.toggle('modal-box--dns-compare-mid',  tier === 2);
  box.classList.toggle('modal-box--dns-compare-wide', tier === 3);
}

// Normalised CNAME chain signature — used to decide whether the two
// detail panels should be expanded by default (a chain divergence is
// worth surfacing even when the terminal A records match).
function _dnsChainSig(r){
  const chain = (r && r.cname_chain) || [];
  return chain.map(n => (n || '').replace(/\.$/, '').toLowerCase()).join('|');
}

// Phase 3C-3 — normalise the diff payload into the v2 array shape
// regardless of source.  Inputs we expect to see in the wild:
//   • v2 (3C-3 server)  : c.diff.side_labels is an Array
//   • v1 (3C-1.5 server): c.diff.side_labels is a {system,custom} dict
//   • pre-1.5 history   : c.diff missing entirely → recompute from
//                         c.system / c.custom by mirroring
//                         src.dns_diag.build_compare_diff_matrix.
// Output is always v2: side_labels: [{key,label}, …],
//                     rows: [{display,qtype,cells:{key:{text,ok}}, …}]
// Renderers downstream only ever see v2.
function _dnsCompareNormalizeDiff(c){
  if(!c) return {side_labels: [], rows: []};

  const fromV1Dict = labelsDict => {
    // labelsDict: {system: "...", custom: "..."}; convert to ordered
    // array using the canonical 3C-1 key order.
    const out = [];
    if(labelsDict.system !== undefined)
      out.push({key: 'system', label: String(labelsDict.system || '系统默认')});
    if(labelsDict.custom !== undefined)
      out.push({key: 'custom', label: String(labelsDict.custom || '自定义')});
    return out;
  };

  const promoteV1Row = (row, sideKeys) => {
    // Convert legacy {system, custom} per-row cells into v2 cells:{key:cell}
    const cells = {};
    sideKeys.forEach(k => {
      const cell = row[k] || row.cells && row.cells[k] || {text: '—', ok: false};
      cells[k] = cell;
    });
    return {
      display: row.display || '',
      qtype:   row.qtype   || '',
      cells:   cells,
    };
  };

  // 1) v2 array form — pass through after a shallow row promotion that
  // adds any missing `cells` dict (older v2 prototypes only carried
  // row[key] aliases for N==2 runs).
  if(c.diff && Array.isArray(c.diff.side_labels)){
    const sl = c.diff.side_labels.map(s => ({
      key: String(s.key || ''), label: String(s.label || '')
    }));
    const sideKeys = sl.map(s => s.key);
    const rows = (c.diff.rows || []).map(r => {
      if(r.cells && typeof r.cells === 'object') return r;
      return promoteV1Row(r, sideKeys);
    });
    return {side_labels: sl, rows};
  }

  // 2) v1 dict form — convert in place.
  if(c.diff && c.diff.side_labels && typeof c.diff.side_labels === 'object'){
    const sl = fromV1Dict(c.diff.side_labels);
    const sideKeys = sl.map(s => s.key);
    const rows = (c.diff.rows || []).map(r => promoteV1Row(r, sideKeys));
    return {side_labels: sl, rows};
  }

  // 3) No diff payload — compute from c.sides (3C-3) or c.system/c.custom
  // (3C-1).  Mirrors src.dns_diag.build_compare_diff_matrix's algorithm
  // closely enough for replay; the server is still the canonical source
  // for newly-recorded runs.
  const sides = _dnsCompareNormalizeSides(c);
  const ingest = side => {
    const m = {};
    ((side && side.answers) || []).forEach(a => {
      const qt = (a.qtype || '').toUpperCase();
      const val = a.value || '';
      let key = val;
      if(qt === 'CNAME' || qt === 'NS') key = val.replace(/\.$/,'').toLowerCase();
      else if(qt === 'MX'){
        let pref = a.preference, target = val;
        if(pref == null){
          const parts = val.split(' ');
          if(parts.length >= 2){
            const n = parseInt(parts[0], 10);
            if(!isNaN(n)){ pref = n; target = parts.slice(1).join(' '); }
          }
        }
        key = (pref != null ? pref : -1) + '|' + target.replace(/\.$/,'').toLowerCase();
      }
      m[key] = {display: val, qtype: qt, ttl: a.ttl};
    });
    return m;
  };
  const maps = sides.map(s => ingest(s.result));
  const sideLabels = sides.map(s => ({key: s.key, label: s.label || ''}));
  const cellFor = (key, map, sideRes) => {
    if(key in map){
      const t = map[key].ttl;
      return {text: (t != null ? `✓ TTL ${t}` : '✓'), ok: true};
    }
    if(sideRes && sideRes.success) return {text: '✗ 缺失', ok: false};
    return {text: '—', ok: false};
  };
  const emptyCell = sideRes => {
    if(!sideRes)         return {text: '—', ok: false};
    if(sideRes.success)  return {text: '✓ 无记录', ok: true};
    return {text: '✗ ' + (sideRes.rcode || '失败'), ok: false};
  };
  const seen = {};
  maps.forEach(m => {
    Object.keys(m).forEach(k => { if(!(k in seen)) seen[k] = m[k]; });
  });
  let rows;
  if(!Object.keys(seen).length){
    const cells = {};
    sides.forEach((s, i) => { cells[s.key] = emptyCell(s.result); });
    rows = [{display: '（解析状态）', qtype: '', cells}];
  } else {
    rows = Object.keys(seen).sort().map(key => {
      const ent = seen[key];
      const cells = {};
      sides.forEach((s, i) => {
        cells[s.key] = cellFor(key, maps[i], s.result);
      });
      return {display: ent.display, qtype: ent.qtype, cells};
    });
  }
  return {side_labels: sideLabels, rows};
}

// Normalise the side list to v2 shape [{key,label,result}, …].  v2
// payloads already carry `sides`; legacy payloads only have
// system / custom dicts which we synthesise into a 2-element list with
// the 3C-1 canonical keys.
function _dnsCompareNormalizeSides(c){
  if(!c) return [];
  if(Array.isArray(c.sides) && c.sides.length){
    return c.sides.map(s => ({
      key:    String(s.key || ''),
      label:  String(s.label || ''),
      result: s.result || {},
    }));
  }
  // 3C-1 fallback — synthesise from c.system / c.custom.
  const out = [];
  if(c.system){
    out.push({
      key: 'system',
      label: (c.system.resolver || '系统默认'),
      result: c.system,
    });
  }
  if(c.custom){
    out.push({
      key: 'custom',
      label: (c.custom.resolver || c.nameserver || '自定义'),
      result: c.custom,
    });
  }
  return out;
}

// Backward-compat alias for the 3C-1.5 fallback name.  Returns v2
// diff shape (side_labels as array).
function _dnsCompareDiffFromSides(c){
  return _dnsCompareNormalizeDiff(c);
}

function _renderDnsCompareDiffTable(diff){
  const sideLabels = diff.side_labels || [];
  const rows = diff.rows || [];
  const N = sideLabels.length;

  // D3: column widths — 2 cols 44/28/28 (3C-1.5 baseline preserved
  // exactly), 3+ cols 34% value + 22% per side.  table-layout:fixed in
  // the parent CSS keeps headers from wandering when long IPv6 strings
  // sit in a row.
  const valPct  = (N <= 2) ? 44 : 34;
  const sidePct = (N <= 2) ? 28 : 22;
  const colgroup = `<col style="width:${valPct}%">` +
                   sideLabels.map(_ => `<col style="width:${sidePct}%">`).join('');

  let html = '<div class="dns-cmp-diff"><table>'
           + `<colgroup>${colgroup}</colgroup>`
           + '<thead><tr><th>值</th>';
  sideLabels.forEach((s, idx) => {
    // Side 0 (always the system baseline) gets the blue col-sys tone;
    // sides 1..N all share the purple col-cus tone.  See CSS comment
    // for the rationale (3-column eye-grouping > N-way colour soup).
    const cls = (idx === 0) ? 'col-sys' : 'col-cus';
    const lbl = esc(s.label || '—');
    html += `<th class="${cls}" title="${lbl}">${lbl}</th>`;
  });
  html += '</tr></thead><tbody>';

  if(!rows.length){
    html += `<tr><td colspan="${N + 1}" class="dns-cmp-empty">(无对比数据)</td></tr>`;
  } else {
    rows.forEach(row => {
      const disp = esc(row.display || '');
      const qt   = row.qtype
        ? `<span class="dns-cmp-qtype">${esc(row.qtype)}</span>` : '';
      html += '<tr>'
            + `<td><span class="dns-cmp-val">${disp}</span>${qt}</td>`;
      sideLabels.forEach(s => {
        const cells = row.cells || {};
        let cell = cells[s.key];
        // v1 row-level alias fallback (system/custom keys directly on row).
        if(cell == null) cell = row[s.key] || {text: '—', ok: false};
        const cls = cell.ok ? 'cell-ok' : 'cell-bad';
        html += `<td class="${cls}">${esc(cell.text || '—')}</td>`;
      });
      html += '</tr>';
    });
  }
  html += '</tbody></table></div>';
  return html;
}

// Render the answer / chain block that lives INSIDE a detail panel.
// Deliberately omits the summary line (RCODE / elapsed) because that
// info already appears in the detail header — duplicating it would
// just waste vertical space in a constrained modal.
function _renderDnsCompareDetailBody(r){
  let html = '';
  const chain = r.cname_chain || [];
  if(chain.length){
    let chainText = '';
    chain.forEach((name, i) => {
      if(i === 0) chainText += name + '\n';
      else        chainText += '  '.repeat(i) + '└─ CNAME ' + name + '\n';
    });
    if(r.qtype === 'A' || r.qtype === 'AAAA'){
      const indent = '  '.repeat(chain.length);
      (r.answers || []).forEach(a => {
        chainText += indent + '└─ ' + a.qtype + ' ' + a.value + '\n';
      });
    }
    html += `<div class="dns-chain">${esc(chainText)}</div>`;
  }
  const answers = r.answers || [];
  html += '<table class="dns-ans-tbl"><thead><tr>'
        + '<th>名称</th><th style="width:60px">类型</th><th>值</th>'
        + '<th style="width:60px;text-align:right">TTL</th>'
        + '<th style="width:100px">操作</th></tr></thead><tbody>';
  if(!answers.length){
    html += '<tr><td colspan="5" style="text-align:center;color:var(--dim);'
          + 'padding:14px">(无返回记录)</td></tr>';
  } else {
    answers.forEach(a => {
      const ttl = (a.ttl != null) ? a.ttl : '—';
      const action = (a.qtype === 'A' || a.qtype === 'AAAA')
        ? `<button class="dns-trace-btn"
                   onclick="traceDnsIp('${esc(a.value)}')">追踪此 IP</button>`
        : '';
      html += `<tr>
        <td>${esc(a.name)}</td>
        <td class="dns-qtype">${esc(a.qtype)}</td>
        <td style="overflow-wrap:anywhere;word-break:normal">${esc(a.value)}</td>
        <td class="dns-ttl">${ttl}</td>
        <td>${action}</td>
      </tr>`;
    });
  }
  html += '</tbody></table>';
  if(r.message){
    html += `<div class="dns-cmp-msg">${esc(r.message)}</div>`;
  }
  return html;
}

function _renderDnsCompareDetail(r, label, kind, openByDefault){
  // `kind` controls accent colour: 'sys' (blue) for side 0,
  // 'cus' (purple) for every other side.  Random suffix scopes the
  // toggle DOM IDs so multiple compare blocks (e.g. live + history
  // preview) don't collide.
  const id = 'dns-cmp-detail-' + kind + '-' + Math.random().toString(36).slice(2, 8);
  const ok = !!(r && r.success);
  const rcode = (r && r.rcode) || '—';
  const ems = (r && r.elapsed_ms != null)
              ? Number(r.elapsed_ms).toFixed(1) + ' ms' : '—';
  const okMark = ok ? '✓' : '✗';
  const okCol  = ok ? 'var(--green)' : 'var(--red)';
  const chev = openByDefault ? '▼' : '▶';
  const bodyCls = openByDefault ? 'dns-cmp-detail-body open' : 'dns-cmp-detail-body';
  return `<div class="dns-cmp-detail">
    <div class="dns-cmp-detail-hdr ${kind}" onclick="toggleDnsCmpDetail('${id}')">
      <span class="dns-cmp-chev" id="${id}-chev">${chev}</span>
      <span class="dns-cmp-resolver">${esc(label)}</span>
      <span class="dns-cmp-resolver-meta">
        <span style="color:${okCol};font-weight:700">${okMark}</span>
        &nbsp;RCODE ${esc(rcode)} · ${ems}
      </span>
    </div>
    <div id="${id}" class="${bodyCls}">${_renderDnsCompareDetailBody(r || {})}</div>
  </div>`;
}

function toggleDnsCmpDetail(id){
  const body = document.getElementById(id);
  const chev = document.getElementById(id + '-chev');
  if(!body) return;
  const open = !body.classList.contains('open');
  body.classList.toggle('open', open);
  if(chev) chev.textContent = open ? '▼' : '▶';
}

// Phase 3C-3 — render a multi-resolver compare result.  Reads the
// canonical v2 shape (sides[] + diff{side_labels:[],rows:[]}); inputs
// in older shapes get normalised first.  Auto-expand rule generalises
// the 3C-1.5 logic: if ANY two sides have differing CNAME chain
// signatures, or the verdict failed, expand all detail panels.
function _renderDnsCompareBlock(c){
  const ok      = !!(c && c.success);
  const cls     = ok ? 'ok' : 'bad';
  const mark    = ok ? '✓' : '✗';
  const elapsed = (c && c.elapsed_ms != null)
                    ? Number(c.elapsed_ms).toFixed(1) + ' ms' : '—';
  let html = `<div class="dns-cmp-verdict ${cls}">
    <span class="dns-cmp-mark">${mark}</span>
    <span>${esc((c && c.conclusion) || '')}</span>
    <span class="dns-cmp-elapsed">总耗时 ${elapsed}</span>
  </div>`;

  html += _renderDnsCompareDiffTable(_dnsCompareNormalizeDiff(c));

  const sides = _dnsCompareNormalizeSides(c);
  if(!sides.length){
    // Should never happen for a valid run — but a stray cancelled /
    // validation-error history row could theoretically reach here.
    // Render an inert hint rather than a half-empty layout.
    html += `<div class="dns-cmp-msg" style="margin-top:8px">（无可展示的解析结果）</div>`;
    return html;
  }

  const chainSigs = sides.map(s => _dnsChainSig(s.result || {}));
  // Auto-expand whenever there's something to inspect: any verdict
  // failure, OR any two sides whose CNAME chains differ even when the
  // final answer set agrees (CDN aliasing drift is itself a finding).
  const allSameChain = chainSigs.every(sig => sig === chainSigs[0]);
  const expand = !ok || !allSameChain;

  sides.forEach((s, idx) => {
    const kind = (idx === 0) ? 'sys' : 'cus';
    const label = s.label
      || (s.result && s.result.resolver)
      || (idx === 0 ? '系统默认' : '自定义');
    html += _renderDnsCompareDetail(s.result || {}, label, kind, expand);
  });
  return html;
}

// ════════════════════════════════════════════════════════════════════
// Phase 3C-4 — authority-trace renderer.
//
// Inputs come from src.dns_diag.dns_auth_to_dict.  Layout:
//   ① verdict bar (reuses .dns-cmp-verdict palette)
//   ② per-hop tree (monospace; one block per hop with optional
//      indented sub-lines for NS list / glue / message / error)
//   ③ final answer table (.dns-ans-tbl, with "追踪此 IP" buttons
//      for A/AAAA so the operator can pivot straight to tracert)
//   ④ inline conclusion echo (verdict bar repeats up top, but a
//      long tree would push it off-screen so we keep a footer copy)
// ════════════════════════════════════════════════════════════════════

// Pick the per-hop CSS class.  ok = success, bad = error_code set
// (timeout / nxdomain / servfail / glue_missing / …), warn = success
// but with a notable message (e.g. glue borrow happened).
function _dnsAuthHopCls(hop){
  if(!hop || hop.success === false) return 'bad';
  // A "warn" hop is success-with-caveat — currently the glue-borrow
  // case where the engine had to ask the system recursive resolver
  // for the NS hostname's A record before continuing.  Identifying
  // it by inspecting message text is brittle but the engine doesn't
  // emit a separate boolean; we accept the slight coupling because
  // the message string is itself part of the public payload.
  const msg = (hop.message || '');
  if(msg.indexOf('glue 缺失') >= 0) return 'warn';
  return 'ok';
}

// Phase 3C-4 rev2 — format ONE glue string "ns1.x → 1.2.3.4" into
// a coloured trio so the IP stands out from the hostname.  Falls
// back to plain escaped text if the engine ever sends an unexpected
// shape (we never want a malformed glue line to break rendering).
function _renderDnsAuthGlueLine(s){
  const m = /^(.+?)\s*(?:→|->)\s*(.+)$/.exec(String(s || ''));
  if(!m) return esc(String(s || ''));
  return '<span class="glue-name">' + esc(m[1].trim()) + '</span>'
       + '<span class="glue-arrow">→</span>'
       + '<span class="glue-ip">' + esc(m[2].trim()) + '</span>';
}

// Render ONE hop as a self-contained <div> block.  Rev2: no literal
// "\n" text nodes anywhere — every line is a real block element so
// the browser handles spacing via CSS, eliminating the stray blank
// row between hops that the rev1 pre-wrap layout produced.  Depth
// indentation is also dropped: the C-side reference renders every
// hop flush-left with the tree-glyph prefix, and that reads better
// than a staircase for traces deeper than 4 hops.
function _renderDnsAuthHop(hop, isLast, idx, totalHops){
  if(!hop) return '';
  const cls     = _dnsAuthHopCls(hop);
  const prefix  = isLast ? '└─' : '├─';
  const zoneTxt = (hop.zone === '.') ? '. (root)' : (hop.zone || '');
  const host    = hop.server_host || '';
  const ip      = hop.server_ip   || '';
  let destHtml  = '';
  if(host && ip){
    destHtml = '→ <b>' + esc(host) + '</b> (' + esc(ip) + ')';
  } else if(host){
    destHtml = '→ <b>' + esc(host) + '</b>';
  } else if(ip){
    destHtml = '→ ' + esc(ip);
  } else {
    destHtml = '→ —';
  }
  const rttTxt  = (hop.rtt_ms != null && hop.rtt_ms >= 0)
                    ? Number(hop.rtt_ms).toFixed(1) + ' ms' : '—';
  const rcodeBit = hop.rcode ? (' · ' + esc(hop.rcode)) : '';

  let html = '<div class="dns-auth-hop ' + cls + '">'
           +   '<div class="dns-auth-hop-head">'
           +     '<span class="dns-auth-prefix">' + esc(prefix) + '</span>'
           +     '<span class="dns-auth-zone">' + esc(zoneTxt) + '</span>'
           +     '<span class="dns-auth-dest">' + destHtml + '</span>'
           +     '<span class="dns-auth-rtt">' + esc(rttTxt) + rcodeBit + '</span>'
           +   '</div>';

  // Body lines — each is its own <div>.  Order: answer summary first
  // (always present), then NS list (when long enough to expand),
  // then glue lines (one per row, IP highlighted), then engine
  // message (only when it adds info beyond the summary).
  const bodyParts = [];

  if(hop.answer_summary){
    bodyParts.push(
      '<div class="dns-auth-line">'
      + '<span class="info">' + esc(hop.answer_summary) + '</span>'
      + '</div>');
  }

  const ref = Array.isArray(hop.referral_ns) ? hop.referral_ns : [];
  if(ref.length > 2){
    const namesHtml = ref.map(n =>
      '<span class="ns-name">' + esc(n.replace(/\.$/, '')) + '</span>'
    ).join(', ');
    bodyParts.push(
      '<div class="dns-auth-line">'
      + '<span class="lbl">NS:</span>' + namesHtml
      + '</div>');
  }

  const glue = Array.isArray(hop.glue) ? hop.glue : [];
  glue.forEach(g => {
    bodyParts.push(
      '<div class="dns-auth-line">'
      + '<span class="lbl">glue:</span>' + _renderDnsAuthGlueLine(g)
      + '</div>');
  });

  if(hop.message && hop.message !== hop.answer_summary){
    const msgCls = (cls === 'bad') ? 'err'
                 : (cls === 'warn' ? 'warn' : 'info');
    bodyParts.push(
      '<div class="dns-auth-line">'
      + '<span class="' + msgCls + '">' + esc(hop.message) + '</span>'
      + '</div>');
  }

  if(bodyParts.length){
    html += '<div class="dns-auth-hop-body">' + bodyParts.join('') + '</div>';
  }
  html += '</div>';
  return html;
}

function _renderDnsAuthBlock(r){
  const ok      = !!(r && r.success);
  const cls     = ok ? 'ok' : 'bad';
  const mark    = ok ? '✓' : '✗';
  const elapsed = (r && r.elapsed_ms != null)
                    ? Number(r.elapsed_ms).toFixed(1) + ' ms' : '—';

  // Verdict bar — reuse the compare palette so the colour language is
  // consistent across the three result modes (single / compare /
  // trace_auth).  Trailing 总耗时 metric mirrors compare exactly.
  let html = `<div class="dns-cmp-verdict ${cls}">
    <span class="dns-cmp-mark">${mark}</span>
    <span>${esc((r && r.message) || '')}</span>
    <span class="dns-cmp-elapsed">总耗时 ${elapsed}</span>
  </div>`;

  // Metadata strip — target / qtype / hop count / CNAME chain tail
  // (when the engine did a one-step follow).  Kept compact so it sits
  // above the tree without competing with the verdict bar.
  const hops   = Array.isArray(r.hops) ? r.hops : [];
  const chain  = Array.isArray(r.cname_chain) ? r.cname_chain : [];
  let cnamePart = '';
  if(chain.length >= 2){
    cnamePart = `<span>CNAME 链: <b>${esc(chain.join(' → '))}</b></span>`;
  }
  html += `<div class="dns-auth-summary">
    <span>目标: <b>${esc((r && r.target) || '')}</b></span>
    <span>类型: <b>${esc((r && r.qtype) || '')}</b></span>
    <span>跳数: <b>${hops.length}</b></span>
    ${cnamePart}
  </div>`;

  // Tree.  Rev2: each hop is a <div class="dns-auth-hop"> with its
  // own header + body blocks; CSS handles inter-hop spacing via the
  // adjacent-sibling dashed-top border, so we just concatenate.
  if(!hops.length){
    html += `<div class="dns-auth-tree" style="color:var(--dim);
        font-style:italic">（未生成任何追踪步骤）</div>`;
  } else {
    let inner = '';
    hops.forEach((h, i) => {
      const isLast = (i === hops.length - 1);
      inner += _renderDnsAuthHop(h, isLast, i, hops.length);
    });
    html += `<div class="dns-auth-tree">${inner}</div>`;
  }

  // Final answer table — only rendered when the run succeeded AND
  // produced concrete answers.  For failures the tree's last (red)
  // hop already carries the error; an empty table would be noise.
  const answers = Array.isArray(r.answers) ? r.answers : [];
  if(ok && answers.length){
    html += '<table class="dns-ans-tbl" style="margin-top:14px"><thead><tr>'
          + '<th>名称</th><th style="width:60px">类型</th><th>值</th>'
          + '<th style="width:60px;text-align:right">TTL</th>'
          + '<th style="width:100px">操作</th></tr></thead><tbody>';
    answers.forEach(a => {
      const ttl = (a.ttl != null) ? a.ttl : '—';
      const action = (a.qtype === 'A' || a.qtype === 'AAAA')
        ? `<button class="dns-trace-btn"
                   onclick="traceDnsIp('${esc(a.value)}')">追踪此 IP</button>`
        : '';
      html += `<tr>
        <td>${esc(a.name)}</td>
        <td class="dns-qtype">${esc(a.qtype)}</td>
        <td style="overflow-wrap:anywhere;word-break:normal;max-width:520px">${esc(a.value)}</td>
        <td class="dns-ttl">${ttl}</td>
        <td>${action}</td>
      </tr>`;
    });
    html += '</tbody></table>';
  }

  // Footer conclusion echo — same wording as the verdict bar so a
  // long tree (10+ hops) doesn't force the operator to scroll back
  // up to see the verdict.
  const conclusion = (r && r.message) || (ok ? '权威追踪成功'
                                              : '权威追踪失败');
  html += `<div class="icmp-conclusion ${cls}" style="margin-top:12px">
    <div style="font-weight:600;margin-bottom:4px">📋 结论</div>
    <div>${esc(conclusion)}</div>
  </div>`;
  return html;
}

// DNS history list (Phase 3A) — UI mirror of the ICMP history block
// above.  Same fetch → render → toggle → replay pattern.
//
// Keep a JS-side cache of the last fetched rows so replayDnsHistory()
// can rebuild a result block from a stable index without re-querying
// the server (avoids a sub-second race where a fresh diag could shift
// rows under the operator's click).
let _dnsHistRows = [];

function loadDnsHistory(){
  if(!currentDiagTid){ _showDnsError('未选中节点'); return; }
  const out = document.getElementById('dns-output');
  if(out) out.innerHTML =
    '<div style="padding:14px;color:var(--dim);text-align:center">⏳ 加载历史...</div>';
  fetch('/api/dns_diag/history?tid=' + encodeURIComponent(currentDiagTid) + '&limit=50')
    .then(r => r.json())
    .then(rows => { _dnsHistRows = rows || []; renderDnsHistory(_dnsHistRows); })
    .catch(e => {
      if(out) out.innerHTML =
        `<div class="icmp-conclusion bad">历史加载失败：${esc(e&&e.message||e)}</div>`;
    });
}

function renderDnsHistory(rows){
  const out = document.getElementById('dns-output');
  if(!out) return;
  if(!rows.length){
    out.innerHTML =
      '<div style="padding:20px;color:var(--dim);text-align:center;font-size:12px">'+
      '暂无 DNS 诊断历史记录<br><span style="font-size:11px">'+
      '运行一次诊断即可在此查看历史</span></div>';
    return;
  }
  let html = `<div style="margin-top:10px;font-size:12px;color:var(--dim)">
    最近 ${rows.length} 条 DNS 诊断记录（含 C 端与 Web 端）
  </div>`;
  rows.forEach((row, idx) => { html += _renderDnsHistRow(row, idx); });
  out.innerHTML = html;
}

function _renderDnsHistRow(row, idx){
  const dt = new Date((row.created_ts || 0) * 1000);
  const timeStr = dt.getFullYear() + '-'
    + String(dt.getMonth()+1).padStart(2, '0') + '-'
    + String(dt.getDate()).padStart(2, '0') + ' '
    + String(dt.getHours()).padStart(2, '0') + ':'
    + String(dt.getMinutes()).padStart(2, '0') + ':'
    + String(dt.getSeconds()).padStart(2, '0');
  const qtype = row.qtype || '?';
  const ok    = !!row.success;
  const rcode = row.rcode || '—';
  const n_ans = row.answer_count || 0;
  const ems   = (row.elapsed_ms != null)
                  ? Number(row.elapsed_ms).toFixed(1) + ' ms' : '—';
  const details = row.details || {};
  const isCompare   = (details.mode === 'compare');
  const isTraceAuth = (details.mode === 'trace_auth');
  // Phase 3C-5: dnssec rows persist rcode as an UPPER-CASE four-state
  // verdict ("SECURE"/"INSECURE"/"BOGUS"/"INDETERMINATE").  The list-
  // row colour mapping diverges from the binary success/fail pill
  // used by all other modes because:
  //   • INSECURE rows have row.success=false (strict per Q2) yet are
  //     NOT incidents — colouring them red would scream "DNS broken"
  //     when reality is "zone未签名".  We override to grey.
  //   • BOGUS rows have row.success=false AND are real incidents — red.
  //   • INDETERMINATE rows are "we can't tell" — amber.
  //   • SECURE rows naturally land on green via the default success
  //     path (which is correct: ok=true).
  const isDnssec    = (details.mode === 'dnssec');
  let stCol, stBg, stTxt;
  if(isDnssec){
    if(rcode === 'SECURE'){
      stCol = 'var(--green)'; stBg = 'rgba(34,197,94,.15)'; stTxt = '已签名';
    } else if(rcode === 'INSECURE'){
      stCol = 'var(--dim)';   stBg = 'rgba(127,127,127,.18)'; stTxt = '未签名';
    } else if(rcode === 'BOGUS'){
      stCol = 'var(--red)';   stBg = 'rgba(239,68,68,.15)';   stTxt = '签名错误';
    } else {
      stCol = 'var(--orange)';stBg = 'rgba(245,158,11,.15)';  stTxt = '未判定';
    }
  } else {
    stCol = ok ? 'var(--green)' : 'var(--red)';
    stBg  = ok ? 'rgba(34,197,94,.15)' : 'rgba(239,68,68,.15)';
    stTxt = ok ? (rcode === 'MISMATCH' ? '一致' : '成功') : '失败';
  }
  // Phase 3B: only show the NS chip when the run was against a custom
  // resolver — system-default runs would clutter every row with the
  // same uninformative tag.  Pre-3B rows have no details.nameserver
  // key, so the falsy fallback restores the legacy look.
  //
  // Phase 3C-1: for compare rows the chip means "对照解析器", not the
  // single-resolver override.  The chip text adapts so the operator
  // can tell at a glance which kind of row they're looking at.
  //
  // Phase 3C-3 D4: N-resolver compare rows show "vs <first> +K" so the
  // chip stays short; full list lives in the `title` for hover.
  let customNss = [];
  if(Array.isArray(details.nameservers) && details.nameservers.length){
    customNss = details.nameservers.slice();
  } else if(details.nameserver){
    customNss = [details.nameserver];
  }
  // Phase 3C-4: trace_auth runs have no operator-supplied resolver,
  // so suppress the NS chip entirely for those rows — the "权威"
  // badge (below) already conveys "this row used the iterative walk".
  if(isTraceAuth) customNss = [];
  let nsChipText = '';
  let nsChipTitle = '';
  if(customNss.length){
    const prefix = isCompare ? 'vs ' : 'NS: ';
    if(customNss.length === 1){
      nsChipText  = prefix + customNss[0];
      nsChipTitle = isCompare ? '对照解析器：' + customNss[0]
                              : '使用了自定义解析器：' + customNss[0];
    } else {
      nsChipText  = prefix + customNss[0] + ' +' + (customNss.length - 1);
      nsChipTitle = '对照解析器：' + customNss.join('、');
    }
  }
  const nsChip = nsChipText
    ? `<span style="font-size:11px;font-weight:600;color:var(--blue2);
              background:rgba(96,165,250,.15);border-radius:3px;
              padding:1px 7px;font-family:monospace"
        title="${esc(nsChipTitle)}"
        >${esc(nsChipText)}</span>`
    : '';
  // Phase 3C-1 compare badge — purple-ish so it's instantly
  // distinguishable from the success/fail status pill.  Phase 3C-4
  // adds a parallel green-ish 权威 badge for trace_auth runs (mutually
  // exclusive with the compare badge by construction — each row
  // carries exactly one mode).
  const compareBadge = isCompare
    ? `<span style="font-size:11px;font-weight:600;color:#a78bfa;
              background:rgba(167,139,250,.15);border-radius:3px;
              padding:1px 7px" title="多解析器对比">对比</span>`
    : '';
  const authBadge = isTraceAuth
    ? `<span style="font-size:11px;font-weight:600;color:#34d399;
              background:rgba(52,211,153,.15);border-radius:3px;
              padding:1px 7px" title="权威追踪（从根迭代）">🌲 权威</span>`
    : '';
  // Phase 3C-5 — fourth mode badge.  Cyan-ish to keep four mutually-
  // exclusive badges (none / 对比 / 权威 / DNSSEC) instantly tellable
  // apart at a glance in a long history list.
  const dnssecBadge = isDnssec
    ? `<span style="font-size:11px;font-weight:600;color:#22d3ee;
              background:rgba(34,211,238,.15);border-radius:3px;
              padding:1px 7px" title="DNSSEC 客户端验证">🔐 DNSSEC</span>`
    : '';
  const bodyId = 'dns-hist-body-' + idx;
  const chevId = 'dns-hist-chev-' + idx;
  // Phase 3C-5 layout fix — the original layout put `flex:1` on the
  // timestamp span which made the time box absorb all spare width.
  // When trailing items grew wider (e.g. the 🔐 DNSSEC badge is wider
  // than 🌲 权威, and the domain column has its own max-width 240),
  // the row was forced to squeeze somewhere — flex picked the time
  // span because of `min-width:0`, and without `overflow:hidden` the
  // nowrap timestamp text bled visually into the next badge.  The
  // fix is structural: pin the timestamp at its natural width, let
  // the DOMAIN absorb spare width (it's the only span here that
  // already has overflow + ellipsis), and use margin-left:auto on
  // the status pill so the right group stays right-aligned even
  // when the domain is short.
  return `<div class="icmp-hist-item">
    <div class="icmp-hist-head" onclick="toggleDnsHist('${bodyId}','${chevId}')">
      <span style="font-size:12px;color:var(--dim);white-space:nowrap;
            flex:0 0 auto">${esc(timeStr)}</span>
      ${compareBadge}
      ${authBadge}
      ${dnssecBadge}
      <span class="icmp-hist-mode">${esc(qtype)}</span>
      <span style="font-size:11px;color:var(--text);font-family:monospace;
            overflow:hidden;text-overflow:ellipsis;white-space:nowrap;
            flex:1 1 auto;min-width:60px;max-width:260px"
            title="${esc(row.domain || '')}">${esc(row.domain || '')}</span>
      <span style="font-size:11px;font-weight:600;color:${stCol};
            background:${stBg};border-radius:3px;padding:1px 7px;
            margin-left:auto;flex:0 0 auto;white-space:nowrap">${stTxt}</span>
      <span style="font-size:11px;color:var(--dim);font-family:monospace;
            flex:0 0 auto;white-space:nowrap">
        ${esc(rcode)} · ${n_ans} 条 · ${ems}
      </span>
      ${nsChip}
      <span id="${chevId}" style="font-size:12px;color:var(--dim);
            transition:transform .2s;display:inline-block;flex:0 0 auto">›</span>
    </div>
    <div id="${bodyId}" class="icmp-hist-body">${_renderDnsHistDetail(row, idx)}</div>
  </div>`;
}

function _renderDnsHistDetail(row, idx){
  // Build a result-shaped object from the stored details snapshot.
  // Phase 3C-1 split: compare rows have details.mode === 'compare'
  // and carry inner system/custom dicts; single rows are the
  // pre-3C-1 shape and render unchanged.  Phase 3C-4 adds the
  // trace_auth shape which carries the hops[] array directly.  Both
  // branches expose the same "▶ 回放到结果区" button which goes
  // through replayDnsHistory.
  const d = row.details || {};
  let html = `<div style="text-align:right;margin-bottom:6px">
    <button class="dns-trace-btn" onclick="replayDnsHistory(${idx})">▶ 回放到结果区</button>
  </div>`;
  if(d.mode === 'compare'){
    html += _renderDnsCompareBlock(_dnsCompareFromDetails(row, d));
  } else if(d.mode === 'trace_auth'){
    html += _renderDnsAuthBlock(_dnsAuthFromDetails(row, d));
  } else if(d.mode === 'dnssec'){
    // Phase 3C-5: dnssec history detail.  The row stores the full
    // dnssec_to_dict payload (chain + answers + status + ad_flag),
    // so inflation is a near-passthrough — same shape as
    // _dnsAuthFromDetails but with a different field set.
    html += _renderDnsSecBlock(_dnsSecFromDetails(row, d));
  } else {
    const r = _dnsSingleFromDetails(row, d);
    html += _renderDnsResultBlock(r);
  }
  return html;
}

// Helpers: build a result-shaped object from either a top-level
// single-result details dict or the system/custom sub-dicts inside
// a compare-result details dict.  Pure functions, no DOM access —
// reused by the history detail panel AND replayDnsHistory.
function _dnsSingleFromDetails(row, d){
  return {
    target:      d.target      || (row && row.domain),
    qtype:       d.qtype       || (row && row.qtype),
    resolver:    d.resolver    || (row && row.resolver),
    success:     (d.success !== undefined) ? d.success
                                            : (row && row.success),
    rcode:       d.rcode       || (row && row.rcode),
    elapsed_ms:  (d.elapsed_ms != null) ? d.elapsed_ms
                                         : (row && row.elapsed_ms),
    cname_chain: d.cname_chain || [],
    answers:     d.answers     || [],
    trace_steps: d.trace_steps || [],
    message:     d.message     || (row && row.message),
  };
}
function _dnsCompareFromDetails(row, d){
  // Phase 3C-3: detect v2 details (compare_version >= 2 + sides[]),
  // otherwise fall back to the 3C-1 system/custom pair.  Output is
  // always the v2 shape — renderers downstream only ever see v2.
  const ver = (d && typeof d.compare_version === 'number')
                ? d.compare_version : 1;
  let sides = [];
  if(ver >= 2 && Array.isArray(d.sides) && d.sides.length){
    sides = d.sides.map((s, idx) => ({
      key:    String(s.key || (idx === 0 ? 'system' : ('r' + idx))),
      label:  String(s.label || ''),
      result: _dnsSingleFromDetails(row, s.result || {}),
    }));
  } else {
    // 3C-1 fallback — synthesise canonical {system, custom} sides.
    if(d.system){
      sides.push({
        key: 'system',
        label: (d.system.resolver || '系统默认'),
        result: _dnsSingleFromDetails(row, d.system || {}),
      });
    }
    if(d.custom){
      sides.push({
        key: 'custom',
        label: (d.custom.resolver || d.nameserver || '自定义'),
        result: _dnsSingleFromDetails(row, d.custom || {}),
      });
    }
  }
  // 3C-3: collect the canonical custom-IP list.  v2 carries it
  // verbatim; v1 fallback synthesises from the single `nameserver`.
  let nameservers = [];
  if(Array.isArray(d.nameservers) && d.nameservers.length){
    nameservers = d.nameservers.slice();
  } else if(d.nameserver){
    nameservers = [d.nameserver];
  }
  return {
    mode:           'compare',
    compare_version: 2,
    target:         d.target     || (row && row.domain),
    qtype:          d.qtype      || (row && row.qtype),
    trace_chain:    !!d.trace_chain,
    nameserver:     d.nameserver || (nameservers[0] || ''),
    nameservers:    nameservers,
    success:        !!d.success,
    conclusion:     d.conclusion || (row && row.message) || '',
    elapsed_ms:     (d.elapsed_ms != null) ? d.elapsed_ms
                                            : (row && row.elapsed_ms),
    sides:          sides,
    // v1 alias keys so any straggling renderer still reads sides[0/1].
    system:         sides.length > 0 ? sides[0].result : undefined,
    custom:         sides.length > 1 ? sides[1].result : undefined,
    // Diff payload: pass through if v2-shape (array side_labels);
    // _dnsCompareNormalizeDiff handles dict / missing cases.
    diff:           (d.diff && Array.isArray(d.diff.side_labels))
                      ? d.diff : null,
  };
}

// Phase 3C-4 — inflate a trace_auth history row's details into a
// fully-shaped DnsAuthResult dict.  The stored payload is already in
// the dns_auth_to_dict shape, so most fields pass through verbatim;
// we only fill in defaults for fields that may be missing on older /
// edge-case rows so the renderer never sees `undefined.hops`.
function _dnsAuthFromDetails(row, d){
  d = d || {};
  return {
    mode:        'trace_auth',
    target:      d.target      || (row && row.domain) || '',
    qtype:       d.qtype       || (row && row.qtype)  || '',
    success:     !!d.success,
    rcode:       d.rcode       || (row && row.rcode)  || '',
    elapsed_ms:  (d.elapsed_ms != null) ? d.elapsed_ms
                                         : (row && row.elapsed_ms),
    message:     d.message     || (row && row.message) || '',
    cname_chain: Array.isArray(d.cname_chain) ? d.cname_chain.slice() : [],
    hops:        Array.isArray(d.hops)    ? d.hops.slice()    : [],
    answers:     Array.isArray(d.answers) ? d.answers.slice() : [],
  };
}

// Phase 3C-5 — inflate a dnssec history row's details into a fully-
// shaped DnsSecResult dict.  Mirrors _dnsAuthFromDetails: the stored
// payload IS the canonical dnssec_to_dict shape, so most fields pass
// through verbatim and we only fill defaults for fields that may be
// absent on edge-case rows (e.g. a future schema bump that drops a
// field — defensive coding rather than expected behaviour).
function _dnsSecFromDetails(row, d){
  d = d || {};
  return {
    mode:        'dnssec',
    target:      d.target      || (row && row.domain) || '',
    qtype:       d.qtype       || (row && row.qtype)  || '',
    resolver:    d.resolver    || (row && row.resolver) || '',
    nameserver:  d.nameserver  || '',
    // Persisted rcode is UPPER-CASE four-state ("SECURE"/"INSECURE"/
    // "BOGUS"/"INDETERMINATE"); status carries the same value in
    // lower-case form.  Prefer details.status when present; fall
    // back to lower-casing the row's rcode for backward-compat.
    status:      (d.status || (row && row.rcode) || 'indeterminate')
                   .toString().toLowerCase(),
    success:     (d.success !== undefined) ? !!d.success
                                            : (row && !!row.success),
    chain:       Array.isArray(d.chain)   ? d.chain.slice()   : [],
    answers:     Array.isArray(d.answers) ? d.answers.slice() : [],
    rrsig_ok:    !!d.rrsig_ok,
    ad_flag:     (d.ad_flag === null || d.ad_flag === undefined)
                   ? null : !!d.ad_flag,
    elapsed_ms:  (d.elapsed_ms != null) ? d.elapsed_ms
                                         : (row && row.elapsed_ms),
    message:     d.message     || (row && row.message) || '',
    cname_chain: Array.isArray(d.cname_chain) ? d.cname_chain.slice() : [],
  };
}

// Phase 3C-5 — render block for a DNSSEC validation result.
//
// Layout (top → bottom):
//   1. Verdict bar — large, colour-coded by 4-state status.  Carries:
//        ✓/○/✗/⚠ glyph  + STATUS tag  + Chinese one-liner  + elapsed
//   2. Resolver + target metadata row (small dim text).
//   3. CNAME chain (only when non-trivial) — same monospace tree
//      style as single mode for visual consistency.
//   4. Verification chain TABLE — one row per zone, root → leaf.
//      Columns: zone | ✓/○/✗ | DS / DNSKEY status | note.
//   5. AD / RRSIG / answer-count metadata strip.
//   6. Answer table (reuses .dns-ans-tbl) when answers are present —
//      even on INSECURE / BOGUS we still show what was returned so
//      the operator sees the actual data.
//   7. Conclusion echo at the bottom (mirrors trace_auth's layout).
function _renderDnsSecBlock(r){
  const status = (r.status || 'indeterminate').toLowerCase();
  const statusMeta = {
    'secure':        { tag: 'SECURE',        mark: '✓',
                        cn: '已签名 · 验证通过' },
    'insecure':      { tag: 'INSECURE',      mark: '○',
                        cn: '未部署 DNSSEC（不算失败）' },
    'bogus':         { tag: 'BOGUS',         mark: '✗',
                        cn: '签名错误（异常）' },
    'indeterminate': { tag: 'INDETERMINATE', mark: '⚠',
                        cn: '未能判定' },
  }[status] || { tag: 'INDETERMINATE', mark: '⚠', cn: '未能判定' };
  const elapsed = (r.elapsed_ms != null)
                    ? Number(r.elapsed_ms).toFixed(1) + ' ms' : '—';

  let html = `<div class="dns-sec-verdict ${esc(status)}">
    <span class="dns-sec-mark">${statusMeta.mark}</span>
    <span class="dns-sec-tag">${statusMeta.tag}</span>
    <span>${esc(r.message || statusMeta.cn)}</span>
    <span class="dns-sec-elapsed">${elapsed}</span>
  </div>`;

  // Metadata row — kept compact; mirrors trace_auth's summary strip.
  // The CNAME chain (when present) is displayed BELOW this strip in
  // the same monospace-pre style as single-mode for visual parity.
  html += `<div class="dns-auth-summary">
    <span>解析器（传输）: <b>${esc(r.resolver || '系统默认')}</b></span>
    <span>目标: <b>${esc(r.target || '')}</b></span>
    <span>类型: <b>${esc(r.qtype || '')}</b></span>
  </div>`;

  // CNAME chain — only render when non-trivial (>1 element).  Reusing
  // the .dns-chain-pre class keeps the visual identical to single
  // mode so an operator's eye doesn't have to re-learn the layout.
  const cn = Array.isArray(r.cname_chain) ? r.cname_chain : [];
  if(cn.length > 1){
    let pre = '';
    for(let i = 0; i < cn.length; i++){
      pre += (i === 0 ? '' : '└─ ') + esc(cn[i]) + '\n';
    }
    html += `<div class="dns-chain-label" style="margin-top:10px">CNAME 链:</div>
      <pre class="dns-chain-pre">${pre}</pre>`;
  }

  // Verification chain table.  Each row shows one zone with its DS /
  // DNSKEY pass/fail combo.  Empty chain (rare — only happens on
  // very early INDETERMINATE) gets a placeholder so the operator
  // doesn't see a bare table header with no rows.
  const chain = Array.isArray(r.chain) ? r.chain : [];
  if(chain.length){
    let rows = '';
    for(let i = 0; i < chain.length; i++){
      const link = chain[i] || {};
      // Decide the row mark + class.  Precedence:
      //   • dnskey_ok && (root OR ds_ok)            → ✓ ok
      //   • has_ds && !ds_ok                        → ✗ bad (DS mismatch)
      //   • !has_ds && i > 0                        → ○ insecure
      //   • has_ds && !has_dnskey                   → ✗ bad
      //   • otherwise                               → ⚠ pending
      let mark = '⚠', markCls = 'sec-mark-pending', rowCls = '';
      if(link.dnskey_ok && (i === 0 || link.ds_ok)){
        mark = '✓'; markCls = 'sec-mark-ok';
      } else if(i > 0 && !link.has_ds){
        mark = '○'; markCls = 'sec-mark-insecure'; rowCls = 'row-insecure';
      } else if(link.has_ds && (!link.ds_ok || !link.dnskey_ok)){
        mark = '✗'; markCls = 'sec-mark-bad';     rowCls = 'row-bad';
      } else if(link.has_ds && !link.has_dnskey){
        mark = '✗'; markCls = 'sec-mark-bad';     rowCls = 'row-bad';
      }

      // Status cell — at-a-glance "DS=✓ DNSKEY=✓" so the operator
      // doesn't have to read the long note column to know which
      // half broke.  Root has no parent → DS column shows "—".
      const dsCell  = (i === 0)
                       ? '<span style="color:var(--dim)">— (root)</span>'
                       : (link.has_ds
                          ? (link.ds_ok
                              ? '<span style="color:var(--green)">DS ✓</span>'
                              : '<span style="color:var(--red)">DS ✗</span>')
                          : '<span style="color:var(--dim)">无 DS</span>');
      const keyCell = link.has_dnskey
                        ? (link.dnskey_ok
                            ? '<span style="color:var(--green)">DNSKEY ✓</span>'
                            : '<span style="color:var(--red)">DNSKEY ✗</span>')
                        : '<span style="color:var(--dim)">无 DNSKEY</span>';

      rows += `<tr class="${rowCls}">
        <td class="sec-mark ${markCls}">${mark}</td>
        <td class="sec-zone">${esc(link.zone || '')}</td>
        <td>${dsCell} ${keyCell}</td>
        <td class="sec-algo">${esc(link.algorithm || '')}</td>
        <td class="sec-note">${esc(link.note || '')}</td>
      </tr>`;
    }
    html += `<div class="dns-sec-chain">
      <table>
        <thead><tr>
          <th style="width:36px"></th>
          <th>Zone</th>
          <th style="width:180px">DS / DNSKEY</th>
          <th style="width:130px">摘要算法</th>
          <th>说明</th>
        </tr></thead>
        <tbody>${rows}</tbody>
      </table>
    </div>`;
  }

  // AD / RRSIG / answer-count strip.  Surfaces the AD bit even when
  // we DIDN'T trust it for the verdict, so an operator comparing
  // "what does the resolver claim" vs "what does my own validator
  // say" can spot disagreements (e.g. resolver AD=1 but our verdict
  // BOGUS = resolver is misbehaving / broken).
  const adRaw = r.ad_flag;
  let adTxt;
  if(adRaw === null || adRaw === undefined){
    adTxt = '<span class="sec-meta-ad-hint">未收到</span>';
  } else {
    adTxt = `<b>${adRaw ? 'AD=1' : 'AD=0'}</b>` +
            ` <span class="sec-meta-ad-hint">(仅参考，验证不依赖此位)</span>`;
  }
  const rrsigTxt = `<b>${r.rrsig_ok ? '✓ 通过' : '✗ 未通过'}</b>`;
  const ansCount = Array.isArray(r.answers) ? r.answers.length : 0;
  html += `<div class="dns-sec-meta">
    <span>解析器返回 AD 位: ${adTxt}</span>
    <span>本机 RRSIG 校验: ${rrsigTxt}</span>
    <span>答案条数: <b>${ansCount}</b></span>
  </div>`;

  // Answer table — reuse the single-mode rendering helper if we have
  // one, otherwise inline a minimal table.  Even on BOGUS / INSECURE
  // we display the records so the operator can see what's being
  // served; the verdict bar carries the security caveat.
  if(ansCount){
    html += `<table class="dns-ans-tbl" style="margin-top:10px">
      <thead><tr><th>NAME</th><th>TYPE</th><th>VALUE</th><th>TTL</th></tr></thead>
      <tbody>`;
    for(const a of r.answers){
      html += `<tr>
        <td>${esc(a.name || '')}</td>
        <td>${esc(a.qtype || '')}</td>
        <td>${esc(a.value || '')}</td>
        <td>${a.ttl != null ? a.ttl : ''}</td>
      </tr>`;
    }
    html += `</tbody></table>`;
  }

  return html;
}

function toggleDnsHist(bodyId, chevId){
  const el = document.getElementById(bodyId);
  const chev = document.getElementById(chevId);
  if(!el) return;
  const open = (el.style.display !== 'block');
  el.style.display = open ? 'block' : 'none';
  if(chev) chev.style.transform = open ? 'rotate(90deg)' : 'rotate(0deg)';
}

function replayDnsHistory(idx){
  const row = _dnsHistRows[idx];
  if(!row) return;
  const d = row.details || {};
  const out = document.getElementById('dns-output');
  if(!out) return;
  if(d.mode === 'compare'){
    out.innerHTML = _renderDnsCompareBlock(_dnsCompareFromDetails(row, d));
  } else if(d.mode === 'trace_auth'){
    out.innerHTML = _renderDnsAuthBlock(_dnsAuthFromDetails(row, d));
  } else if(d.mode === 'dnssec'){
    // Phase 3C-5 replay: same shape as the history detail panel —
    // _dnsSecFromDetails inflates the persisted dict.  We do NOT
    // re-execute the validation; replay is purely visual reconstruction
    // of the past run (consistent with single / compare / trace_auth).
    out.innerHTML = _renderDnsSecBlock(_dnsSecFromDetails(row, d));
  } else {
    out.innerHTML = _renderDnsResultBlock(_dnsSingleFromDetails(row, d));
  }
}

// (dns block end)

// Statistics
//
// Uptime stats are now AUTHORITATIVELY tracked on the server. The browser
// receives the cumulative totals (total, success, latency_sum, latency_n)
// in the SSE init event, then increments them locally on each update event.
// On page refresh, the next init carries the fresh server-side totals.
//
// This fixes the previous behaviour where the donut "uptime %" was actually
// "uptime % since you opened this browser tab" — refreshing the page or
// reconnecting always reset the counters to 0.
function accumStat(t){
  if(t.status==='paused')return;
  if(!stats[t.tid])stats[t.tid]={total:0,success:0,latency_avg:0,latency_n:0,
                                  latencies:[],label:t.label,ip:t.ip};
  const s=stats[t.tid];s.total++;s.label=t.label;s.ip=t.ip;
  if(t.status==='green'||t.status==='orange')s.success++;
  if(t.latency_ms!=null){
    s.latency_n+=1;
    // CMA: keeps average near the true value without summing large numbers
    if(!s.latency_avg)s.latency_avg=0;
    s.latency_avg+=(t.latency_ms-s.latency_avg)/s.latency_n;
    s.latencies.push(t.latency_ms);
    if(s.latencies.length>MAX_LAT)s.latencies.shift();}}

// Apply server-provided cumulative stats. Called on every init SSE event.
// Server-side totals OVERWRITE any in-browser totals — the server is the
// source of truth for cumulative uptime since program start.
function applyServerStats(serverStats){
  if(!serverStats)return;
  Object.keys(serverStats).forEach(tid=>{
    const srv=serverStats[tid];
    if(!stats[tid])stats[tid]={total:0,success:0,latency_sum:0,latency_n:0,
                                latencies:[],label:'',ip:''};
    // Sync label/ip from targets (HTTP nodes have long intervals so accumStat
    // may not have run yet; preload from targets as fallback)
    if(targets[tid]&&targets[tid].label){stats[tid].label=targets[tid].label;
                                         stats[tid].ip   =targets[tid].ip||'';}
    stats[tid].total       = srv.total       || 0;
    stats[tid].success     = srv.success     || 0;
    stats[tid].latency_n   = srv.latency_n   || 0;
    // latency_avg from server is the CMA; browser accumulates using same formula
    stats[tid].latency_avg = srv.latency_avg || 0;
    // keep latency_sum/latency_n for backward compat if server sends old format
    if(srv.latency_sum&&!srv.latency_avg&&srv.latency_n)
      stats[tid].latency_avg=(srv.latency_sum/srv.latency_n)||0;
  });
}


// ── 时间范围状态（全局，供所有统计功能共享）──────────────────────────────
let _statsStart = Math.floor(Date.now()/1000) - 3600;   // 默认近1小时
let _statsEnd   = Math.floor(Date.now()/1000);

function _fmt(ts){
  const d=new Date(ts*1000);
  const pad=n=>String(n).padStart(2,'0');
  return `${d.getFullYear()}-${pad(d.getMonth()+1)}-${pad(d.getDate())} `+
         `${pad(d.getHours())}:${pad(d.getMinutes())}`;
}
function _toInputVal(ts){
  // datetime-local 需要 "YYYY-MM-DDTHH:MM" 格式（本地时间）
  const d=new Date(ts*1000);
  const pad=n=>String(n).padStart(2,'0');
  return `${d.getFullYear()}-${pad(d.getMonth()+1)}-${pad(d.getDate())}`+
         `T${pad(d.getHours())}:${pad(d.getMinutes())}`;
}

// 点击快捷按钮：设置时间范围并立刻刷新
function setQuick(btn){
  document.querySelectorAll('.tbtn').forEach(b=>b.classList.remove('tsel'));
  btn.classList.add('tsel');
  const sec=parseInt(btn.dataset.sec);
  _statsEnd   = Math.floor(Date.now()/1000);
  _statsStart = _statsEnd - sec;
  // 同步写入 datetime-local 输入框
  const ds=document.getElementById('dt-start');
  const de=document.getElementById('dt-end');
  if(ds) ds.value=_toInputVal(_statsStart);
  if(de) de.value=_toInputVal(_statsEnd);
  _refreshStatsCharts();
  loadWaveform();   // refresh waveform with same time window
}

// 点击"查询"：从 datetime-local 读取并刷新
function applyTimeRange(){
  const ds=document.getElementById('dt-start');
  const de=document.getElementById('dt-end');
  if(!ds||!de||!ds.value||!de.value){alert('请选择完整的开始和结束时间');return;}
  _statsStart=Math.floor(new Date(ds.value).getTime()/1000);
  _statsEnd  =Math.floor(new Date(de.value).getTime()/1000);
  if(_statsStart>=_statsEnd){alert('开始时间必须早于结束时间');return;}
  // 取消快捷选中状态
  document.querySelectorAll('.tbtn').forEach(b=>b.classList.remove('tsel'));
  _refreshStatsCharts();
  loadWaveform();
}

function exportWaveform(){
  const sel   = document.getElementById('waveform-tid');
  const tid   = sel ? sel.value : '';
  if(!tid){ alert('请先选择节点'); return; }
  const label = sel.options[sel.selectedIndex]?.text || tid;
  const start = _statsStart || (Math.floor(Date.now()/1000)-3600);
  const end   = _statsEnd   || Math.floor(Date.now()/1000);
  const url   = `/api/waveform/export?tid=${encodeURIComponent(tid)}`
              + `&label=${encodeURIComponent(label)}`
              + `&start=${start}&end=${end}`;
  // Check for empty-data 404 before triggering download
  fetch(url, {method:'HEAD'}).then(r=>{
    if(r.status===404){ alert('该时间范围内无可导出数据'); return; }
    const a=document.createElement('a');
    a.href=url; a.download=''; a.click();
  }).catch(()=>{ const a=document.createElement('a'); a.href=url; a.download=''; a.click(); });
}

// ── Waveform chart ─────────────────────────────────────────────────────
let _waveChart = null;
let _waveReqId = 0;   // incremented on each loadWaveform call; stale responses discard themselves

// Chart.js plugin: paint outage bands behind the data and transition markers above it.
// Reads from chart.$alertOverlay (set by _renderWaveform when alert data arrives).
// Uses chart.$pointsTs (RTT-point timestamp array) to map UNIX ts → category-axis
// pixel positions via binary search.  Chart.js' category x-axis stores indices, not
// times, so we manually project; this also lets outage bands extend cleanly to the
// chart edges when the segment overlaps the time window boundary.
const waveformAlertOverlayPlugin = {
  id: 'waveformAlertOverlay',
  beforeDatasetsDraw(chart) {
    const data = chart.$alertOverlay;
    const ts   = chart.$pointsTs;
    if (!data || !ts || !ts.length) return;
    const {ctx, chartArea, scales} = chart;
    const xs = scales.x;
    if (!xs || !chartArea) return;

    const ts2x = (t) => {
      if (t <= ts[0])              return chartArea.left;
      if (t >= ts[ts.length - 1])  return chartArea.right;
      let lo = 0, hi = ts.length - 1;
      while (lo < hi) {
        const mid = (lo + hi) >> 1;
        if (ts[mid] < t) lo = mid + 1; else hi = mid;
      }
      // Linearly interpolate between two indices for smoother band edges
      const i = lo;
      if (i === 0) return xs.getPixelForValue(0);
      const t0 = ts[i - 1], t1 = ts[i];
      const x0 = xs.getPixelForValue(i - 1);
      const x1 = xs.getPixelForValue(i);
      const frac = t1 === t0 ? 0 : (t - t0) / (t1 - t0);
      return x0 + (x1 - x0) * frac;
    };

    ctx.save();
    // Red bands (severe — availability outage)
    ctx.fillStyle = 'rgba(239,68,68,0.13)';
    (data.red || []).forEach(seg => {
      const x1 = ts2x(seg.start);
      const x2 = Math.max(x1 + 1, ts2x(seg.end));
      ctx.fillRect(x1, chartArea.top, x2 - x1, chartArea.bottom - chartArea.top);
    });
    // Orange bands (performance warning) — drawn first below the red would
    // be incorrect because reds are higher severity; but since red & orange
    // segments are mutually exclusive in time, paint order is irrelevant.
    ctx.fillStyle = 'rgba(251,191,36,0.10)';
    (data.orange || []).forEach(seg => {
      const x1 = ts2x(seg.start);
      const x2 = Math.max(x1 + 1, ts2x(seg.end));
      ctx.fillRect(x1, chartArea.top, x2 - x1, chartArea.bottom - chartArea.top);
    });
    ctx.restore();
  },
  afterDatasetsDraw(chart) {
    const data = chart.$alertOverlay;
    const ts   = chart.$pointsTs;
    if (!data || !ts || !ts.length) return;
    const {ctx, chartArea, scales} = chart;
    const xs = scales.x;
    if (!xs || !chartArea) return;

    // Inline ts→x (same logic as above, kept local to avoid closure dep)
    const ts2x = (t) => {
      if (t <= ts[0])              return chartArea.left;
      if (t >= ts[ts.length - 1])  return chartArea.right;
      let lo = 0, hi = ts.length - 1;
      while (lo < hi) {
        const mid = (lo + hi) >> 1;
        if (ts[mid] < t) lo = mid + 1; else hi = mid;
      }
      const i = lo;
      if (i === 0) return xs.getPixelForValue(0);
      const t0 = ts[i - 1], t1 = ts[i];
      const x0 = xs.getPixelForValue(i - 1);
      const x1 = xs.getPixelForValue(i);
      const frac = t1 === t0 ? 0 : (t - t0) / (t1 - t0);
      return x0 + (x1 - x0) * frac;
    };

    ctx.save();
    // Two-row marker layout so orange ▼ doesn't visually merge with red ▼
    // when they occur 2 ping intervals apart (typical orange→red gap):
    //   top row    : ▼red (critical) and ▲green (recovery)
    //   lower row  : ▼orange (warn) — pushed down by `markerRowGap` pixels
    const sz = 5;
    const markerRowGap = 10;
    (data.events || []).forEach(ev => {
      // Marker semantics:
      //   enter red    → red ▼ at top
      //   exit red     → green ▲ at top
      //   enter orange → orange ▼ at lower row
      //   exit orange  → no marker (handled by closing band only)
      let color = null, marker = null, rowOffset = 0;
      if (ev.new === 'red') {
        color = '#ef4444'; marker = 'down'; rowOffset = 0;
      } else if (ev.old === 'red' && ev.new !== 'red') {
        color = '#10b981'; marker = 'up';   rowOffset = 0;
      } else if (ev.new === 'orange') {
        color = '#f59e0b'; marker = 'down'; rowOffset = markerRowGap;
      }
      if (!color) return;

      const x  = ts2x(ev.ts);
      const yT = chartArea.top + rowOffset;
      const yB = chartArea.bottom;

      // Vertical dashed line (subtle, doesn't clutter) — full chart height
      ctx.strokeStyle = color;
      ctx.lineWidth   = 1;
      ctx.setLineDash([3, 3]);
      ctx.beginPath();
      ctx.moveTo(x, chartArea.top);
      ctx.lineTo(x, yB);
      ctx.stroke();
      ctx.setLineDash([]);

      // Triangle marker
      ctx.fillStyle = color;
      ctx.beginPath();
      if (marker === 'down') {
        ctx.moveTo(x, yT + sz);
        ctx.lineTo(x - sz, yT);
        ctx.lineTo(x + sz, yT);
      } else {
        ctx.moveTo(x, yT);
        ctx.lineTo(x - sz, yT + sz);
        ctx.lineTo(x + sz, yT + sz);
      }
      ctx.closePath();
      ctx.fill();
    });
    ctx.restore();
  }
};
try { Chart.register(waveformAlertOverlayPlugin); } catch(_) { /* Chart not loaded yet */ }

// Populate node selector from current targets list
function _populateWaveformTids(){
  const sel = document.getElementById('waveform-tid');
  if(!sel) return;
  const current = sel.value;
  sel.innerHTML = '<option value="">— 选择节点 —</option>';
  Object.values(targets).forEach(t=>{
    const opt = document.createElement('option');
    opt.value = t.tid;
    opt.textContent = `${esc(t.label||t.tid)} (${esc(t.ip||'')})`;
    if(t.tid === current) opt.selected = true;
    sel.appendChild(opt);
  });
  // Auto-select first node if none selected
  if(!sel.value && sel.options.length > 1) sel.selectedIndex = 1;
}

function loadWaveform(){
  const sel = document.getElementById('waveform-tid');
  if(!sel) return;
  _populateWaveformTids();
  const tid = sel.value;
  const emptyEl = document.getElementById('waveform-empty');
  const loadEl  = document.getElementById('waveform-loading');
  if(!tid){
    if(emptyEl){ emptyEl.textContent='请选择节点后点击刷新'; emptyEl.style.display='flex'; }
    if(loadEl) loadEl.style.display='none';
    return;
  }
  if(emptyEl) emptyEl.style.display='none';
  if(loadEl)  loadEl.style.display='flex';
  const start = _statsStart || (Math.floor(Date.now()/1000)-3600);
  const end   = _statsEnd   || Math.floor(Date.now()/1000);
  const myId  = ++_waveReqId;   // capture before async gap

  // Parallel fetch: RTT points + alert overlay (red/orange bands + transitions).
  // Two endpoints not one so the chart can re-render outage bands without
  // re-pulling the (possibly large) RTT point array, and so a future
  // "hide alerts" toggle becomes one-line.
  const pWave   = fetch(`/api/waveform?tid=${encodeURIComponent(tid)}&start=${start}&end=${end}`)
                   .then(r=>r.json());
  const pAlerts = fetch(`/api/waveform/alerts?tid=${encodeURIComponent(tid)}&start=${start}&end=${end}`)
                   .then(r=>r.json())
                   .catch(()=>({red:[],orange:[],events:[]}));   // tolerate alert-API failure

  Promise.all([pWave, pAlerts])
    .then(([data, alerts])=>{
      if(myId !== _waveReqId) return;   // newer request in flight — discard
      _renderWaveform(data, alerts);
    })
    .catch(()=>{
      if(myId !== _waveReqId) return;
      if(emptyEl){ emptyEl.textContent='加载失败'; emptyEl.style.display='flex'; }
    })
    .finally(()=>{
      // Hide spinner only if this is still the latest request — prevents
      // a fast cancel-then-restart from leaving the spinner stuck off.
      if(myId === _waveReqId && loadEl) loadEl.style.display='none';
    });
}

function _renderWaveform(data, alerts){
  // Chart.js missing -> skip the chart render entirely; the rest of
  // the waveform panel (controls, summary line) still functions.
  if(typeof Chart==='undefined')return;
  const resEl = document.getElementById('waveform-resolution');
  const resMap = {'raw':'秒级','1m':'分钟级','1h':'小时级'};
  if(resEl){
    let resText = `粒度：${resMap[data.resolution]||data.resolution}`;
    // 7d/30d views use hourly aggregation: most recent incomplete hour won't
    // appear until the flush runs — expected design constraint, not a bug.
    if(data.resolution==='1h')
      resText += '  ※ 长周期视图按整点聚合，最近约 1 小时数据可能延后显示';
    resEl.textContent = resText;
  }

  // Summary row
  const s = data.summary || {};
  const sumEl = document.getElementById('waveform-summary');
  if(sumEl){
    const pct = s.loss_rate != null ? (s.loss_rate*100).toFixed(1)+'%' : '—';
    sumEl.innerHTML = [
      s.avg_ms   != null ? `<span>均值 <b>${s.avg_ms}</b> ms</span>` : '',
      s.min_ms   != null ? `<span>最小 <b>${s.min_ms}</b> ms</span>` : '',
      s.max_ms   != null ? `<span>最大 <b>${s.max_ms}</b> ms</span>` : '',
      s.loss_rate!= null ? `<span style="color:${s.loss_rate>0.05?'var(--red)':'var(--green)'}">丢包 <b>${pct}</b></span>` : '',
      s.sample_n != null ? `<span>样本 ${s.sample_n}</span>` : '',
    ].filter(Boolean).join('<span style="color:var(--border2);margin:0 4px">|</span>');
  }

  // Build chart data using bucket_state for richer visual semantics (V4)
  // bucket_state: "ok" | "partial_fail" | "all_fail"
  // Gaps in ts sequence = "no_data" (not returned by API, shown as line breaks)
  const pts      = data.points || [];
  const isDark   = document.documentElement.getAttribute('data-theme')==='dark';
  const gridC    = isDark ? 'rgba(255,255,255,.08)' : 'rgba(0,0,0,.06)';
  const textC    = isDark ? '#9ca3af' : '#6b7280';
  const labels    = pts.map(p=>{
    const d=new Date(p.ts*1000);
    return d.toLocaleTimeString('zh-CN',{hour:'2-digit',minute:'2-digit'});
  });
  const rttData      = pts.map(p=>p.ok ? p.rtt_ms : null);
  // all_fail: red X markers; partial_fail: orange diamond markers
  const allFailData  = pts.map(p=>(p.bucket_state==='all_fail') ? 0 : null);
  const partialData  = pts.map(p=>(p.bucket_state==='partial_fail' && p.rtt_ms!=null)
                                   ? p.rtt_ms : null);

  const canvas = document.getElementById('waveform-chart');
  if(!canvas) return;
  if(_waveChart){ _waveChart.destroy(); _waveChart=null; }
  _waveChart = new Chart(canvas, {
    type: 'line',
    data: {
      labels,
      datasets: [
        // RTT line (ok points only)
        { label:'RTT (ms)', data:rttData,
          borderColor:'#3b82f6', backgroundColor:'rgba(59,130,246,.15)',
          borderWidth:1.5, pointRadius:0, pointHoverRadius:4,
          fill:true, tension:0.2, spanGaps:false },
        // partial_fail: orange diamond markers (connected but some probes failed)
        { label:'部分丢包', data:partialData,
          borderColor:'rgba(245,158,11,.0)', backgroundColor:'rgba(245,158,11,.85)',
          borderWidth:0, pointRadius:5, pointStyle:'rectRot',
          pointHoverRadius:7, fill:false, showLine:false },
        // all_fail: red X markers (entire bucket failed)
        { label:'全部丢包', data:allFailData,
          borderColor:'rgba(239,68,68,.8)', backgroundColor:'rgba(239,68,68,.7)',
          borderWidth:0, pointRadius:5, pointStyle:'crossRot',
          pointHoverRadius:7, fill:false, showLine:false }
      ]
    },
    options: {
      responsive:true, maintainAspectRatio:false, animation:false,
      interaction:{ mode:'index', intersect:false },
      plugins:{
        legend:{ display:false },
        tooltip:{
          callbacks:{
            label:ctx=>{
              if(ctx.datasetIndex===0 && ctx.parsed.y!=null)
                return `RTT: ${ctx.parsed.y} ms`;
              if(ctx.datasetIndex===1 && ctx.parsed.y!=null)
                return '⚠ 部分丢包（桶内有失败）';
              if(ctx.datasetIndex===2 && ctx.parsed.y!=null)
                return '✕ 全部丢包';
              return null;
            },
            // afterBody appends one line after the per-dataset labels —
            // used to surface an "in outage" hint when the hovered point
            // falls inside a red/orange interval (parallels the C-side
            // hover tooltip).  ctx is the tooltip-items array; first item's
            // dataIndex maps back to the points_ts array we stash on the
            // chart instance in _renderWaveform.
            afterBody:items=>{
              if(!items || !items.length) return '';
              const chart = items[0].chart;
              const ts    = (chart.$pointsTs || [])[items[0].dataIndex];
              const ov    = chart.$alertOverlay || {};
              if(ts == null) return '';
              const inRange = (segs) =>
                (segs || []).some(s => ts >= s.start && ts <= s.end);
              if(inRange(ov.red))    return '⚠ 故障期间';
              if(inRange(ov.orange)) return '△ 性能告警期间';
              return '';
            }
          }
        }
      },
      scales:{
        x:{ ticks:{color:textC,maxTicksLimit:8,maxRotation:0},
            grid:{color:gridC} },
        y:{ min:0, ticks:{color:textC},
            grid:{color:gridC},
            title:{display:true,text:'ms',color:textC,font:{size:11}} }
      }
    }
  });

  // Hand alert overlay + ts mapping to the plugin via chart-instance fields.
  // Done after chart creation so plugin's beforeDatasetsDraw on the first
  // render already sees the data (Chart.js draws once synchronously here).
  _waveChart.$alertOverlay = alerts || {red:[],orange:[],events:[]};
  _waveChart.$pointsTs     = pts.map(p=>p.ts);
  _waveChart.update('none');   // redraw with overlay applied, no animation
}

// 核心：向后端请求该时间段的聚合统计，重渲所有图表
async function _refreshStatsCharts(){
  const info=document.getElementById('time-info');
  if(info) info.textContent=`查询中... ${_fmt(_statsStart)} — ${_fmt(_statsEnd)}`;

  try{
    const resp=await fetch(
      `/api/stats?start=${_statsStart}&end=${_statsEnd}`);
    if(!resp.ok)throw new Error(`HTTP ${resp.status}`);
    const rangeStats=await resp.json();

    // 用时间段数据替换全局 stats（用于图表渲染），渲染完恢复
    const savedStats={...stats};
    // 把返回的 {tid:{total,success,lat_avg,...}} 合并成图表需要的格式
    Object.keys(rangeStats).forEach(tid=>{
      // Skip tids not in current active targets (Strategy B: keep DB data,
      // never render deleted nodes on frontend).
      if(!targets[tid])return;
      const rs=rangeStats[tid];
      if(!stats[tid]) stats[tid]={latencies:[],label:'',ip:''};
      stats[tid]._range_total  =rs.total;
      stats[tid]._range_success=rs.success;
      stats[tid]._range_lat_avg=rs.lat_avg;
      stats[tid]._range_lat_max=rs.lat_max;
      stats[tid]._range_lat_p95=rs.lat_p95;
    });

    // 重绘图表（使用 _range_* 字段）
    buildDonutGrid(true);
    buildBarCharts(true);

    // 恢复全局 stats
    Object.keys(stats).forEach(tid=>{
      delete stats[tid]._range_total;
      delete stats[tid]._range_success;
      delete stats[tid]._range_lat_avg;
      delete stats[tid]._range_lat_max;
      delete stats[tid]._range_lat_p95;
    });

    const span=_statsEnd-_statsStart;
    const gran=span<=86400?'秒级原始':'小时级聚合';
    if(info) info.textContent=
      `${_fmt(_statsStart)} — ${_fmt(_statsEnd)}（${gran}数据）`;
  }catch(e){
    if(info) info.textContent=`查询失败: ${e.message}`;
  }
}

// 导出 Excel（带 loading 遮罩）
async function doExportExcel(){
  const overlay=document.getElementById('stats-loading');
  const msg=document.getElementById('loading-msg');
  if(overlay){overlay.style.display='flex';if(msg)msg.textContent='正在生成 Excel 报表，请稍候...';}
  try{
    const tids=Object.keys(targets);
    if(!tids.length){alert('暂无节点数据');return;}
    const url=`/api/export?tids=${tids.join(',')}&start=${_statsStart}&end=${_statsEnd}`;
    const resp=await fetch(url);
    if(!resp.ok)throw new Error(`服务端错误 HTTP ${resp.status}`);
    const blob=await resp.blob();
    const a=document.createElement('a');
    a.href=URL.createObjectURL(blob);
    const dt=new Date().toISOString().slice(0,16).replace('T','_').replace(':','');
    a.download=`monitor_export_${dt}.xlsx`;
    document.body.appendChild(a);a.click();
    document.body.removeChild(a);
    URL.revokeObjectURL(a.href);
  }catch(e){
    alert(`导出失败：${e.message}`);
  }finally{
    if(overlay) overlay.style.display='none';
  }
}

// 初始化统计 tab：设置默认时间段并刷新
function initStatsTab(){
  // 默认选"近1小时"
  const firstBtn=document.querySelector('.tbtn[data-sec="3600"]');
  if(firstBtn) setQuick(firstBtn);
  else _refreshStatsCharts();
}

function buildStats(){buildDonutGrid();buildBarCharts();}
function buildDonutGrid(){
  const grid=document.getElementById('donut-grid');
  if(!grid)return;
  // Defence in depth: when Chart.js failed to load (offline / blocked
  // CDN) the donut grid stays empty; the rest of the dashboard (live
  // status cards + SSE) keeps working.  Without this guard the
  // unconditional `new Chart(...)` below would throw ReferenceError
  // and the dashboard's stats panel would render half-built every
  // refresh.
  if(typeof Chart==='undefined')return;
  Object.keys(stats).forEach(tid=>{
    // Ghost filter: skip tids that are not in the current active targets.
    // Deleted nodes may leave stale entries in stats via cum_stats seed.
    if(!targets[tid]||!targets[tid].label)return;
    const s=stats[tid];
    // Use time-range data (_range_*) when available (set by _refreshStatsCharts
    // before calling this function).  Fall back to cumulative server totals.
    // Previously this function always used s.total/s.success regardless,
    // so the donut charts showed all-time uptime even when the user selected
    // "last 1 hour" — the donut and the bar chart were inconsistent.
    const totalProbes   = s._range_total   != null ? s._range_total   : s.total;
    const successProbes = s._range_success != null ? s._range_success : s.success;
    const failProbes    = Math.max(0, totalProbes - successProbes);
    const uptime        = totalProbes ? (successProbes / totalProbes * 100) : 100;
    const uptimePct     = Math.round(uptime*100)/100;
    const lossPct       = Math.round((100-uptime)*100)/100;
    // Avg latency: prefer range value, then server CMA, then bounded local array
    const avg = s._range_lat_avg != null
      ? (+s._range_lat_avg).toFixed(1)
      : s.latency_n>0 ? (s.latency_avg||0).toFixed(1)
      : s.latencies.length ? (s.latencies.reduce((a,b)=>a+b,0)/s.latencies.length).toFixed(1)
      : '--';
    const t=targets[tid];const sc=t?(st(t.status).color||'var(--gray)'):'var(--gray)';
    const cid='donut-'+tid;
    if(!document.getElementById(cid)){const div=document.createElement('div');div.className='stat-card';
      div.innerHTML=`<div class="stat-title"><span style="color:${sc}">●</span> ${esc(t&&t.label?t.label:s.label)}
        <span style="color:var(--dim);font-weight:400;font-size:11px">${esc(t&&t.ip?t.ip:s.ip)}</span></div>
        <div class="chart-wrap"><canvas id="${cid}"></canvas></div>
        <div style="text-align:center;margin-top:8px;font-size:12px;color:var(--dim)">
          连通率 <strong style="color:${sc};font-size:16px" id="uptime-${tid}">${uptimePct}%</strong>
          &nbsp; 均延时 <strong style="color:var(--blue2)" id="avg-${tid}">${avg} ms</strong></div>`;
      grid.appendChild(div);}
    // Update existing labels (server stats may have advanced since last build)
    const upEl=document.getElementById('uptime-'+tid);
    const avEl=document.getElementById('avg-'+tid);
    if(upEl)upEl.textContent=`${uptimePct}%`;
    if(avEl)avEl.textContent=`${avg} ms`;
    const canvas=document.getElementById(cid);if(!canvas)return;
    if(donutCharts[tid]){
      donutCharts[tid].data.datasets[0].data=[uptimePct,Math.max(0,100-uptimePct)];
      // Update tooltip context for the new totals
      donutCharts[tid].$tooltipContext={totalProbes,successProbes,failProbes,uptimePct,lossPct};
      donutCharts[tid].update('none');}
    else{
      const chart=new Chart(canvas,{type:'doughnut',
        data:{labels:['连通','丢失'],datasets:[{data:[uptimePct,Math.max(0,100-uptimePct)],
          backgroundColor:(()=>{
            const ctx2=canvas.getContext('2d');
            const g=ctx2.createLinearGradient(0,0,0,100);
            g.addColorStop(0,'#22d3ee');   // 科技蓝
            g.addColorStop(1,'#22c55e');   // 荧光绿
            return [g,'rgba(239,68,68,.18)'];
          })(),
          borderColor:['rgba(34,211,238,.6)','rgba(239,68,68,.3)'],
          borderWidth:1,
          borderRadius:6,
          hoverOffset:6,
        }]},
        options:{responsive:true,maintainAspectRatio:false,cutout:'72%',
          plugins:{legend:{display:false},
            tooltip:{
              enabled:true,
              displayColors:false,
              backgroundColor:'rgba(0,0,0,.85)',
              padding:10,titleFont:{size:12,weight:'bold'},bodyFont:{size:11},
              callbacks:{
                title:function(items){
                  // Use the most recent context captured during update()
                  return items[0].label;
                },
                label:function(ctx){
                  const c=ctx.chart.$tooltipContext||{};
                  const lines=[];
                  if(ctx.dataIndex===0){
                    lines.push(`连通率: ${c.uptimePct!=null?c.uptimePct:ctx.parsed}%`);
                    lines.push(`成功探测: ${c.successProbes||0} 次`);
                  }else{
                    lines.push(`丢失率: ${c.lossPct!=null?c.lossPct:ctx.parsed}%`);
                    lines.push(`失败探测: ${c.failProbes||0} 次`);
                  }
                  lines.push(`总探测: ${c.totalProbes||0} 次`);
                  return lines;
                }
              }
            }
          },animation:{duration:600}}});
      chart.$tooltipContext={totalProbes,successProbes,failProbes,uptimePct,lossPct};
      donutCharts[tid]=chart;
    }})}

// Bar chart with log/linear toggle
// When HTTP latency (~1000ms) coexists with ICMP (~5ms), linear scale
// makes ICMP bars invisible. Log scale shows relative differences properly.
function buildBarCharts(){
  // Same Chart.js missing-library guard as buildDonutGrid.  The bar
  // chart panel just stays empty rather than aborting the dashboard's
  // periodic stats refresh.
  if(typeof Chart==='undefined')return;
  // Horizontal bar chart: Y-axis = node names, X-axis = latency (ms)
  // Per-type independent charts; auto height based on node count.
  const TYPE_GROUPS={
    icmp:{types:['icmp'],        id:'bar-icmp', wrap:'bar-icmp-wrap', group:'bar-icmp-group'},
    tcp: {types:['tcp'],         id:'bar-tcp',  wrap:'bar-tcp-wrap',  group:'bar-tcp-group'},
    http:{types:['http','https'],id:'bar-http', wrap:'bar-http-wrap', group:'bar-http-group'},
    dns: {types:['dns'],         id:'bar-dns',  wrap:'bar-dns-wrap',  group:'bar-dns-group'},
  };
  const gradients={
    green:['rgba(34,197,94,.75)','rgba(34,197,94,.35)'],
    orange:['rgba(245,158,11,.75)','rgba(245,158,11,.35)'],
    red:['rgba(239,68,68,.75)','rgba(239,68,68,.35)'],
    gray:['rgba(107,114,128,.5)','rgba(107,114,128,.2)'],
    paused:['rgba(107,114,128,.3)','rgba(107,114,128,.1)'],
  };
  let row2Visible=false;
  Object.entries(TYPE_GROUPS).forEach(([key,{types,id,wrap,group}])=>{
    const tids=Object.keys(stats).filter(tid=>{
      const t=targets[tid];if(!t)return false;
      return types.includes(t.ping_type||'icmp');
    });
    const grpEl=document.getElementById(group);if(!grpEl)return;
    if(!tids.length){grpEl.style.display='none';return;}
    grpEl.style.display='';
    if(key!=='icmp')row2Visible=true;

    const labels=tids.map(tid=>targets[tid]?.label||tid);
    const avgs=tids.map(tid=>{
      const s=stats[tid];if(!s)return 0;
      if(s._range_lat_avg!=null)return+s._range_lat_avg.toFixed(1);
      if(s.latency_n>0)return+((s.latency_avg||0).toFixed(1));
      return s.latencies?.length?+(s.latencies.reduce((a,b)=>a+b,0)/s.latencies.length).toFixed(1):0;
    });

    // Dynamic height: 32px per bar, min 60, max 400
    const wrapEl=document.getElementById(wrap);
    const h=Math.max(60,Math.min(400,tids.length*34+20));
    if(wrapEl)wrapEl.style.height=h+'px';

    // Gradient colors based on node status
    const canvas=document.getElementById(id);if(!canvas)return;
    const getColors=()=>{
      return tids.map(tid=>{
        const t=targets[tid];const st=t?t.status:'gray';
        const g=gradients[st]||gradients.gray;
        try{
          const ctx=canvas.getContext('2d');
          const grd=ctx.createLinearGradient(avgs[tids.indexOf(tid)]*1.2,0,0,0);
          grd.addColorStop(0,g[0]);grd.addColorStop(1,g[1]);
          return grd;
        }catch{return g[0];}
      });
    };

    if(barCharts[key]){
      barCharts[key].data.labels=labels;
      barCharts[key].data.datasets[0].data=avgs;
      barCharts[key].data.datasets[0].backgroundColor=getColors();
      barCharts[key].update('active');
    }else{
      barCharts[key]=new Chart(canvas,{
        type:'bar',
        data:{labels,datasets:[{
          label:'平均延时',data:avgs,
          backgroundColor:getColors(),
          borderRadius:6,borderSkipped:false,
          maxBarThickness:26,
        }]},
        options:{
          indexAxis:'y',  // horizontal bars
          responsive:true,maintainAspectRatio:false,
          plugins:{
            legend:{display:false},
            tooltip:{callbacks:{label:ctx=>`${ctx.raw} ms`}},
          },
          scales:{
            x:{type:'linear',min:0,
               ticks:{font:{size:10},
                 callback:(v,i,all)=>i===all.length-1?v+' ms':v},
               grid:{color:'rgba(128,128,128,.1)'}},
            y:{ticks:{font:{size:11}},
               grid:{display:false}},
          },
          animation:{duration:500,easing:'easeOutQuart'},
        },
      });
    }
  });

  const r2=document.getElementById('bar-row2');
  if(r2)r2.style.display=row2Visible?'flex':'none';
  const emptyEl=document.getElementById('bar-empty');
  const anyVisible=Object.values(TYPE_GROUPS).some(
    ({group})=>document.getElementById(group)?.style.display!=='none');
  if(emptyEl)emptyEl.style.display=anyVisible?'none':'';
}

// Card selection
let chipSelection=[];
function openCardSel(){chipSelection=pinnedTargets?[...pinnedTargets]:Object.keys(targets);renderChips();document.getElementById('sel-overlay').classList.remove('hidden');}
function closeCardSel(){document.getElementById('sel-overlay').classList.add('hidden');}
function renderChips(){const grid=document.getElementById('chip-grid');
  grid.innerHTML=Object.values(targets).map(t=>{const s=st(t.status);const sel=chipSelection.includes(t.tid);
    return`<div class="chip${sel?' selected':''}" onclick="toggleChip('${esc(t.tid)}')">
      <span class="chip-dot" style="background:${s.color}"></span>
      <span>${esc(t.label)}</span><small style="color:var(--dim)">${esc(t.ip)}</small>
    </div>`;}).join('');}
function toggleChip(tid){const i=chipSelection.indexOf(tid);if(i>=0)chipSelection.splice(i,1);else chipSelection.push(tid);renderChips();}
function selectAllChips(){chipSelection=Object.keys(targets);renderChips();}
function selectNoneChips(){chipSelection=[];renderChips();}
function saveCardSel(){
  if(chipSelection.length===Object.keys(targets).length){pinnedTargets=null;localStorage.removeItem('pinnedTargets');}
  else{pinnedTargets=[...chipSelection];localStorage.setItem('pinnedTargets',JSON.stringify(pinnedTargets));}
  closeCardSel();renderCards();checkAlerts();}

// SSE
let es=null,retryTimer=null;
function connect(){
  if(es){es.close();es=null;}
  es=new EventSource('/events');
  es.onopen=()=>{const p=document.getElementById('conn-pill');p.textContent='● 已连接';p.className='';
    if(retryTimer){clearTimeout(retryTimer);retryTimer=null;}};
  es.onmessage=e=>{
    let msg;try{msg=JSON.parse(e.data);}catch{return;}
    if(msg.type==='heartbeat')return;
    if(msg.type==='init'){
      // Treat init as the authoritative server truth — full reset first,
      // then rebuild from snapshot.  Merge-only would leave ghost nodes/stats
      // from any delete/clear events that happened while the browser was offline.
      Object.keys(targets).forEach(k=>delete targets[k]);
      Object.keys(portStatus).forEach(k=>delete portStatus[k]);
      applyServerStats(msg.stats);
      msg.targets.forEach(t=>{targets[t.tid]=t;});
      if(msg.port_statuses) Object.assign(portStatus,msg.port_statuses);
      // Always rebuild hist from server, even if server hist is empty
      hist.length=0;
      if(msg.history) msg.history.forEach(ev=>hist.push({t:ev.time,label:ev.label,ip:ev.ip,oldSt:ev.old_status,newSt:ev.new_status}));
      renderCards();renderHist();checkAlerts();
      document.getElementById('upd').textContent='已连接';return;}
    if(msg.type==='update'){const t=msg.target,old=targets[t.tid];
      if(old&&old.status!==t.status&&soundEnabled&&t.status!=='paused'){
        if(t.status==='red'){
          playAlarm();
          pushNotification('🔴 网络告警',`${t.label || t.ip} 连接中断！`);
        }
        else if(t.status==='orange'&&(old.status==='green'||old.status==='gray')){
          playWarning();
          pushNotification('🟠 网络警告',`${t.label || t.ip} 延迟或质量劣化`);
        }
        else if(t.status==='green'&&(old.status==='red'||old.status==='orange'))playRecovery();}
      if(old&&old.status!==t.status)addHist(t.label,t.ip,old.status,t.status);
      // Drop the cached port_status snapshot when the node's IP
      // changes -- the snapshot belonged to the OLD identity (it
      // was captured against the OLD ip's tcp_check results) and
      // would otherwise keep showing on the repurposed card until
      // the next traceroute cycle (~5min by default) landed a
      // fresh snapshot.  Server already clears its tracer cache on
      // IP edit; this is the matching client-side eviction.
      if(old&&old.ip!==t.ip)delete portStatus[t.tid];
      targets[t.tid]=t;accumStat(t);
      patchCard(t);   // in-place: avoids animation restart & click failures
      updateSummary();checkAlerts();
      document.getElementById('upd').textContent='更新 '+new Date().toTimeString().slice(0,8);}
    if(msg.type==='remove'){delete targets[msg.tid];delete stats[msg.tid];
      if(donutCharts[msg.tid]){donutCharts[msg.tid].destroy();delete donutCharts[msg.tid];}
      renderCards();checkAlerts();}
    // port_status: background scheduler finished a tracert cycle for one target
    if(msg.type==='port_status'){
      const t=targets[msg.tid];
      // Reject SSE port_status messages whose `ip` field does not
      // match the current identity.  Server already filters under
      // the cache lock + a final pre-push recheck, but the recheck
      // and _push_fn() call are still not a single atomic
      // operation: an identity flip landing in the residual window
      // between recheck and broadcast slips through.  msg.ip
      // carries the snapshot the server saw at recheck time, so
      // comparing it to t.ip closes the tail of that race entirely
      // on the consumer side.  Same defence covers a slow / queued
      // SSE that lands after the user has already edited the IP.
      if(!t||(msg.ip&&t.ip&&msg.ip!==t.ip))return;
      // Cache the ip alongside the ports so buildPortsHtml can do
      // one more validity check at render time (e.g. an
      // update-without-explicit-IP-change path that didn't fire
      // the delete above).
      portStatus[msg.tid]={ip:msg.ip||t.ip,ports:msg.ports,updated_at:msg.updated_at};
      const el=document.getElementById('ports-'+msg.tid);
      if(el)el.innerHTML=buildPortsHtml(t);
    }};
  es.onerror=()=>{const p=document.getElementById('conn-pill');p.textContent='● 断开';p.className='disc';
    document.getElementById('upd').textContent='重连中...';
    es.close();es=null;retryTimer=setTimeout(connect,3000);};}
initNotifications();
connect();

// SLA Report
let _slaData=[],_slaSortCol=1,_slaSortAsc=true;
function initSLATab(){
  const fmt=d=>`${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,"0")}-${String(d.getDate()).padStart(2,"0")}`;
  const now=new Date();
  document.getElementById("sla-e").value=fmt(now);
  const s=new Date(+now-7*86400000);
  document.getElementById("sla-s").value=fmt(s);
  loadSLA();
}
function setSLAPeriod(days){
  document.querySelectorAll("[id^=slab-]").forEach(b=>b.classList.remove("active"));
  if(days===7)document.getElementById("slab-7d").classList.add("active");
  else if(days===30)document.getElementById("slab-30d").classList.add("active");
  const fmt=d=>`${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,"0")}-${String(d.getDate()).padStart(2,"0")}`;
  const now=new Date();
  document.getElementById("sla-s").value=fmt(new Date(+now-days*86400000));
  document.getElementById("sla-e").value=fmt(now);
  loadSLA();
}
function setSLACurrentMonth(){
  document.querySelectorAll("[id^=slab-]").forEach(b=>b.classList.remove("active"));
  document.getElementById("slab-mo").classList.add("active");
  const now=new Date(),y=now.getFullYear(),m=now.getMonth();
  const fmt=d=>`${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,"0")}-${String(d.getDate()).padStart(2,"0")}`;
  document.getElementById("sla-s").value=fmt(new Date(y,m,1));
  document.getElementById("sla-e").value=fmt(now);
  loadSLA();
}
async function loadSLA(){
  const sv=document.getElementById("sla-s").value,ev=document.getElementById("sla-e").value;
  if(!sv||!ev)return;
  const start=new Date(sv+" 00:00:00").getTime()/1000;
  const end  =new Date(ev+" 23:59:59").getTime()/1000;
  document.getElementById("sla-hint").textContent="计算中...";
  document.getElementById("sla-tbody").innerHTML='<tr><td colspan="7" style="text-align:center;padding:30px;color:var(--dim)">⏳ 计算中...</td></tr>';
  try{
    const data=await(await fetch(`/api/sla?start=${start}&end=${end}`)).json();
    _slaData=Object.values(data);
    document.getElementById("sla-hint").textContent=`${sv} — ${ev}，共 ${_slaData.length} 个节点`;
    renderSLATable();
  }catch(e){
    document.getElementById("sla-tbody").innerHTML='<tr><td colspan="7" style="text-align:center;padding:30px;color:var(--red)">加载失败</td></tr>';
  }
}
function sortSLA(col){
  if(_slaSortCol===col)_slaSortAsc=!_slaSortAsc;else{_slaSortCol=col;_slaSortAsc=col===0;}
  renderSLATable();
}
function renderSLATable(){
  const rows=[..._slaData];
  const vals=r=>[r.label,r.uptime_pct,r.lat_avg,r.lat_p95,r.lat_max,r.outage_count,r.outage_minutes];
  rows.sort((a,b)=>{const av=vals(a)[_slaSortCol],bv=vals(b)[_slaSortCol];
    if(av==null)return 1;if(bv==null)return -1;
    return _slaSortAsc?(av>bv?1:-1):(av<bv?1:-1);});
  const ms=v=>v==null||v===0?"--":`${v} ms`;
  const pct=v=>{
    if(v==null)return'<span style="color:var(--dim)">--</span>';
    const c=v>=99.9?"var(--green)":v>=99?"var(--orange)":"var(--red)";
    return `<b style="color:${c}">${v.toFixed(3)}%</b>`;
  };
  const fmtMin=m=>{
    if(!m)return'<span style="color:var(--green)">无中断</span>';
    if(m<1)return `${Math.round(m*60)} 秒`;
    if(m<60)return `${m.toFixed(1)} 分钟`;
    return `${(m/60).toFixed(1)} 小时`;
  };
  const tbody=rows.map(r=>{
    const bg=r.uptime_pct!=null&&r.uptime_pct<99?"rgba(239,68,68,.05)":"";
    return `<tr style="border-bottom:1px solid var(--border)${bg?";background:"+bg:""}">
      <td style="padding:9px 14px"><b style="color:var(--text)">${esc(r.label)}</b><br><span style="font-size:11px;color:var(--dim)">${esc(r.ip)}</span></td>
      <td style="text-align:center;padding:9px 14px">${pct(r.uptime_pct)}</td>
      <td style="text-align:right;padding:9px 14px;color:var(--text)">${ms(r.lat_avg)}</td>
      <td style="text-align:right;padding:9px 14px;color:var(--text)">${ms(r.lat_p95)}</td>
      <td style="text-align:right;padding:9px 14px;color:var(--dim)">${ms(r.lat_max)}</td>
      <td style="text-align:center;padding:9px 14px;color:${r.outage_count?"var(--red)":"var(--green)"};">${r.outage_count} 次</td>
      <td style="text-align:right;padding:9px 14px">${fmtMin(r.outage_minutes)}</td>
    </tr>`;
  }).join("")||'<tr><td colspan="7" style="text-align:center;padding:30px;color:var(--dim)">暂无数据（监控时间不足或该周期无探测记录）</td></tr>';
  document.getElementById("sla-tbody").innerHTML=tbody;
}
function exportSLAExcel(){
  const sv=document.getElementById("sla-s").value,ev=document.getElementById("sla-e").value;
  if(!_slaData.length){alert("请先查询 SLA 数据");return;}
  const start=new Date(sv+" 00:00:00").getTime()/1000;
  const end  =new Date(ev+" 23:59:59").getTime()/1000;
  window.open(`/api/export/sla?start=${start}&end=${end}`);
}
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────
class _SSEBroadcaster:
    def __init__(self):
        self._lock = threading.Lock()
        self._queues: list[queue.Queue] = []
    def subscribe(self):
        q: queue.Queue = queue.Queue(maxsize=2000)
        with self._lock: self._queues.append(q)
        return q
    def unsubscribe(self, q):
        with self._lock:
            try: self._queues.remove(q)
            except ValueError: pass
    def push(self, data: str):
        with self._lock:
            for q in list(self._queues):
                try: q.put_nowait(data)
                except queue.Full: pass


def _is_ip_literal(s: str) -> bool:
    """Strict IPv4 / IPv6 literal validator for the quick-tracert ?ip= path.

    socket.inet_pton is preferred over a regex because it handles all
    IPv6 corner cases (compressed zeros, dual-stack mapped addresses,
    zone IDs).  Length cap keeps the regex-free path bounded.
    """
    if not s or len(s) > 45:    # max IPv6 string length is 45 chars
        return False
    if "%" in s:                # reject zone IDs (fe80::1%eth0); tracert
        return False            # doesn't accept them anyway
    import socket as _sk
    for fam in (_sk.AF_INET, _sk.AF_INET6):
        try:
            _sk.inet_pton(fam, s)
            return True
        except (OSError, ValueError):
            continue
    return False


def _add_firewall_rule(port: int):
    """
    在 Windows 防火墙中为本程序添加 TCP 入站规则。
    - 只在 Windows 上执行，其他平台直接跳过。
    - 需要管理员权限；没有权限时静默失败（不影响本机访问）。
    - 规则名称含端口号，重复添加时 netsh 会报错但不影响运行。
    打包后的 exe 第一次运行如果没有权限，用户会看到 Windows UAC 弹窗
    或防火墙询问弹窗，点"允许"即可；之后规则永久保留。
    """
    import sys, subprocess, os
    if sys.platform != "win32":
        return
    rule_name = f"NetMonitor-Port-{port}"
    try:
        # 先检查规则是否已存在，避免重复添加
        check = subprocess.run(
            ["netsh", "advfirewall", "firewall", "show", "rule", f"name={rule_name}"],
            capture_output=True, text=True,
            creationflags=subprocess.CREATE_NO_WINDOW)
        if "No rules match" not in check.stdout and rule_name in check.stdout:
            return   # 规则已存在，无需重复添加
        # 添加入站规则：允许 TCP，所有配置文件（域/公共/专用）
        subprocess.run(
            ["netsh", "advfirewall", "firewall", "add", "rule",
             f"name={rule_name}",
             "dir=in", "action=allow", "protocol=TCP",
             f"localport={port}",
             "profile=any",
             "description=NetMonitor web dashboard"],
            capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW)
    except Exception:
        pass   # 无权限或 netsh 不可用时静默忽略


class WebServer:
    MAX_HISTORY = 60

    def __init__(self, port: int = 8765, log_dir: str = None, sound_dir: str = None,
                 allow_lan: bool = False):
        self.port          = port
        self._log_dir      = log_dir    # None = auto-detect in _run()
        self._sound_dir    = sound_dir  # writable WAV dir (DATA_DIR/assets)
        # Default to loopback-only binding.  Pre-2025 the server bound
        # 0.0.0.0 unconditionally AND netsh-added a Windows firewall
        # rule for every profile (domain/private/public), so the
        # dashboard + every API endpoint -- including
        # /api/traceroute/quick which runs subprocess.Popen on an
        # operator-supplied IP -- was reachable to any host that
        # could route to the box, with no authentication of any
        # kind.  Loopback-only by default keeps the dashboard usable
        # without exposing it.  Operators who genuinely need LAN
        # access can flip web_allow_lan in settings; that path also
        # gates the Windows firewall add.
        self._allow_lan    = bool(allow_lan)
        self._running      = False
        self._server       = None       # werkzeug BaseWSGIServer; set in _run()
        # Used by start() / restart() to wait for the background _run()
        # thread to ACTUALLY bind (or fail to bind) the listening port
        # before reporting success.  Without this handshake, start()
        # returns True the instant the thread is launched -- the caller
        # then advertises the Web URL on the UI even if the port was
        # already in use and the bind threw EADDRINUSE inside _run.
        self._bind_done    = threading.Event()
        self._bind_ok      = False
        self._targets: dict[str, dict] = {}
        self._stats: dict[str, dict] = {}   # see comment below
        self._history: list[dict] = []
        self._paused: set[str] = set()   # tids currently paused
        # Cumulative uptime stats per target, kept on the SERVER side so
        # browser refreshes don\'t reset the counters. The previous
        # implementation accumulated stats only in the browser, which
        # meant uptime% always showed "based on this browser session"
        # and any page reload wiped the history. By keeping the running
        # totals here in WebServer (which lives for the whole program
        # lifetime), the donut charts reflect actual monitored uptime.
        #
        # Schema: {tid: {"total": int,        # probes seen
        #                "success": int,      # green-or-orange probes
        #                "latency_sum": float,# rolling sum for avg
        #                "latency_n":   int}} # samples in sum
        self._lock         = threading.Lock()
        self._broadcaster  = _SSEBroadcaster()
        self._tracer_cache: dict[str, dict] = {}
        self._tracer_lock  = threading.Lock()
        self._scheduler    = _TracerouteScheduler(
            cache=self._tracer_cache, cache_lock=self._tracer_lock,
            push_fn=self._broadcaster.push)
        self._data_store   = None   # set via set_data_store()
        # Concurrency guard for /api/traceroute/quick — each request spawns a
        # subprocess; 3 concurrent is more than enough for a single-operator tool.
        self._quick_trace_sem = threading.BoundedSemaphore(3)
        # ── ICMP diagnostic task table (Task B) ──────────────────────────
        # Process-local state: each browser-initiated diagnostic run gets a
        # task_id and a state dict kept here so subsequent /status polls can
        # read incremental progress. Cleared on server restart by design —
        # any persistent record lives in DataStore via _persist_result().
        #
        # Concurrency model: at most ONE active task on the Web side
        # (_icmp_active_task_id != None), enforced atomically under the
        # lock at /start time. The icmp_diag module's _TASK_SEM(1) provides
        # the global cross-side guard (C-side 🔬 ↔ Web both block on it),
        # which gives the 409 response when the C-side is busy.
        self._icmp_tasks: dict = {}        # task_id -> state dict
        self._icmp_active_task_id = None    # str | None — current running task
        self._icmp_active_task_obj = None   # DiagTask | None — for cancel()
        self._icmp_tasks_lock = threading.Lock()
        # How long to keep finished task entries in memory so a late-arriving
        # /status poll can still see them. After this they're gc'd to bound
        # memory. The browser stops polling on done/error/cancelled anyway,
        # so 30 min is plenty of safety margin.
        self._icmp_task_ttl_s = 1800
        # ── DNS diagnostic task table (Phase 2) ──────────────────────────
        # Mirror of the ICMP task table above.  DNS queries don't have
        # incremental progress (single dnspython call), so the only
        # mutation is status: running → done|error|cancelled with the
        # full result attached on completion.
        self._dns_tasks: dict = {}         # task_id -> state dict
        self._dns_active_task_id = None     # str | None
        self._dns_active_task_obj = None    # DnsDiagTask | None — for cancel()
        self._dns_tasks_lock = threading.Lock()
        self._dns_task_ttl_s = 1800
        # ── Cross-Tab busy lock (Phase 2 §4.3) ───────────────────────────
        # Held only by /api/icmp_diag/start and /api/dns_diag/start to
        # atomically read both *_active_task_id fields and reserve the
        # active slot — without this, two tabs racing /start can each
        # see the other's slot as "free" before either has written its
        # own.  Lock order: _web_diag_busy_lock outer, _xxx_tasks_lock
        # inner; clearing the slot in callbacks/cancel keeps using only
        # the per-kind lock (single-pointer assignment to None is
        # monotonic, so stale reads only cause at most one extra 409).
        self._web_diag_busy_lock = threading.Lock()
        if FLASK_AVAILABLE:
            # Explicitly pass root_path so Flask does not try to infer
            # it via importlib in frozen (PyInstaller) mode.
            # In frozen mode, __file__ resolves to _MEIPASS/src/web_server.pyc
            # which is read-only; passing the abspath here is safe because
            # we serve no templates or static files through Flask directly.
            import os as _os
            self._app = Flask(__name__,
                              root_path=_os.path.dirname(_os.path.abspath(__file__)))
            self._setup_routes()
            logging.getLogger("werkzeug").setLevel(logging.ERROR)

    def start(self) -> bool:
        if not FLASK_AVAILABLE: return False
        if self._running: return True
        # Reset the bind-done handshake before launching the thread so
        # repeated start() calls (e.g. after a prior failure) don't see
        # a stale "yes I bound!" event from a previous attempt.
        self._bind_done.clear()
        self._bind_ok = False
        self._running = True
        # Windows: only add the netsh firewall rule when LAN access
        # is explicitly enabled.  In loopback-only mode the bind is
        # 127.0.0.1 so the OS firewall already drops every remote
        # packet -- opening the public profile would be misleading
        # (and a misconfiguration risk).
        if self._allow_lan:
            _add_firewall_rule(self.port)
        threading.Thread(target=self._run, daemon=True, name="web-server").start()
        # Wait for the background thread to actually bind the port (or
        # fail to).  make_server() either returns immediately (success)
        # or raises (port-in-use / permission denied), so 3 s is far
        # more than enough on any reasonable machine.  The previous
        # "return True" returned BEFORE the bind, so the caller's UI
        # advertised the URL even on EADDRINUSE -- browser then refused
        # the connection while the user saw "started successfully".
        if not self._bind_done.wait(timeout=3.0):
            # Timed out waiting -- treat as failure and force the flag
            # back so a follow-up start()/restart() can try again.
            self._running = False
            return False
        return self._bind_ok

    def stop(self):
        """Signal the web server to stop. Returns immediately; shutdown is async."""
        self._running = False
        srv = self._server
        if srv:
            # shutdown() blocks until serve_forever() exits — run it in a
            # daemon thread so we don't stall the Tkinter main thread.
            threading.Thread(target=srv.shutdown, daemon=True,
                             name="web-stop").start()

    def restart(self, new_port: int = None) -> bool:
        """Stop the server and start again, optionally on a new port.
        Uses make_server() shutdown, so no port-in-use race on same port.

        Returns True on a successful re-bind, False if the new port
        could not be acquired (already in use / permission denied /
        Werkzeug crash, etc.).  Mirrors start()'s synchronous-result
        contract so the UI caller can show a clear error dialog
        instead of advertising a new URL that nothing is actually
        listening on.
        """
        import time as _t
        old_port = self.port
        self.stop()
        if new_port is not None:
            self.port = new_port
        # Wait up to 2 s for the old server thread to release the port.
        # serve_forever() checks the shutdown flag every 0.5 s, so 1 s is
        # normally enough.  We only need to wait when reusing the same port.
        if new_port is None or new_port == old_port:
            # Wait for the old server's _run() finally-block to set _server=None,
            # which is the only reliable signal that the port has been released.
            # Waiting on self._running is useless because stop() already set it
            # to False before this loop even starts.
            deadline = _t.time() + 2.0
            while self._server is not None and _t.time() < deadline:
                _t.sleep(0.1)
        # Same loopback-only / LAN-only gate as start() above.
        if self._allow_lan:
            _add_firewall_rule(self.port)
        # Reset the bind-done handshake before relaunching (see start()
        # for the rationale).
        self._bind_done.clear()
        self._bind_ok = False
        self._running = True
        threading.Thread(target=self._run, daemon=True,
                         name="web-server").start()
        # Synchronously wait for the new bind result, same as start().
        # The previous fire-and-forget contract advertised the new URL
        # on the UI even when the port was already in use, leaving the
        # operator clicking a non-functional link with no error feedback.
        if not self._bind_done.wait(timeout=3.0):
            self._running = False
            return False
        if not self._bind_ok:
            # _run's finally only clears self._running when its srv
            # local matches self._server -- on a make_server() raise,
            # srv was never assigned and the identity check can go
            # either way depending on whether the OLD server's
            # finally has caught up.  Force the state explicitly so
            # is_running() can't lie to the caller.
            self._running = False
            return False
        return True

    def set_data_store(self, ds):
        """Attach a DataStore. Called from main() after both are created."""
        self._data_store = ds
        self._scheduler._data_store = ds   # scheduler needs it to persist traceroute history
        # Load historical cum_stats → seed in-memory _stats so first browser
        # connection shows real uptime, not 0%, even right after a restart.
        if ds:
            for tid, s in ds.get_cum_stats().items():
                if tid not in self._stats:
                    self._stats[tid] = {
                        "total":       s.get("total", 0),
                        "success":     s.get("success", 0),
                        "latency_avg": s.get("latency_avg", 0.0),
                        "latency_n":   s.get("latency_n", 0),
                    }

    def set_traceroute_interval(self, seconds: int):
        self._scheduler.interval = max(60, int(seconds))

    def set_traceroute_max_hops(self, hops: int):
        self._scheduler.max_hops = max(5, min(64, int(hops)))

    def _build_scheduler_snapshot(self) -> dict:
        """
        Build the active-target snapshot for the traceroute scheduler.
        Must be called with self._lock already held.
        Paused targets are excluded so the scheduler doesn't run traceroutes
        for nodes the operator has explicitly silenced.
        """
        all_t = {}
        for t in self._targets.values():
            if t.get("status") == "paused":
                continue
            # `t.get(..., "80,443")` only defaults when the KEY IS ABSENT.
            # The edit dialog stores an empty string when the user leaves
            # the "diagnostic ports" field blank (see dialogs.py:300-310 +
            # main_window:1507), so the key IS present with value "" -- the
            # dict-default never fires.  An empty string then splits to
            # [""] which the filter drops, leaving ports=[] and disabling
            # the TCP-probe path entirely.  UI promised "leave empty =
            # default 80/443"; honour that with `or` so any falsy value
            # (None, "") falls back to the default.
            probe_raw = t.get("tcp_probe_ports") or "80,443"
            try:
                ports = [int(p.strip()) for p in probe_raw.split(",") if p.strip()][:6]
            except Exception:
                ports = [80, 443]
            # Re-apply the default if parsing succeeded but produced an
            # empty list (e.g. "  ,  " → no usable ints).  This keeps the
            # "leave empty = default" promise even with whitespace-only
            # noise that the integer parse swallowed silently.
            if not ports:
                ports = [80, 443]
            all_t[t["tid"]] = {
                "ip":          t["ip"],
                "label":       t.get("label", t["tid"]),   # for DB record
                "probe_ports": ports,
            }
        return all_t

    def pause_target(self, tid: str):
        """Mark a target as paused — web card shows ⏸, no alerts, not counted in stats."""
        with self._lock:
            self._paused.add(tid)
            self._targets[tid] = {**self._targets.get(tid, {}), "status": "paused"}
            # Snapshot inside the lock so the push payload can't race against
            # a concurrent remove_target / update_target that mutates _targets.
            to_push = dict(self._targets[tid])
            snap    = self._build_scheduler_snapshot()
        # Immediately update scheduler so paused target stops getting traced
        self._scheduler.set_targets(snap)
        self._broadcaster.push(json.dumps({"type": "update", "target": to_push}))

    def resume_target(self, tid: str):
        """Clear paused state; next real ping result will show the true status.

        Mirrors pause_target() by also rebuilding the traceroute-scheduler
        snapshot.  Without this, a resumed node stayed absent from the
        scheduler's _targets until its next real probe state arrived --
        an immediate /api/traceroute request would be queued via
        request_now() but the scheduler tick would then look up the
        tid in its (still-stale) snapshot, find nothing, and drop the
        urgent request on the floor.  For HTTP / DNS targets with long
        probe intervals this turned "resume + traceroute" into a UX
        deadlock for minutes.
        """
        with self._lock:
            self._paused.discard(tid)
            if tid in self._targets:
                self._targets[tid]["status"] = "gray"
                to_push = dict(self._targets[tid])   # snapshot inside lock
            else:
                to_push = None
            snap = self._build_scheduler_snapshot()
        # Outside lock: scheduler.set_targets + SSE push.  Doing both
        # outside the lock keeps the critical section minimal and
        # matches pause_target() above.
        self._scheduler.set_targets(snap)
        if to_push:
            self._broadcaster.push(json.dumps({"type": "update", "target": to_push}))

    def update_target(self, tid, label, ip, status, latency_ms, jitter_ms,
                      loss_rate, ping_type="icmp",
                      alert_category="ok", status_code=None,
                      failure_reason=None, timings=None,
                      cert_days_left=None, tcp_probe_ports="",
                      dns_domain="",
                      keyword_ok=None,
                      is_probe_result: bool = True):
        """Push a status snapshot into the Web-side _targets dict.

        is_probe_result: True (default) means this call corresponds to
        a real ping/HTTP/DNS probe result and the per-target rolling
        statistics (_stats.total / success / latency_avg) should be
        updated.  Callers using this method only to sync UI metadata
        (initial snapshot on Web server start, post-edit label /
        tcp_probe_ports / dns_domain sync) MUST pass False so the
        meta-sync isn't double-counted as a fake probe -- the previous
        unconditional s["total"] += 1 inflated the dashboard's
        cumulative probe count by one per Web-start and one per edit,
        and if the synced status happened to be green/orange it also
        inflated the "success" counter, polluting uptime%.
        """
        if not self._running: return
        # Build data dict outside the lock (pure computation, no shared state)
        data = {"tid": tid, "label": label, "ip": ip, "status": status,
                "latency_ms": round(latency_ms, 1) if latency_ms is not None else None,
                "jitter_ms":  round(jitter_ms,  1) if jitter_ms  is not None else None,
                "loss_rate":  loss_rate, "ping_type": ping_type,
                "alert_category": alert_category,
                "status_code": status_code,
                "failure_reason": failure_reason,
                "timings": timings,
                "cert_days_left":   cert_days_left,
                "tcp_probe_ports":  tcp_probe_ports,
                "dns_domain":       dns_domain,
                # buildSubHtml reads t.keyword_ok to render the
                # ✓/⚠ "关键字" badge for HTTP nodes; the field has
                # been there since the http-keyword feature shipped
                # but was never forwarded to the browser, so the
                # frontend branch was unreachable.  Tri-state:
                # None = node not configured for keyword check,
                # True/False = match result.
                "keyword_ok":       keyword_ok}
        # Single atomic lock section — paused check, write, history, and
        # building all_t all happen under one lock so there is no window
        # between the paused check and the _targets write.
        with self._lock:
            if tid in self._paused:
                if is_probe_result:
                    # Real probe arriving while node is paused: dropping
                    # it keeps the SLA / waveform / "paused" status
                    # from being overwritten by stale measurement
                    # data.  This is the original guard.
                    return
                # Meta-sync (is_probe_result=False) for a paused node:
                # let the edit propagate through (label / ip /
                # tcp_probe_ports / dns_domain), but force the visible
                # status back to "paused" so an accidentally-non-
                # paused snapshot can't un-pause the node.  Without
                # this branch a node edited while paused stayed
                # advertising the OLD ip / label on the Web side
                # until the next real probe -- and crucially, the
                # traceroute scheduler snapshot rebuilt on resume
                # used the stale _targets entry and scheduled a
                # tracert against the old IP.
                data["status"] = "paused"
            old = self._targets.get(tid)
            old_status = old["status"] if old else None
            # If the target's IP changed (rename / repurpose edit
            # propagating through the steady-state update path), drop
            # any cached traceroute hops -- they belong to the previous
            # IP and would otherwise be served by /api/traceroute as if
            # they were the new node's path.  Mirrors the eviction
            # already done in remove_target so cache lifecycle stays
            # tied to target identity.
            if old is not None and old.get("ip") != ip:
                with self._tracer_lock:
                    # bump_generation FIRST -- atomic via _gen_lock,
                    # and observed by any stale tracert that subsequently
                    # acquires _tracer_lock for its in-lock re-check
                    # in _run_one.  Without this ordering, a stale task
                    # that arrived between "cache.pop" and
                    # "bump_generation" would observe gen still equal
                    # to its captured_generation and write its OLD-ip
                    # entry back into the freshly-cleared cache; the
                    # next bump_generation would arrive too late to
                    # prevent a stale write.  Doing the bump under the
                    # cache lock pairs it with the subsequent pop so
                    # _run_one's in-lock gen check is authoritative.
                    self._scheduler.bump_generation(tid)
                    self._tracer_cache.pop(tid, None)
            self._targets[tid] = data
            if is_probe_result and status != "paused":
                if tid not in self._stats:
                    self._stats[tid] = {"total": 0, "success": 0,
                                         "latency_avg": 0.0, "latency_n": 0}
                s = self._stats[tid]
                s["total"] += 1
                if status in ("green", "orange"):
                    s["success"] += 1
                if latency_ms is not None:
                    s["latency_n"] += 1
                    s["latency_avg"] += (latency_ms - s["latency_avg"]) / s["latency_n"]
            if old_status not in (None, "paused") and old_status != status:
                if not (old_status == "gray" and status == "green"):
                    self._history.append({"time": datetime.now().strftime("%H:%M:%S"),
                        "label": label, "ip": ip,
                        "old_status": old_status, "new_status": status})
            if len(self._history) > self.MAX_HISTORY: self._history.pop(0)
            snap = self._build_scheduler_snapshot()   # uses shared helper, lock already held

        # Lock-free: schedule traceroute, push targets snapshot, broadcast
        if status == "red" and (old_status not in ("red", "paused")):
            self._scheduler.request_now(tid)
        self._scheduler.set_targets(snap)
        self._broadcaster.push(json.dumps({"type": "update", "target": data}))

    def remove_target(self, tid: str):
        with self._lock:
            self._targets.pop(tid, None)
            self._paused.discard(tid)
            self._stats.pop(tid, None)
            snap = self._build_scheduler_snapshot()
        # Tell scheduler immediately so it stops tracing the removed node.
        # Without this, if all nodes are deleted the scheduler keeps its stale
        # target list and continues running traceroutes indefinitely.
        self._scheduler.set_targets(snap)
        # Pair bump_generation + cache pop under one lock acquisition
        # so a stale tracert can't acquire _tracer_lock between the
        # two and re-insert its old-ip entry into a cache we just
        # cleared.  See update_target's matching block for the full
        # ordering rationale.
        with self._tracer_lock:
            self._scheduler.bump_generation(tid)
            self._tracer_cache.pop(tid, None)
        self._broadcaster.push(json.dumps({"type": "remove", "tid": tid}))

    def invalidate_traceroute(self, tid: str) -> None:
        """Public hook for callers (e.g. main_window before
        wipe_target_history) to invalidate any in-flight traceroute
        for `tid` -- on return it will skip the cache / DataStore
        writes, preventing the freshly-wiped history from being
        immediately re-populated by a tracert that was already
        running when the user clicked "更换监控对象（清空历史）".

        Paired bump + cache pop under one lock so any concurrent
        stale tracert about to acquire _tracer_lock for its in-lock
        re-check observes the new generation before it can write.
        """
        with self._tracer_lock:
            self._scheduler.bump_generation(tid)
            self._tracer_cache.pop(tid, None)

    # Diagnostic-task helpers (Task B / DNS Phase 2)
    # ---------------------------------------------------------------
    # These run on the Flask worker thread (for /start, /status, /cancel)
    # and on the icmp_diag / dns_diag worker threads (progress/done/error
    # callbacks).  All shared-state access goes through _xxx_tasks_lock
    # with short, simple critical sections so the diag worker is never
    # blocked on the GIL by a slow Flask client.  Cross-Tab busy checks
    # in /start additionally take _web_diag_busy_lock (see __init__).

    def _web_diag_busy(self) -> bool:
        """Return True if either ICMP- or DNS-tab has an active task.

        MUST be called with _web_diag_busy_lock held — both fields are
        plain attributes whose write protocol guarantees correctness only
        under that lock.  Callers in /start use this to fail-fast with
        409 before they even build a task entry.
        """
        return (self._icmp_active_task_id is not None or
                self._dns_active_task_id  is not None)

    def _icmp_gc_old_tasks(self, now_ts: float):
        """Purge finished task entries older than _icmp_task_ttl_s.
        Caller MUST hold _icmp_tasks_lock. The active task is never gc'd
        regardless of age - only entries in a terminal status are evicted.
        """
        cutoff = now_ts - self._icmp_task_ttl_s
        stale = []
        for tid_, st in self._icmp_tasks.items():
            if st.get("status") in ("done", "error", "cancelled") \
               and st.get("updated_ts", 0) < cutoff:
                stale.append(tid_)
        for tid_ in stale:
            self._icmp_tasks.pop(tid_, None)

    def _icmp_make_callbacks(self, task_id: str):
        """Build the (on_progress, on_pmtu_progress, on_done, on_error)
        callbacks bound to a specific task_id.

        Each callback only mutates state under _icmp_tasks_lock and does
        the minimum work necessary - it runs on the diag worker thread,
        so a heavy callback would stretch wall-clock between probes.
        """
        from .icmp_diag import icmp_sample_to_dict, icmp_result_to_dict

        def _on_progress(sample, completed, total):
            with self._icmp_tasks_lock:
                st = self._icmp_tasks.get(task_id)
                if st is None or st.get("status") != "running":
                    return
                prog = st["progress"]
                prog["samples"].append(icmp_sample_to_dict(sample))
                prog["completed"] = completed
                prog["total"]     = total
                st["updated_ts"]  = time.time()

        def _on_pmtu_progress(payload, success, error_code):
            with self._icmp_tasks_lock:
                st = self._icmp_tasks.get(task_id)
                if st is None or st.get("status") != "running":
                    return
                prog = st["progress"]
                prog["pmtu_steps"].append({
                    "payload":    payload,
                    "success":    success,
                    "error_code": error_code,
                })
                prog["completed"] = len(prog["pmtu_steps"])
                # PMTU total is unknown in advance (binary search); leave
                # `total` as None so the UI renders a "测试中..." indicator
                # rather than a fake 0/N progress bar.
                st["updated_ts"]  = time.time()

        def _on_done(result):
            # _persist_result already ran inside run_icmp_diag_async, so DB
            # write is complete before this callback fires.
            with self._icmp_tasks_lock:
                st = self._icmp_tasks.get(task_id)
                if st is None:
                    return
                # If user already cancelled, keep the cancelled status -
                # don't overwrite with a partial "done" result from a probe
                # that finished after cancel() but before the loop saw it.
                if st.get("status") == "cancelled":
                    if self._icmp_active_task_id == task_id:
                        self._icmp_active_task_id  = None
                        self._icmp_active_task_obj = None
                    return
                st["status"]     = "done"
                st["result"]     = icmp_result_to_dict(result)
                st["updated_ts"] = time.time()
                if self._icmp_active_task_id == task_id:
                    self._icmp_active_task_id  = None
                    self._icmp_active_task_obj = None

        def _on_error(msg):
            with self._icmp_tasks_lock:
                st = self._icmp_tasks.get(task_id)
                if st is None:
                    return
                if st.get("status") == "cancelled":
                    if self._icmp_active_task_id == task_id:
                        self._icmp_active_task_id  = None
                        self._icmp_active_task_obj = None
                    return
                st["status"]     = "error"
                st["error"]      = str(msg)
                st["updated_ts"] = time.time()
                if self._icmp_active_task_id == task_id:
                    self._icmp_active_task_id  = None
                    self._icmp_active_task_obj = None

        return _on_progress, _on_pmtu_progress, _on_done, _on_error

    # DNS diagnostic task helpers (Phase 2)
    # ---------------------------------------------------------------
    # Conceptually identical to the ICMP helpers above but simpler: DNS
    # queries are single dnspython calls with no incremental progress,
    # so we only flip status running → done|error|cancelled and attach
    # the full result on completion.

    def _dns_gc_old_tasks(self, now_ts: float):
        """Purge finished DNS task entries older than _dns_task_ttl_s.
        Caller MUST hold _dns_tasks_lock.  The active task is never gc'd.
        """
        cutoff = now_ts - self._dns_task_ttl_s
        stale = []
        for tid_, st in self._dns_tasks.items():
            if st.get("status") in ("done", "error", "cancelled") \
               and st.get("updated_ts", 0) < cutoff:
                stale.append(tid_)
        for tid_ in stale:
            self._dns_tasks.pop(tid_, None)

    def _dns_make_callbacks(self, task_id: str):
        """Build (on_done, on_error) callbacks bound to a DNS task_id.

        on_done fires both for genuine completion AND for a synthetic
        "cancelled" DnsDiagResult emitted by run_dns_diag_async when the
        worker observes task._cancel before the result lands.  We
        distinguish via result.rcode == ERROR_CANCELLED.

        Phase 3C-1: the result can now be either a DnsDiagResult (single
        mode, including the synthetic cancelled sentinel) or a
        DnsCompareResult (compare mode).  dns_run_to_dict dispatches by
        isinstance so we get one shape on the wire either way — single
        results carry mode="single", compare results mode="compare".
        Cancelled compare runs surface as a DnsDiagResult with rcode ==
        ERROR_CANCELLED (see run_dns_diag_async's _build_cancelled
        rationale), and that path falls through THIS function's
        cancelled branch unchanged.
        """
        from .dns_diag import (dns_run_to_dict, DnsCompareResult,
                                ERROR_CANCELLED)

        def _on_done(result):
            with self._dns_tasks_lock:
                st = self._dns_tasks.get(task_id)
                if st is None:
                    return
                # If /api/dns_diag/cancel already flipped us to
                # cancelled, keep that — don't overwrite with the
                # synthetic cancelled result that's about to arrive
                # (or, in the very unlikely race, with a genuine
                # result that beat the cancel).
                if st.get("status") == "cancelled":
                    if self._dns_active_task_id == task_id:
                        self._dns_active_task_id  = None
                        self._dns_active_task_obj = None
                    return
                # Detect cancelled-sentinel for single mode (and the
                # compare-cancelled path which also returns one of
                # these).  Compare results never carry rcode directly,
                # so the isinstance check excludes them from the
                # cancellation branch.
                is_cancelled = (not isinstance(result, DnsCompareResult)
                                and getattr(result, "rcode", "")
                                    == ERROR_CANCELLED)
                if is_cancelled:
                    # Cancel was observed by the dns_diag worker thread
                    # but the Flask /cancel route either hasn't been
                    # called or hasn't won the lock yet.  Surface as
                    # cancelled status without a result payload so the
                    # UI's terminal rendering branches stay clean.
                    st["status"]     = "cancelled"
                    st["result"]     = None
                    st["updated_ts"] = time.time()
                else:
                    st["status"]     = "done"
                    st["result"]     = dns_run_to_dict(result)
                    st["updated_ts"] = time.time()
                if self._dns_active_task_id == task_id:
                    self._dns_active_task_id  = None
                    self._dns_active_task_obj = None

        def _on_error(msg):
            with self._dns_tasks_lock:
                st = self._dns_tasks.get(task_id)
                if st is None:
                    return
                if st.get("status") == "cancelled":
                    if self._dns_active_task_id == task_id:
                        self._dns_active_task_id  = None
                        self._dns_active_task_obj = None
                    return
                st["status"]     = "error"
                st["error"]      = str(msg)
                st["updated_ts"] = time.time()
                if self._dns_active_task_id == task_id:
                    self._dns_active_task_id  = None
                    self._dns_active_task_obj = None

        return _on_done, _on_error


    def get_url(self) -> str:
        """LAN URL — shown as text so other devices can connect."""
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80)); ip = s.getsockname()[0]; s.close()
        except Exception: ip = "127.0.0.1"
        return f"http://{ip}:{self.port}"



    @property
    def is_running(self) -> bool: return self._running

    def _run(self):
        """
        Flask server thread.

        Diagnostics:  writes  web_server.log  next to the exe (frozen) or
        project root (dev) **before** app.run() so we know if the issue is
        pre-bind (import/init) or post-bind (runtime crash).
        """
        import sys as _sys, os as _os, traceback as _tb

        # --- resolve a writable log dir regardless of frozen/dev mode -------
        if self._log_dir:
            _log_dir = self._log_dir   # explicitly set by caller (e.g. DATA_DIR)
        elif getattr(_sys, 'frozen', False):
            # Fallback: LOCALAPPDATA/NetMonitor/logs — consistent with main.py DATA_DIR
            _lad = _os.environ.get("LOCALAPPDATA") or _os.path.expanduser("~")
            _log_dir = _os.path.join(_lad, "NetMonitor", "logs")
        else:
            _log_dir = _os.path.dirname(_os.path.dirname(
                           _os.path.abspath(__file__)))     # project root
        _os.makedirs(_log_dir, exist_ok=True)

        _log = _os.path.join(_log_dir, 'web_server.log')

        def _write(msg: str):
            try:
                with open(_log, 'a', encoding='utf-8') as _f:
                    import datetime as _dt
                    _f.write(f"[{_dt.datetime.now().isoformat()}] {msg}\n")
            except Exception:
                pass

        _write(f"Flask thread starting  port={self.port}  frozen={getattr(_sys,'frozen',False)}")
        _write(f"Python {_sys.version}  exe={_sys.executable}")

        try:
            import flask as _flask
            import werkzeug as _wz
            _write(f"Flask {_flask.__version__}  Werkzeug {_wz.__version__}")
        except Exception as _e:
            _write(f"VERSION CHECK FAILED: {_e}")

        # Flask calls click.echo() for the startup banner, which requires a
        # writable sys.stdout. With console=False in PyInstaller they are None.
        # NEVER use io.StringIO() here: it is an unbounded in-memory buffer that
        # accumulates every printed byte for the lifetime of the process, causing
        # OOM after weeks of 7x24 operation. Use os.devnull — a true black-hole.
        if getattr(_sys, 'frozen', False):
            import os as _os
            _devnull = open(_os.devnull, 'w', encoding='utf-8', errors='replace')
            if _sys.stdout is None:
                _sys.stdout = _devnull
            if _sys.stderr is None:
                _sys.stderr = _devnull

        try:
            from werkzeug.serving import make_server as _make_wsgi_server
            # Loopback by default; bind 0.0.0.0 only when the operator
            # explicitly opted into LAN access via web_allow_lan.
            bind_host = "0.0.0.0" if self._allow_lan else "127.0.0.1"
            srv = _make_wsgi_server(bind_host, self.port, self._app,
                                    threaded=True)
            self._server = srv
            # Re-assert _running in case an OLD _run thread's finally
            # block managed to slip in between restart()'s
            # `_running=True` and our make_server() return.  Pair this
            # with the `self._server is srv` ownership check in the
            # finally below: together they keep restart(new_port) clean
            # whether the old thread exits before, during, or after the
            # new thread finishes binding.
            self._running = True
            _write(f"serving on http://{bind_host}:{self.port}")
            # Bind succeeded — signal start() so it can stop blocking and
            # report success.  serve_forever() blocks indefinitely so the
            # signal MUST happen before it; otherwise start() times out
            # even on a perfectly healthy launch.
            self._bind_ok = True
            self._bind_done.set()
            srv.serve_forever()   # blocks until srv.shutdown() is called
            _write("server stopped cleanly")
        except Exception:
            _write("server RAISED EXCEPTION:")
            try:
                with open(_log, 'a', encoding='utf-8') as _f:
                    _tb.print_exc(file=_f)
            except Exception:
                pass
        finally:
            # Only clear shared state if WE still own self._server.  When
            # restart(new_port=different) spawns a NEW _run thread before
            # our serve_forever() returns, that new thread has already
            # replaced self._server with its own srv handle; unconditionally
            # clearing here would orphan the live new server (port keeps
            # listening but is_running()=False, _server=None — UI thinks
            # the service is down while it's actually serving requests).
            srv_local = locals().get("srv")
            if self._server is srv_local:
                self._running = False
                self._server  = None
            # Wake any start() still blocked on bind_done so failure paths
            # (port-in-use, permission denied, import error, …) don't make
            # the caller wait the full 3 s timeout for a result it could
            # already know is False.
            self._bind_done.set()

    def _setup_routes(self):
        app = self._app

        @app.teardown_request
        def _release_request_db(exc):
            if self._data_store is not None:
                try:
                    self._data_store.release_read_conn()
                except Exception:
                    pass

        # Shared formula-injection sanitiser -- defined at module
        # scope in data_store so both the SLA / waveform /
        # general-export paths and DataStore.export_excel share one
        # rule set.  Previously this lived as a closure here and only
        # /api/waveform/export benefited; /api/export/sla and
        # /api/export still wrote untrusted labels verbatim.
        #
        # Package-relative import: the project is run as `python main.py`
        # with `from src.web_server import WebServer`, so the top-level
        # `data_store` module does not exist on sys.path.  The previous
        # absolute import `from data_store import ...` raised
        # ModuleNotFoundError on every WebServer() construction,
        # blocking the desktop app's startup chain entirely.  All other
        # cross-module imports in this file already use the
        # `from .icmp_diag import …` / `from .dns_diag import …` form
        # so this just matches the existing convention.
        from .data_store import excel_safe_text as _excel_safe_text

        def _body_str(body, key, default=""):
            """Read a string field from a parsed JSON body, falling
            back to ``default`` if the value is missing OR not a
            string.

            ``body.get(k) or default).strip()`` raises AttributeError
            when the caller put a non-string (int, bool, list, dict,
            None — wait, ``None or ""`` is ``""``, but ``1 or ""`` is
            ``1``) under the key, which Flask returns as a 500.  The
            isinstance gate collapses every non-string shape to the
            default so the subsequent .strip() and .upper() chain on
            the caller side stays infallible.  Empty/whitespace
            strings still strip to "" and the route's own "required
            field" check handles them as missing input -- the same
            behaviour legitimate clients have seen all along.
            """
            v = body.get(key, default)
            if not isinstance(v, str):
                return default
            return v.strip()

        def _qnum(name, default, cast=int):
            """Parse a numeric query-string arg, falling back to
            ``default`` on missing / malformed / wrong-type input.

            Defence for the public time-range API surface: the bare
            ``int(request.args.get("start", default))`` form raises
            ValueError on any non-numeric string the caller sends
            (e.g. ``?start=abc``) and Flask turns that into a 500
            with a stack trace in the log.  These endpoints (/api/sla,
            /api/stats, /api/waveform, /api/waveform/alerts,
            /api/history, /api/export/sla) are reachable by anything
            that can hit the dashboard's port, so a single curl with
            a bad parameter could repeatedly flood the log and
            return server errors to the dashboard's regular polling.
            Falling back to the same default the route would use
            when the arg is omitted keeps the call usable while
            silently dropping the bad input.
            """
            try:
                v = request.args.get(name)
                if v is None or v == "":
                    return default
                return cast(v)
            except (TypeError, ValueError):
                return default

        @app.route("/")
        def index():
            return DASHBOARD_HTML, 200, {"Content-Type": "text/html; charset=utf-8"}

        @app.route("/api/status")
        def api_status():
            with self._lock: data = list(self._targets.values())
            return json.dumps(data), 200, {"Content-Type": "application/json"}

        # ── Quick-tracert window (Feature B) ─────────────────────────────
        # Accepts either:
        #   ?tid=<monitored-target-id>   (original Task B path)
        #   ?ip=<ipv4-or-ipv6-literal>   (Phase 2 — DNS "追踪此 IP" entry)
        # When `ip` is provided without `tid`, we skip the targets lookup
        # and let the operator trace any literal IP (LAN-tool semantics:
        # equivalent to running `tracert <ip>` in cmd; not recorded in
        # traceroute_logs).  The IP regex is strict — domains and shell
        # metacharacters are rejected long before reaching tracert.exe.
        @app.route("/traceroute/window")
        def traceroute_window():
            """Return the standalone quick-tracert window HTML page."""
            tid   = request.args.get("tid", "").strip()
            ip_q  = request.args.get("ip",  "").strip()
            import html as _html
            if tid:
                with self._lock:
                    target = self._targets.get(tid)
                if not target:
                    return "Target not found", 404
                label = _html.escape(target.get("label", tid))
                ip    = _html.escape(target.get("ip", ""))
                qs    = "tid=" + tid
            elif ip_q:
                if not _is_ip_literal(ip_q):
                    return "Invalid IP", 400
                label = "DNS 解析 IP"
                ip    = _html.escape(ip_q)
                qs    = "ip=" + ip_q
            else:
                return "Target not found", 404
            html  = QUICK_TRACERT_HTML.format(qs=qs, label=label, ip=ip)
            return html, 200, {"Content-Type": "text/html; charset=utf-8"}

        @app.route("/api/traceroute/quick")
        def api_quick_traceroute():
            """
            Streaming SSE endpoint for quick (-d) traceroute.

            Each hop is emitted as soon as the process prints its line, so
            the browser sees results in real-time — not all at once at the end.
            Uses subprocess.Popen (not run) so we can iterate stdout line by
            line while the process is still running.

            Selector params (exactly one of these is required):
              tid=<target_id>  – trace the IP of an existing monitored node
              ip=<ip_literal>  – trace an arbitrary IPv4/IPv6 literal,
                                 used by the DNS-diag "追踪此 IP" buttons
                                 (Phase 2).  The literal is validated
                                 strictly before reaching the subprocess.

            Security:
              • tid path: only IPs that correspond to an existing target.
              • ip path: only validated IPv4/IPv6 literals (no hostnames).
              • Max 3 concurrent quick-traces per server (enforced below).
            """
            tid   = request.args.get("tid", "").strip()
            ip_q  = request.args.get("ip",  "").strip()
            if tid:
                with self._lock:
                    target = self._targets.get(tid)
                if not target:
                    return json.dumps({"error": "target not found"}), 404,                            {"Content-Type": "application/json"}
                ip_str = target.get("ip", "")
                label  = target.get("label", ip_str)
                # Resolve domain → IPv4 for the subprocess call
                if re.match(r"^\d{1,3}(\.\d{1,3}){3}$", ip_str):
                    resolve_ip = ip_str
                else:
                    try:
                        resolve_ip = socket.gethostbyname(ip_str)
                    except socket.gaierror:
                        resolve_ip = ip_str
            elif ip_q:
                # Phase 2: DNS "追踪此 IP" path — operator-supplied IP
                # literal, no target lookup.  Reject anything that's not
                # a valid v4/v6 literal before it can reach the
                # subprocess argv (defence in depth — argv is a list, not
                # a shell string, so injection is already impossible,
                # but bad IPs cause useless 120s tracert hangs).
                if not _is_ip_literal(ip_q):
                    return json.dumps({"error": "invalid ip"}), 400,                            {"Content-Type": "application/json"}
                resolve_ip = ip_q
                label      = "DNS 解析 IP"
            else:
                return json.dumps({"error": "target not found"}), 404,                        {"Content-Type": "application/json"}

            def generate():
                is_win   = platform.system().lower() == "windows"
                max_hops = str(self._scheduler.max_hops)
                proc     = None
                # Acquire semaphore — reject immediately if 3 are already running
                if not self._quick_trace_sem.acquire(blocking=False):
                    yield (b"data: " +
                           json.dumps({"type": "error",
                                       "message": "当前快速追踪任务过多（最多3个并发），请稍后再试"})
                           .encode() + b"\n\n")
                    return
                try:
                    if is_win:
                        sysroot = os.environ.get("SystemRoot", "C:\\Windows")
                        tracert = os.path.join(sysroot, "System32", "tracert.exe")
                        if not os.path.isfile(tracert):
                            tracert = "tracert"
                        si = subprocess.STARTUPINFO()
                        si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                        si.wShowWindow = subprocess.SW_HIDE
                        cmd = [tracert, "-d", "-h", max_hops, "-w", "800", resolve_ip]
                        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE,
                                                stderr=subprocess.PIPE,
                                                startupinfo=si)
                    else:
                        cmd  = ["traceroute", "-n", "-m", max_hops, "-w", "1", resolve_ip]
                        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE,
                                                stderr=subprocess.PIPE)

                    # Stream lines as they arrive
                    for raw_line in iter(proc.stdout.readline, b""):
                        line = raw_line.decode("gbk" if is_win else "utf-8",
                                               errors="replace").strip()
                        m = re.match(r"^(\d+)\s+", line)
                        if not m:
                            continue
                        hop_n = int(m.group(1))
                        rest  = line[m.end():]

                        has_ms    = bool(re.search(r"\d+\s*ms", rest, re.I))
                        asterisks = rest.count("*")
                        is_timeout = asterisks >= 3 and not has_ms

                        if is_timeout:
                            hop = {"type": "hop", "hop": hop_n,
                                   "ip": None, "ms": None, "is_target": False}
                        else:
                            rtt_m = re.search(r"[<]?(\d+(?:\.\d+)?)\s*ms", rest, re.I)
                            ms    = round(float(rtt_m.group(1)), 1) if rtt_m else None
                            # In -d mode, no hostname — just look for bare IP or IPv6
                            # Windows: "   1    3 ms    2 ms    1 ms  192.168.1.1"
                            # Remove time values to isolate IP at end
                            ip_part = re.sub(r"[<]?\d+\s*ms\s*", "", rest).strip()
                            # Match IPv4 or IPv6
                            ip_m = re.search(
                                r"([0-9a-fA-F:]+(?::[0-9a-fA-F]+)+|"
                                r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", ip_part)
                            hop_ip = ip_m.group(1) if ip_m else None
                            is_target = (hop_ip == resolve_ip)
                            hop = {"type": "hop", "hop": hop_n,
                                   "ip": hop_ip, "ms": ms,
                                   "is_target": is_target}

                        yield f"data: {json.dumps(hop)}\n\n"

                    proc.wait(timeout=5)
                    yield f"data: {json.dumps({'type': 'done'})}\n\n"

                except GeneratorExit:
                    pass   # client disconnected — finally block handles cleanup
                except Exception as exc:
                    try:
                        yield f"data: {json.dumps({'type':'error','message':str(exc)})}\n\n"
                    except Exception:
                        pass
                finally:
                    self._quick_trace_sem.release()
                    if proc and proc.poll() is None:
                        try: proc.kill()
                        except Exception: pass

            from flask import stream_with_context, Response
            return Response(
                stream_with_context(generate()),
                content_type="text/event-stream",
                headers={"Cache-Control": "no-cache",
                         "X-Accel-Buffering": "no"})

        @app.route("/api/sla")
        def api_sla():
            if not self._data_store:
                return json.dumps({}), 200, {"Content-Type": "application/json"}
            start = _qnum("start", time.time() - 86400*30, cast=float)
            end   = _qnum("end",   time.time(),            cast=float)
            with self._lock:
                tids = {tid: {"label": t.get("label",""), "ip": t.get("ip","")}
                        for tid, t in self._targets.items()}
            # Batch fetch all targets at once: previously this did
            # len(tids) serial DB hits (2 SQL each).  Now it's 2 SQL total.
            batch = self._data_store.get_sla_stats_batch(
                list(tids.keys()), start, end)
            result = {}
            for tid, meta in tids.items():
                s = batch.get(tid) or self._data_store.get_sla_stats(tid, start, end)
                s["label"] = meta["label"]; s["ip"] = meta["ip"]
                result[tid] = s
            return json.dumps(result), 200, {"Content-Type": "application/json"}

        @app.route("/api/stats")
        def api_stats():
            """
            GET /api/stats?start=<unix_ts>&end=<unix_ts>
            Returns per-target aggregated stats for a time range.
            Auto-selects pings_raw (span<=24h) or pings_hourly (span>24h).
            Used by the frontend charts for time-range-aware rendering.
            """
            if not self._data_store:
                return json.dumps({}), 200, {"Content-Type":"application/json"}
            start = _qnum("start", int(time.time()) - 86400)
            end   = _qnum("end",   int(time.time()))
            span  = end - start
            use_hourly = span > 86400
            result = {}
            # Only query active targets — deleted nodes stay in DB (soft-delete)
            # but must never appear in API responses.
            with self._lock:
                tids = list(self._targets.keys())
            import math as _math
            for tid in tids:
                rows = self._data_store.get_history(
                    tid, start, end, use_hourly=use_hourly)
                if not rows:
                    continue
                if use_hourly:
                    total   = sum(r["sample_n"] for r in rows)
                    success = sum(r["success_n"] for r in rows)
                    # Weight per-hour averages by success_n -- the same
                    # weighting get_sla_stats does in SQL.  The previous
                    # unweighted `sum(lats)/len(lats)` treated one
                    # 1-sample hour the same as a 3600-sample hour, so
                    # a single bad outlier sample could shift the
                    # dashboard's reported average by orders of
                    # magnitude on long ranges.
                    lat_pairs = [(r["lat_avg"], r["success_n"]) for r in rows
                                  if r["lat_avg"] is not None and r["success_n"]]
                    p95_pairs = [(r["lat_p95"], r["success_n"]) for r in rows
                                  if r["lat_p95"] is not None and r["success_n"]]
                    maxlats   = [r["lat_max"] for r in rows
                                  if r["lat_max"] is not None]
                    w_n  = sum(n for _v, n in lat_pairs)
                    lat_avg = (sum(v * n for v, n in lat_pairs) / w_n
                                if w_n else None)
                    w_n_p95 = sum(n for _v, n in p95_pairs)
                    lat_p95 = (sum(v * n for v, n in p95_pairs) / w_n_p95
                                if w_n_p95 else None)
                    lat_max = max(maxlats) if maxlats else None
                else:
                    total   = len(rows)
                    success = sum(1 for r in rows
                                  if r["status"] in ("green","orange"))
                    lats    = [r["latency_ms"] for r in rows
                                if r["latency_ms"] is not None]
                    # Nearest-rank P95 with ceiling -- matches the fix
                    # already applied to data_store._hourly_agg.  The
                    # previous `int(...) - 1` floor returned the MIN
                    # value for n=2 and was off-by-one for many other
                    # small sample counts (e.g. n=19 picked the 18th
                    # value instead of the 19th).
                    if lats:
                        idx = min(len(lats) - 1,
                                  max(0, _math.ceil(len(lats) * 0.95) - 1))
                        lat_p95 = sorted(lats)[idx]
                    else:
                        lat_p95 = None
                    lat_avg = sum(lats) / len(lats) if lats else None
                    lat_max = max(lats) if lats else None
                result[tid] = {
                    "total":   total,
                    "success": success,
                    "lat_avg": round(lat_avg, 1) if lat_avg is not None else None,
                    "lat_max": round(lat_max, 1) if lat_max is not None else None,
                    "lat_p95": round(lat_p95, 1) if lat_p95 is not None else None,
                }
            return json.dumps(result), 200, {"Content-Type":"application/json"}

        @app.route("/api/waveform/export")
        def api_waveform_export():
            """
            GET /api/waveform/export?tid=X&start=X&end=X
            Export waveform data as Excel for the selected node and time range.
            """
            if not self._data_store:
                return "无数据", 503
            tid   = request.args.get("tid","").strip()
            now   = int(time.time())
            start = _qnum("start", now - 3600)
            end   = _qnum("end",   now)
            if not tid:
                return "缺少 tid 参数", 400
            # Resolve label from the trusted server-side snapshot,
            # not from the URL query.  The previous code did
            # `label = request.args.get("label","节点").strip()`,
            # which let any caller seed the Excel summary sheet with
            # arbitrary text -- including formula-trigger strings
            # like "=1+1" or "=HYPERLINK(...)" that openpyxl saves
            # verbatim as cells with data_type='f', so opening the
            # workbook would evaluate the attacker-supplied formula
            # (CWE-1236, "CSV/Excel formula injection").  We accept
            # the URL label only as a hint and fall back to the
            # server-known label for that tid; the writer also runs
            # _excel_safe_text on the value before persisting (defence
            # in depth covers admin-typed labels like "=production").
            with self._lock:
                tgt = self._targets.get(tid)
            label = (tgt.get("label", "节点").strip()
                     if tgt else "节点")

            data = self._data_store.get_waveform(tid, start, end)
            pts  = data.get("points", [])
            summ = data.get("summary", {})
            res  = data.get("resolution", "raw")

            # Return clear error instead of an empty/confusing file
            if not pts:
                return json.dumps({"error": "该时间范围内无可导出数据"}), 404, \
                       {"Content-Type": "application/json"}

            res_cn = {"raw":"秒级", "1m":"分钟级", "1h":"小时级"}.get(res, res)

            try:
                import openpyxl, io, datetime as _dt
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "波形数据"

                # Header row
                hdr_fill = PatternFill("solid", fgColor="1E40AF")
                hdr_font = Font(bold=True, color="FFFFFF")
                headers  = ["时间", "延时 (ms)", "状态", "bucket_state"]
                for col, h in enumerate(headers, 1):
                    c = ws.cell(1, col, h)
                    c.fill, c.font = hdr_fill, hdr_font
                    c.alignment = Alignment(horizontal="center")
                ws.column_dimensions["A"].width = 20
                ws.column_dimensions["B"].width = 14
                ws.column_dimensions["C"].width = 12
                ws.column_dimensions["D"].width = 16

                # Data rows
                state_cn = {"ok":"正常", "partial_fail":"部分丢包", "all_fail":"全部丢包"}
                for p in pts:
                    dt_str = _dt.datetime.fromtimestamp(p["ts"]).strftime("%Y-%m-%d %H:%M:%S")
                    ws.append([dt_str,
                                p["rtt_ms"] if p["rtt_ms"] is not None else "",
                                state_cn.get(p.get("bucket_state","ok"), ""),
                                p.get("bucket_state","")])

                # Summary sheet
                ws2 = wb.create_sheet("摘要")
                ws2.column_dimensions["A"].width = 20
                ws2.column_dimensions["B"].width = 16
                pairs = [
                    ("节点",      _excel_safe_text(label)),
                    ("数据粒度",  res_cn),
                    ("时间范围",  f"{_dt.datetime.fromtimestamp(start):%Y-%m-%d %H:%M} ~ "
                                   f"{_dt.datetime.fromtimestamp(end):%Y-%m-%d %H:%M}"),
                    ("样本数",    summ.get("sample_n","")),
                    ("平均延时",  f"{summ['avg_ms']} ms" if summ.get("avg_ms") is not None else "—"),
                    ("最小延时",  f"{summ['min_ms']} ms" if summ.get("min_ms") is not None else "—"),
                    ("最大延时",  f"{summ['max_ms']} ms" if summ.get("max_ms") is not None else "—"),
                    ("丢包率",    f"{summ['loss_rate']*100:.2f}%"
                                   if summ.get("loss_rate") is not None else "—"),
                ]
                for row in pairs:
                    ws2.append(row)

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)

                import urllib.parse as _up
                start_str = _dt.datetime.fromtimestamp(start).strftime("%Y%m%d")
                end_str   = _dt.datetime.fromtimestamp(end).strftime("%Y%m%d")
                filename  = _up.quote(f"波形_{label}_{start_str}_{end_str}.xlsx")
                return buf.getvalue(), 200, {
                    "Content-Type":        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
            except Exception as exc:
                return f"导出失败: {exc}", 500

        @app.route("/api/waveform")
        def api_waveform():
            """
            GET /api/waveform?tid=X&range=1h|24h|7d|30d
            or  GET /api/waveform?tid=X&start=<ts>&end=<ts>
            Auto-selects data resolution by time span.
            Returns {resolution, points:[{ts,rtt_ms,ok},...], summary:{...}}
            """
            if not self._data_store:
                return json.dumps({"resolution":"raw","points":[],"summary":{}}), 200, \
                       {"Content-Type":"application/json"}
            tid = request.args.get("tid","").strip()
            now = int(time.time())
            quick = request.args.get("range","")
            span_map = {"1h": 3600, "24h": 86400, "7d": 604800, "30d": 2592000}
            if quick in span_map:
                start, end = now - span_map[quick], now
            else:
                start = _qnum("start", now - 3600)
                end   = _qnum("end",   now)
            data = self._data_store.get_waveform(tid, start, end)
            return json.dumps(data), 200, {"Content-Type":"application/json"}

        @app.route("/api/waveform/alerts")
        def api_waveform_alerts():
            """
            GET /api/waveform/alerts?tid=X&start=<ts>&end=<ts>
            Returns merged red/orange outage intervals plus raw transition
            events for waveform-chart overlay rendering.

            Decoupled from /api/waveform so the chart can re-render
            outage bands without re-fetching the (possibly large) RTT point
            series, and so a future "show only RTT" toggle can simply skip
            this request.
            """
            if not self._data_store:
                return json.dumps({"red":[],"orange":[],"events":[]}), 200, \
                       {"Content-Type":"application/json"}
            tid = request.args.get("tid","").strip()
            now = int(time.time())
            quick = request.args.get("range","")
            span_map = {"1h": 3600, "24h": 86400, "7d": 604800, "30d": 2592000}
            if quick in span_map:
                start, end = now - span_map[quick], now
            else:
                start = _qnum("start", now - 3600)
                end   = _qnum("end",   now)
            data = self._data_store.get_outage_intervals(tid, start, end)
            return json.dumps(data), 200, {"Content-Type":"application/json"}

        @app.route("/api/history")
        def api_history():
            """
            GET /api/history?tid=...&start=<unix_ts>&end=<unix_ts>&type=raw|hourly
            Returns JSON array of data points for one target.
            """
            if not self._data_store:
                return json.dumps([]), 200, {"Content-Type":"application/json"}
            tid   = request.args.get("tid","").strip()
            start = _qnum("start", int(time.time()) - 86400)
            end   = _qnum("end",   int(time.time()))
            dtype = request.args.get("type", "raw")
            # Auto-downsample for over-long raw windows.  A 1Hz target
            # with 7d raw retention can produce ~600k rows = ~70 MB of
            # JSON for a single request; serialising that on the Flask
            # worker thread blocks every other request and can OOM the
            # process under concurrent load.  Anything wider than 24h
            # is forced to the hourly aggregate (same threshold
            # /api/waveform already uses for its 1m -> 1h switch).
            # Callers that want truly fine-grained data over a long
            # window must paginate by issuing multiple narrower
            # requests.
            _RAW_SPAN_CAP = 86400   # 24 h
            if dtype == "raw" and (end - start) > _RAW_SPAN_CAP:
                dtype = "hourly"
            rows  = self._data_store.get_history(
                        tid, start, end, use_hourly=(dtype == "hourly"))
            # Defence in depth: even within the 24h raw window a
            # corrupt request with sub-second probes could blow past
            # a reasonable response size.  Slice to the most-recent
            # 50k rows (the time series's tail is what the dashboard
            # actually plots).
            _ROW_CAP = 50000
            if len(rows) > _ROW_CAP:
                rows = rows[-_ROW_CAP:]
            return json.dumps(rows), 200, {"Content-Type":"application/json"}

        @app.route("/api/export/sla")
        def api_export_sla():
            """Export SLA report as Excel."""
            if not self._data_store:
                return "No data", 404
            start = _qnum("start", time.time() - 86400*30, cast=float)
            end   = _qnum("end",   time.time(),            cast=float)
            with self._lock:
                tids = {tid: {"label": t.get("label",""), "ip": t.get("ip","")}
                        for tid, t in self._targets.items()}
            # Use the batch API: get_sla_stats() looped per-target opens N
            # DB connections / runs N hourly-aggregate queries and was
            # already known-bad enough for the dashboard JSON API to
            # switch (see /api/sla above).  Fall back to the per-target
            # call only if a tid is missing from the batch result (e.g.
            # added between snapshot and query).
            stats = self._data_store.get_sla_stats_batch(
                list(tids.keys()), start, end)
            rows = []
            for tid, meta in tids.items():
                s = stats.get(tid) or self._data_store.get_sla_stats(tid, start, end)
                # Sanitise label / ip BEFORE building the row dict --
                # ws_xl.append below pulls cell values verbatim from
                # this dict.  An admin-typed (or attacker-supplied)
                # label/ip starting with =/+/-/@ would otherwise hit
                # openpyxl's autodetect and be stored as a formula
                # cell, executed on workbook open.  Numeric / formatted
                # columns are server-generated and pass through
                # unchanged.
                rows.append({
                    "节点名称": _excel_safe_text(meta["label"]),
                    "IP / 域名": _excel_safe_text(meta["ip"]),
                    "连通率(%)": s["uptime_pct"],
                    "均延时(ms)": s["lat_avg"], "P95延时(ms)": s["lat_p95"],
                    "最大延时(ms)": s["lat_max"],
                    "故障次数": s["outage_count"],
                    "累计中断(分钟)": s["outage_minutes"],
                })
            try:
                import openpyxl, io
                from datetime import datetime as _dt
                wb = openpyxl.Workbook()
                ws_xl = wb.active
                ws_xl.title = "SLA报告"
                headers = list(rows[0].keys()) if rows else []
                ws_xl.append(headers)
                for r in rows:
                    ws_xl.append([r.get(h) for h in headers])
                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                ts_str = _dt.now().strftime("%Y%m%d")
                # HTTP headers are ASCII-only; use RFC 5987 for UTF-8 filename
                import urllib.parse as _up
                fname_ascii = f"SLA_Report_{ts_str}.xlsx"
                fname_utf8  = _up.quote(f"SLA报告_{ts_str}.xlsx")
                return buf.getvalue(), 200, {
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "Content-Disposition": (
                        f'attachment; filename="{fname_ascii}"; '
                        f"filename*=UTF-8''{fname_utf8}"
                    )
                }
            except Exception as e:
                return f"Export error: {e}", 500

        @app.route("/api/export")
        def api_export():
            """
            GET /api/export?tids=t1,t2&start=<unix_ts>&end=<unix_ts>
            Generates and streams an Excel workbook for download.
            Auto-selects raw/hourly based on time span.
            """
            if not self._data_store:
                return "DataStore not available", 503
            tids_raw = request.args.get("tids","").strip()
            requested = [t.strip() for t in tids_raw.split(",") if t.strip()]
            # Filter the requested tids against the trusted server-side
            # snapshot so caller-controlled junk (control characters,
            # arbitrary user-supplied IDs, attempts to enumerate tids
            # that never existed) can never reach DataStore.export_excel,
            # where the unknown tid was previously used as the sheet's
            # label cell and crashed openpyxl with IllegalCharacterError
            # on bytes like %00.
            with self._lock:
                valid_tids = set(self._targets)
            if requested:
                tids = [t for t in requested if t in valid_tids]
                if not tids:
                    return (json.dumps({"error": "no valid tid"}), 400,
                            {"Content-Type": "application/json"})
            else:
                # Empty tids= preserves "export every current node"
                # behaviour the dashboard relies on for the global
                # export button.
                tids = list(valid_tids)
            start = _qnum("start", int(time.time()) - 86400)
            end   = _qnum("end",   int(time.time()))
            import tempfile, os
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                tmp_path = f.name
            try:
                with self._lock:
                    meta = {t["tid"]: {"label": t.get("label",""),
                                       "ip":    t.get("ip",""),
                                       "ping_type": t.get("ping_type","icmp")}
                            for t in self._targets.values()}
                self._data_store.export_excel(tmp_path, tids, start, end, meta)
                dt = datetime.now().strftime("%Y%m%d_%H%M")
                fname = f"monitor_export_{dt}.xlsx"
                with open(tmp_path, "rb") as fh:
                    data = fh.read()
                return data, 200, {
                    "Content-Type":
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet",
                    "Content-Disposition": f'attachment; filename="{fname}"'}
            finally:
                try: os.unlink(tmp_path)
                except Exception: pass

        @app.route("/api/sound/<path:name>")
        def api_sound(name):
            """Serve alert WAV files from the assets directory.
            The web dashboard uses these instead of synthesised oscillator
            tones, which some browsers/policies silently suppress.
            Only .wav and .mp3 files are served (no path traversal)."""
            import mimetypes, pathlib
            safe = pathlib.PurePosixPath(name).name   # strip any dir components
            if not re.match(r"^[a-zA-Z0-9_\-]+\.(wav|mp3|ogg)$", safe):
                return "", 404
            # Anchor to project root (parent of src/) so the path is correct
            # regardless of whether Flask's root_path or os.getcwd() is used.
            # send_file() with a relative path resolves against Flask's root_path
            # (= src/), but assets/ lives one level up at the project root.
            project_root  = pathlib.Path(__file__).parent.parent
            project_assets = (project_root / "assets").resolve()
            # (base_dir, candidate_path) pairs.  Path traversal protection is
            # enforced by checking that the resolved candidate is contained in
            # its declared base_dir — not in `candidate.parent`, which was a
            # tautological check that did nothing.
            candidates = []
            if self._sound_dir:
                sd = pathlib.Path(self._sound_dir).resolve()
                candidates.append((sd, (sd / safe).resolve()))
            candidates.append((project_assets, (project_assets / safe).resolve()))

            fp = None
            for base_dir, candidate in candidates:
                # Containment check: candidate must live under base_dir.
                # Redundant given the filename regex above already forbids
                # path separators, but spelled out so any future loosening
                # of `safe` doesn't silently open a traversal hole.
                try:
                    candidate.relative_to(base_dir)
                except ValueError:
                    continue
                if candidate.exists():
                    fp = candidate
                    break
            if fp is None:
                return "", 404
            mt = mimetypes.guess_type(str(fp))[0] or "audio/wav"
            from flask import send_file
            return send_file(fp, mimetype=mt, max_age=3600)

        @app.route("/api/traceroute/history")
        def api_traceroute_history():
            tid   = request.args.get("tid", "")
            # Mirror the defensive pattern already used in
            # api_icmp_diag_history / api_dns_diag_history: a bare
            # int() on user-controlled input raises on garbage like
            # ?limit=abc, surfacing as a Flask 500 + stack trace --
            # one-shot DoS on a public GET endpoint.  Default to 5,
            # clamp into a sane window.
            try:
                limit = int(request.args.get("limit", 5))
            except (TypeError, ValueError):
                limit = 5
            limit = max(1, min(500, limit))
            if not tid or not self._data_store:
                return json.dumps([]), 200, {"Content-Type": "application/json"}
            try:
                rows = self._data_store.get_traceroute_history(tid, limit)
            except Exception:
                rows = []
            return json.dumps(rows), 200, {"Content-Type": "application/json"}

        @app.route("/api/traceroute")
        def api_traceroute():
            tid   = request.args.get("tid", "").strip()
            force = request.args.get("force", "0") == "1"
            with self._lock: target = self._targets.get(tid)
            if not target:
                return json.dumps({"error": "target not found"}), 404
            ip = target["ip"]
            if force:
                # Clear stale cache so the fresh result replaces it
                with self._tracer_lock:
                    self._tracer_cache.pop(tid, None)
            with self._tracer_lock:
                cached = self._tracer_cache.get(tid)
                # Defence in depth against the cache-vs-current-ip
                # mismatch the update_target path already evicts on.
                # If anything ever bypasses that eviction (manual cache
                # population, future code path that mutates ip without
                # going through update_target, etc.), treat the cached
                # entry as a miss rather than serving someone else's
                # hop list under this tid.
                if cached is not None and cached.get("ip") != ip:
                    self._tracer_cache.pop(tid, None)
                    cached = None
            if cached is None:
                # Delegate to the background scheduler — never block a Flask
                # worker thread for up to 135s (120s traceroute + 15s TCP).
                # The frontend will poll /api/traceroute until the result lands.
                self._scheduler.request_now(tid)
                return json.dumps({"status": "pending"}), 202, \
                    {"Content-Type": "application/json"}
            return json.dumps(cached), 200, {"Content-Type": "application/json"}

        # ICMP advanced diagnostic API (Task B)
        # POST /api/icmp_diag/start      - kick off a diagnostic run
        # GET  /api/icmp_diag/status     - poll progress / final result
        # POST /api/icmp_diag/cancel     - abort a running task
        # GET  /api/icmp_diag/history    - read DB-persisted history rows
        #
        # The Web layer enforces at-most-one-active task on its own side
        # (_icmp_active_task_id, atomic under _icmp_tasks_lock). The
        # cross-side guarantee with C-side comes from icmp_diag._TASK_SEM(1)
        # inside run_icmp_diag_async; we surface its on_error("...") message
        # as a 409 via the same task state flow as a regular validation error.
        @app.route("/api/icmp_diag/start", methods=["POST"])
        def api_icmp_diag_start():
            import uuid
            from .icmp_diag import (IcmpDiagRequest, validate_request,
                                    run_icmp_diag_async, run_pmtu_async)
            try:
                body = request.get_json(force=True, silent=True)
            except Exception:
                body = None
            # `... or {}` alone treats only the falsy-shaped JSON
            # values (null, [], "", 0, {}) as missing; non-empty
            # arrays / numbers / strings pass through unchanged, and
            # the immediate body.get(...) below then crashes with
            # AttributeError -> 500.  Explicit isinstance check
            # collapses every non-object JSON to an empty dict so the
            # field defaults below apply uniformly.  Callers that
            # depended on the silent coercion (legitimate clients
            # always send an object) see the same defaults as before.
            if not isinstance(body, dict):
                body = {}
            # _body_str rejects non-string values (e.g. {"host": 1})
            # which the previous `(body.get(k) or "").strip()` chain
            # crashed on with AttributeError -> 500; see helper above.
            tid     = _body_str(body, "target_id") or None
            host    = _body_str(body, "host")
            mode    = _body_str(body, "mode", "single") or "single"
            if not host:
                return (json.dumps({"error": "host is required"}),
                        400, {"Content-Type": "application/json"})
            if mode not in ("single", "stability", "pmtu"):
                return (json.dumps({"error": "invalid mode"}),
                        400, {"Content-Type": "application/json"})

            # Coerce numeric params with defensive defaults — JS may send
            # them as strings or omit them entirely.
            def _num(key, default, cast=float):
                v = body.get(key, default)
                try:    return cast(v)
                except (TypeError, ValueError): return default
            payload_size = _num("payload_size", 32, int)
            count        = _num("count",        4,  int)
            timeout_s    = _num("timeout_s",    2.0)
            interval_s   = _num("interval_s",   1.0)
            df_flag      = bool(body.get("df", False))

            # Build the request object up front so validate_request() can
            # reject bad params *before* we touch the active-task slot.
            # max_runtime_s: same formula the desktop diag window uses
            # (icmp_diag_window.py:355).  Without this, the default
            # 30 s ceiling cuts any non-trivial stability run short
            # (e.g. count=50 × interval=5 s exceeds 30 s an order of
            # magnitude over) -- the loop breaks on deadline, but
            # _build_result still reports the original req.count, so
            # the Web dashboard shows a 50-probe stability test that
            # finished with 5 samples and lies about loss_rate.  The
            # 300 s upper bound matches the desktop cap.
            max_runtime_s = min(300.0,
                                 count * (timeout_s + interval_s) + 10.0)
            req = IcmpDiagRequest(
                host=host, mode=mode,
                payload_size=payload_size, df=df_flag,
                count=count, timeout_s=timeout_s, interval_s=interval_s,
                max_runtime_s=max_runtime_s)
            verr = validate_request(req)
            if verr is not None:
                return (json.dumps({"error": verr}),
                        400, {"Content-Type": "application/json"})

            task_id = uuid.uuid4().hex
            now_ts  = time.time()
            # Cross-Tab busy check: holding _web_diag_busy_lock outer
            # gives an atomic snapshot of both ICMP and DNS active flags,
            # closing the TOCTOU window where two tabs' /start could each
            # see the other's slot as still-free.  Lock order is always
            # busy-lock outer, per-kind lock inner.
            with self._web_diag_busy_lock, self._icmp_tasks_lock:
                self._icmp_gc_old_tasks(now_ts)
                if self._web_diag_busy():
                    # Either ICMP or DNS tab already busy.  Return 409
                    # immediately so the UI shows the standard
                    # "请稍后再试" message without creating a doomed
                    # task entry.
                    return (json.dumps({
                                "error": "当前有诊断任务正在进行，请稍后再试"
                            }),
                            409, {"Content-Type": "application/json"})
                # Pre-register state BEFORE launching the worker so any
                # progress callback that fires immediately can find it.
                self._icmp_tasks[task_id] = {
                    "task_id":   task_id,
                    "target_id": tid,
                    "host":      host,
                    "mode":      mode,
                    "status":    "running",
                    "progress":  {
                        "completed":  0,
                        "total":      None if mode == "pmtu" else int(count),
                        "samples":    [],
                        "pmtu_steps": [],
                    },
                    "result":     None,
                    "error":      None,
                    "created_ts": now_ts,
                    "updated_ts": now_ts,
                }
                self._icmp_active_task_id = task_id

            on_prog, on_pmtu, on_done, on_err = \
                self._icmp_make_callbacks(task_id)

            try:
                if mode == "pmtu":
                    diag_task = run_pmtu_async(
                        host=host, timeout_s=timeout_s,
                        on_progress=on_pmtu, on_done=on_done,
                        on_error=on_err,
                        data_store=self._data_store, target_id=tid)
                else:
                    diag_task = run_icmp_diag_async(
                        req,
                        on_progress=on_prog, on_done=on_done,
                        on_error=on_err,
                        data_store=self._data_store, target_id=tid)
            except Exception as e:
                # Failed to even spin up the task — release the active slot
                # and surface the error to the caller.
                with self._icmp_tasks_lock:
                    st = self._icmp_tasks.get(task_id)
                    if st is not None:
                        st["status"]     = "error"
                        st["error"]      = f"启动失败: {e}"
                        st["updated_ts"] = time.time()
                    if self._icmp_active_task_id == task_id:
                        self._icmp_active_task_id  = None
                        self._icmp_active_task_obj = None
                return (json.dumps({"error": f"启动失败: {e}"}),
                        500, {"Content-Type": "application/json"})

            with self._icmp_tasks_lock:
                # Only stash the handle if the task is still our active task —
                # on_error from _TASK_SEM contention may have fired before we
                # got here, in which case status is already 'error' and the
                # slot has been released.
                if self._icmp_active_task_id == task_id:
                    self._icmp_active_task_obj = diag_task

            return (json.dumps({"task_id": task_id}),
                    200, {"Content-Type": "application/json"})

        @app.route("/api/icmp_diag/status")
        def api_icmp_diag_status():
            task_id = (request.args.get("task_id") or "").strip()
            if not task_id:
                return (json.dumps({"error": "task_id is required"}),
                        400, {"Content-Type": "application/json"})
            with self._icmp_tasks_lock:
                st = self._icmp_tasks.get(task_id)
                if st is None:
                    return (json.dumps({"error": "task not found"}),
                            404, {"Content-Type": "application/json"})
                # Snapshot inside the lock so the diag worker can't mutate
                # samples/pmtu_steps between dict-copy and JSON serialise.
                # Lists are duplicated; small dicts inside them are shared
                # but the values themselves are immutable primitives.
                snap = {
                    "task_id":   st["task_id"],
                    "target_id": st.get("target_id"),
                    "host":      st["host"],
                    "mode":      st["mode"],
                    "status":    st["status"],
                    "progress": {
                        "completed":  st["progress"]["completed"],
                        "total":      st["progress"]["total"],
                        "samples":    list(st["progress"]["samples"]),
                        "pmtu_steps": list(st["progress"]["pmtu_steps"]),
                    },
                    "result":     st.get("result"),
                    "error":      st.get("error"),
                    "created_ts": st.get("created_ts"),
                    "updated_ts": st.get("updated_ts"),
                }
            return (json.dumps(snap, ensure_ascii=False),
                    200, {"Content-Type": "application/json"})

        @app.route("/api/icmp_diag/cancel", methods=["POST"])
        def api_icmp_diag_cancel():
            try:
                body = request.get_json(force=True, silent=True)
            except Exception:
                body = None
            # `... or {}` alone treats only the falsy-shaped JSON
            # values (null, [], "", 0, {}) as missing; non-empty
            # arrays / numbers / strings pass through unchanged, and
            # the immediate body.get(...) below then crashes with
            # AttributeError -> 500.  Explicit isinstance check
            # collapses every non-object JSON to an empty dict so the
            # field defaults below apply uniformly.  Callers that
            # depended on the silent coercion (legitimate clients
            # always send an object) see the same defaults as before.
            if not isinstance(body, dict):
                body = {}
            # _body_str: drops the row if {"task_id": 1} sneaks in
            # (legitimate clients always send a string).
            task_id = _body_str(body, "task_id")
            if not task_id:
                return (json.dumps({"error": "task_id is required"}),
                        400, {"Content-Type": "application/json"})
            with self._icmp_tasks_lock:
                st = self._icmp_tasks.get(task_id)
                if st is None:
                    return (json.dumps({"error": "task not found"}),
                            404, {"Content-Type": "application/json"})
                # Snapshot the DiagTask handle under the lock — same handle
                # is used by both Web and (in principle) any future code
                # paths that want to abort. After flipping status to
                # 'cancelled' the _on_done/_on_error callbacks will short
                # circuit and not overwrite it.
                obj = self._icmp_active_task_obj \
                    if self._icmp_active_task_id == task_id else None
                if st.get("status") == "running":
                    st["status"]     = "cancelled"
                    st["updated_ts"] = time.time()
                if self._icmp_active_task_id == task_id:
                    self._icmp_active_task_id  = None
                    self._icmp_active_task_obj = None
            if obj is not None:
                try: obj.cancel()
                except Exception: pass
            return (json.dumps({"status": "cancelled",
                                "task_id": task_id}),
                    200, {"Content-Type": "application/json"})

        @app.route("/api/icmp_diag/history")
        def api_icmp_diag_history():
            tid = (request.args.get("tid") or "").strip()
            try:
                limit = int(request.args.get("limit", 50))
            except (TypeError, ValueError):
                limit = 50
            limit = max(1, min(500, limit))
            if not self._data_store:
                return (json.dumps([]),
                        200, {"Content-Type": "application/json"})
            try:
                rows = self._data_store.get_diag_history(
                    target_id=tid or None, limit=limit)
            except Exception:
                rows = []
            return (json.dumps(rows, ensure_ascii=False),
                    200, {"Content-Type": "application/json"})

        # DNS diagnostic API (Phases 2 + 3A)
        # POST /api/dns_diag/start    - start a query, returns task_id
        # GET  /api/dns_diag/status   - poll for result (500ms cadence in UI)
        # POST /api/dns_diag/cancel   - abort a running query
        # GET  /api/dns_diag/history  - read DB-persisted history rows
        #
        # Mirror of /api/icmp_diag/* with the same shape; no incremental
        # progress (single dnspython call ≈ tens of ms; result is
        # delivered all-at-once via on_done).
        @app.route("/api/dns_diag/start", methods=["POST"])
        def api_dns_diag_start():
            import uuid
            from .dns_diag import (DnsDiagRequest, validate_dns_request,
                                    run_dns_diag_async)
            try:
                body = request.get_json(force=True, silent=True)
            except Exception:
                body = None
            # `... or {}` alone treats only the falsy-shaped JSON
            # values (null, [], "", 0, {}) as missing; non-empty
            # arrays / numbers / strings pass through unchanged, and
            # the immediate body.get(...) below then crashes with
            # AttributeError -> 500.  Explicit isinstance check
            # collapses every non-object JSON to an empty dict so the
            # field defaults below apply uniformly.  Callers that
            # depended on the silent coercion (legitimate clients
            # always send an object) see the same defaults as before.
            if not isinstance(body, dict):
                body = {}
            # _body_str rejects non-string values (e.g. {"qtype": 1})
            # which the previous `(body.get(k) or "").strip()` chain
            # crashed on with AttributeError -> 500.
            tid    = _body_str(body, "target_id") or None
            domain = _body_str(body, "domain")
            qtype  = (_body_str(body, "qtype", "A") or "A").upper()
            if not domain:
                return (json.dumps({"error": "domain is required"}),
                        400, {"Content-Type": "application/json"})

            def _num(key, default, cast=float):
                v = body.get(key, default)
                try:    return cast(v)
                except (TypeError, ValueError): return default
            timeout_s   = _num("timeout_s", 2.0)
            trace_chain = bool(body.get("trace_chain", True))
            # Phase 3B: optional custom resolver.  Body may send null,
            # an empty string, or whitespace — all coalesce to None
            # (system default).  validate_dns_request enforces the
            # IPv4/IPv6-literal rule when non-empty.
            raw_ns = body.get("nameserver")
            ns = (raw_ns or "").strip() if isinstance(raw_ns, str) else None
            if ns == "":
                ns = None

            # Phase 3C-3 D7: `nameservers` (plural) takes precedence over
            # `nameserver` (singular) — the array is the canonical 3C-3
            # path; the scalar field is retained as a 3C-1-compatible
            # fallback for older clients (mobile / scripts) that only
            # know about single-IP compare.  validate_dns_request enforces
            # the 1..3 length range, IP-literal-ness, and dedup; we just
            # do shape-coercion here.
            raw_nss = body.get("nameservers")
            nss = None
            if isinstance(raw_nss, list):
                cleaned: list = []
                for x in raw_nss:
                    if isinstance(x, str):
                        s = x.strip()
                        if s:
                            cleaned.append(s)
                if cleaned:
                    nss = cleaned

            # Phase 3C-1: optional mode field.  Defaults to "single" so
            # pre-3C-1 clients continue to work without code changes.
            # validate_dns_request enforces the compare-needs-NS rule, so
            # we don't double-check it here — just pass it through.
            raw_mode = body.get("mode")
            mode = raw_mode.strip().lower() \
                   if isinstance(raw_mode, str) else "single"
            # Phase 3C-4 / 3C-5: accept "trace_auth" and "dnssec" as the
            # third + fourth modes.  trace_auth ignores ns / nss /
            # trace_chain in the engine; dnssec uses ns (system or
            # custom IP) the same way single mode does but ignores
            # nss / trace_chain / auth_max_hops.  validate_dns_request
            # enforces the per-mode rules so we forward verbatim.
            if mode not in ("single", "compare", "trace_auth", "dnssec"):
                return (json.dumps({"error": f"unsupported mode: {mode}"}),
                        400, {"Content-Type": "application/json"})

            # Phase 3C-4: pull auth_max_hops if the client sent it
            # (optional; engine clamps to 4.._MAX_AUTH_HOPS via
            # validate_dns_request).  Default is _MAX_AUTH_HOPS.
            from .dns_diag import _MAX_AUTH_HOPS as _DEFAULT_AUTH_HOPS
            try:
                auth_max_hops = int(body.get("auth_max_hops",
                                              _DEFAULT_AUTH_HOPS))
            except (TypeError, ValueError):
                auth_max_hops = _DEFAULT_AUTH_HOPS

            req = DnsDiagRequest(
                target        = domain,
                qtype         = qtype,
                timeout_s     = timeout_s,
                trace_chain   = trace_chain,
                nameserver    = ns,
                nameservers   = nss,
                mode          = mode,
                auth_max_hops = auth_max_hops,
            )
            verr = validate_dns_request(req)
            if verr is not None:
                return (json.dumps({"error": verr}),
                        400, {"Content-Type": "application/json"})

            # Phase 3C-3: compute the canonical custom-IP list AFTER
            # validation so the task-state snapshot we save matches what
            # the engine will actually run.  For single mode this is
            # always empty (`nameserver` lives separately in `ns`).
            from .dns_diag import _normalize_compare_nameservers
            ns_list_for_task = (_normalize_compare_nameservers(req)
                                if mode == "compare" else [])
            # Human-readable summary used by /status's progress hint
            # ("正在对比解析（系统 → 8.8.8.8, +1）...") — kept short so
            # the spinner line doesn't wrap inside the modal.
            if mode == "compare" and ns_list_for_task:
                if len(ns_list_for_task) == 1:
                    ns_display = ns_list_for_task[0]
                else:
                    ns_display = (f"{ns_list_for_task[0]}, "
                                  f"+{len(ns_list_for_task) - 1}")
            else:
                ns_display = ns or ""

            task_id = uuid.uuid4().hex
            now_ts  = time.time()
            # Cross-Tab busy check — see ICMP /start comment.  Lock order
            # is always _web_diag_busy_lock outer, _dns_tasks_lock inner.
            with self._web_diag_busy_lock, self._dns_tasks_lock:
                self._dns_gc_old_tasks(now_ts)
                if self._web_diag_busy():
                    return (json.dumps({
                                "error": "当前有诊断任务正在进行，请稍后再试"
                            }),
                            409, {"Content-Type": "application/json"})
                # Pre-register task state BEFORE launching the worker so
                # any on_done that fires synchronously (e.g. validation
                # error inside the thread) can find it.
                self._dns_tasks[task_id] = {
                    "task_id":   task_id,
                    "target_id": tid,
                    "domain":    domain,
                    "qtype":     qtype,
                    # Phase 3C-1 / 3C-3: surface mode + the custom-NS
                    # summary through /status so the polling UI can pick
                    # its renderer the moment a task is created, before
                    # the first result arrives.  `nameserver` stays the
                    # display-friendly summary string (used by the
                    # progress hint); `nameservers` is the full list
                    # available to UIs that want chip-level fidelity.
                    "mode":        mode,
                    "nameserver":  ns_display,
                    "nameservers": list(ns_list_for_task),
                    "status":    "running",
                    "result":    None,
                    "error":     None,
                    "created_ts": now_ts,
                    "updated_ts": now_ts,
                }
                self._dns_active_task_id = task_id

            on_done, on_err = self._dns_make_callbacks(task_id)

            try:
                diag_task = run_dns_diag_async(
                    req, on_done=on_done, on_error=on_err,
                    data_store=self._data_store, target_id=tid)
            except Exception as e:
                with self._dns_tasks_lock:
                    st = self._dns_tasks.get(task_id)
                    if st is not None:
                        st["status"]     = "error"
                        st["error"]      = f"启动失败: {e}"
                        st["updated_ts"] = time.time()
                    if self._dns_active_task_id == task_id:
                        self._dns_active_task_id  = None
                        self._dns_active_task_obj = None
                return (json.dumps({"error": f"启动失败: {e}"}),
                        500, {"Content-Type": "application/json"})

            with self._dns_tasks_lock:
                if self._dns_active_task_id == task_id:
                    self._dns_active_task_obj = diag_task

            return (json.dumps({"task_id": task_id}),
                    200, {"Content-Type": "application/json"})

        @app.route("/api/dns_diag/status")
        def api_dns_diag_status():
            task_id = (request.args.get("task_id") or "").strip()
            if not task_id:
                return (json.dumps({"error": "task_id is required"}),
                        400, {"Content-Type": "application/json"})
            with self._dns_tasks_lock:
                st = self._dns_tasks.get(task_id)
                if st is None:
                    return (json.dumps({"error": "task not found"}),
                            404, {"Content-Type": "application/json"})
                # Snapshot inside the lock so the worker can't mutate
                # result/status between dict-read and JSON serialise.
                snap = {
                    "task_id":   st["task_id"],
                    "target_id": st.get("target_id"),
                    "domain":    st["domain"],
                    "qtype":     st["qtype"],
                    # Phase 3C-1 / 3C-3: mode + nameserver(s) carry
                    # forward to the polling UI even before the first
                    # result arrives, so the "running" placeholder can
                    # say "正在对比解析（系统 → 8.8.8.8, +1）..." for
                    # compare runs instead of the generic spinner.
                    "mode":        st.get("mode") or "single",
                    "nameserver":  st.get("nameserver") or "",
                    "nameservers": list(st.get("nameservers") or []),
                    "status":    st["status"],
                    "result":    st.get("result"),
                    "error":     st.get("error"),
                    "created_ts": st.get("created_ts"),
                    "updated_ts": st.get("updated_ts"),
                }
            return (json.dumps(snap, ensure_ascii=False),
                    200, {"Content-Type": "application/json"})

        @app.route("/api/dns_diag/cancel", methods=["POST"])
        def api_dns_diag_cancel():
            try:
                body = request.get_json(force=True, silent=True)
            except Exception:
                body = None
            # `... or {}` alone treats only the falsy-shaped JSON
            # values (null, [], "", 0, {}) as missing; non-empty
            # arrays / numbers / strings pass through unchanged, and
            # the immediate body.get(...) below then crashes with
            # AttributeError -> 500.  Explicit isinstance check
            # collapses every non-object JSON to an empty dict so the
            # field defaults below apply uniformly.  Callers that
            # depended on the silent coercion (legitimate clients
            # always send an object) see the same defaults as before.
            if not isinstance(body, dict):
                body = {}
            # _body_str: drops the row if {"task_id": 1} sneaks in
            # (legitimate clients always send a string).
            task_id = _body_str(body, "task_id")
            if not task_id:
                return (json.dumps({"error": "task_id is required"}),
                        400, {"Content-Type": "application/json"})
            with self._dns_tasks_lock:
                st = self._dns_tasks.get(task_id)
                if st is None:
                    return (json.dumps({"error": "task not found"}),
                            404, {"Content-Type": "application/json"})
                obj = self._dns_active_task_obj \
                    if self._dns_active_task_id == task_id else None
                if st.get("status") == "running":
                    st["status"]     = "cancelled"
                    st["updated_ts"] = time.time()
                if self._dns_active_task_id == task_id:
                    self._dns_active_task_id  = None
                    self._dns_active_task_obj = None
            if obj is not None:
                try: obj.cancel()
                except Exception: pass
            return (json.dumps({"status": "cancelled",
                                "task_id": task_id}),
                    200, {"Content-Type": "application/json"})

        @app.route("/api/dns_diag/history")
        def api_dns_diag_history():
            tid = (request.args.get("tid") or "").strip()
            try:
                limit = int(request.args.get("limit", 50))
            except (TypeError, ValueError):
                limit = 50
            limit = max(1, min(500, limit))
            if not self._data_store:
                return (json.dumps([]),
                        200, {"Content-Type": "application/json"})
            try:
                rows = self._data_store.get_dns_diag_history(
                    target_id=tid or None, limit=limit)
            except Exception:
                rows = []
            return (json.dumps(rows, ensure_ascii=False),
                    200, {"Content-Type": "application/json"})

        @app.route("/events")
        def sse_events():
            def generate():
                q = self._broadcaster.subscribe()
                try:
                    # All snapshots taken inside _lock so update_target /
                    # remove_target cannot mutate _targets or _stats mid-iteration
                    # (which previously risked RuntimeError: dictionary changed
                    # size during iteration, or partial-state init payloads).
                    with self._lock:
                        init_data = list(self._targets.values())
                        init_hist = list(reversed(self._history))
                        # cum_stats may contain entries for deleted nodes;
                        # filtering here prevents "ghost" charts on the stats page.
                        init_stats = {tid: dict(s)
                                      for tid, s in self._stats.items()
                                      if tid in self._targets}
                    # Include any cached port-status so browser shows
                    # indicators immediately without waiting for the next
                    # background tracert cycle.
                    with self._tracer_lock:
                        init_ports = {
                            tid: {"ports": c["tcp_checks"],
                                  "updated_at": c.get("updated_at", "")}
                            for tid, c in self._tracer_cache.items()
                            if c.get("tcp_checks")
                        }
                    yield "data: " + json.dumps({"type": "init",
                                                  "targets":     init_data,
                                                  "history":     init_hist,
                                                  "stats":       init_stats,
                                                  "port_statuses": init_ports}) + "\n\n"
                    while True:
                        try: yield "data: " + q.get(timeout=20) + "\n\n"
                        except queue.Empty: yield 'data: {"type":"heartbeat"}\n\n'
                except GeneratorExit: pass
                finally: self._broadcaster.unsubscribe(q)
            return Response(generate(), mimetype="text/event-stream",
                            headers={"Cache-Control": "no-cache",
                                     "X-Accel-Buffering": "no", "Connection": "keep-alive"})