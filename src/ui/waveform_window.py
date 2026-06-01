"""
waveform_window.py — independent large-chart detail window (Phase A2)

Design principles:
- Completely decoupled from main window layout/geometry
- Data sources:  5 min  → in-memory TargetHistory (get_history callback)
                15 min  → DataStore.get_waveform()
                 1 hour → DataStore.get_waveform()
- Refresh loop uses after() on THIS window; stops on close/minimize
- Uses same matplotlib colour palette as ChartPanel for visual consistency
"""

from __future__ import annotations

import time
import threading
import customtkinter as ctk
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from typing import Optional, Callable

from src.ui.fonts import make as F

# ── colour helpers (match ChartPanel) ──────────────────────────────────
_DARK_PALETTE  = {"line": "#60a5fa", "fill": "#2563eb", "loss": "#f87171",
                  "grid": "#2d3748", "bg":   "#1a1f2e", "text": "#9ca3af",
                  "warn": "#fbbf24"}
_LIGHT_PALETTE = {"line": "#2563eb", "fill": "#3b82f6", "loss": "#ef4444",
                  "grid": "#e5e7eb", "bg":   "#f9fafb", "text": "#6b7280",
                  "warn": "#d97706"}


def _palette() -> dict:
    return _DARK_PALETTE if ctk.get_appearance_mode() == "Dark" else _LIGHT_PALETTE


class WaveformDetailWindow(ctk.CTkToplevel):
    """Standalone waveform chart window for a single monitored target."""

    # Seconds for each time-range option
    _RANGES = {"5分钟": 300, "15分钟": 900, "1小时": 3600}
    # Refresh interval per range (seconds)
    _REFRESH = {300: 2, 900: 10, 3600: 30}
    # Minimum refresh when minimized (slow down but keep alive)
    _MINIMIZED_REFRESH = 60
    # Remembered geometry across instances (class-level, shared by all nodes)
    _saved_geometry: str = "780x480"

    def __init__(self,
                 parent,
                 target_id: str,
                 label: str,
                 ip: str,
                 get_history: Callable,        # () -> TargetHistory
                 data_store=None):
        super().__init__(parent)
        self.target_id   = target_id
        self._label      = label
        self._ip         = ip
        self._get_hist   = get_history         # in-memory 5-min window
        self._ds         = data_store          # DataStore for 15m/1h queries

        self._range_secs  = 300               # current time range in seconds
        self._refresh_job = None
        self._fig         = None
        self._canvas      = None
        self._running     = True
        self._render_seq  = 0          # stale async fetch guard (range buttons)
        self._range_btns  = {}

        # ── Alert-overlay + hover state ──────────────────────────────────
        # _pts_cache: last-drawn points list, used by hover handler to map
        #             mouse x → data point without re-querying the backend.
        # _red_ivals / _orange_ivals: snapshotted in _draw so hover can show
        #             "in outage" hint without rescanning the alert structure.
        # _tooltip:   matplotlib annotation re-created every _draw (cla() wipes
        #             it); hover handler updates xy/text/visibility.
        # _hover_cid: mpl_connect id, disconnected on _on_close.
        # _last_hover_ts: throttle gate (monotonic seconds) — caps redraw rate
        #             at ~20 Hz so 5-min range with fast cursor stays smooth.
        self._pts_cache: list      = []
        self._red_ivals: list      = []
        self._orange_ivals: list   = []
        self._tooltip              = None
        self._hover_cid            = None
        self._last_hover_ts        = 0.0

        # Last successful data dict — kept so the CTK appearance-mode hook
        # can repaint the figure (which is a raw matplotlib canvas, NOT a
        # CTk widget that recolours itself automatically) without making
        # another DB / memory fetch.
        self._last_data: dict | None = None

        self.title(f"延时趋势 — {label}")
        # Restore saved geometry, then clamp to visible screen area.
        # Prevents the window from opening off-screen when a secondary
        # monitor used during the previous session is no longer connected.
        self._restore_geometry()
        self.minsize(560, 380)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _restore_geometry(self):
        """Apply saved geometry and clamp window position to visible screen."""
        try:
            self.geometry(WaveformDetailWindow._saved_geometry)
            self.update_idletasks()   # ensure winfo_* values are current
            # Parse current position
            win_x = self.winfo_x()
            win_y = self.winfo_y()
            win_w = self.winfo_width()
            win_h = self.winfo_height()
            scr_w = self.winfo_screenwidth()
            scr_h = self.winfo_screenheight()
            # Clamp: ensure at least a 100×50 portion is visible
            margin_x, margin_y = 100, 50
            clamped_x = max(-win_w + margin_x, min(win_x, scr_w - margin_x))
            clamped_y = max(0, min(win_y, scr_h - margin_y))
            if clamped_x != win_x or clamped_y != win_y:
                self.geometry(f"{win_w}x{win_h}+{clamped_x}+{clamped_y}")
        except Exception:
            self.geometry("780x480")   # fallback to safe default

        # Bring window to front
        self.lift()
        self.focus_force()

        self._build_ui()
        self._start_refresh()

    def update_target(self, label: str, ip: str) -> None:
        """Refresh the target's display name / address while the window
        stays open.

        Called by main_window._open_waveform_detail when the operator
        edits the underlying node and then re-focuses this window.
        Without this hook, _label / _ip were locked at construction
        and the next CSV / PNG / Excel export (and the in-window
        title + header) carried the OLD identity, while the chart
        data itself silently moved on to the new IP because both
        are queried by target_id.  The cosmetic mismatch is what
        usually trips operators -- "this is the WRONG hostname for
        this server" -- so keep label / ip / title / header all in
        lockstep.  In-flight refresh loops are left running; they
        don't depend on _label / _ip.
        """
        self._label = label
        self._ip    = ip
        try:
            self.title(f"延时趋势 — {label}")
        except Exception:
            pass
        try:
            header = getattr(self, "_header_label", None)
            if header is not None:
                header.configure(text=f"📈  {label}  ({ip})")
        except Exception:
            pass

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # ── header bar ────────────────────────────────────────────────
        header = ctk.CTkFrame(self, corner_radius=0, fg_color=("gray88","gray15"))
        header.grid(row=0, column=0, sticky="ew")
        # Stash a reference to the header text label so update_target()
        # can refresh it in place when the operator edits the node's
        # name or IP while this window is open.  Without the
        # reference the original text was baked in at construction
        # and the next CSV / PNG / Excel export would still carry
        # the OLD label even though the rest of the UI moved on.
        self._header_label = ctk.CTkLabel(
            header,
            text=f"📈  {self._label}  ({self._ip})",
            font=F(14, "bold"), anchor="w")
        self._header_label.pack(side="left", padx=14, pady=8)

        # ── toolbar: range buttons + refresh ──────────────────────────
        toolbar = ctk.CTkFrame(self, fg_color="transparent")
        toolbar.grid(row=1, column=0, sticky="ew", padx=12, pady=(8, 2))

        ctk.CTkLabel(toolbar, text="时间范围：", font=F(12), text_color="gray50"
                     ).pack(side="left")

        for label, secs in self._RANGES.items():
            btn = ctk.CTkButton(
                toolbar, text=label, width=64, height=26,
                font=F(11),
                fg_color=("#5B7FA6","gray25"),
                hover_color=("#4A6E94","gray35"),
                command=lambda s=secs: self._set_range(s))
            btn.pack(side="left", padx=2)
            self._range_btns[secs] = btn

        ctk.CTkButton(
            toolbar, text="🔄 刷新", width=64, height=26,
            font=F(11),
            fg_color=("#5B7FA6","gray25"),
            hover_color=("#4A6E94","gray35"),
            command=self._render_now
        ).pack(side="left", padx=(10, 2))

        # Export button — uses a tk.Menu popup because CTk has no native
        # dropdown menu and inline 3-button cluster crowds the toolbar.
        # tk.Menu doesn't follow the CTk theme but only appears briefly
        # on click, so the visual impact is minimal.
        ctk.CTkButton(
            toolbar, text="💾 导出 ▾", width=84, height=26,
            font=F(11),
            fg_color=("#5B7FA6","gray25"),
            hover_color=("#4A6E94","gray35"),
            command=self._show_export_menu
        ).pack(side="left", padx=(6, 2))

        self._status_lbl = ctk.CTkLabel(
            toolbar, text="", font=F(11), text_color="gray50")
        self._status_lbl.pack(side="right", padx=8)

        # ── statistics summary bar ─────────────────────────────────────
        self._stats_frame = ctk.CTkFrame(self, fg_color=("gray93","gray18"),
                                          corner_radius=6)
        self._stats_frame.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 6))
        self._stat_labels = {}
        for key, text in [("avg","均值"), ("min","最小"), ("max","最大"),
                           ("loss","丢包率"), ("n","样本")]:
            frm = ctk.CTkFrame(self._stats_frame, fg_color="transparent")
            frm.pack(side="left", padx=16, pady=6)
            ctk.CTkLabel(frm, text=text, font=F(10), text_color="gray50"
                         ).pack()
            lbl = ctk.CTkLabel(frm, text="—", font=F(12, "bold"))
            lbl.pack()
            self._stat_labels[key] = lbl

        # ── chart area ─────────────────────────────────────────────────
        self._chart_frame = ctk.CTkFrame(self, fg_color="transparent")
        self._chart_frame.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0,10))
        self._chart_frame.grid_columnconfigure(0, weight=1)
        self._chart_frame.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        self._build_figure()
        self._highlight_range_btn(self._range_secs)

    def _build_figure(self):
        C = _palette()
        self._fig = Figure(figsize=(7, 3.2), dpi=96, facecolor=C["bg"])
        self._ax  = self._fig.add_subplot(111, facecolor=C["bg"])
        self._fig.subplots_adjust(left=0.07, right=0.97, top=0.92, bottom=0.12)
        self._ax.tick_params(colors=C["text"], labelsize=9)
        for sp in self._ax.spines.values():
            sp.set_color(C["grid"])
        self._ax.grid(True, color=C["grid"], linewidth=0.6, alpha=0.8)
        self._ax.set_xlabel("时间", color=C["text"], fontsize=9)
        self._ax.set_ylabel("延时 (ms)", color=C["text"], fontsize=9)

        self._canvas = FigureCanvasTkAgg(self._fig, master=self._chart_frame)
        self._canvas.get_tk_widget().grid(row=0, column=0, sticky="nsew")
        self._canvas.draw()

        # Wire hover handler. Stored cid so we can disconnect on close —
        # otherwise matplotlib keeps the callback alive on the global
        # figure manager after window destruction (slow leak).
        try:
            self._hover_cid = self._canvas.mpl_connect(
                "motion_notify_event", self._on_hover)
        except Exception:
            self._hover_cid = None

    # ------------------------------------------------------------------
    # Range selection
    # ------------------------------------------------------------------

    def _set_range(self, secs: int):
        self._range_secs = secs
        self._highlight_range_btn(secs)
        self._render_now()

    def _highlight_range_btn(self, secs: int):
        for s, btn in self._range_btns.items():
            if s == secs:
                btn.configure(fg_color=("#2563eb","#1d4ed8"))
            else:
                btn.configure(fg_color=("#5B7FA6","gray25"))

    # ------------------------------------------------------------------
    # Data fetching
    # ------------------------------------------------------------------

    def _fetch_data(self) -> dict:
        """Fetch RTT series + alert overlay for the current time range.
        Returns {"points":[...], "summary":{...}, "alerts":{red,orange,events}}.

        Alert fetch is unconditional (also for 5-min in-memory mode) because:
          (a) the alert query is tiny — only events table is scanned, typically
              0–2 rows for a 5-min window;
          (b) keeping the API uniform avoids special-casing _draw().
        """
        if self._range_secs == 300:
            d = self._fetch_from_memory()
        else:
            d = self._fetch_from_datastore()
        d["alerts"] = self._fetch_alerts()
        return d

    def _fetch_alerts(self) -> dict:
        empty = {"red": [], "orange": [], "events": []}
        if not self._ds:
            return empty
        try:
            now   = time.time()
            start = now - self._range_secs
            return self._ds.get_outage_intervals(self.target_id, start, now)
        except Exception:
            return empty

    def _fetch_from_memory(self) -> dict:
        try:
            hist = self._get_hist()
            if hist is None or hist.is_empty():
                return {"points": [], "summary": {}}
            times, latencies = hist.get_plot_data()
            now = time.time()
            points = [{"ts": now + t,
                       "rtt_ms": lat,
                       "ok": lat is not None,
                       "bucket_state": "ok" if lat is not None else "all_fail"}
                      for t, lat in zip(times, latencies)]
            return {"points": points, "summary": self._calc_summary(points)}
        except Exception:
            return {"points": [], "summary": {}}

    def _fetch_from_datastore(self) -> dict:
        if not self._ds:
            return {"points": [], "summary": {}}
        try:
            now   = int(time.time())
            start = now - self._range_secs
            return self._ds.get_waveform(self.target_id, start, now)
        except Exception:
            return {"points": [], "summary": {}}

    @staticmethod
    def _calc_summary(points: list) -> dict:
        valid = [p["rtt_ms"] for p in points if p.get("rtt_ms") is not None]
        n_ok  = sum(1 for p in points if p.get("ok"))
        return {
            "avg_ms":    round(sum(valid)/len(valid), 2) if valid else None,
            "min_ms":    round(min(valid), 2) if valid else None,
            "max_ms":    round(max(valid), 2) if valid else None,
            "loss_rate": round(1-n_ok/len(points), 4) if points else None,
            "sample_n":  len(points),
        }

    # ------------------------------------------------------------------
    # Rendering
    # ------------------------------------------------------------------

    def _render_now(self):
        """Fetch data and redraw in a background thread to keep UI responsive."""
        self._render_seq += 1
        current_seq = self._render_seq

        def _work():
            data = self._fetch_data()
            if self._running and self._render_seq == current_seq:
                try:
                    self.after(0, lambda d=data, seq=current_seq: self._draw(d, seq))
                except Exception:
                    pass
        threading.Thread(target=_work, daemon=True, name="waveform-fetch").start()

    def _draw(self, data: dict, render_seq: int | None = None):
        if not self._running:
            return
        if render_seq is not None and render_seq != self._render_seq:
            return
        try:
            self._draw_impl(data)
        except Exception:
            pass

    def _draw_impl(self, data: dict):
        # Snapshot for theme-change re-render (see _set_appearance_mode).
        self._last_data = data
        C = _palette()
        pts = data.get("points", [])
        alerts = data.get("alerts") or {"red": [], "orange": [], "events": []}

        # ── update stats bar ──────────────────────────────────────────
        s = data.get("summary", {})
        self._stat_labels["avg"].configure(
            text=f"{s['avg_ms']} ms" if s.get("avg_ms") is not None else "—")
        self._stat_labels["min"].configure(
            text=f"{s['min_ms']} ms" if s.get("min_ms") is not None else "—")
        self._stat_labels["max"].configure(
            text=f"{s['max_ms']} ms" if s.get("max_ms") is not None else "—")
        loss_pct = f"{s['loss_rate']*100:.1f}%" if s.get("loss_rate") is not None else "—"
        loss_color = "#ef4444" if (s.get("loss_rate") or 0) > 0.05 else ("gray70" if ctk.get_appearance_mode()=="Dark" else "gray40")
        self._stat_labels["loss"].configure(text=loss_pct, text_color=loss_color)
        self._stat_labels["n"].configure(
            text=str(s.get("sample_n", "—")))

        # ── redraw figure ─────────────────────────────────────────────
        self._ax.cla()
        self._tooltip = None      # cla() wipes annotation; recreate later
        self._ax.set_facecolor(C["bg"])
        self._ax.tick_params(colors=C["text"], labelsize=9)
        for sp in self._ax.spines.values():
            sp.set_color(C["grid"])
        self._ax.grid(True, color=C["grid"], linewidth=0.6, alpha=0.8)
        self._ax.set_xlabel("时间", color=C["text"], fontsize=9)
        self._ax.set_ylabel("延时 (ms)", color=C["text"], fontsize=9)

        # Cache for hover handler (always set, even on empty data so hover
        # over an empty chart short-circuits cleanly).
        self._pts_cache    = pts
        self._red_ivals    = alerts.get("red") or []
        self._orange_ivals = alerts.get("orange") or []

        if not pts:
            self._ax.text(0.5, 0.5, "暂无数据", transform=self._ax.transAxes,
                          ha="center", va="center", color=C["text"], fontsize=13)
            self._canvas.draw_idle()
            self._update_status()
            return

        # Build time axis as formatted strings, RTT values, loss markers
        import datetime
        def _fmt(ts):
            dt = datetime.datetime.fromtimestamp(ts)
            if self._range_secs <= 900:
                return dt.strftime("%H:%M:%S")
            return dt.strftime("%H:%M")

        x_vals   = list(range(len(pts)))
        x_labels = [_fmt(p["ts"]) for p in pts]
        rtt_vals     = [p["rtt_ms"] if p.get("ok") else None for p in pts]
        loss_x       = [i for i, p in enumerate(pts)
                        if p.get("bucket_state") == "all_fail"]
        partial_x    = [i for i, p in enumerate(pts)
                        if p.get("bucket_state") == "partial_fail"]
        partial_y    = [pts[i]["rtt_ms"] for i in partial_x]
        loss_y       = [0] * len(loss_x)

        # ── Alert overlay: red/orange bands drawn UNDER the RTT line ──
        # Uses linear ts→index mapping for smooth band edges that don't snap
        # to the nearest data point. zorder<line ensures the line stays on top.
        for seg in self._orange_ivals:
            x1 = self._ts_to_idx(seg.get("start"), pts)
            x2 = self._ts_to_idx(seg.get("end"),   pts)
            if x2 > x1:
                self._ax.axvspan(x1, x2, color=C["warn"], alpha=0.10, zorder=0)
        for seg in self._red_ivals:
            x1 = self._ts_to_idx(seg.get("start"), pts)
            x2 = self._ts_to_idx(seg.get("end"),   pts)
            if x2 > x1:
                self._ax.axvspan(x1, x2, color=C["loss"], alpha=0.14, zorder=0)

        # RTT line with gaps at loss points
        # Plot connected segments between valid points
        seg_x, seg_y = [], []
        for i, v in enumerate(rtt_vals):
            if v is not None:
                seg_x.append(x_vals[i]); seg_y.append(v)
            else:
                if seg_x:
                    self._ax.plot(seg_x, seg_y, color=C["line"],
                                  linewidth=1.5, solid_capstyle="round")
                    if len(seg_x) > 1:
                        self._ax.fill_between(seg_x, seg_y, 0,
                                              color=C["fill"], alpha=0.12)
                    seg_x, seg_y = [], []
        if seg_x:
            self._ax.plot(seg_x, seg_y, color=C["line"],
                          linewidth=1.5, solid_capstyle="round")
            if len(seg_x) > 1:
                self._ax.fill_between(seg_x, seg_y, 0,
                                      color=C["fill"], alpha=0.12)

        # Loss markers (all_fail: red X at y=0)
        if loss_x:
            self._ax.scatter(loss_x, loss_y, color=C["loss"],
                             marker="x", s=40, linewidths=1.5, zorder=5,
                             label="全失败")

        # Partial-fail markers (orange diamond at actual RTT height)
        if partial_x:
            self._ax.scatter(partial_x, partial_y, color=C["warn"],
                             marker="D", s=25, linewidths=0, zorder=6,
                             alpha=0.85, label="部分丢包")

        # X-axis ticks: show ~8 evenly spaced labels
        n = len(pts)
        step = max(1, n // 8)
        tick_pos    = x_vals[::step]
        tick_labels = x_labels[::step]
        self._ax.set_xticks(tick_pos)
        self._ax.set_xticklabels(tick_labels, rotation=0, fontsize=8,
                                  color=C["text"])
        self._ax.set_xlim(-0.5, len(pts) - 0.5)

        # Y-axis min at 0
        self._ax.set_ylim(bottom=0)

        # ── Alert transition markers (drawn AFTER the line so they sit on top)
        # ▼ red    = enter red (critical) — placed at TOP row
        # ▼ orange = enter orange (warn)  — placed at LOWER row to avoid
        #            visually merging with adjacent ▼red (orange→red gap is
        #            typically only 2 ping intervals ≈ 2 sec)
        # ▲ green  = recovery             — placed at TOP row
        events = (alerts.get("events") or [])
        if events:
            y_top = self._ax.get_ylim()[1]
            y_top_row    = y_top * 0.97 if y_top > 0 else 1.00
            y_lower_row  = y_top * 0.90 if y_top > 0 else 0.93
            for ev in events:
                new_s, old_s = ev.get("new"), ev.get("old")
                if new_s == "red":
                    color, marker, y_m = C["loss"],  "v", y_top_row
                elif old_s == "red" and new_s != "red":
                    color, marker, y_m = "#10b981", "^", y_top_row
                elif new_s == "orange":
                    color, marker, y_m = C["warn"],  "v", y_lower_row
                else:
                    continue
                xi = self._ts_to_idx(ev.get("ts"), pts)
                self._ax.axvline(xi, color=color, linestyle="--",
                                 linewidth=0.9, alpha=0.7, zorder=4)
                self._ax.plot([xi], [y_m], marker=marker, color=color,
                              markersize=8, markeredgewidth=0, zorder=5)

        # Recreate tooltip annotation — cla() at start of _draw wipes it.
        # Hidden until the first hover event sets visible=True and updates text.
        self._tooltip = self._ax.annotate(
            "", xy=(0, 0), xytext=(12, 12), textcoords="offset points",
            fontsize=9, color="#1f2937",
            bbox=dict(boxstyle="round,pad=0.4", fc="#fef3c7",
                      ec="#d97706", lw=0.8, alpha=0.95),
            arrowprops=dict(arrowstyle="->", color="#d97706", lw=0.8),
            zorder=10, visible=False)

        self._canvas.draw_idle()
        self._update_status()

    @staticmethod
    def _ts_to_idx(ts, pts: list) -> float:
        """Linear-interpolated mapping UNIX ts → fractional point index.
        Used by overlay band edges & event markers so they sit precisely on
        the time axis instead of snapping to the nearest sample."""
        if not pts or ts is None:
            return 0.0
        n = len(pts)
        first = pts[0]["ts"]
        last  = pts[-1]["ts"]
        if ts <= first: return -0.5            # clamp to chart left edge
        if ts >= last:  return n - 0.5          # clamp to chart right edge
        lo, hi = 0, n - 1
        while lo < hi:
            mid = (lo + hi) >> 1
            if pts[mid]["ts"] < ts:
                lo = mid + 1
            else:
                hi = mid
        i = lo
        if i == 0:
            return 0.0
        t0, t1 = pts[i - 1]["ts"], pts[i]["ts"]
        if t1 == t0:
            return float(i)
        frac = (ts - t0) / (t1 - t0)
        return (i - 1) + frac

    # ------------------------------------------------------------------
    # Theme switch
    # ------------------------------------------------------------------

    def _set_appearance_mode(self, mode_string: str) -> None:
        """CTK lifecycle hook — called by CustomTkinter on theme change
        for every widget in the tree, INCLUDING this Toplevel.

        CTk widgets (frame, label, button) recolour themselves automatically
        via super(); the matplotlib Figure does NOT — it's a raw canvas of
        pixels owned by the FigureCanvasTkAgg backend, invisible to CTk's
        palette machinery.  We therefore re-render with the new palette.

        Guards:
        - `getattr(..., None)` instead of attribute access — this hook can
          fire during super().__init__() before our own __init__ body has
          assigned the figure / data attributes.
        - Skip if window is closed (avoids draw on destroyed widgets).
        """
        try:
            super()._set_appearance_mode(mode_string)
        except Exception:
            pass
        if not getattr(self, "_running", False):
            return
        if getattr(self, "_fig", None) is None:
            return

        # Update Figure & Axes facecolor so the area BEHIND the plot also
        # follows the theme (axes.cla() doesn't reset facecolor).
        C = _palette()
        try:
            self._fig.set_facecolor(C["bg"])
            self._ax.set_facecolor(C["bg"])
        except Exception:
            pass

        last = getattr(self, "_last_data", None)
        if last is not None:
            # Re-render with new palette using cached data — no backend hit.
            self._draw(last)
        else:
            # First paint hasn't happened yet (rare race during startup)
            # — just request a redraw so any default styling reflects.
            try:
                self._canvas.draw_idle()
            except Exception:
                pass

    # ------------------------------------------------------------------
    # Hover tooltip (matplotlib motion_notify_event)
    # ------------------------------------------------------------------

    def _on_hover(self, event):
        """Update tooltip text/position on cursor motion.

        Throttled to ~20 Hz: matplotlib fires motion_notify_event tens of
        times per second on fast cursor movement; rerendering at that rate
        in a Tk-backed canvas drops UI frames noticeably.
        """
        if not self._running or self._tooltip is None:
            return
        # 20 Hz throttle
        now = time.monotonic()
        if now - self._last_hover_ts < 0.05:
            return
        self._last_hover_ts = now

        pts = self._pts_cache
        try:
            if event.inaxes is not self._ax or not pts or event.xdata is None:
                if self._tooltip.get_visible():
                    self._tooltip.set_visible(False)
                    self._canvas.draw_idle()
                return

            idx = int(round(event.xdata))
            if idx < 0 or idx >= len(pts):
                if self._tooltip.get_visible():
                    self._tooltip.set_visible(False)
                    self._canvas.draw_idle()
                return

            p = pts[idx]
            import datetime
            ts_str = datetime.datetime.fromtimestamp(p["ts"]).strftime("%H:%M:%S")
            rtt    = p.get("rtt_ms")
            state  = p.get("bucket_state", "ok")
            if state == "all_fail":
                line2 = "✕ 全部丢包"
                y_pos = 0
            elif state == "partial_fail":
                line2 = f"⚠ 部分丢包  RTT≈{rtt} ms" if rtt is not None else "⚠ 部分丢包"
                y_pos = rtt if rtt is not None else 0
            elif rtt is not None:
                line2 = f"RTT = {rtt} ms  ✓"
                y_pos = rtt
            else:
                line2 = "无数据"
                y_pos = 0

            # Add "in outage" hint if this point falls within a red interval
            ts_now = p["ts"]
            in_red = any(seg.get("start", 0) <= ts_now <= seg.get("end", 0)
                         for seg in self._red_ivals)
            in_org = any(seg.get("start", 0) <= ts_now <= seg.get("end", 0)
                         for seg in self._orange_ivals)
            extra = ""
            if in_red:
                extra = "\n⚠ 故障期间"
            elif in_org:
                extra = "\n△ 性能告警期间"

            self._tooltip.xy = (idx, y_pos)
            self._tooltip.set_text(f"{ts_str}\n{line2}{extra}")
            self._tooltip.set_visible(True)
            self._canvas.draw_idle()
        except Exception:
            # Annotation may have been recycled by a concurrent _draw —
            # silently ignore; next motion event will recover.
            pass

    def _update_status(self):
        import datetime
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        try:
            self._status_lbl.configure(text=f"更新于 {ts}")
        except Exception:
            pass

    # ------------------------------------------------------------------
    # Refresh loop
    # ------------------------------------------------------------------

    def _start_refresh(self):
        self._render_now()
        self._schedule_next()

    def _schedule_next(self):
        if not self._running:
            return
        try:
            # Slow down when minimized to avoid wasted renders
            minimized = self.state() in ("iconic", "withdrawn")
            interval  = (self._MINIMIZED_REFRESH if minimized
                         else self._REFRESH.get(self._range_secs, 10)) * 1000
            self._refresh_job = self.after(interval, self._tick)
        except Exception:
            pass

    def _tick(self):
        if not self._running:
            return
        self._render_now()
        self._schedule_next()

    # ------------------------------------------------------------------
    # Export (CSV / PNG / Excel) — Phase 3
    # ------------------------------------------------------------------

    def _show_export_menu(self):
        """Tk popup menu — three export formats.  Disabled-style if no data."""
        import tkinter as tk
        m = tk.Menu(self, tearoff=0)
        has_data = bool(self._pts_cache)
        m.add_command(label="📄  CSV (数据)",
                      command=self._export_csv,
                      state=("normal" if has_data else "disabled"))
        m.add_command(label="🖼   PNG (图片)",
                      command=self._export_png,
                      state=("normal" if (self._fig is not None) else "disabled"))
        m.add_separator()
        m.add_command(label="📊  Excel (数据+图表)",
                      command=self._export_excel,
                      state=("normal" if has_data else "disabled"))
        try:
            # Anchor popup near the button.  Using winfo_pointerxy() is more
            # reliable than locating the button widget after CTk wraps it.
            x = self.winfo_pointerx()
            y = self.winfo_pointery()
            m.tk_popup(x, y)
        finally:
            m.grab_release()

    def _suggest_filename(self, ext: str) -> str:
        """Build a safe default filename: {label}_{range}_{YYYYMMDD-HHMMSS}.ext"""
        import datetime, re
        safe_label = re.sub(r"[^\w\-.()（）\u4e00-\u9fff]", "_", self._label)[:40]
        range_str  = {300:"5min", 900:"15min", 3600:"1h"}.get(self._range_secs,
                                                               f"{self._range_secs}s")
        ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        return f"waveform_{safe_label}_{range_str}_{ts}{ext}"

    def _ask_save_path(self, ext: str, filetypes: list) -> Optional[str]:
        """Wrap filedialog.asksaveasfilename — returns None on cancel."""
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            parent           = self,
            defaultextension = ext,
            initialfile      = self._suggest_filename(ext),
            filetypes        = filetypes)
        return path or None

    def _export_csv(self):
        """Dump _pts_cache + alert overlay metadata to a CSV file.

        Schema:
          timestamp_iso, unix_ts, rtt_ms, ok, bucket_state, alert_band
        alert_band column is 'red' / 'orange' / '' depending on whether
        that sample falls in an outage interval.
        """
        path = self._ask_save_path(".csv", [("CSV", "*.csv"), ("所有", "*.*")])
        if not path: return
        import csv, datetime
        try:
            pts = self._pts_cache or []
            red = self._red_ivals or []
            org = self._orange_ivals or []

            def _band(ts):
                if any(s.get("start", 0) <= ts <= s.get("end", 0) for s in red):
                    return "red"
                if any(s.get("start", 0) <= ts <= s.get("end", 0) for s in org):
                    return "orange"
                return ""

            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                # utf-8-sig adds BOM so Excel double-click correctly auto-detects UTF-8
                w = csv.writer(f)
                w.writerow(["timestamp", "unix_ts", "rtt_ms",
                            "ok", "bucket_state", "alert_band"])
                for p in pts:
                    ts = p.get("ts")
                    w.writerow([
                        datetime.datetime.fromtimestamp(ts).isoformat(
                            sep=" ", timespec="seconds") if ts else "",
                        int(ts) if ts else "",
                        p.get("rtt_ms", "") if p.get("rtt_ms") is not None else "",
                        "1" if p.get("ok") else "0",
                        p.get("bucket_state", ""),
                        _band(ts) if ts else "",
                    ])
            self._set_status(f"已导出 CSV：{len(pts)} 行")
        except Exception as e:
            self._set_status(f"CSV 导出失败：{e}", error=True)

    def _export_png(self):
        """Save the current matplotlib figure to PNG, preserving the
        on-screen theme (dark/light) and alert overlay.  Uses 150 dpi
        for legibility when zoomed in a viewer / docs."""
        path = self._ask_save_path(".png", [("PNG", "*.png"), ("所有", "*.*")])
        if not path: return
        try:
            C = _palette()
            self._fig.savefig(path, dpi=150,
                              facecolor=C["bg"], edgecolor="none",
                              bbox_inches="tight")
            self._set_status(f"已导出 PNG → {path.split('/')[-1].split(chr(92))[-1]}")
        except Exception as e:
            self._set_status(f"PNG 导出失败：{e}", error=True)

    def _export_excel(self):
        """Write a .xlsx with two sheets:
          - '数据': per-sample rows (same schema as CSV)
          - '摘要': stats summary + embedded PNG screenshot of the figure

        Image embedding uses the matplotlib figure rendered to an
        in-memory PNG; this matches what the user is currently looking
        at (theme, alert bands, ▼▲ markers) so the Excel report is a
        faithful reproduction rather than a recreated Excel-native chart
        (which would require redoing all the alert-overlay logic in
        openpyxl's chart API — high cost, low fidelity).
        """
        path = self._ask_save_path(".xlsx", [("Excel", "*.xlsx"),
                                              ("所有", "*.*")])
        if not path: return
        try:
            import io, datetime
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.drawing.image import Image as XLImage
            # Share the same formula-injection sanitiser the Web
            # export routes use: an operator-typed label of "=1+1"
            # (the add-node dialog accepts any non-empty string) is
            # otherwise saved as an Excel formula cell (CWE-1236) and
            # evaluated on workbook open.  Mirrors what
            # /api/waveform/export, /api/export, /api/export/sla, and
            # DataStore.export_excel do for the Web-side sinks.
            from src.data_store import excel_safe_text

            pts = self._pts_cache or []
            red = self._red_ivals or []
            org = self._orange_ivals or []

            def _band(ts):
                if any(s.get("start", 0) <= ts <= s.get("end", 0) for s in red):
                    return "red"
                if any(s.get("start", 0) <= ts <= s.get("end", 0) for s in org):
                    return "orange"
                return ""

            wb = Workbook()

            # ── Sheet 1: 摘要 (also serves as the default sheet on open) ──
            ws_sum = wb.active
            ws_sum.title = "摘要"
            ws_sum.column_dimensions["A"].width = 18
            ws_sum.column_dimensions["B"].width = 28

            hdr_font = Font(bold=True, size=12)
            ws_sum["A1"] = "波形导出报告"
            ws_sum["A1"].font = Font(bold=True, size=14)
            ws_sum.merge_cells("A1:B1")

            range_lbl = {300:"近 5 分钟", 900:"近 15 分钟",
                          3600:"近 1 小时"}.get(self._range_secs,
                                                f"近 {self._range_secs} 秒")
            now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Use _last_data summary if available (already computed by _draw)
            last = getattr(self, "_last_data", None) or {}
            summary = last.get("summary", {}) or {}
            n_red    = len(red)
            n_orange = len(org)

            rows_meta = [
                ("节点",       excel_safe_text(self._label)),
                ("地址",       excel_safe_text(self._ip)),
                ("时间范围",   range_lbl),
                ("导出时间",   now_str),
                ("样本数",     summary.get("sample_n", len(pts))),
                ("均值 (ms)", summary.get("avg_ms", "")),
                ("最小 (ms)", summary.get("min_ms", "")),
                ("最大 (ms)", summary.get("max_ms", "")),
                ("丢包率",
                 f"{summary['loss_rate']*100:.2f}%"
                 if summary.get("loss_rate") is not None else ""),
                ("故障段数",   n_red),
                ("性能告警段数", n_orange),
            ]
            for i, (k, v) in enumerate(rows_meta, start=3):
                ws_sum.cell(row=i, column=1, value=k).font = hdr_font
                ws_sum.cell(row=i, column=2, value=v)

            # ── Embed the current figure as PNG ────────────────────────
            # Render to BytesIO at 150 dpi to keep file size moderate
            # (a 300 dpi 7×3.2-inch figure is ~400KB; 150dpi ≈ 120KB).
            C = _palette()
            buf = io.BytesIO()
            self._fig.savefig(buf, format="png", dpi=150,
                              facecolor=C["bg"], edgecolor="none",
                              bbox_inches="tight")
            buf.seek(0)
            img = XLImage(buf)
            # Cap embed width to ~1000 px so it doesn't blow up the sheet.
            # MUST cache orig_width BEFORE overwriting img.width -- otherwise
            # the height ratio uses (clamped_width / clamped_width) = 1.0 and
            # height never scales, producing squashed images (e.g. 2000x1000
            # becomes 1000x1000 instead of the correct 1000x500).
            orig_width = img.width
            img.width  = min(orig_width, 1000)
            img.height = int(img.height * (img.width / max(orig_width, 1)))
            ws_sum.add_image(img, "D3")

            # ── Sheet 2: per-sample data ────────────────────────────────
            ws = wb.create_sheet("数据")
            ws.append(["时间", "Unix 时间戳", "RTT (ms)",
                       "成功", "桶状态", "告警区间"])
            for c, w in enumerate([20, 14, 12, 8, 14, 12], start=1):
                ws.column_dimensions[chr(64 + c)].width = w
            head_fill = PatternFill("solid", fgColor="D6E4FF")
            for c in range(1, 7):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True)
                cell.fill = head_fill
                cell.alignment = Alignment(horizontal="center")
            for p in pts:
                ts = p.get("ts")
                ws.append([
                    datetime.datetime.fromtimestamp(ts).strftime(
                        "%Y-%m-%d %H:%M:%S") if ts else "",
                    int(ts) if ts else "",
                    p.get("rtt_ms") if p.get("rtt_ms") is not None else None,
                    1 if p.get("ok") else 0,
                    p.get("bucket_state", ""),
                    _band(ts) if ts else "",
                ])
            ws.freeze_panes = "A2"

            wb.save(path)
            self._set_status(f"已导出 Excel：{len(pts)} 行 + 图表")
        except Exception as e:
            self._set_status(f"Excel 导出失败：{e}", error=True)

    def _set_status(self, msg: str, error: bool = False):
        """Update toolbar status text — used by export methods.  Mirrors
        _update_status but accepts an explicit message and error flag for
        success/failure feedback."""
        try:
            color = "#ef4444" if error else "gray50"
            self._status_lbl.configure(text=msg, text_color=color)
        except Exception:
            pass

    # ------------------------------------------------------------------
    # Lifecycle
    # ------------------------------------------------------------------

    def _on_close(self):
        # Save geometry so next window opens at same size
        try:
            geo = self.geometry()
            if geo and "x" in geo:
                WaveformDetailWindow._saved_geometry = geo
        except Exception:
            pass
        self._running = False
        if self._refresh_job:
            try:
                self.after_cancel(self._refresh_job)
            except Exception:
                pass
        # Disconnect mpl hover callback BEFORE closing the figure —
        # otherwise the canvas may still dispatch one last motion event
        # after the Tk widget is gone, triggering a Tcl error.
        if self._hover_cid is not None:
            try:
                self._canvas.mpl_disconnect(self._hover_cid)
            except Exception:
                pass
            self._hover_cid = None
        try:
            if self._fig:
                self._fig.clf()
                plt.close(self._fig)
                self._fig = None
        except Exception:
            pass
        self.destroy()
