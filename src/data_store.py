from __future__ import annotations
"""
data_store.py — SQLite-backed persistent storage for monitoring data.

Architecture
────────────
写入：单一 write thread 独占写连接（无锁争用）。
      探测结果放入内存队列，每 60 秒批量 INSERT 一次（极低 IO 开销）。
      告警事件优先级更高，同样进队列但每次 flush 都清空。
读取：每个查询开新的只读连接（WAL 模式下读写完全并发）。

Tables
──────
pings_raw     每次探测的原始记录（默认保留 7 天）
pings_hourly  每小时聚合（avg/max/p95/丢包率，保留 90 天）
alert_events  状态变化事件（保留 365 天；ts 为浮点 Unix 秒，保留亚秒精度）
targets_meta  节点快照（name/ip/type，永久）
cum_stats     跨重启累计统计（total/success/lat_avg，永久）

Schema version via PRAGMA user_version，支持未来迁移。
"""

import math
import os
import queue
import sqlite3
import threading
import time
from collections import defaultdict
from datetime import datetime
from typing import Optional

# SQLite: 1 = probe succeeded, 0 = failed (NULL probe_success = legacy row).
_PROBE_OK_SQL = (
    "CASE WHEN probe_success IS NOT NULL THEN probe_success "
    "WHEN failure_reason IS NOT NULL THEN 0 "
    "WHEN latency_ms IS NULL THEN 0 ELSE 1 END"
)
# Successful probes with a latency sample (for avg/max/min/p95 only).
_PROBE_LAT_OK_SQL = f"({_PROBE_OK_SQL}) = 1 AND latency_ms IS NOT NULL"


_XLSX_ILLEGAL_RE = None


def excel_safe_sheet_name(value, fallback: str = "节点") -> str:
    """Sanitise a string for use as an openpyxl worksheet title.

    Excel's worksheet-name constraints are tighter than its cell
    constraints:
      * the characters \\ / * ? : [ ] are explicitly forbidden;
      * the same XML 1.0 control-char set is rejected (otherwise the
        workbook saves but openpyxl can't reopen it -- the failure
        mode the /api/export?tids=... path tripped when a label
        contained \\x00);
      * the title must be 1-31 chars long;
      * cannot be empty / whitespace-only.

    The function is intentionally separate from excel_safe_text:
    formula-injection escape (leading apostrophe) is harmful in a
    sheet title because Excel renders the apostrophe in the tab and
    treats it as a literal name character.  Pre-2025 this routine
    lived inline in DataStore.export_excel and only handled the
    forbidden-character substitution; lifting it into a helper
    centralises the XML-illegal-char strip too and keeps every
    export path aligned with the same rules.
    """
    import re as _re
    s = value if isinstance(value, str) else ""
    # Drop XML-illegal control chars (\x00-\x08, \x0b-\x0c, \x0e-\x1f)
    # before the forbidden-character pass so an injected NUL between
    # otherwise-valid letters doesn't survive into the title.
    global _XLSX_ILLEGAL_RE
    if _XLSX_ILLEGAL_RE is None:
        _XLSX_ILLEGAL_RE = _re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
    s = _XLSX_ILLEGAL_RE.sub("", s)
    # Replace the Excel-forbidden filename-like characters.
    s = _re.sub(r'[\\/*?:\[\]]', '_', s)
    # Strip leading/trailing whitespace (Excel rejects names ending
    # with whitespace).
    s = s.strip()
    if not s:
        s = fallback
    # Excel caps the title at 31 characters.  Cut hard rather than
    # smart-cutting so the user can see the exact prefix that survived.
    return s[:31]


def excel_safe_text(s):
    """Neutralise CSV/Excel formula-injection prefixes (=, +, -, @,
    plus tab/CR) AND strip XML-illegal control characters that
    openpyxl would otherwise reject with IllegalCharacterError on
    every workbook write.

    Excel evaluates a leading =/+/-/@ as a formula on open; openpyxl
    preserves the marker so a downstream user opening the file ends
    up running attacker-supplied content.  Leading-apostrophe escape
    is the standard mitigation (OWASP "Formula Injection"): Excel
    displays the value as literal text but the cell stores the
    apostrophe verbatim, so consumers reading the workbook
    programmatically just see the original string with one leading
    "'".

    Second layer: strip the XML 1.0 forbidden control characters
    (\\x00-\\x08, \\x0B, \\x0C, \\x0E-\\x1F).  openpyxl checks for
    exactly this set and raises IllegalCharacterError otherwise; an
    attacker who could route an arbitrary string into a cell (e.g.
    via the /api/export?tids=%00 path before it was filtered) could
    use that to wedge the export endpoint with a 500.  Pre-stripping
    here means every export sink stays robust even if a future code
    path forgets the upstream filter.

    Non-string inputs and the empty string pass through unchanged.

    Module-level (was previously a closure inside WebServer's
    /api/waveform/export route) so DataStore.export_excel, the SLA
    export route, and the desktop waveform window can share the
    same sanitiser without duplicating the rule set.
    """
    if not isinstance(s, str) or not s:
        return s
    global _XLSX_ILLEGAL_RE
    if _XLSX_ILLEGAL_RE is None:
        import re as _re
        _XLSX_ILLEGAL_RE = _re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")
    cleaned = _XLSX_ILLEGAL_RE.sub("", s)
    if not cleaned:
        return cleaned
    return ("'" + cleaned) if cleaned[0] in ("=", "+", "-", "@", "\t", "\r") else cleaned


def fmt_ts_ms(ts) -> str:
    """Format Unix timestamp (float seconds) for Excel with millisecond precision."""
    if ts is None:
        return "—"
    dt = datetime.fromtimestamp(float(ts))
    return dt.strftime("%Y-%m-%d %H:%M:%S.") + f"{dt.microsecond // 1000:03d}"


class DataStore:
    SCHEMA_VERSION = 5
    FLUSH_INTERVAL       = 60       # 秒，批量写入间隔
    HOURLY_AGG_INTERVAL  = 3600     # 秒，每小时聚合一次
    VACUUM_INTERVAL      = 86400 * 7  # 秒，每周渐进式 VACUUM

    def __init__(self,
                 db_path: str = "data/monitor.db",
                 raw_retention_days: int = 8,
                 hourly_retention_days: int = 90,
                 alert_retention_days: int = 365,
                 traceroute_retention_days: int = 30,
                 diag_retention_days: int = 180):

        self._db_path                = db_path
        self._raw_retention_days          = raw_retention_days
        self._hourly_retention_days        = hourly_retention_days
        self._alert_retention_days         = alert_retention_days
        self._traceroute_retention_days    = traceroute_retention_days
        # ICMP + DNS diagnostic history retention.  Diagnostic rows are
        # small (~1–2 KB each incl. details_json) and infrequent
        # (user-triggered only), so default 180 d is generous without
        # bloating the DB.  Shared between icmp_diagnostic_history and
        # dns_diagnostic_history — the user has one knob for "diagnostic
        # history" rather than separate knobs per protocol.
        self._diag_retention_days          = diag_retention_days

        os.makedirs(os.path.dirname(os.path.abspath(db_path)), exist_ok=True)

        self._queue        = queue.Queue()
        self._schema_ready = threading.Event()

        # Thread-local read-connection cache.
        # Old behaviour: every get_*() / api_* call invoked _read_conn() which
        # did sqlite3.connect(file:?mode=ro) + PRAGMA journal_mode=WAL, then
        # the caller's finally-block close()d the connection.  An /api/sla
        # request on 50 nodes loops get_sla_stats() 50 times → 50 connect/
        # close pairs + 100 PRAGMA round-trips for a single HTTP request.
        # SQLite connection setup is fast but not free; on machines with slow
        # AV scanning of new file handles it dominates response time.
        #
        # New behaviour: keep one read-only connection per thread, hand the
        # SAME instance back on subsequent calls.  Caller code that still
        # calls conn.close() must be updated to NOT close — see the docstring
        # on _read_conn() for the rule.  WAL mode is set once at open.
        self._tls = threading.local()

        self._write_thread = threading.Thread(
            target=self._write_loop, daemon=True, name="data-store-writer")
        self._write_thread.start()
        self._active_incidents: dict[str, str] = {}  # tid → current incident_id
        # Per-target identity generation.  Bumped by wipe_target_history
        # (and on_target_removed) so an in-flight ICMP / DNS diagnostic
        # task started against the OLD target can compare its captured
        # generation against the current value at persist-time and skip
        # the DataStore write -- otherwise the diag result returns AFTER
        # wipe completes and re-inserts the very rows the user just
        # asked us to delete (same asynchronous-writeback class of bug
        # the traceroute scheduler's _gen already guards against).
        self._target_generation: dict[str, int] = {}
        self._target_gen_lock = threading.Lock()
        # Per-target "commit lock" that serialises
        # bump_target_generation with the MainWindow commit sequence
        # (record_ping -> record_alert -> _last_statuses update).
        # Without it the producer-side record_ping return value can
        # only tell the caller "generation was valid at the moment
        # of queue.put"; an edit landing in the gap between that
        # return and the _last_statuses assignment leaves
        # _last_statuses polluted with a value the writer thread
        # will later drop -- and the NEXT genuine probe synthesises
        # a phantom transition against the polluted seed.  Held by
        # MainWindow._on_state_update around the whole commit;
        # acquired briefly by bump_target_generation so concurrent
        # bumps fan out per-tid without serialising unrelated nodes.
        self._commit_locks: dict[str, threading.Lock] = {}
        self._commit_locks_lock = threading.Lock()
        # Optional hook: invoked on the write thread after a wipe_target
        # job commits (so WebServer can notify browsers only after DB rows
        # are actually deleted, not when the wipe is merely enqueued).
        self._wipe_complete_callback = None
        self._outbox_lock = threading.Lock()
        self._outbox_tls = threading.local()

        threading.Thread(
            target=self._maybe_vacuum, daemon=True,
            name="data-store-vacuum").start()

        # 等待 schema 初始化完成再返回（最多 10 秒）
        self._queue.put(("_init_schema", None))
        if not self._schema_ready.wait(timeout=10):
            raise RuntimeError("DataStore: schema init timed out")

    # ── 公开写入 API ────────────────────────────────────────────────

    @staticmethod
    def row_probe_success(latency_ms, status: str,
                          failure_reason=None,
                          probe_success=None) -> bool:
        """True when this raw row represents a successful probe."""
        if probe_success is not None:
            return bool(probe_success)
        if failure_reason:
            return False
        if status == "red":
            return False
        return latency_ms is not None

    def record_ping(self, *, target_id: str, label: str, ip: str,
                    ping_type: str, ts: float, status: str,
                    latency_ms: Optional[float], loss_rate: float,
                    probe_success: bool = True,
                    status_code: Optional[int] = None,
                    failure_reason: Optional[str] = None,
                    captured_generation: Optional[int] = None) -> bool:
        """把探测结果放入写队列，立即返回，不阻塞监控线程。

        captured_generation: per-target identity generation captured
        when the probe started.  Embedded in the queued payload and
        re-checked by the writer thread (on enqueue) and by _flush
        (right before INSERT) against current_target_generation().  A
        mismatch -- caused by wipe_target_history / on_target_removed
        landing between probe start and INSERT -- drops the row so a
        stale measurement under the OLD identity can never resurrect
        under the NEW identity's history.  Same shape as the
        protection already in place for record_traceroute /
        record_diag_result / record_dns_diag_result; None disables
        the check (e.g. legacy / test paths that don't capture it).

        Returns True when the result was queued, False when the
        producer-side fast-path generation check dropped it.  Callers
        in _on_state_update consult the return value before
        advancing _last_statuses -- without that gate a stale-but-
        dropped probe still mutates the in-memory transition
        bookkeeping and the NEXT genuine probe is recorded as a
        phantom state change.
        """
        # Producer-side fast-path: skip queueing entirely if we can
        # see the generation has already moved on at the moment of
        # call.  Cheap O(1) lock-free read in the common case.
        if (captured_generation is not None and target_id is not None
            and self.current_target_generation(target_id)
                != captured_generation):
            return False
        self._queue.put(("ping", {
            "target_id": target_id, "label": label, "ip": ip,
            "ping_type": ping_type, "ts": float(ts), "status": status,
            "latency_ms": latency_ms, "loss_rate": loss_rate,
            "probe_success": 1 if probe_success else 0,
            "status_code": status_code, "failure_reason": failure_reason,
            "captured_generation": captured_generation,
        }))
        return True

    def record_alert(self, *, target_id: str, label: str, ip: str,
                     ts: float, old_status: str, new_status: str,
                     category: str, failure_reason: str = "",
                     ping_type: str = "",
                     captured_generation: Optional[int] = None):
        """记录状态变化事件（含故障原因，自动分配 incident_id）。

        captured_generation: per-target identity generation captured
        when the probe that produced this transition started.  Mirrors
        record_ping's protection -- a stale red->green transition
        produced by an OLD-config probe that survived through the
        commit-side identity gate would otherwise plant a bogus
        recovery event in alert_events.  The writer thread checks
        the generation immediately before INSERT and drops the row
        when current_target_generation has moved on.  None disables
        the check for legacy / test callers.
        """
        if (captured_generation is not None and target_id is not None
            and self.current_target_generation(target_id)
                != captured_generation):
            return
        self._queue.put(("alert", {
            "target_id": target_id, "label": label, "ip": ip,
            "ts": float(ts), "old_status": old_status,
            "new_status": new_status, "category": category,
            "failure_reason": failure_reason or "",
            "ping_type": (ping_type or "").strip().lower() or "icmp",
            "captured_generation": captured_generation,
        }))

    def on_target_removed(self, *, target_id: str, label: str, ip: str,
                          old_status: str, ping_type: str = ""):
        """Close any open incident and drop in-memory state when a node is deleted."""
        # Bump generation BEFORE queuing the removal so any in-flight
        # diag task captured the OLD generation; its persist call will
        # see a mismatch and skip the write.
        self.bump_target_generation(target_id)
        self._queue.put(("target_removed", {
            "target_id": target_id,
            "label": label,
            "ip": ip,
            "old_status": old_status,
            "ping_type": (ping_type or "").strip().lower() or "icmp",
            "ts": int(time.time()),
        }))

    def commit_lock_for(self, target_id: str) -> threading.Lock:
        """Return the per-target commit lock used to serialise
        bump_target_generation with the MainWindow commit sequence.

        Lazily creates the lock on first access so callers don't have
        to coordinate.  Held by MainWindow._on_state_update around
        the whole record_ping -> record_alert -> _last_statuses
        update sequence; acquired briefly by bump_target_generation
        so an edit bumping generation can't slip between the producer-
        side record_ping check and the in-memory bookkeeping update.
        """
        with self._commit_locks_lock:
            lock = self._commit_locks.get(target_id)
            if lock is None:
                lock = threading.Lock()
                self._commit_locks[target_id] = lock
            return lock

    def bump_target_generation(self, target_id: str) -> None:
        """Invalidate any in-flight diagnostic task for this target.

        Call from wipe_target_history / on_target_removed (and any
        future identity-changing path).  Diag run helpers capture
        the generation at start; their persist functions re-read it
        and skip the DB write on mismatch.  Symmetric with
        _TracerouteScheduler.bump_generation in web_server.

        Acquires the per-target commit lock around the increment so
        a MainWindow commit currently inside its record_ping ->
        _last_statuses sequence completes before this bump becomes
        visible -- without that ordering, the commit could update
        _last_statuses with a value the writer thread will later
        drop (gen mismatch), and the next genuine probe would
        synthesise a phantom transition against the polluted seed.
        """
        with self.commit_lock_for(target_id):
            with self._target_gen_lock:
                self._target_generation[target_id] = \
                    self._target_generation.get(target_id, 0) + 1

    def current_target_generation(self, target_id: str) -> int:
        """Read current target generation (for capture at task start)."""
        with self._target_gen_lock:
            return self._target_generation.get(target_id, 0)

    # ── 公开读取 API ────────────────────────────────────────────────

    def update_retention(self, raw: int | None = None,
                         hourly: int | None = None,
                         alert: int | None = None,
                         traceroute: int | None = None,
                         diag: int | None = None) -> None:
        """Live-update retention thresholds without restarting the process."""
        if raw        is not None: self._raw_retention_days          = raw
        if hourly     is not None: self._hourly_retention_days        = hourly
        if alert      is not None: self._alert_retention_days         = alert
        if traceroute is not None: self._traceroute_retention_days    = traceroute
        if diag       is not None: self._diag_retention_days          = diag

    def get_cum_stats(self) -> dict:
        """
        读取所有节点的累计统计。
        返回 {target_id: {total, success, latency_avg, latency_n}}
        供 WebServer 在 SSE init 时将历史统计发送给浏览器。
        """
        conn = self._read_conn()
        try:
            rows = conn.execute(
                "SELECT target_id, total, success, lat_avg, lat_n "
                "FROM cum_stats").fetchall()
            return {r[0]: {"total": r[1], "success": r[2],
                           "latency_avg": r[3] or 0.0, "latency_n": r[4]}
                    for r in rows}
        except Exception:
            return {}
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def get_open_incidents(self, target_ids: list[str]) -> dict[str, dict]:
        """Return open (unrecovered) incidents for the given target ids.

        Derived from alert_events + incident_id continuity restored at
        startup via _restore_active_incidents.  Paused targets are
        excluded — pause is operator silence, not an open outage.
        """
        if not target_ids:
            return {}
        conn = self._read_conn()
        try:
            placeholders = ",".join("?" * len(target_ids))
            rows = conn.execute(
                f"""
                SELECT ae.target_id, ae.incident_id, ae.new_status,
                       ae.label, ae.ip
                FROM alert_events ae
                INNER JOIN (
                    SELECT target_id, MAX(id) AS max_id
                    FROM alert_events
                    WHERE target_id IN ({placeholders})
                    GROUP BY target_id
                ) latest ON ae.id = latest.max_id
                WHERE ae.incident_id IS NOT NULL
                  AND ae.new_status NOT IN ('green', 'paused')
                """,
                target_ids,
            ).fetchall()
            out: dict[str, dict] = {}
            for tid, iid, ns, label, ip in rows:
                start_row = conn.execute(
                    "SELECT MIN(ts) FROM alert_events "
                    "WHERE incident_id=? AND new_status IN ('red', 'orange')",
                    (iid,),
                ).fetchone()
                started_at = start_row[0] if start_row and start_row[0] is not None else None
                if started_at is None:
                    fb = conn.execute(
                        "SELECT MIN(ts) FROM alert_events WHERE incident_id=?",
                        (iid,),
                    ).fetchone()
                    started_at = fb[0] if fb and fb[0] is not None else time.time()
                probe_row = conn.execute(
                    "SELECT ping_type, failure_reason FROM alert_events "
                    "WHERE incident_id=? AND new_status='red' "
                    "ORDER BY id DESC LIMIT 1",
                    (iid,),
                ).fetchone()
                if probe_row is None:
                    probe_row = conn.execute(
                        "SELECT ping_type, failure_reason FROM alert_events "
                        "WHERE incident_id=? ORDER BY id DESC LIMIT 1",
                        (iid,),
                    ).fetchone()
                ping_type = self._resolve_alert_ping_type(
                    conn, tid, probe_row[0] if probe_row else None)
                failure_reason = (probe_row[1] or "") if probe_row else ""
                out[tid] = {
                    "incident_id": iid,
                    "started_at": float(started_at),
                    "last_status": ns,
                    "label": label or "",
                    "ip": ip or "",
                    "ping_type": ping_type,
                    "failure_reason": failure_reason,
                }
            return out
        except Exception:
            return {}

    def record_webhook_ack(self, incident_id: str, target_id: str,
                           ts: float | None = None) -> None:
        """Persist operator ACK for one open incident (survives restart)."""
        iid = (incident_id or "").strip()
        if not iid or not target_id:
            return
        self._queue.put(("webhook_ack", {
            "action": "set",
            "incident_id": iid,
            "target_id": target_id,
            "ts": float(ts or time.time()),
        }))

    def clear_webhook_ack(self, incident_id: str) -> None:
        """Drop persisted ACK when the incident closes."""
        iid = (incident_id or "").strip()
        if not iid:
            return
        self._queue.put(("webhook_ack", {
            "action": "clear",
            "incident_id": iid,
        }))

    def is_webhook_acknowledged(self, incident_id: str) -> bool:
        """Whether operators ACK'd this incident_id (webhook reminder silence)."""
        iid = (incident_id or "").strip()
        if not iid:
            return False
        conn = self._read_conn()
        try:
            row = conn.execute(
                "SELECT 1 FROM webhook_ack WHERE incident_id=? LIMIT 1",
                (iid,)).fetchone()
            return row is not None
        except Exception:
            return False

    # ── Webhook outbox (reliable delivery) ───────────────────────────

    def _outbox_write_conn(self) -> sqlite3.Connection:
        """Per-thread outbox write connection (dispatcher runs off main thread)."""
        conn = getattr(self._outbox_tls, "conn", None)
        if conn is not None:
            return conn
        conn = sqlite3.connect(
            self._db_path, timeout=30, check_same_thread=False)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        self._outbox_tls.conn = conn
        return conn

    def enqueue_webhook_outbox(
            self, *, delivery_id: str, target_id: str, incident_id: str,
            incident_seq: int | None, event: str, order_key: str,
            payload: dict, event_ts: float, max_attempts: int = 0) -> str:
        """Insert a pending webhook delivery row (synchronous, durable)."""
        import json as _json
        now = time.time()
        with self._outbox_lock:
            conn = self._outbox_write_conn()
            if event in ("alert_reminder", "diagnostic_update"):
                conn.execute(
                    "UPDATE webhook_outbox SET delivery_state='dropped_stale', "
                    "last_error='coalesced', updated_at=? "
                    "WHERE delivery_state='pending' AND event=? "
                    "AND order_key=? AND incident_id=? AND delivery_id!=?",
                    (now, event, order_key, incident_id or "", delivery_id))
            conn.execute(
                "INSERT INTO webhook_outbox "
                "(delivery_id, target_id, incident_id, incident_seq, event, "
                " order_key, payload_json, event_ts, first_queued_ts, "
                " next_attempt_ts, attempt_count, max_attempts, delivery_state, "
                " last_error, created_at, updated_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,0,?, 'pending', '', ?, ?)",
                (
                    delivery_id,
                    target_id or "",
                    incident_id or "",
                    incident_seq,
                    event,
                    order_key,
                    _json.dumps(payload, ensure_ascii=False),
                    float(event_ts),
                    now,
                    now,
                    int(max_attempts or 0),
                    now,
                    now,
                ))
            conn.commit()
        return delivery_id

    def fetch_deliverable_webhook_outbox(self, now: float,
                                         limit: int = 50) -> list[dict]:
        """Head-of-line due messages: one per order_key."""
        conn = self._read_conn()
        try:
            rows = conn.execute(
                """
                WITH head AS (
                    SELECT order_key, MIN(id) AS min_id
                    FROM webhook_outbox
                    WHERE delivery_state IN ('pending', 'sending')
                    GROUP BY order_key
                )
                SELECT o.*
                FROM webhook_outbox o
                JOIN head h ON o.order_key = h.order_key AND o.id = h.min_id
                WHERE o.delivery_state = 'pending'
                  AND o.next_attempt_ts <= ?
                ORDER BY o.first_queued_ts, o.id
                LIMIT ?
                """,
                (float(now), int(limit)),
            ).fetchall()
            return [self._outbox_row_to_dict(r) for r in rows]
        except Exception:
            return []

    def recover_orphaned_sending_webhook_outbox(
            self, now: float | None = None) -> int:
        """After process restart, in-flight sending rows are orphaned."""
        now = float(now or time.time())
        with self._outbox_lock:
            conn = self._outbox_write_conn()
            cur = conn.execute(
                "UPDATE webhook_outbox SET delivery_state='pending', "
                "last_error='recovered_after_restart', next_attempt_ts=?, "
                "updated_at=? WHERE delivery_state='sending'",
                (now, now))
            conn.commit()
            return cur.rowcount

    def recover_stale_sending_webhook_outbox(
            self, now: float, lease_sec: float) -> int:
        """Reclaim sending rows whose delivery lease has expired."""
        now = float(now)
        cutoff = now - float(lease_sec)
        with self._outbox_lock:
            conn = self._outbox_write_conn()
            cur = conn.execute(
                "UPDATE webhook_outbox SET delivery_state='pending', "
                "last_error='stale_sending_recovered', next_attempt_ts=?, "
                "updated_at=? "
                "WHERE delivery_state='sending' AND last_attempt_ts <= ?",
                (now, now, cutoff))
            conn.commit()
            return cur.rowcount

    def drop_red_blocked_closed_summary(self, now: float,
                                        delay_sec: float) -> int:
        """Drop head alert_red rows blocking due recovery past closed-summary delay.

        When alert_red is in long backoff (next_attempt_ts in the future) it
        still occupies head-of-line, preventing recovery from being fetched.
        If the red has been pending longer than delay_sec and a matching
        recovery row is due, drop the red so closed-summary can proceed.
        """
        now = float(now)
        delay_sec = float(delay_sec)
        with self._outbox_lock:
            conn = self._outbox_write_conn()
            blockers = conn.execute(
                """
                WITH head AS (
                    SELECT order_key, MIN(id) AS min_id
                    FROM webhook_outbox
                    WHERE delivery_state IN ('pending', 'sending')
                    GROUP BY order_key
                )
                SELECT o.delivery_id, o.order_key, o.incident_id
                FROM webhook_outbox o
                JOIN head h ON o.order_key = h.order_key AND o.id = h.min_id
                WHERE o.event = 'alert_red'
                  AND o.delivery_state = 'pending'
                  AND o.next_attempt_ts > ?
                  AND (? - o.first_queued_ts) >= ?
                """,
                (now, now, delay_sec),
            ).fetchall()
            dropped = 0
            for delivery_id, order_key, incident_id in blockers:
                due_recovery = conn.execute(
                    "SELECT 1 FROM webhook_outbox "
                    "WHERE event='recovery' AND order_key=? "
                    "AND incident_id=? AND delivery_state='pending' "
                    "LIMIT 1",
                    (order_key, incident_id or ""),
                ).fetchone()
                if not due_recovery:
                    continue
                cur = conn.execute(
                    "UPDATE webhook_outbox SET delivery_state='dropped_stale', "
                    "last_error='superseded_by_closed_summary', updated_at=? "
                    "WHERE delivery_id=? AND delivery_state='pending'",
                    (now, delivery_id),
                )
                if cur.rowcount:
                    conn.execute(
                        "UPDATE webhook_outbox SET next_attempt_ts=?, "
                        "updated_at=? "
                        "WHERE event='recovery' AND order_key=? "
                        "AND incident_id=? AND delivery_state='pending' "
                        "AND next_attempt_ts > ?",
                        (now, now, order_key, incident_id or "", now),
                    )
                dropped += cur.rowcount
            conn.commit()
            return dropped

    @staticmethod
    def _outbox_row_to_dict(row) -> dict:
        cols = (
            "id", "delivery_id", "target_id", "incident_id", "incident_seq",
            "event", "order_key", "payload_json", "event_ts",
            "first_queued_ts", "next_attempt_ts", "last_attempt_ts",
            "delivered_ts", "attempt_count", "max_attempts", "delivery_state",
            "last_error", "created_at", "updated_at",
        )
        return dict(zip(cols, row))

    def claim_webhook_outbox(self, delivery_id: str, now: float) -> bool:
        with self._outbox_lock:
            conn = self._outbox_write_conn()
            cur = conn.execute(
                "UPDATE webhook_outbox SET delivery_state='sending', "
                "last_attempt_ts=?, updated_at=? "
                "WHERE delivery_id=? AND delivery_state='pending'",
                (float(now), float(now), delivery_id))
            conn.commit()
            return cur.rowcount == 1

    def finish_webhook_outbox(
            self, delivery_id: str, *, state: str, error: str = "",
            now: float | None = None, attempt_count: int | None = None,
            next_attempt_ts: float | None = None,
            only_if_sending: bool = False) -> bool:
        """Update outbox row; return True if a row was updated."""
        now = float(now or time.time())
        with self._outbox_lock:
            conn = self._outbox_write_conn()
            row = conn.execute(
                "SELECT attempt_count FROM webhook_outbox WHERE delivery_id=?",
                (delivery_id,)).fetchone()
            if row is None:
                return False
            attempts = int(attempt_count if attempt_count is not None
                           else row[0])
            sending_clause = " AND delivery_state='sending'" if only_if_sending else ""
            if state == "delivered":
                cur = conn.execute(
                    "UPDATE webhook_outbox SET delivery_state='delivered', "
                    "delivered_ts=?, last_attempt_ts=?, attempt_count=?, "
                    "last_error='', updated_at=? WHERE delivery_id=?"
                    + sending_clause,
                    (now, now, attempts, now, delivery_id))
            elif state == "pending":
                cur = conn.execute(
                    "UPDATE webhook_outbox SET delivery_state='pending', "
                    "last_attempt_ts=?, attempt_count=?, last_error=?, "
                    "next_attempt_ts=?, updated_at=? WHERE delivery_id=?"
                    + sending_clause,
                    (now, attempts, error or "", float(next_attempt_ts or now),
                     now, delivery_id))
            else:
                cur = conn.execute(
                    "UPDATE webhook_outbox SET delivery_state=?, "
                    "last_attempt_ts=?, attempt_count=?, last_error=?, "
                    "updated_at=? WHERE delivery_id=?"
                    + sending_clause,
                    (state, now, attempts, error or "", now, delivery_id))
            conn.commit()
            return cur.rowcount == 1

    def get_webhook_outbox_delivery_state(self, delivery_id: str) -> str | None:
        if not delivery_id:
            return None
        conn = self._read_conn()
        try:
            row = conn.execute(
                "SELECT delivery_state FROM webhook_outbox WHERE delivery_id=?",
                (delivery_id,)).fetchone()
            return row[0] if row else None
        except Exception:
            return None

    def list_sending_webhook_delivery_ids(
            self, target_id: str,
            events: tuple[str, ...] | None = None) -> list[str]:
        if not target_id:
            return []
        conn = self._read_conn()
        try:
            if events:
                placeholders = ",".join("?" * len(events))
                rows = conn.execute(
                    f"SELECT delivery_id FROM webhook_outbox "
                    f"WHERE target_id=? AND delivery_state='sending' "
                    f"AND event IN ({placeholders})",
                    (target_id, *events)).fetchall()
            else:
                rows = conn.execute(
                    "SELECT delivery_id FROM webhook_outbox "
                    "WHERE target_id=? AND delivery_state='sending'",
                    (target_id,)).fetchall()
            return [r[0] for r in rows if r and r[0]]
        except Exception:
            return []

    def find_undelivered_alert_red(
            self, order_key: str, incident_id: str,
            exclude_id: str = "") -> dict | None:
        conn = self._read_conn()
        try:
            row = conn.execute(
                "SELECT * FROM webhook_outbox "
                "WHERE event='alert_red' AND order_key=? "
                "AND incident_id=? "
                "AND delivery_id!=? AND delivered_ts IS NULL "
                "AND delivery_state != 'delivered' "
                "ORDER BY id LIMIT 1",
                (order_key, incident_id or "", exclude_id or ""),
            ).fetchone()
            return self._outbox_row_to_dict(row) if row else None
        except Exception:
            return None

    def drop_pending_webhook_for_target(self, target_id: str, reason: str) -> None:
        if not target_id:
            return
        now = time.time()
        with self._outbox_lock:
            conn = self._outbox_write_conn()
            conn.execute(
                "UPDATE webhook_outbox SET delivery_state='dropped_stale', "
                "last_error=?, updated_at=? "
                "WHERE target_id=? AND delivery_state IN ('pending', 'sending')",
                (reason or "target_removed", now, target_id))
            conn.commit()

    def get_webhook_deliveries(
            self, *, incident_id: str | None = None,
            target_id: str | None = None, limit: int = 100) -> list[dict]:
        import json as _json
        conn = self._read_conn()
        clauses = ["1=1"]
        params: list = []
        if incident_id:
            clauses.append("incident_id=?")
            params.append(incident_id)
        if target_id:
            clauses.append("target_id=?")
            params.append(target_id)
        params.append(int(limit))
        try:
            rows = conn.execute(
                f"SELECT * FROM webhook_outbox WHERE {' AND '.join(clauses)} "
                f"ORDER BY id DESC LIMIT ?",
                params).fetchall()
            out = []
            for row in rows:
                d = self._outbox_row_to_dict(row)
                try:
                    d["payload"] = _json.loads(d.pop("payload_json") or "{}")
                except Exception:
                    d["payload"] = {}
                out.append(d)
            return out
        except Exception:
            return []

    def get_webhook_delivery_stats(self) -> dict:
        conn = self._read_conn()
        try:
            rows = conn.execute(
                "SELECT delivery_state, COUNT(*) FROM webhook_outbox "
                "GROUP BY delivery_state").fetchall()
            return {st: cnt for st, cnt in rows}
        except Exception:
            return {}

    def get_last_webhook_failures(self, limit: int = 5) -> list[dict]:
        conn = self._read_conn()
        try:
            rows = conn.execute(
                "SELECT delivery_id, target_id, event, delivery_state, "
                "last_error, updated_at FROM webhook_outbox "
                "WHERE delivery_state IN ('failed_permanent', 'pending') "
                "AND last_error!='' "
                "ORDER BY updated_at DESC LIMIT ?",
                (int(limit),)).fetchall()
            return [
                {"delivery_id": r[0], "target_id": r[1], "event": r[2],
                 "state": r[3], "error": r[4], "updated_at": r[5]}
                for r in rows
            ]
        except Exception:
            return []

    def drop_pending_webhook_events(
            self, target_id: str, events: tuple[str, ...],
            reason: str) -> None:
        if not target_id or not events:
            return
        now = time.time()
        placeholders = ",".join("?" * len(events))
        with self._outbox_lock:
            conn = self._outbox_write_conn()
            conn.execute(
                f"UPDATE webhook_outbox SET delivery_state='dropped_stale', "
                f"last_error=?, updated_at=? "
                f"WHERE target_id=? AND delivery_state IN ('pending','sending') "
                f"AND event IN ({placeholders})",
                (reason or "dropped", now, target_id, *events))
            conn.commit()

    def find_pending_recovery(
            self, order_key: str, incident_id: str = "") -> dict | None:
        conn = self._read_conn()
        try:
            row = conn.execute(
                "SELECT * FROM webhook_outbox "
                "WHERE event='recovery' AND order_key=? "
                "AND incident_id=? "
                "AND delivery_state='pending' "
                "ORDER BY id LIMIT 1",
                (order_key, incident_id or ""),
            ).fetchone()
            return self._outbox_row_to_dict(row) if row else None
        except Exception:
            return None

    def get_last_known_statuses(self) -> dict:
        """
        Return {target_id: last_new_status} — the latest alert transition per
        target, including ``paused`` (and any other stored new_status).

        Used to seed ``_last_statuses`` on startup so the first post-restart
        probe compares against the real last persisted state:

          - crashed while RED -> first green writes red->green (closes outage)
          - paused while RED -> first red writes paused->red (opens new outage)

        The previous filter that ignored ``paused`` caused a target paused in
        RED to seed as ``red``; a restart where the node was still down then
        skipped recording because old_st == new_st == ``red``.
        """
        conn = self._read_conn()
        try:
            rows = conn.execute(
                "SELECT target_id, new_status FROM alert_events "
                "WHERE id IN ("
                "  SELECT MAX(id) FROM alert_events GROUP BY target_id"
                ")"
            ).fetchall()
            return {tid: ns for tid, ns in rows}
        except Exception:
            return {}
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def get_last_persisted_status(self, target_id: str) -> Optional[str]:
        """Return the most recently persisted ping status for one target,
        or None if no ping rows exist.

        Used by MainWindow._on_edit to reseed _last_statuses across a
        config edit, so the FIRST probe under the new config can
        write a proper transition event (e.g. red->green to close a
        prior outage, or green->red to open a new one).

        Why pings_raw rather than alert_events:
          alert_events only records TRANSITIONS, not steady state.
          A target that has been green since startup with no transition
          would return None from get_last_known_statuses, and the
          edit's reseed would have nothing to anchor against -- the
          first post-edit transition would still be lost.
          pings_raw, in contrast, is the source of truth for "what
          was the status of the very last probe accepted by the
          DataStore writer thread".  That value is the only baseline
          that lets the next probe synthesise an SLA-correct
          transition.
        """
        conn = self._read_conn()
        try:
            row = conn.execute(
                "SELECT status FROM pings_raw WHERE target_id = ? "
                "ORDER BY ts DESC, id DESC LIMIT 1",
                (target_id,)).fetchone()
            return row[0] if row is not None else None
        except Exception:
            return None
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def get_targets_meta(self) -> dict:
        """返回 {target_id: {label, ip, ping_type}}"""
        conn = self._read_conn()
        try:
            rows = conn.execute(
                "SELECT target_id, label, ip, ping_type FROM targets_meta"
            ).fetchall()
            return {r[0]: {"label": r[1], "ip": r[2], "ping_type": r[3]}
                    for r in rows}
        except Exception:
            return {}
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def should_use_hourly(self, start_ts: int, end_ts: int,
                          raw_span_limit: int) -> bool:
        """
        True when get_history() must use pings_hourly for [start_ts, end_ts].

        Any window starting before pings_raw retention, or longer than
        raw_span_limit seconds, cannot rely on pings_raw alone.
        """
        start_ts = int(start_ts)
        end_ts   = int(end_ts)
        if end_ts < start_ts:
            return False
        raw_cutoff = int(time.time()) - self._raw_retention_days * 86400
        return (start_ts < raw_cutoff
                or (end_ts - start_ts) > int(raw_span_limit))

    def _waveform_points_hourly(self, conn, target_id: str,
                                start_ts: int, end_ts: int) -> list:
        """Build waveform points from _bounded_hourly_buckets."""
        points = []
        for b in self._bounded_hourly_buckets(conn, target_id, start_ts, end_ts):
            lat_min = b.get("lat_min")
            points.append({
                "ts": b["hour_ts"],
                "rtt_ms": round(b["lat_avg"], 2) if b["lat_avg"] is not None else None,
                "max_ms": round(b["lat_max"], 2) if b["lat_max"] is not None else None,
                "min_ms": round(lat_min, 2) if lat_min is not None else None,
                "sample_n":  b["sample_n"],
                "success_n": b["success_n"],
                "ok": b["lat_avg"] is not None,
                "bucket_state": (
                    "all_fail" if b["lat_avg"] is None
                    else "partial_fail" if b["success_n"] < b["sample_n"]
                    else "ok"),
            })
        return points

    def get_waveform(self, target_id: str, start_ts: int, end_ts: int) -> dict:
        """
        Return waveform data auto-aggregated by span and raw retention.

        If any part of [start_ts, end_ts] is older than pings_raw retention,
        use hourly buckets (pings_hourly) for the whole window so the chart
        does not silently omit the pre-cutoff portion.  Otherwise:
          span ≤ 2h   → raw second-level data
          span ≤ 48h  → minute-level aggregation from pings_raw
          span > 48h  → hourly data from pings_hourly

        Returns:
          { "resolution": "raw"|"1m"|"1h",
            "points": [{"ts": int, "rtt_ms": float|null, "ok": bool}, ...],
            "summary": {"avg_ms","min_ms","max_ms","loss_rate","sample_n"} }
        """
        span = end_ts - start_ts
        use_hourly = self.should_use_hourly(start_ts, end_ts, 172800)
        conn = self._read_conn()
        try:
            if use_hourly:
                points = self._waveform_points_hourly(
                    conn, target_id, start_ts, end_ts)
                resolution = "1h"

            elif span <= 7200:                        # ≤ 2 hours → raw
                rows = conn.execute(
                    "SELECT ts, latency_ms, status, failure_reason, "
                    "probe_success FROM pings_raw "
                    "WHERE target_id=? AND ts BETWEEN ? AND ? "
                    "ORDER BY ts ASC, id ASC",
                    (target_id, start_ts, end_ts)).fetchall()
                # max_ms/min_ms == rtt_ms here: each raw row IS a single sample,
                # so the per-point min and max collapse to the sample's latency.
                points = []
                for r in rows:
                    ok = self.row_probe_success(r[1], r[2], r[3], r[4])
                    lat = r[1] if ok else None
                    points.append({
                        "ts": r[0],
                        "rtt_ms": lat,
                        "max_ms": lat,
                        "min_ms": lat,
                        "ok": ok,
                        "sample_n": 1,
                        "success_n": 1 if ok else 0,
                        "bucket_state": "ok" if ok else "all_fail",
                    })
                resolution = "raw"

            elif span <= 172800:                    # ≤ 48 hours → 1-minute buckets
                # MAX/MIN naturally skip NULL latencies in SQLite, so they
                # only reflect successful probes within the bucket — the same
                # population AVG uses.  Without these the summary's max_ms
                # would collapse to max-of-averages and hide real spikes
                # (e.g. a 3000ms 30-second burst inside an otherwise-quiet
                # minute averages to ~25ms).
                rows = conn.execute(
                    f"SELECT CAST(ts / 60 AS INTEGER) * 60 AS minute, "
                    f"AVG(CASE WHEN {_PROBE_LAT_OK_SQL} THEN latency_ms END), "
                    f"COUNT(*), "
                    f"SUM(CASE WHEN {_PROBE_OK_SQL} = 1 THEN 1 ELSE 0 END), "
                    f"MAX(CASE WHEN {_PROBE_LAT_OK_SQL} THEN latency_ms END), "
                    f"MIN(CASE WHEN {_PROBE_LAT_OK_SQL} THEN latency_ms END) "
                    f"FROM pings_raw "
                    f"WHERE target_id=? AND ts BETWEEN ? AND ? "
                    f"GROUP BY minute ORDER BY minute",
                    (target_id, start_ts, end_ts)).fetchall()
                points = [{"ts": r[0],
                           "rtt_ms": round(r[1], 2) if r[1] is not None else None,
                           "max_ms": round(r[4], 2) if r[4] is not None else None,
                           "min_ms": round(r[5], 2) if r[5] is not None else None,
                           "sample_n":  r[2],
                           "success_n": r[3],
                           # ok = bucket has valid RTT data (majority not failed).
                           # Consistent with raw resolution where ok=True for orange/green.
                           "ok": r[1] is not None and r[3] > 0,
                           # V4: partial_fail when some probes in bucket failed
                           "bucket_state": (
                               "all_fail"     if r[1] is None
                               else "partial_fail" if r[3] < r[2]   # success_n < total_n
                               else "ok")}
                          for r in rows]
                resolution = "1m"

            # Summary - probe-weighted across buckets (matches export_excel / SLA).
            total_n   = sum(p.get("sample_n") or 0 for p in points)
            success_n = sum(p.get("success_n") or 0 for p in points)
            # Weight by success_n (samples that actually contributed a latency),
            # NOT sample_n (total probes including failures).  rtt_ms here is
            # the mean over latencies of successful probes only -- a bucket
            # with 2 successes out of 3600 probes has rtt_ms based on 2
            # samples; weighting it by 3600 would mask severe outages by
            # giving the lucky 2 successes the same statistical weight as a
            # fully-healthy 3600-success bucket.  Fall back to None (rather
            # than a misleading number) when no successful samples exist.
            w_pairs   = [(p["rtt_ms"], p["success_n"]) for p in points
                         if p.get("rtt_ms") is not None and p.get("success_n")]
            # min/max use per-bucket real extremes (raw=sample value;
            # 1m/1h=SQL MAX/MIN over successful probes), NOT the bucket
            # average — that would collapse to "max of averages" and hide
            # short bursts inside an otherwise-quiet bucket.
            valid_max = [p["max_ms"] for p in points if p.get("max_ms") is not None]
            valid_min = [p["min_ms"] for p in points if p.get("min_ms") is not None]
            w_n       = sum(n for _, n in w_pairs)
            w_sum     = sum(rtt * n for rtt, n in w_pairs)
            summary = {
                "avg_ms":    round(w_sum / w_n, 2) if w_n else None,
                "min_ms":    round(min(valid_min), 2) if valid_min else None,
                "max_ms":    round(max(valid_max), 2) if valid_max else None,
                "loss_rate": round(1 - success_n / total_n, 4) if total_n else None,
                "sample_n":  total_n,
                "resolution": resolution,
            }
            return {"resolution": resolution, "points": points, "summary": summary}
        except Exception:
            return {"resolution": "raw", "points": [], "summary": {}}
        # _read_conn() returns a thread-local cached connection — do NOT close.

    @staticmethod
    def _current_hour_start(now: Optional[int] = None) -> int:
        t = int(now if now is not None else time.time())
        return (t // 3600) * 3600

    @staticmethod
    def _compute_hourly_bucket(lat_status_rows) -> Optional[dict]:
        """Aggregate pings_raw rows into one hourly bucket.

        Each row is (latency_ms, status) or
        (latency_ms, status, failure_reason, probe_success).
        """
        if not lat_status_rows:
            return None
        sample_n  = len(lat_status_rows)
        success_n = 0
        for row in lat_status_rows:
            if len(row) >= 4:
                lat, st, fr, ps = row[0], row[1], row[2], row[3]
            elif len(row) == 3:
                lat, st, fr, ps = row[0], row[1], row[2], None
            else:
                lat, st, fr, ps = row[0], row[1], None, None
            if DataStore.row_probe_success(lat, st, fr, ps):
                success_n += 1
        lats = []
        for row in lat_status_rows:
            if len(row) >= 4:
                lat, st, fr, ps = row[0], row[1], row[2], row[3]
            elif len(row) == 3:
                lat, st, fr, ps = row[0], row[1], row[2], None
            else:
                lat, st, fr, ps = row[0], row[1], None, None
            if DataStore.row_probe_success(lat, st, fr, ps) and lat is not None:
                lats.append(lat)
        lats.sort()
        lat_avg   = sum(lats) / len(lats) if lats else None
        lat_max   = lats[-1] if lats else None
        lat_min   = lats[0]  if lats else None
        lat_p95   = (lats[min(len(lats) - 1,
                              max(0, math.ceil(len(lats) * 0.95) - 1))]
                     if lats else None)
        loss_rate = 1 - success_n / sample_n
        return {
            "sample_n":  sample_n,
            "success_n": success_n,
            "lat_avg":   lat_avg,
            "lat_max":   lat_max,
            "lat_min":   lat_min,
            "lat_p95":   lat_p95,
            "loss_rate": loss_rate,
        }

    def _raw_window_bucket(self, conn, target_id: str, hour_ts: int,
                           win_start: float, win_end: float) -> Optional[dict]:
        """Aggregate pings_raw for [win_start, win_end] into one hourly-shaped bucket.

        Uses SQL COUNT/AVG/MAX so 0.05s sampling in a one-hour live window
        does not materialize millions of rows in Python.
        """
        if win_start > win_end:
            return None
        row = conn.execute(
            f"SELECT COUNT(*), "
            f"SUM(CASE WHEN {_PROBE_OK_SQL} = 1 THEN 1 ELSE 0 END), "
            f"AVG(CASE WHEN {_PROBE_LAT_OK_SQL} THEN latency_ms END), "
            f"MAX(CASE WHEN {_PROBE_LAT_OK_SQL} THEN latency_ms END), "
            f"MIN(CASE WHEN {_PROBE_LAT_OK_SQL} THEN latency_ms END) "
            f"FROM pings_raw "
            f"WHERE target_id=? AND ts >= ? AND ts <= ?",
            (target_id, win_start, win_end)).fetchone()
        if not row or not row[0]:
            return None
        sample_n, success_n, lat_avg, lat_max, lat_min = row
        lat_p95 = self._global_lat_p95_sql(
            conn, target_id, win_start, win_end)
        loss_rate = 1 - success_n / sample_n if sample_n else 0.0
        return {
            "hour_ts":   hour_ts,
            "sample_n":  sample_n,
            "success_n": success_n,
            "lat_avg":   lat_avg,
            "lat_max":   lat_max,
            "lat_min":   lat_min,
            "lat_p95":   lat_p95,
            "loss_rate": loss_rate,
        }

    def _raw_covers_window_start(self, start_ts) -> bool:
        """True when pings_raw still retains samples from the query window start.

        Strict cutoff only — no slack (Bug100).  Tolerance would mark P95 as
        \"exact\" while the first minutes of the window may already be purged.
        Default 8-day raw retention + 7→8 config migration covers rolling
        \"近 7 天\"; users who set 7-day retention may see N/A for that preset.
        """
        return int(start_ts) >= self._raw_retention_cutoff()

    def _raw_probe_count_in_window(self, conn, target_id: str,
                                  win_start: float, win_end: float) -> int:
        """All pings_raw rows for a target in [win_start, win_end]."""
        row = conn.execute(
            "SELECT COUNT(*) FROM pings_raw "
            "WHERE target_id=? AND ts >= ? AND ts <= ?",
            (target_id, win_start, win_end)).fetchone()
        return int(row[0]) if row else 0

    def _global_lat_p95_sql(self, conn, target_id: str,
                            win_start: float, win_end: float) -> Optional[float]:
        """Nearest-rank P95 over successful probes with latency in a window."""
        if win_start > win_end:
            return None
        lat_n = conn.execute(
            f"SELECT COUNT(latency_ms) FROM pings_raw "
            f"WHERE target_id=? AND ts >= ? AND ts <= ? "
            f"AND {_PROBE_LAT_OK_SQL}",
            (target_id, win_start, win_end)).fetchone()[0]
        if not lat_n:
            return None
        offset = min(lat_n - 1, max(0, math.ceil(lat_n * 0.95) - 1))
        p95_row = conn.execute(
            f"SELECT latency_ms FROM pings_raw "
            f"WHERE target_id=? AND ts >= ? AND ts <= ? "
            f"AND {_PROBE_LAT_OK_SQL} "
            f"ORDER BY latency_ms ASC LIMIT 1 OFFSET ?",
            (target_id, win_start, win_end, offset)).fetchone()
        return p95_row[0] if p95_row else None

    def _rollup_latency_for_window(self, conn, target_id: str,
                                  buckets: list, start_ts, end_ts) -> dict:
        """Hourly-bucket rollup for totals/avg/max; global P95 from raw when possible."""
        merged = self._rollup_hourly_buckets(buckets)
        merged["lat_p95"] = None
        merged["p95_from_raw"] = False
        total = merged.get("total") or 0
        win_start = float(start_ts)
        win_end = float(end_ts)
        if (total > 0
                and self._raw_covers_window_start(start_ts)
                and self._raw_probe_count_in_window(
                    conn, target_id, win_start, win_end) == total):
            merged["lat_p95"] = self._global_lat_p95_sql(
                conn, target_id, win_start, win_end)
            merged["p95_from_raw"] = True
        return merged

    def _hourly_row_to_bucket(self, row) -> dict:
        return {"hour_ts":  row[0], "sample_n": row[1],
                "success_n": row[2], "lat_avg":  row[3],
                "lat_max":   row[4], "lat_min":   row[5],
                "lat_p95":  row[6], "loss_rate": row[7]}

    def _fetch_hourly_map(self, conn, target_id: str,
                          hour_lo: int, hour_hi: int) -> dict:
        """All pings_hourly buckets for one target in [hour_lo, hour_hi]."""
        if hour_lo > hour_hi:
            return {}
        rows = conn.execute(
            "SELECT hour_ts, sample_n, success_n, lat_avg, lat_max, "
            "lat_min, lat_p95, loss_rate FROM pings_hourly "
            "WHERE target_id=? AND hour_ts >= ? AND hour_ts <= ?",
            (target_id, hour_lo, hour_hi)).fetchall()
        return {r[0]: self._hourly_row_to_bucket(r) for r in rows}

    def _fetch_hourly_map_batch(self, conn, target_ids: list,
                                hour_lo: int, hour_hi: int) -> dict:
        """{(target_id, hour_ts): bucket} for many targets, one SQL round-trip."""
        if not target_ids or hour_lo > hour_hi:
            return {}
        ph = ",".join("?" * len(target_ids))
        rows = conn.execute(
            f"SELECT target_id, hour_ts, sample_n, success_n, lat_avg, "
            f"lat_max, lat_min, lat_p95, loss_rate FROM pings_hourly "
            f"WHERE target_id IN ({ph}) "
            f"AND hour_ts >= ? AND hour_ts <= ?",
            (*target_ids, hour_lo, hour_hi)).fetchall()
        out = {}
        for r in rows:
            out[(r[0], r[1])] = {
                "hour_ts":  r[1], "sample_n": r[2], "success_n": r[3],
                "lat_avg":  r[4], "lat_max":   r[5], "lat_min":   r[6],
                "lat_p95":  r[7], "loss_rate": r[8],
            }
        return out

    def _raw_buckets_for_hours(self, conn, target_id: str,
                               hours: list, range_lo: int,
                               range_hi: int) -> dict:
        """One pings_raw SELECT, bucket rows into full-hour shapes."""
        if not hours:
            return {}
        # Half-open hour [h, h+3600): float ts can land in the last
        # sub-second (e.g. h+3599.50); +3599 / <= would drop those rows.
        qlo = max(float(range_lo), min(hours))
        qhi = min(float(range_hi), max(hours) + 3600)
        rows = conn.execute(
            "SELECT ts, latency_ms, status, failure_reason, probe_success "
            "FROM pings_raw "
            "WHERE target_id=? AND ts >= ? AND ts < ?",
            (target_id, qlo, qhi)).fetchall()
        by_hour = defaultdict(list)
        for ts, lat, st, fr, ps in rows:
            by_hour[int(ts // 3600) * 3600].append((lat, st, fr, ps))
        out = {}
        for h in hours:
            bucket = self._compute_hourly_bucket(by_hour.get(h, []))
            if bucket:
                bucket["hour_ts"] = h
                out[h] = bucket
        return out

    def _raw_retention_cutoff(self) -> int:
        """Earliest ts still present in pings_raw (exclusive lower bound)."""
        return int(time.time()) - self._raw_retention_days * 86400

    def _partial_boundary_bucket(self, conn, target_id: str, hour_ts: int,
                                 win_start: float, win_end: float,
                                 hourly_by_hour: dict) -> Optional[dict]:
        """
        Partial-hour bucket: precise pings_raw slice when available.

        One pings_raw read for the whole hour, then slice in Python.

          * window starts before raw retention cutoff → hourly approx
            (purged slice cannot be rebuilt from later in-hour raw);
          * else slice has rows → precise bucket;
          * else hour still has raw → None (Bug 21, no out-of-window bleed);
          * else → hourly approx (Bug 19, raw fully purged).
        """
        raw_cutoff = self._raw_retention_cutoff()
        if int(win_start) < raw_cutoff:
            return hourly_by_hour.get(hour_ts)
        rows = conn.execute(
            "SELECT ts, latency_ms, status, failure_reason, probe_success "
            "FROM pings_raw "
            "WHERE target_id=? AND ts >= ? AND ts < ?",
            (target_id, hour_ts, hour_ts + 3600)).fetchall()
        slice_rows = [(lat, st, fr, ps) for ts, lat, st, fr, ps in rows
                      if win_start <= ts <= win_end and ts < hour_ts + 3600]
        if slice_rows:
            bucket = self._compute_hourly_bucket(slice_rows)
            if bucket:
                bucket["hour_ts"] = hour_ts
                return bucket
            return None
        if rows:
            return None
        return hourly_by_hour.get(hour_ts)

    def _bounded_hourly_buckets(self, conn, target_id: str,
                                start_ts, end_ts,
                                hourly_by_hour: Optional[dict] = None) -> list:
        """
        Hourly buckets aligned to [start_ts, end_ts].

        Full interior hours come from pings_hourly (one range query).
        Partial hours at boundaries prefer pings_raw; if raw is gone,
        approximate with the stored pings_hourly bucket.
        """
        start_f = float(start_ts)
        end_f   = float(end_ts)
        if end_f < start_f:
            return []
        ch = self._current_hour_start()
        h_first = (int(start_f) // 3600) * 3600
        h_last  = (int(min(end_f, ch - 1)) // 3600) * 3600
        if hourly_by_hour is None:
            hourly_by_hour = (self._fetch_hourly_map(conn, target_id, h_first, h_last)
                              if h_first <= h_last else {})
        out = []
        missing_full = []
        h  = h_first
        while h <= end_f and h < ch:
            if h + 3600 <= start_f:
                h += 3600
                continue
            win_start = max(start_f, float(h))
            win_end   = min(end_f, float(h + 3600))
            if win_start > win_end:
                h += 3600
                continue
            if start_f <= h and end_f >= h + 3600:
                b = hourly_by_hour.get(h)
                if b:
                    out.append(b)
                else:
                    missing_full.append(h)
            else:
                b = self._partial_boundary_bucket(
                    conn, target_id, h, win_start, win_end, hourly_by_hour)
                if b:
                    out.append(b)
            h += 3600
        if missing_full:
            for h, b in self._raw_buckets_for_hours(
                    conn, target_id, missing_full, start_f, end_f).items():
                out.append(b)
        out.sort(key=lambda b: b["hour_ts"])
        live = self._current_hour_bucket(conn, target_id, start_f, end_f)
        if live:
            out.append(live)
        return out

    def _bounded_hourly_buckets_for_targets(
            self, conn, target_ids: list,
            start_ts, end_ts) -> dict:
        """Per-target bounded buckets; one batch pings_hourly SELECT."""
        start_f = float(start_ts)
        end_f   = float(end_ts)
        if not target_ids or end_f < start_f:
            return {tid: [] for tid in target_ids}
        ch = self._current_hour_start()
        h_first = (int(start_f) // 3600) * 3600
        h_last  = (int(min(end_f, ch - 1)) // 3600) * 3600
        batch = (self._fetch_hourly_map_batch(conn, target_ids, h_first, h_last)
                 if h_first <= h_last else {})
        per_tid = defaultdict(dict)
        for (tid, h), bucket in batch.items():
            per_tid[tid][h] = bucket
        return {
            tid: self._bounded_hourly_buckets(
                conn, tid, start_f, end_f,
                hourly_by_hour=per_tid.get(tid))
            for tid in target_ids
        }

    @staticmethod
    def _rollup_hourly_buckets(buckets: list) -> dict:
        """Fold hourly-shaped buckets into SLA latency totals."""
        if not buckets:
            return {"total": 0, "success": 0,
                    "lat_avg": None, "lat_max": None, "lat_p95": None}
        total   = sum(b["sample_n"] for b in buckets)
        success = sum(b["success_n"] for b in buckets)
        lat_pairs = [(b["lat_avg"], b["success_n"]) for b in buckets
                     if b["lat_avg"] is not None and b["success_n"]]
        maxlats   = [b["lat_max"] for b in buckets if b["lat_max"] is not None]
        w_n  = sum(n for _v, n in lat_pairs)
        lat_avg = (sum(v * n for v, n in lat_pairs) / w_n if w_n else None)
        lat_max = max(maxlats) if maxlats else None
        # lat_p95 is not statistically valid across buckets; callers use
        # _rollup_latency_for_window() for global P95 when raw is available.
        return {"total": total, "success": success,
                "lat_avg": lat_avg, "lat_max": lat_max, "lat_p95": None}

    def _current_hour_bucket(self, conn, target_id: str,
                             start_ts: float, end_ts: float) -> Optional[dict]:
        """In-memory hourly bucket for the still-open current hour."""
        now = time.time()
        ch  = self._current_hour_start(now)
        if float(end_ts) < ch:
            return None
        return self._raw_window_bucket(
            conn, target_id, ch,
            max(float(start_ts), float(ch)), min(float(end_ts), now))

    def _merge_sla_latency(self, base: dict, extra: Optional[dict]) -> dict:
        """Combine pings_hourly SQL rollup with a current-hour raw bucket."""
        if not extra:
            return base
        total   = (base.get("total") or 0) + extra["sample_n"]
        success = (base.get("success") or 0) + extra["success_n"]
        old_sn  = base.get("success") or 0
        new_sn  = extra["success_n"]
        lat_avg = base.get("lat_avg")
        if extra["lat_avg"] is not None and new_sn:
            if lat_avg is not None and old_sn:
                lat_avg = (((lat_avg * old_sn) + (extra["lat_avg"] * new_sn))
                           / (old_sn + new_sn))
            else:
                lat_avg = extra["lat_avg"]
        lat_max = base.get("lat_max")
        if extra["lat_max"] is not None:
            lat_max = max(lat_max, extra["lat_max"]) if lat_max is not None else extra["lat_max"]
        lat_p95 = base.get("lat_p95")
        if extra["lat_p95"] is not None and new_sn:
            if lat_p95 is not None and old_sn:
                lat_p95 = (((lat_p95 * old_sn) + (extra["lat_p95"] * new_sn))
                           / (old_sn + new_sn))
            else:
                lat_p95 = extra["lat_p95"]
        return {
            "total":   total,
            "success": success,
            "lat_avg": lat_avg,
            "lat_max": lat_max,
            "lat_p95": lat_p95,
        }

    def get_history(self, target_id: str, start_ts: int, end_ts: int,
                    use_hourly: bool = False) -> list:
        """
        查询单节点历史数据。
        use_hourly=False → pings_raw（秒级）
        use_hourly=True  → pings_hourly（小时级）+ 当前未结束小时的临时桶
        """
        conn = self._read_conn()
        try:
            if use_hourly:
                return self._bounded_hourly_buckets(
                    conn, target_id, start_ts, end_ts)
            else:
                rows = conn.execute(
                    "SELECT ts, status, latency_ms, loss_rate, "
                    "status_code, failure_reason, probe_success FROM pings_raw "
                    "WHERE target_id=? AND ts BETWEEN ? AND ? "
                    "ORDER BY ts ASC, id ASC",
                    (target_id, start_ts, end_ts)).fetchall()
                return [{"ts": r[0], "status": r[1], "latency_ms": r[2],
                         "loss_rate": r[3], "status_code": r[4],
                         "failure_reason": r[5], "probe_success": r[6]}
                        for r in rows]
        except Exception:
            return []
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def get_alert_history(self, target_id: Optional[str] = None,
                          start_ts: int = 0, limit: int = 200) -> list:
        """查询告警事件历史。"""
        conn = self._read_conn()
        try:
            if target_id:
                rows = conn.execute(
                    "SELECT id, target_id, label, ip, ts, old_status, "
                    "new_status, category, failure_reason, incident_id, "
                    "ping_type FROM alert_events "
                    "WHERE target_id=? AND ts>=? "
                    "ORDER BY ts DESC, id DESC LIMIT ?",
                    (target_id, start_ts, limit)).fetchall()
            else:
                rows = conn.execute(
                    "SELECT id, target_id, label, ip, ts, old_status, "
                    "new_status, category, failure_reason, incident_id, "
                    "ping_type FROM alert_events "
                    "WHERE ts>=? ORDER BY ts DESC, id DESC LIMIT ?",
                    (start_ts, limit)).fetchall()
            # id + ORDER BY ts DESC, id DESC: export_excel re-sorts by
            # (ts, id) ascending; stable tie-break avoids same-second
            # events appearing in reverse insertion order.
            # incident_id is included so export_excel can do the Level-2
            # exact JOIN into traceroute_logs.  May be NULL on historical
            # rows that pre-date the v1→v2 backfill — get_incident_traceroute
            # then falls back to its time-window match.
            return [{"id": r[0], "target_id": r[1], "label": r[2], "ip": r[3],
                     "ts": r[4], "old_status": r[5],
                     "new_status": r[6], "category": r[7],
                     "failure_reason": r[8] if len(r) > 8 else "",
                     "incident_id": r[9] if len(r) > 9 else None,
                     "ping_type": r[10] if len(r) > 10 else ""}
                    for r in rows]
        except Exception:
            return []
        # _read_conn() returns a thread-local cached connection — do NOT close.

    @staticmethod
    def _alert_event_from_row(row) -> dict:
        return {"id": row[0], "target_id": row[1], "label": row[2], "ip": row[3],
                "ts": row[4], "old_status": row[5], "new_status": row[6],
                "category": row[7], "failure_reason": row[8] if len(row) > 8 else "",
                "incident_id": row[9] if len(row) > 9 else None,
                "ping_type": row[10] if len(row) > 10 else ""}

    def _get_export_alert_events(self, conn, target_id: str,
                                 start_ts: int, end_ts: int) -> list:
        """Alert events for Excel timeline: boundary state + in-window stream.

        Mirrors get_sla_stats() boundary lookup: the last red/orange-touching
        event before start_ts seeds ongoing incidents that began before the
        export window (Case C).  Window events are [start_ts, end_ts].
        """
        boundary = conn.execute(
            "SELECT id, target_id, label, ip, ts, old_status, new_status, "
            "category, failure_reason, incident_id, ping_type FROM alert_events "
            "WHERE target_id=? "
            "AND (new_status IN ('red','orange') OR old_status IN ('red','orange')) "
            "AND ts < ? "
            "ORDER BY ts DESC, id DESC LIMIT 1",
            (target_id, start_ts)).fetchone()
        window_rows = conn.execute(
            "SELECT id, target_id, label, ip, ts, old_status, new_status, "
            "category, failure_reason, incident_id, ping_type FROM alert_events "
            "WHERE target_id=? AND ts >= ? AND ts <= ? "
            "ORDER BY ts ASC, id ASC",
            (target_id, start_ts, end_ts)).fetchall()
        out = []
        seen = set()
        if boundary and boundary[6] in ("red", "orange"):
            ev = self._alert_event_from_row(boundary)
            out.append(ev)
            seen.add(ev["id"])
        for row in window_rows:
            ev = self._alert_event_from_row(row)
            if ev["id"] not in seen:
                out.append(ev)
                seen.add(ev["id"])
        return out

    def export_excel(self, output_path: str, target_ids: list,
                     start_ts: int, end_ts: int,
                     targets_meta: Optional[dict] = None) -> str:
        """
        导出 Excel 报表。
        窗口起点早于 pings_raw 保留期，或跨度 > 24 小时 → 小时级聚合；
        否则 → 亚秒级原始数据（时间列显示到毫秒）。每个节点一个 Sheet，首页汇总。
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl 未安装，请运行 pip install openpyxl")

        from src.trace_policy import describe_failure, ping_type_label

        if targets_meta is None:
            targets_meta = self.get_targets_meta()

        use_hourly = self.should_use_hourly(start_ts, end_ts, 24 * 3600)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ── 汇总 Sheet ──────────────────────────────────────────────
        ws_sum = wb.create_sheet("汇总")
        HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
        HDR_FONT = Font(color="FFFFFF", bold=True)

        sum_headers = ["节点名称", "IP 地址", "探测类型", "探测次数",
                       "成功次数", "连通率 %", "平均延时 ms", "最大延时 ms"]
        for col, h in enumerate(sum_headers, 1):
            cell = ws_sum.cell(1, col, h)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT

        for row_i, tid in enumerate(target_ids, 2):
            meta    = targets_meta.get(tid, {})
            label   = meta.get("label", tid)
            ip      = meta.get("ip", "")
            ptype   = meta.get("ping_type", "icmp")
            rows    = self.get_history(tid, start_ts, end_ts, use_hourly)

            if use_hourly:
                total   = sum(r["sample_n"] for r in rows)
                success = sum(r["success_n"] for r in rows)
                maxlats = [r["lat_max"] for r in rows if r["lat_max"] is not None]
                # Weight by success_n -- the count of samples that actually
                # contributed a latency to lat_avg.  See _hourly_agg: lat_avg
                # is computed as sum(latencies)/count(non-null latencies),
                # so weighting by sample_n (total probes including failures)
                # would let a 2-success/3600-probe hour dominate the average
                # as if it had 3600 measurements, hiding severe outages in
                # the SLA report.
                w_pairs = [(r["lat_avg"], r["success_n"]) for r in rows
                             if r["lat_avg"] is not None and r["success_n"]]
                if w_pairs:
                    w_n   = sum(n for _, n in w_pairs)
                    avg_lat = round(sum(avg * n for avg, n in w_pairs) / w_n, 1)
                else:
                    avg_lat = ""
            else:
                total   = len(rows)
                success = sum(
                    1 for r in rows
                    if self.row_probe_success(
                        r["latency_ms"], r["status"],
                        r.get("failure_reason"), r.get("probe_success")))
                lats = [
                    r["latency_ms"] for r in rows
                    if self.row_probe_success(
                        r["latency_ms"], r["status"],
                        r.get("failure_reason"), r.get("probe_success"))
                    and r["latency_ms"] is not None]
                maxlats = lats
                avg_lat = round(sum(lats) / len(lats), 1) if lats else ""

            rate    = round(success / total * 100, 2) if total else 0
            max_lat = round(max(maxlats), 1) if maxlats else ""

            # excel_safe_text on every string column before append() --
            # operator-typed labels / ips that start with =/+/-/@ would
            # otherwise be saved as Excel formulas (CWE-1236), executed
            # when the recipient opens the workbook.  The summary
            # sheet is the most visible sink so it gets the explicit
            # treatment; the alert-timeline sheet below escapes its
            # external-string cells the same way.
            ws_sum.append([excel_safe_text(label),
                           excel_safe_text(ip),
                           excel_safe_text(ptype),
                           total, success, rate, avg_lat, max_lat])

        ws_sum.column_dimensions["A"].width = 20
        ws_sum.column_dimensions["B"].width = 16

        # ── 每节点 Sheet ────────────────────────────────────────────
        for tid in target_ids:
            meta  = targets_meta.get(tid, {})
            raw_label = (meta.get("label", tid) or tid)
            # excel_safe_sheet_name handles the full constraint set
            # (\\ / * ? : [ ] forbidden chars + XML-illegal control
            # chars + length cap + empty-name fallback) so a label
            # containing e.g. \\x00 produces a reopenable workbook.
            # Previously only the filename-character substitution ran
            # and openpyxl saved a workbook whose XML had raw NUL in a
            # sheet title -- the file looked fine, returned HTTP 200,
            # but ParseError'd on every subsequent open.
            label = excel_safe_sheet_name(raw_label, fallback=tid[:31] or "node")
            ws    = wb.create_sheet(label)
            rows  = self.get_history(tid, start_ts, end_ts, use_hourly)

            if use_hourly:
                headers = ["时间（小时）", "探测次数", "成功次数",
                           "连通率 %", "平均延时 ms", "最大延时 ms",
                           "P95 延时 ms", "丢包率 %"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(1, col, h)
                    cell.fill = HDR_FILL; cell.font = HDR_FONT

                for r in rows:
                    dt   = datetime.fromtimestamp(r["hour_ts"]).strftime("%Y-%m-%d %H:%M")
                    rate = round(r["success_n"] / r["sample_n"] * 100, 1) \
                           if r["sample_n"] else 0
                    loss = round((r["loss_rate"] or 0) * 100, 1)
                    # NULL latencies mean "no successful probe in this hour"
                    # (e.g. 100% packet loss).  `... or 0` would coerce that
                    # to 0.0ms in the export, suggesting an absurd
                    # instantaneous response time during a total outage --
                    # leave the cell blank instead.
                    ws.append([dt, r["sample_n"], r["success_n"], rate,
                               round(r["lat_avg"], 1) if r["lat_avg"] is not None else "",
                               round(r["lat_max"], 1) if r["lat_max"] is not None else "",
                               round(r["lat_p95"], 1) if r["lat_p95"] is not None else "",
                               loss])
            else:
                headers = ["时间", "状态", "延时 ms", "丢包率 %", "HTTP 状态码", "失败原因"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(1, col, h)
                    cell.fill = HDR_FILL; cell.font = HDR_FONT

                for r in rows:
                    dt   = fmt_ts_ms(r["ts"])
                    loss = round((r["loss_rate"] or 0) * 100, 1)
                    # `if r["latency_ms"] else ""` would silently drop a
                    # legitimate 0.0 ms reading (sub-millisecond LAN /
                    # loopback / mocked targets) as if it were missing.
                    # The hourly-sheet path above (~line 478) already uses
                    # `is not None`; match that semantic here.
                    lat = r["latency_ms"]
                    fr = r.get("failure_reason") or ""
                    ws.append([dt, r["status"],
                               round(lat, 1) if lat is not None else "",
                               loss, r["status_code"] or "",
                               describe_failure(fr) if fr else ""])

            ws.column_dimensions["A"].width = 20


        # ── 告警时间线 Sheet ─────────────────────────────────────────
        # Per-target boundary + in-window events (not a fixed 24h lookback).
        from collections import defaultdict
        events_by_tid = defaultdict(list)
        conn_al = self._read_conn()
        try:
            tids = target_ids or []
            if not tids:
                rows = conn_al.execute(
                    "SELECT DISTINCT target_id FROM alert_events "
                    "WHERE ts >= ? AND ts <= ?",
                    (start_ts, end_ts)).fetchall()
                tids = [r[0] for r in rows]
            for tid in tids:
                events_by_tid[tid] = self._get_export_alert_events(
                    conn_al, tid, start_ts, end_ts)
        except Exception:
            pass

        # ── 告警时间线 sheet (完整故障事件流) ───────────────────────────
        ws_al = wb.create_sheet("告警时间线")

        # Column layout
        # A: 序号  B: 节点  C: IP  D: 探测类型  E: 事件时间  F: 状态变化
        # G: 分类  H: 原因/详情  I: 距上次事件
        COL_W = {"A": 6, "B": 18, "C": 16, "D": 10, "E": 20, "F": 16,
                 "G": 12, "H": 38, "I": 14}
        for col, w in COL_W.items():
            ws_al.column_dimensions[col].width = w

        # ── Style constants ──────────────────────────────────────────
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        _thin = Side(style='thin', color="CCCCCC")
        _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        INCIDENT_HDR_FILL = _fill("1F4E79")   # dark blue – incident header
        DETAIL_HDR_FILL   = _fill("D6E4F0")   # light blue – column headers
        GREEN_FILL        = _fill("E8F5E9")   # light green – recovery events
        ORANGE_FILL       = _fill("FFF3E0")   # light orange – warning events
        RED_FILL          = _fill("FFEBEE")   # light red – critical events
        GRAY_FILL         = _fill("F5F5F5")   # pause/transitional events
        WHITE_FILL        = _fill("FFFFFF")

        STATUS_LABEL = {
            'green': '正常', 'orange': '警告', 'red': '严重',
            'gray': '离线', 'paused': '暂停',
        }
        STATUS_FILL = {
            'green': GREEN_FILL, 'orange': ORANGE_FILL, 'red': RED_FILL,
            'gray': GRAY_FILL, 'paused': GRAY_FILL,
        }
        SEVERITY = {'green': 0, 'gray': 1, 'paused': 1, 'orange': 2, 'red': 3}

        def _fmt_ts(ts):
            return fmt_ts_ms(ts)

        def _fmt_dur(seconds):
            if seconds is None: return "—"
            h2, r = divmod(int(abs(seconds)), 3600)
            m2, s2 = divmod(r, 60)
            if h2:
                return f"{h2}时{m2:02d}分{s2:02d}秒"
            elif m2:
                return f"{m2}分{s2:02d}秒"
            return f"{s2}秒"

        def _st_label(s):
            return STATUS_LABEL.get(s, s or '未知')

        def _cat_label(cat):
            """Return human-readable alert category."""
            cat_map = {
                'availability': '可用性告警',
                'performance':  '性能告警',
                'ok':           '恢复正常',
                'paused':       '手动暂停',
                'resumed':      '手动恢复',
                'deleted':      '节点已删除',
            }
            return cat_map.get(cat, cat or '—')

        def _write_row(ws, row_num, values, fill=None, bold=False,
                       font_color="000000", merge_to=None, align="left"):
            for col, val in enumerate(values, 1):
                c = ws.cell(row_num, col, val)
                c.border = _border
                if fill:
                    c.fill = fill
                c.font = Font(bold=bold, color=font_color,
                              name="微软雅黑", size=9)
                c.alignment = Alignment(horizontal=align, vertical="center",
                                        wrap_text=(col == 8))
            if merge_to:
                ws.merge_cells(
                    start_row=row_num, start_column=1,
                    end_row=row_num,   end_column=merge_to)

        # ── Group events into incidents ──────────────────────────────
        def _build_incidents(events):
            """
            Returns list of dicts:
              {'label', 'ip', 'events': [...], 'closed': bool,
               'start_truncated': bool}
            Each 'events' list contains the full transition stream
            (may start mid-incident if it began before the query window).
            """
            incidents = []
            current_events = None
            start_truncated = False

            for ev in sorted(events, key=lambda e: (e['ts'], e.get('id', 0))):
                ns  = ev.get('new_status', '')
                os_ = ev.get('old_status', '')
                cat = ev.get('category', '')

                if cat in ('paused', 'resumed'):
                    if current_events is not None:
                        current_events.append(ev)
                    # outside incident: show as standalone entry below
                    continue

                if current_events is None:
                    if ns in ('orange', 'red'):
                        current_events = [ev]
                        # Boundary events from _get_export_alert_events()
                        # carry ts < start_ts — mark so the header and
                        # duration reflect the query window, not the full
                        # pre-window fault span.
                        start_truncated = ev['ts'] < start_ts
                    elif os_ in ('orange', 'red') and ns == 'green':
                        # Recovery with no fault-start in this window
                        current_events = [ev]
                        start_truncated = True
                        incidents.append({
                            'events': current_events,
                            'closed': True,
                            'start_truncated': start_truncated,
                        })
                        current_events = None
                else:
                    current_events.append(ev)
                    if ns == 'green':
                        incidents.append({
                            'events': current_events,
                            'closed': True,
                            'start_truncated': start_truncated,
                        })
                        current_events = None

            if current_events:
                incidents.append({
                    'events': current_events,
                    'closed': False,
                    'start_truncated': start_truncated,
                })
            return incidents

        # ── Render the sheet ─────────────────────────────────────────
        row_al = 1
        # Sheet-level header
        _write_row(ws_al, row_al,
                   ["", "节点", "IP", "探测类型", "事件时间", "状态变化",
                    "分类", "原因/详情", "距上次事件"],
                   fill=HDR_FILL, bold=True, font_color="FFFFFF")
        row_al += 1

        incident_seq = 0
        for tid in (target_ids or list(events_by_tid.keys())):
            meta   = targets_meta.get(tid, {})
            label  = meta.get("label", tid)
            ip     = meta.get("ip", "")
            events = events_by_tid.get(tid, [])

            incidents = _build_incidents(events)

            for inc in incidents:
                evts = inc['events']
                closed = inc['closed']
                truncated = inc['start_truncated']

                if not evts:
                    continue

                # Only show if the incident overlaps the query window
                inc_start_ts = evts[0]['ts']
                inc_end_ts   = evts[-1]['ts'] if closed else end_ts
                if inc_end_ts < start_ts or inc_start_ts > end_ts:
                    continue

                incident_seq += 1

                # Calculate incident metrics
                fault_start = evts[0]['ts']
                fault_end   = evts[-1]['ts'] if closed else None
                effective_start = (
                    max(fault_start, start_ts) if truncated else fault_start)
                total_secs  = (
                    (fault_end - effective_start) if fault_end else None)
                max_sev     = max(
                    (SEVERITY.get(e.get('new_status', ''), 0) for e in evts),
                    default=0)
                sev_label   = {0: '正常', 1: '离线/暂停',
                               2: '⚠ 性能警告', 3: '🔴 严重告警'}[max_sev]
                dur_str     = (_fmt_dur(total_secs)
                               if total_secs is not None else "仍在持续")
                if truncated:
                    prefix = f"⚠ 故障事件 #{incident_seq}（开始于查询范围外）"
                elif not closed:
                    prefix = f"⚠ 故障事件 #{incident_seq}（查询结束时仍未恢复）"
                else:
                    prefix = f"故障事件 #{incident_seq}"

                hdr_text = (f"{prefix}  │  {label}  ({ip})  │  "
                            f"历时: {dur_str}  │  最高级别: {sev_label}")

                # Incident header row -- excel_safe_text on the whole
                # composed line because label/ip can each start with a
                # formula-trigger character; even though the prefix
                # ("故障事件 #N  │  ") makes the cell start with a
                # safe character, the wrapping helper still scans the
                # leading byte so this is a no-op for the normal
                # case but covers an adversarial label like
                # "=cmd|...".  Same reasoning applies to the per-row
                # cells below.
                _write_row(ws_al, row_al,
                           [excel_safe_text(hdr_text)] + [""] * 8,
                           fill=INCIDENT_HDR_FILL, bold=True,
                           font_color="FFFFFF", merge_to=9)
                ws_al.row_dimensions[row_al].height = 18
                row_al += 1

                # ── Forensic auto-traceroute snapshot row ─────────────────
                # The auto-tracert fires the instant a node turns red and is
                # persisted to traceroute_logs.  Surface "where the path
                # broke" here so the report is incident-complete even after
                # the link recovered and the live tracert looks healthy.
                # red_ts = first red transition (auto-tracert only fires on
                # red, never on orange-only performance warnings).
                red_ts = next((e['ts'] for e in evts
                               if e.get('new_status') == 'red'), None)
                if red_ts is not None:
                    from src.trace_policy import (
                        should_request_traceroute,
                        summarize_for_alert,
                        trace_skip_summary,
                    )
                    red_ev = next((e for e in evts
                                   if e.get('new_status') == 'red'), None)
                    ptype = meta.get("ping_type", "icmp")
                    red_fr = (red_ev.get("failure_reason") or "") if red_ev else ""
                    if not should_request_traceroute(ptype, red_fr):
                        diag_text = (
                            "🔬 " + trace_skip_summary(ptype, red_fr)["text"])
                    else:
                        # incident_id is None on the Level-1 schema
                        # (get_alert_history does not yet return it);
                        # get_incident_traceroute then uses the time-window
                        # fallback.  Wired here for Level-2 forward-compat.
                        inc_iid = next((e.get('incident_id') for e in evts
                                        if e.get('new_status') == 'red'), None)
                        snap = self.get_incident_traceroute(
                            tid, red_ts, incident_id=inc_iid)
                        if snap:
                            brk = summarize_for_alert(
                                snap.get("hops", []),
                                ping_type=ptype,
                                failure_reason=red_fr,
                            )
                            trace_t = _fmt_ts(snap.get("ts"))
                            chg = ("；路径较上次有变化"
                                   if snap.get("route_changed") else "")
                            diag_text = (
                                f"🔬 故障时自动追踪 @ {trace_t}："
                                f"{brk['text']}{chg}")
                        else:
                            diag_text = (
                                "🔬 故障时自动追踪：（无追踪记录，"
                                "可能保留期已过或追踪未及执行）")
                    _write_row(ws_al, row_al, [diag_text] + [""] * 8,
                               fill=_fill("E1F5FE"), bold=False,
                               font_color="01579B", merge_to=9)
                    row_al += 1

                # Event rows
                prev_ts = None
                for ev in evts:
                    ts   = ev['ts']
                    ns   = ev.get('new_status', '')
                    os_  = ev.get('old_status', '')
                    cat  = ev.get('category', '')
                    fr   = ev.get('failure_reason', '')
                    ev_pt = (ev.get('ping_type') or meta.get("ping_type", "icmp"))

                    # Keep pre-window transitions that belong to this incident;
                    # only drop rows far after the export window.
                    if ts > end_ts + 86400:
                        continue

                    gap = _fmt_dur(ts - prev_ts) if prev_ts else "—"
                    prev_ts = ts

                    fr_text = describe_failure(fr) if fr else "—"
                    row_fill = STATUS_FILL.get(ns, WHITE_FILL)
                    _write_row(ws_al, row_al, [
                        "  ↳",
                        excel_safe_text(label),
                        excel_safe_text(ip),
                        ping_type_label(ev_pt),
                        _fmt_ts(ts),
                        f"{_st_label(os_)} → {_st_label(ns)}",
                        _cat_label(cat),
                        excel_safe_text(fr_text),
                        gap,
                    ], fill=row_fill)
                    row_al += 1

                # Blank separator row between incidents
                for col in range(1, 10):
                    ws_al.cell(row_al, col, "").fill = WHITE_FILL
                row_al += 1

        if incident_seq == 0:
            _write_row(ws_al, row_al,
                       ["（查询时段内无故障事件记录）"] + [""] * 8,
                       fill=GREEN_FILL, merge_to=9)
            row_al += 1

        ws_al.freeze_panes = "B2"
        # ── end alert sheet ─────────────────────────────────────────

        wb.save(output_path)
        return output_path

    # ── 显式控制 ────────────────────────────────────────────────────

    def flush(self):
        """同步 flush，阻塞直到写入完成（用于程序退出前）。"""
        done = threading.Event()
        self._queue.put(("_flush_sync", done))
        done.wait(timeout=15)

    def _maybe_vacuum(self):
        """VACUUM the database at most once every 7 days.
        Waits 5s on startup so the write-thread WAL connection is
        established first, avoiding a "database is locked" error.
        """
        import time as _t, sqlite3 as _sq
        _t.sleep(5)   # let the write thread complete its initial setup
        INTERVAL = 7 * 86400
        KEY = "last_vacuum_ts"
        try:
            conn = _sq.connect(self._db_path, check_same_thread=True)
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("CREATE TABLE IF NOT EXISTS app_meta(key TEXT PRIMARY KEY, value TEXT)")
            row  = conn.execute("SELECT value FROM app_meta WHERE key=?", (KEY,)).fetchone()
            last = float(row[0]) if row else 0.0
            if _t.time() - last >= INTERVAL:
                conn.execute("VACUUM")
                sql = "INSERT OR REPLACE INTO app_meta(key,value) VALUES(?,?)"
                conn.execute(sql, (KEY, str(_t.time())))
                conn.commit()
                print("[DataStore] VACUUM completed")
        except Exception as e:
            print(f"[DataStore] VACUUM skipped: {e}")
        finally:
            try: conn.close()
            except Exception: pass

    def shutdown(self):
        """flush 后关闭写线程（优雅退出）。"""
        done = threading.Event()
        self._queue.put(("_shutdown", done))
        done.wait(timeout=30)

    # ── 内部：写线程 ────────────────────────────────────────────────

    def _write_loop(self):
        conn = sqlite3.connect(self._db_path, check_same_thread=True)
        # 性能 + 并发安全配置
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA auto_vacuum=INCREMENTAL")
        conn.execute("PRAGMA cache_size=-8192")   # 8MB 页缓存

        ping_buf  = []
        alert_buf = []
        last_flush   = time.time()
        last_hourly  = time.time()
        last_cleanup = time.time() - 86400   # trigger cleanup on first run
        last_vacuum  = time.time()

        while True:   # outer: retry on transient error, never let thread exit
            try:
                while True:
                    try:
                        item = self._queue.get(timeout=5)
                    except queue.Empty:
                        item = None

                    if item is not None:
                        kind, data = item

                        if kind == "_init_schema":
                            self._init_schema(conn)
                            self._schema_ready.set()

                        elif kind == "ping":
                            # Final-stop generation guard at the queue
                            # boundary -- producer-side check in
                            # record_ping is a fast path, but the
                            # check + put aren't atomic with a
                            # concurrent bump_target_generation +
                            # wipe_target enqueue.  Drop here so a
                            # stale ping queued just before a wipe
                            # ran can't slip into the buffer.  _flush
                            # below applies the SAME filter again so
                            # a stale ping that sat in the buffer
                            # through a later wipe is also dropped
                            # right before INSERT.
                            cg = data.get("captured_generation")
                            tid = data.get("target_id")
                            if (cg is None or tid is None
                                or self.current_target_generation(tid) == cg):
                                ping_buf.append(data)

                        elif kind == "alert":
                            # Same generation guard as the "ping" branch
                            # above -- a stale red->green produced by an
                            # OLD-config probe that snuck past the
                            # commit-side identity check would otherwise
                            # land in alert_events as a phantom recovery
                            # event and silently close the user's real
                            # incident.  _flush re-checks alert_buf the
                            # same way ping_buf is filtered.
                            cg = data.get("captured_generation")
                            tid = data.get("target_id")
                            if not (cg is None or tid is None
                                    or self.current_target_generation(tid) == cg):
                                # Stale -- drop without buffering AND skip
                                # the immediate flush, since the alert
                                # branch's whole point was "force a flush
                                # now because this event matters".  A
                                # dropped event isn't a real event.
                                pass
                            else:
                                alert_buf.append(data)
                                # Flush alert events immediately — SLA/outage queries
                                # read directly from SQLite. Buffering for 60s means
                                # any SLA query made within 60s of an outage shows
                                # stale data. Flush now so the DB reflects reality.
                                if self._flush(conn, ping_buf, alert_buf):
                                    ping_buf.clear(); alert_buf.clear()
                                    last_flush = time.time()
                                # else: buffers preserved; next periodic flush retries

                        elif kind == "webhook_ack":
                            iid = data.get("incident_id")
                            if iid:
                                if data.get("action") == "clear":
                                    conn.execute(
                                        "DELETE FROM webhook_ack "
                                        "WHERE incident_id=?",
                                        (iid,))
                                else:
                                    conn.execute(
                                        "INSERT OR REPLACE INTO webhook_ack "
                                        "(incident_id, target_id, ack_ts) "
                                        "VALUES (?,?,?)",
                                        (iid, data["target_id"], data["ts"]))
                                conn.commit()

                        elif kind == "target_removed":
                            tid = data["target_id"]
                            # Close an open incident in DB before popping memory
                            # state — _assign_incident_ids needs the stored iid.
                            if (tid in self._active_incidents
                                    or data.get("old_status") in ("red", "orange")):
                                alert_buf.append({
                                    "target_id": tid,
                                    "label": data.get("label", ""),
                                    "ip": data.get("ip", ""),
                                    "ts": data["ts"],
                                    "old_status": data.get("old_status", "gray"),
                                    "new_status": "green",
                                    "category": "deleted",
                                    "failure_reason": "",
                                    "ping_type": data.get("ping_type", "icmp"),
                                })
                                if self._flush(conn, ping_buf, alert_buf):
                                    ping_buf.clear(); alert_buf.clear()
                                    last_flush = time.time()
                            else:
                                self._active_incidents.pop(tid, None)

                        elif kind == "traceroute":
                            row_data, captured_gen = data
                            target_id, label, ip, ts, hops_json = row_data
                            # Final-stop generation guard: the producer
                            # (TracerouteScheduler._run_one) already
                            # captures generation, but its capture + put
                            # is NOT atomic with a concurrent
                            # bump_target_generation + wipe_target enqueue.
                            # Wipe queues a DELETE first; this stale INSERT
                            # would otherwise repopulate traceroute_logs.
                            # Re-checking here under the writer thread --
                            # immediately before the INSERT -- closes the
                            # hole.  Same pattern repeated for "diag" /
                            # "dns_diag" below.
                            if (captured_gen is None or target_id is None
                                or self.current_target_generation(target_id)
                                    == captured_gen):
                                # Level-2 incident binding: stamp the
                                # currently-open incident_id (if any) onto
                                # this snapshot.  Both _assign_incident_ids
                                # (writer of _active_incidents) and this
                                # INSERT run in the same write thread, and
                                # the queue is single-consumer FIFO — the
                                # red alert event was queued the instant
                                # the node turned red and is flushed
                                # immediately on enqueue (see the "alert"
                                # branch above), so by the time the tracert
                                # result lands here (~5–125 s later) the
                                # incident is already open in
                                # self._active_incidents.  The rare
                                # "extremely brief outage that recovered
                                # before tracert INSERT" race is
                                # intentionally handled by leaving
                                # incident_id = NULL here and letting
                                # export_excel's Level-1 time-window
                                # fallback catch it.  Older rows written
                                # before the column existed also stay NULL
                                # — same fallback.
                                if self._traceroute_has_incident_col(conn):
                                    iid = self._active_incidents.get(target_id)
                                    conn.execute(
                                        "INSERT INTO traceroute_logs"
                                        "(target_id,label,ip,ts,hops,incident_id) "
                                        "VALUES(?,?,?,?,?,?)",
                                        (target_id, label, ip, ts, hops_json, iid))
                                else:
                                    # Defensive: schema migration somehow didn't run.
                                    # Keep the old INSERT shape so writes never fail.
                                    conn.execute(
                                        "INSERT INTO traceroute_logs"
                                        "(target_id,label,ip,ts,hops) "
                                        "VALUES(?,?,?,?,?)",
                                        (target_id, label, ip, ts, hops_json))
                                conn.commit()

                        elif kind == "wipe_target":
                            # Purge every historical row for one target_id.
                            # Triggered by the "更换监控对象" branch in
                            # MainWindow._on_edit when the IP/probe-type
                            # change is a repurposing rather than a node
                            # renumbering.  See data_store.wipe_target_history
                            # for the full rationale.
                            #
                            # Before wiping, flush ANY pending writes (ping
                            # AND alert) so a write that landed in the
                            # in-memory buffer just before the wipe doesn't
                            # INSERT after the DELETE -- the old `if
                            # alert_buf:` guard ignored ping_buf, so
                            # historic probes for the now-repurposed
                            # target survived the DELETE and reappeared
                            # on the next periodic flush, polluting the
                            # new target's history with ghost data from
                            # the previous one.
                            tid = data["target_id"]
                            if ping_buf or alert_buf:
                                if self._flush(conn, ping_buf, alert_buf):
                                    ping_buf.clear(); alert_buf.clear()
                                    last_flush = time.time()
                            # If the flush above failed (DB locked / disk
                            # full / etc.), _flush DELIBERATELY preserves
                            # the buffers for retry on the next cycle
                            # (see _flush docstring).  But we are about
                            # to DELETE every historical row for `tid` —
                            # any buffered row for THAT target must die
                            # too, or the next successful periodic flush
                            # would INSERT them right back, polluting the
                            # repurposed target's history with ghost data
                            # from the previous one.  Other targets'
                            # buffered rows still retry as intended.
                            ping_buf[:]  = [p for p in ping_buf
                                            if p.get("target_id") != tid]
                            alert_buf[:] = [a for a in alert_buf
                                            if a.get("target_id") != tid]
                            for table in ("pings_raw", "pings_hourly",
                                          "alert_events", "traceroute_logs",
                                          "icmp_diagnostic_history",
                                          "dns_diagnostic_history",
                                          "cum_stats", "webhook_ack"):
                                try:
                                    conn.execute(
                                        f"DELETE FROM {table} WHERE target_id=?",
                                        (tid,))
                                except Exception as e:
                                    # Per-table tolerance: one missing/locked
                                    # table shouldn't abort the whole wipe.
                                    print(f"[wipe_target] {table}: {e}")
                            conn.commit()
                            # Pop in-memory incident pointer so the next
                            # red transition for this tid opens a fresh
                            # incident rather than re-binding to a stale id
                            # from the (now-deleted) prior fault history.
                            # Safe to mutate from this thread -- write
                            # thread is the sole writer of _active_incidents.
                            self._active_incidents.pop(tid, None)
                            cb = self._wipe_complete_callback
                            if cb is not None:
                                try:
                                    cb(tid)
                                except Exception:
                                    pass

                        elif kind == "diag":
                            # ICMP advanced-diagnostic result.  Inserted
                            # immediately (not batched) because results are
                            # rare and a stale UI is very visible: history
                            # window opened right after a probe should see it.
                            # Final-stop generation guard -- see the
                            # "traceroute" branch above for the full
                            # rationale.  target_id is the first tuple
                            # element of the INSERT payload.
                            row_data, captured_gen = data
                            target_id_in_msg = row_data[0]
                            if (captured_gen is None
                                or target_id_in_msg is None
                                or self.current_target_generation(target_id_in_msg)
                                    == captured_gen):
                                conn.execute(
                                    "INSERT INTO icmp_diagnostic_history("
                                    "target_id,host,mode,payload_size,df,count,"
                                    "success_count,loss_count,avg_rtt_ms,"
                                    "min_rtt_ms,max_rtt_ms,estimated_path_mtu,"
                                    "max_df_payload,error_summary,conclusion,"
                                    "details_json,created_ts) "
                                    "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                    row_data)
                                conn.commit()

                        elif kind == "dns_diag":
                            # DNS diagnostic result.  Same immediate-insert
                            # pattern as ICMP "diag" — opening 📜 right after
                            # a probe must see the new row.  Identical
                            # generation guard.
                            row_data, captured_gen = data
                            target_id_in_msg = row_data[0]
                            if (captured_gen is None
                                or target_id_in_msg is None
                                or self.current_target_generation(target_id_in_msg)
                                    == captured_gen):
                                conn.execute(
                                    "INSERT INTO dns_diagnostic_history("
                                    "target_id,domain,qtype,trace_chain,resolver,"
                                    "success,rcode,elapsed_ms,answer_count,"
                                    "error_summary,message,details_json,created_ts) "
                                    "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                                    row_data)
                                conn.commit()

                        elif kind == "_flush_sync":
                            # Whether _flush succeeded or not, signal the
                            # waiter so flush() doesn't block for 15s.
                            # On failure the buffers stay so a subsequent
                            # periodic flush / _shutdown can still try.
                            if self._flush(conn, ping_buf, alert_buf):
                                ping_buf.clear(); alert_buf.clear()
                                last_flush = time.time()
                            if data: data.set()

                        elif kind == "_shutdown":
                            # Best-effort final flush; on failure the data is
                            # lost (process is exiting) but we still close
                            # the connection cleanly and unblock shutdown().
                            self._flush(conn, ping_buf, alert_buf)
                            conn.close()
                            if data: data.set()
                            return

                    now = time.time()

                    if now - last_flush >= self.FLUSH_INTERVAL:
                        if ping_buf or alert_buf:
                            if self._flush(conn, ping_buf, alert_buf):
                                ping_buf.clear(); alert_buf.clear()
                            # On failure: keep last_flush stale so we try again
                            # on the next loop tick (5s) rather than waiting
                            # another full FLUSH_INTERVAL (60s).
                            else:
                                last_flush = now - (self.FLUSH_INTERVAL - 5)
                                continue
                        last_flush = now

                    if now - last_hourly >= self.HOURLY_AGG_INTERVAL:
                        if ping_buf or alert_buf:
                            if self._flush(conn, ping_buf, alert_buf):
                                ping_buf.clear(); alert_buf.clear()
                                last_flush = now
                        self._hourly_agg(conn)  # routine: no retention-wide gap scan
                        last_hourly = now

                    if now - last_cleanup >= 86400:
                        # MUST run the hourly aggregation BEFORE the
                        # cleanup that deletes expired pings_raw rows.
                        # Without this ordering, the very first
                        # post-startup iteration fires _cleanup (because
                        # last_cleanup was initialised to now-86400 to
                        # trigger immediately) but skips _hourly_agg
                        # (because last_hourly was initialised to now),
                        # so any unsummed raw rows older than the raw-
                        # retention window get DELETE'd without ever
                        # being folded into pings_hourly -- the SLA /
                        # waveform record for that window is then
                        # permanently lost.  Re-aggregating here is a
                        # cheap no-op when there's nothing pending, so
                        # making it unconditional before every cleanup
                        # closes the gap without complicating the
                        # periodic-aggregation logic.
                        if ping_buf or alert_buf:
                            if self._flush(conn, ping_buf, alert_buf):
                                ping_buf.clear(); alert_buf.clear()
                                last_flush = now
                        self._hourly_agg(conn, backfill_internal_gaps=True)
                        last_hourly = now
                        self._cleanup(conn)
                        last_cleanup = now

                    if now - last_vacuum >= self.VACUUM_INTERVAL:
                        try:
                            conn.execute("PRAGMA incremental_vacuum(2000)")
                            conn.commit()
                        except Exception:
                            pass
                        last_vacuum = now

            except Exception as e:
                # Transient error (disk full, DB locked, etc.) — log and retry
                # after 5s. Thread stays alive; queued data is preserved.
                import traceback
                print(f"[DataStore] write loop error, retrying in 5s: {e}")
                traceback.print_exc()
                try: time.sleep(5)
                except Exception: pass
                # outer while True will restart the inner loop

        # Only reached if outer while somehow ends — close connection
        try: conn.close()
        except Exception: pass

    def _init_schema(self, conn: sqlite3.Connection):
        ver = conn.execute("PRAGMA user_version").fetchone()[0]
        if ver < 1:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS pings_raw (
                    id             INTEGER PRIMARY KEY,
                    target_id      TEXT    NOT NULL,
                    ts             REAL    NOT NULL,
                    status         TEXT    NOT NULL,
                    latency_ms     REAL,
                    loss_rate      REAL,
                    probe_success  INTEGER,
                    status_code    INTEGER,
                    failure_reason TEXT,
                    ping_type      TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_raw_target_ts
                    ON pings_raw(target_id, ts);

                CREATE TABLE IF NOT EXISTS pings_hourly (
                    id        INTEGER PRIMARY KEY,
                    target_id TEXT    NOT NULL,
                    hour_ts   INTEGER NOT NULL,
                    sample_n  INTEGER NOT NULL,
                    success_n INTEGER NOT NULL,
                    lat_avg   REAL,
                    lat_max   REAL,
                    lat_min   REAL,
                    lat_p95   REAL,
                    loss_rate REAL
                );
                CREATE UNIQUE INDEX IF NOT EXISTS idx_hourly_target_hour
                    ON pings_hourly(target_id, hour_ts);

                CREATE TABLE IF NOT EXISTS alert_events (
                    id             INTEGER PRIMARY KEY,
                    target_id      TEXT    NOT NULL,
                    label          TEXT,
                    ip             TEXT,
                    ts             REAL    NOT NULL,
                    old_status     TEXT,
                    new_status     TEXT,
                    category       TEXT,
                    failure_reason TEXT,
                    incident_id    TEXT,
                    ping_type      TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_alert_target_ts
                    ON alert_events(target_id, ts);
                CREATE INDEX IF NOT EXISTS idx_alert_incident
                    ON alert_events(incident_id);

                -- Single-column timestamp indexes for cleanup DELETEs.
                -- The composite (target_id, ts) indexes above CANNOT serve
                -- "WHERE ts < ?" without a target_id prefix (SQLite would
                -- fall back to a full table scan or skip-scan), which made
                -- weekly cleanup hold a long writer lock once the tables
                -- grew to millions of rows.
                CREATE INDEX IF NOT EXISTS idx_raw_ts_only
                    ON pings_raw(ts);
                CREATE INDEX IF NOT EXISTS idx_hourly_ts_only
                    ON pings_hourly(hour_ts);
                CREATE INDEX IF NOT EXISTS idx_alert_ts_only
                    ON alert_events(ts);

                CREATE TABLE IF NOT EXISTS targets_meta (
                    target_id  TEXT PRIMARY KEY,
                    label      TEXT,
                    ip         TEXT,
                    ping_type  TEXT,
                    updated_at INTEGER
                );

                CREATE TABLE IF NOT EXISTS cum_stats (
                    target_id  TEXT PRIMARY KEY,
                    total      INTEGER DEFAULT 0,
                    success    INTEGER DEFAULT 0,
                    lat_avg    REAL    DEFAULT 0,
                    lat_n      INTEGER DEFAULT 0,
                    updated_at INTEGER
                );

                CREATE TABLE IF NOT EXISTS traceroute_logs (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    target_id   TEXT    NOT NULL,
                    label       TEXT,
                    ip          TEXT,
                    ts          REAL    NOT NULL,
                    hops        TEXT,
                    incident_id TEXT
                );
                -- Index for incident<->traceroute time-window correlation
                -- (export_excel.get_incident_traceroute time-window lookup).
                CREATE INDEX IF NOT EXISTS idx_tracert_target_ts
                    ON traceroute_logs(target_id, ts);
                -- Index for exact incident_id JOIN (Level 2 forensic binding).
                CREATE INDEX IF NOT EXISTS idx_tracert_incident
                    ON traceroute_logs(incident_id);

                CREATE TABLE IF NOT EXISTS icmp_diagnostic_history (
                    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
                    target_id          TEXT,
                    host               TEXT    NOT NULL,
                    mode               TEXT    NOT NULL,
                    payload_size       INTEGER NOT NULL,
                    df                 INTEGER NOT NULL,
                    count              INTEGER NOT NULL,
                    success_count      INTEGER NOT NULL,
                    loss_count         INTEGER NOT NULL,
                    avg_rtt_ms         REAL,
                    min_rtt_ms         REAL,
                    max_rtt_ms         REAL,
                    estimated_path_mtu INTEGER,
                    max_df_payload     INTEGER,
                    error_summary      TEXT,
                    conclusion         TEXT,
                    details_json       TEXT,
                    created_ts         REAL    NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_diag_target_ts
                    ON icmp_diagnostic_history(target_id, created_ts DESC);
                CREATE INDEX IF NOT EXISTS idx_diag_host_ts
                    ON icmp_diagnostic_history(host, created_ts DESC);

                CREATE TABLE IF NOT EXISTS dns_diagnostic_history (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    target_id     TEXT,
                    domain        TEXT    NOT NULL,
                    qtype         TEXT    NOT NULL,
                    trace_chain   INTEGER NOT NULL,
                    resolver      TEXT,
                    success       INTEGER NOT NULL,
                    rcode         TEXT,
                    elapsed_ms    REAL,
                    answer_count  INTEGER NOT NULL,
                    error_summary TEXT,
                    message       TEXT,
                    details_json  TEXT,
                    created_ts    REAL    NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_dns_diag_target_ts
                    ON dns_diagnostic_history(target_id, created_ts DESC);
                CREATE INDEX IF NOT EXISTS idx_dns_diag_domain_ts
                    ON dns_diagnostic_history(domain, created_ts DESC);

                PRAGMA user_version = 7;
            """)
            conn.commit()

        # Fresh CREATE sets user_version in SQL but leaves the Python
        # `ver` variable stale (still 0); re-read so v7→v10 migrations run.
        ver = conn.execute("PRAGMA user_version").fetchone()[0]

        # ── Incremental migrations ──────────────────────────────────────
        # v1 → v2: add incident_id and failure_reason to alert_events
        if ver == 1:
            existing = {r[1] for r in conn.execute(
                "PRAGMA table_info(alert_events)").fetchall()}
            if 'incident_id' not in existing:
                conn.execute(
                    "ALTER TABLE alert_events ADD COLUMN incident_id TEXT")
                conn.execute(
                    "CREATE INDEX IF NOT EXISTS idx_alert_incident "
                    "ON alert_events(incident_id)")
            if 'failure_reason' not in existing:
                conn.execute(
                    "ALTER TABLE alert_events ADD COLUMN failure_reason TEXT")
            conn.execute("PRAGMA user_version = 2")
            conn.commit()
            # Back-fill incident_id for historical events
            self._backfill_incident_ids(conn)
            ver = 2

        # v2 → v3: add icmp_diagnostic_history table (introduced for
        # ICMP advanced diagnostics Phase 3 — store-and-recall of past
        # diagnostic runs).  CREATE TABLE IF NOT EXISTS is idempotent so
        # re-running this on a freshly-built v3 schema is a safe no-op.
        if ver == 2:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS icmp_diagnostic_history (
                    id                 INTEGER PRIMARY KEY AUTOINCREMENT,
                    target_id          TEXT,
                    host               TEXT    NOT NULL,
                    mode               TEXT    NOT NULL,
                    payload_size       INTEGER NOT NULL,
                    df                 INTEGER NOT NULL,
                    count              INTEGER NOT NULL,
                    success_count      INTEGER NOT NULL,
                    loss_count         INTEGER NOT NULL,
                    avg_rtt_ms         REAL,
                    min_rtt_ms         REAL,
                    max_rtt_ms         REAL,
                    estimated_path_mtu INTEGER,
                    max_df_payload     INTEGER,
                    error_summary      TEXT,
                    conclusion         TEXT,
                    details_json       TEXT,
                    created_ts         REAL    NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_diag_target_ts
                    ON icmp_diagnostic_history(target_id, created_ts DESC);
                CREATE INDEX IF NOT EXISTS idx_diag_host_ts
                    ON icmp_diagnostic_history(host, created_ts DESC);
                PRAGMA user_version = 3;
            """)
            conn.commit()
            ver = 3

        # v3 → v4: add dns_diagnostic_history table (DNS advanced
        # diagnostics Phase 3A — store-and-recall of past DNS resolution
        # runs, mirror of icmp_diagnostic_history).  Shares the
        # db_diag_retention_days knob with the ICMP table.  CREATE TABLE
        # IF NOT EXISTS is idempotent so re-running this on a freshly
        # built v4 schema is a safe no-op.
        if ver == 3:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS dns_diagnostic_history (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    target_id     TEXT,
                    domain        TEXT    NOT NULL,
                    qtype         TEXT    NOT NULL,
                    trace_chain   INTEGER NOT NULL,
                    resolver      TEXT,
                    success       INTEGER NOT NULL,
                    rcode         TEXT,
                    elapsed_ms    REAL,
                    answer_count  INTEGER NOT NULL,
                    error_summary TEXT,
                    message       TEXT,
                    details_json  TEXT,
                    created_ts    REAL    NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_dns_diag_target_ts
                    ON dns_diagnostic_history(target_id, created_ts DESC);
                CREATE INDEX IF NOT EXISTS idx_dns_diag_domain_ts
                    ON dns_diagnostic_history(domain, created_ts DESC);
                PRAGMA user_version = 4;
            """)
            conn.commit()
            ver = 4

        # v4 → v5: add incident_id column to traceroute_logs for exact
        # incident<->snapshot binding (Level 2 forensic correlation).
        # Old rows keep incident_id = NULL; export_excel automatically
        # falls back to the time-window match (Level 1) for those, so
        # historical reports remain complete.  CREATE INDEX is idempotent
        # via IF NOT EXISTS so re-runs on a freshly-built v5 schema are
        # a safe no-op.  PRAGMA table_info guards the ALTER against
        # double-add if a previous run partially completed.
        # Note: also (re)creates idx_tracert_target_ts here — it was
        # introduced alongside Level 1 in the fresh-DB CREATE block but
        # never given its own migration step, so a pre-Level-1 v4
        # database upgrading straight to v5 would otherwise miss it.
        if ver == 4:
            existing = {r[1] for r in conn.execute(
                "PRAGMA table_info(traceroute_logs)").fetchall()}
            if 'incident_id' not in existing:
                conn.execute(
                    "ALTER TABLE traceroute_logs ADD COLUMN incident_id TEXT")
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_tracert_target_ts "
                "ON traceroute_logs(target_id, ts)")
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_tracert_incident "
                "ON traceroute_logs(incident_id)")
            conn.execute("PRAGMA user_version = 5")
            conn.commit()
            ver = 5

        # v5 → v6: add lat_min column to pings_hourly so the waveform
        # summary's "min latency" reflects the real per-hour minimum
        # instead of min(of hourly averages), which previously masked
        # short low-latency dips when looking at multi-day spans.
        # Pre-migration rows keep lat_min = NULL; get_waveform's summary
        # tolerates None entries (those points are skipped), so old
        # historical ranges show min only from post-migration hours.
        if ver == 5:
            existing = {r[1] for r in conn.execute(
                "PRAGMA table_info(pings_hourly)").fetchall()}
            if 'lat_min' not in existing:
                conn.execute(
                    "ALTER TABLE pings_hourly ADD COLUMN lat_min REAL")
            conn.execute("PRAGMA user_version = 6")
            conn.commit()
            ver = 6

        # v6 → v7: add single-column timestamp indexes so weekly cleanup
        # ("DELETE FROM pings_raw WHERE ts<?", etc.) doesn't fall back to
        # a full table scan.  The existing composite (target_id, ts)
        # indexes can't serve a ts-only range predicate efficiently;
        # without these, a multi-million-row pings_raw cleanup could
        # hold the writer lock for tens of seconds and back up live
        # probe writes in memory.
        if ver == 6:
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_raw_ts_only ON pings_raw(ts)")
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_hourly_ts_only ON pings_hourly(hour_ts)")
            conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_alert_ts_only ON alert_events(ts)")
            conn.execute("PRAGMA user_version = 7")
            conn.commit()
            ver = 7

        # v7 → v8: alert_events.ts stored as REAL (sub-second precision).
        # Existing INTEGER columns accept REAL values in SQLite; new
        # installs get REAL NOT NULL from CREATE TABLE above.
        if ver == 7:
            conn.execute("PRAGMA user_version = 8")
            conn.commit()
            ver = 8

        # v8 → v9: pings_raw.ts stored as REAL (sub-second precision).
        # Existing INTEGER columns accept REAL values in SQLite; new
        # installs get REAL NOT NULL from CREATE TABLE above.
        if ver == 8:
            conn.execute("PRAGMA user_version = 9")
            conn.commit()
            ver = 9

        # v9 → v10: per-probe success flag (distinct from alert status).
        if ver == 9:
            cols = {r[1] for r in conn.execute(
                "PRAGMA table_info(pings_raw)").fetchall()}
            if "probe_success" not in cols:
                conn.execute(
                    "ALTER TABLE pings_raw ADD COLUMN probe_success INTEGER")
            conn.execute("PRAGMA user_version = 10")
            conn.commit()
            ver = 10

        # v10 → v11: rebuild cum_stats from pings_raw (Bug91).
        # Pre-Bug90 builds could persist failed-probe RTT into lat_avg/lat_n.
        # Recompute from the raw rows still in retention so Web restart loads
        # probe_success-correct cumulative latency instead of stale pollution.
        if ver == 10:
            self._rebuild_cum_stats_from_raw(conn)
            conn.execute("PRAGMA user_version = 11")
            conn.commit()
            ver = 11

        # v11 → v12: alert_events.ping_type for per-event probe context.
        if ver == 11:
            cols = {r[1] for r in conn.execute(
                "PRAGMA table_info(alert_events)").fetchall()}
            if "ping_type" not in cols:
                conn.execute(
                    "ALTER TABLE alert_events ADD COLUMN ping_type TEXT")
            self._backfill_alert_ping_types(conn)
            conn.execute("PRAGMA user_version = 12")
            conn.commit()
            ver = 12

        # v12 → v13: backfill NULL alert_events.ping_type from targets_meta
        # for DBs that upgraded to v12 before the backfill existed (Bug 154).
        if ver == 12:
            self._backfill_alert_ping_types(conn)
            conn.execute("PRAGMA user_version = 13")
            conn.commit()
            ver = 13

        # v13 → v14: persist webhook incident ACK across restarts (Bug 156).
        if ver == 13:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS webhook_ack (
                    incident_id TEXT PRIMARY KEY,
                    target_id   TEXT NOT NULL,
                    ack_ts      REAL NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_webhook_ack_target
                    ON webhook_ack(target_id);
            """)
            conn.execute("PRAGMA user_version = 14")
            conn.commit()
            ver = 14

        # v14 → v15: reliable webhook outbox (persistent delivery queue).
        if ver == 14:
            conn.executescript("""
                CREATE TABLE IF NOT EXISTS webhook_outbox (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    delivery_id TEXT UNIQUE NOT NULL,
                    target_id TEXT,
                    incident_id TEXT,
                    incident_seq INTEGER,
                    event TEXT NOT NULL,
                    order_key TEXT NOT NULL,
                    payload_json TEXT NOT NULL,
                    event_ts REAL NOT NULL,
                    first_queued_ts REAL NOT NULL,
                    next_attempt_ts REAL NOT NULL,
                    last_attempt_ts REAL,
                    delivered_ts REAL,
                    attempt_count INTEGER NOT NULL DEFAULT 0,
                    max_attempts INTEGER NOT NULL DEFAULT 0,
                    delivery_state TEXT NOT NULL,
                    last_error TEXT,
                    created_at REAL NOT NULL,
                    updated_at REAL NOT NULL
                );
                CREATE INDEX IF NOT EXISTS idx_webhook_outbox_due
                    ON webhook_outbox(delivery_state, next_attempt_ts);
                CREATE INDEX IF NOT EXISTS idx_webhook_outbox_order
                    ON webhook_outbox(order_key, first_queued_ts, id);
                CREATE INDEX IF NOT EXISTS idx_webhook_outbox_incident
                    ON webhook_outbox(incident_id);
                CREATE INDEX IF NOT EXISTS idx_webhook_outbox_target
                    ON webhook_outbox(target_id);
            """)
            conn.execute("PRAGMA user_version = 15")
            conn.commit()
            ver = 15

        # Restore in-memory active-incident state from DB
        self._restore_active_incidents(conn)

    @staticmethod
    def _resolve_alert_ping_type(conn: sqlite3.Connection, target_id: str,
                                 alert_ping_type: Optional[str]) -> str:
        """Resolve probe type for an alert row.

        Prefer the per-event ping_type when present (v12+ writes).  Legacy
        rows migrated from v11 keep NULL — fall back to targets_meta so
        HTTP/TCP/DNS open incidents are not misclassified as ICMP on restart.
        """
        pt = (alert_ping_type or "").strip().lower()
        if pt:
            return pt
        try:
            row = conn.execute(
                "SELECT ping_type FROM targets_meta WHERE target_id=?",
                (target_id,)).fetchone()
            if row and (row[0] or "").strip():
                return (row[0] or "").strip().lower()
        except Exception:
            pass
        return "icmp"

    def _backfill_alert_ping_types(self, conn: sqlite3.Connection) -> None:
        """Fill NULL alert_events.ping_type from targets_meta (v11 legacy)."""
        cols = {r[1] for r in conn.execute(
            "PRAGMA table_info(alert_events)").fetchall()}
        if "ping_type" not in cols:
            return
        conn.execute("""
            UPDATE alert_events
            SET ping_type = (
                SELECT tm.ping_type FROM targets_meta tm
                WHERE tm.target_id = alert_events.target_id
            )
            WHERE ping_type IS NULL
              AND EXISTS (
                SELECT 1 FROM targets_meta tm
                WHERE tm.target_id = alert_events.target_id
                  AND tm.ping_type IS NOT NULL AND tm.ping_type != ''
              )
        """)

    def _backfill_incident_ids(self, conn: sqlite3.Connection):
        """Assign incident_ids to historical alert_events that lack them."""
        rows = conn.execute(
            "SELECT id, target_id, ts, old_status, new_status, category "
            "FROM alert_events WHERE incident_id IS NULL "
            "ORDER BY target_id, ts, id"
        ).fetchall()

        active: dict[str, str] = {}   # target_id → current incident_id
        updates = []
        from datetime import datetime as _dt

        for row in rows:
            # Renamed `os` -> `old_st` to avoid shadowing the imported `os`
            # module at the top of this file. Functionality unchanged.
            rid, tid, ts, old_st, ns, cat = row
            if cat in ('paused', 'resumed'):
                iid = active.get(tid)  # inherit incident if active
                if iid:
                    updates.append((iid, rid))
                # Pause terminates the open incident (operator silence).
                if cat == 'paused' or ns == 'paused':
                    active.pop(tid, None)
                continue

            iid = active.get(tid)
            if ns in ('orange', 'red'):
                if not iid:
                    # Match live _assign_incident_ids semantics: second-only
                    # keys collapse independent same-second outages on one
                    # target.  Use row id suffix so backfill is stable and
                    # each new outage gets a distinct incident_id.
                    ts_str = _dt.fromtimestamp(ts).strftime('%Y%m%d%H%M%S')
                    iid = f"INC-{ts_str}-{tid[:6].upper()}-{rid:08X}"
                    active[tid] = iid
                updates.append((iid, rid))
            elif ns == 'green':
                if iid:
                    updates.append((iid, rid))
                    del active[tid]
            # gray/others: no incident_id

        if updates:
            conn.executemany(
                "UPDATE alert_events SET incident_id=? WHERE id=?", updates)
            conn.commit()

    def _restore_active_incidents(self, conn: sqlite3.Connection):
        """
        On startup, restore in-memory _active_incidents from the last DB state
        per target so that incident continuity is preserved across restarts.
        """
        rows = conn.execute("""
            SELECT target_id, incident_id, new_status
            FROM alert_events
            WHERE id IN (
                SELECT MAX(id) FROM alert_events GROUP BY target_id
            )
            AND incident_id IS NOT NULL
            AND new_status IN ('red', 'orange')
        """).fetchall()
        for tid, iid, _ in rows:
            self._active_incidents[tid] = iid

    def _flush(self, conn, ping_buf: list, alert_buf: list) -> bool:
        """Persist all buffered probe and alert events to SQLite.

        Returns:
            True  — all rows committed; caller may clear the buffers.
            False — transaction was rolled back; caller MUST preserve the
                    buffers so the same batch is retried on the next flush
                    cycle.  In-memory _active_incidents is restored to its
                    pre-call state so the retry produces identical
                    incident_id assignments.

        Why the snapshot/rollback matters:
            _assign_incident_ids() mutates self._active_incidents BEFORE the
            INSERT runs (it needs the side-effects to thread the same
            incident_id across the events of one batch).  If the INSERT
            then fails, SQLite rolls back the transaction — but the
            in-memory dict has already been advanced.  Without restoring
            it, a green event that "closed" an incident on the first
            attempt would no longer find that incident on retry, causing
            the recovery row to be persisted with a NULL incident_id and
            permanently breaking the incident's open→close pairing.
        """
        if not ping_buf and not alert_buf:
            return True
        # Final generation pass over ping_buf: a ping that survived the
        # enqueue-time check could still have been overtaken by a wipe
        # / removal while sitting in the buffer (ping_buf flushes at
        # FLUSH_INTERVAL, so a stale row can wait up to ~60 s here).
        # Filter in-place so the executemany below only sees rows
        # whose captured generation still matches the current target
        # generation.  Rows with no captured_generation (legacy /
        # detached test paths) pass through unchanged.
        if ping_buf:
            kept = []
            for p in ping_buf:
                cg = p.get("captured_generation")
                tid = p.get("target_id")
                if (cg is None or tid is None
                    or self.current_target_generation(tid) == cg):
                    kept.append(p)
            ping_buf[:] = kept
        # Same final pass over alert_buf -- a stale recovery event
        # that snuck through the enqueue-time guard (or pre-dates the
        # guard for any reason) gets dropped here rather than landing
        # as a phantom red->green transition.  Events synthesised by
        # the writer thread itself (target_removed branch above)
        # carry no captured_generation key and pass through unchanged.
        if alert_buf:
            kept_a = []
            for a in alert_buf:
                cg = a.get("captured_generation")
                tid = a.get("target_id")
                if (cg is None or tid is None
                    or self.current_target_generation(tid) == cg):
                    kept_a.append(a)
            alert_buf[:] = kept_a
        if not ping_buf and not alert_buf:
            return True
        # Shallow snapshot is sufficient: values are immutable str tokens.
        incidents_snapshot = dict(self._active_incidents)
        try:
            with conn:
                if ping_buf:
                    conn.executemany(
                        "INSERT INTO pings_raw "
                        "(target_id,ts,status,latency_ms,loss_rate,"
                        " probe_success,status_code,failure_reason,ping_type) "
                        "VALUES(:target_id,:ts,:status,:latency_ms,:loss_rate,"
                        ":probe_success,:status_code,:failure_reason,:ping_type)",
                        ping_buf)
                    self._update_cum_stats(conn, ping_buf)
                    # 更新 targets_meta 快照（只保留每个 tid 的最新值）
                    seen = {p["target_id"]: p for p in ping_buf}
                    now  = int(time.time())
                    for p in seen.values():
                        conn.execute(
                            "INSERT OR REPLACE INTO targets_meta "
                            "(target_id,label,ip,ping_type,updated_at) "
                            "VALUES(?,?,?,?,?)",
                            (p["target_id"], p.get("label",""),
                             p.get("ip",""), p.get("ping_type","icmp"), now))
                if alert_buf:
                    self._assign_incident_ids(alert_buf)
                    conn.executemany(
                        "INSERT INTO alert_events "
                        "(target_id,label,ip,ts,old_status,new_status,"
                        " category,failure_reason,incident_id,ping_type) "
                        "VALUES(:target_id,:label,:ip,:ts,"
                        ":old_status,:new_status,:category,"
                        ":failure_reason,:incident_id,:ping_type)",
                        alert_buf)
            return True
        except Exception as e:
            # Restore incident state so the retry path is deterministic.
            # Mutate in place rather than rebinding so any external code
            # holding a reference to the original dict (unlikely but
            # cheap insurance) still sees the rolled-back state.
            self._active_incidents.clear()
            self._active_incidents.update(incidents_snapshot)
            print(f"[DataStore] _flush error (buffers preserved for retry): {e}")
            return False

    def _assign_incident_ids(self, alert_buf: list):
        """
        Assign incident_id to each event dict in alert_buf (in-place).

        Rules:
        - orange / red  → start or continue an incident
        - green         → close an incident (recovery event gets same ID)
        - paused        → tag the closing incident on the pause row, then
                          drop the active pointer (operator silence ends the
                          open incident for reseed / traceroute / duration)
        - gray / resumed → transitional; inherit incident_id only if still
                          active (normally None after pause)

        _active_incidents persists across flushes so incident continuity
        survives program restarts (restored from DB in _init_schema).
        """
        from datetime import datetime as _dt

        # Process in chronological order so multi-event flushes stay consistent
        for ev in sorted(alert_buf, key=lambda e: e['ts']):
            tid = ev['target_id']
            ns  = ev.get('new_status', '')
            cat = ev.get('category', '')
            iid = self._active_incidents.get(tid)

            if ns in ('orange', 'red'):
                if not iid:
                    # The (ts_str, tid[:6]) key collapses to the same
                    # value when two independent outages happen in the
                    # SAME WALL-CLOCK SECOND on the SAME target -- a
                    # green→red→green→red→green flap within one
                    # second produced four events that all shared one
                    # incident_id, breaking the Level-2 traceroute
                    # join (forensic snapshots for the second outage
                    # would be matched against the first).  Append 4
                    # hex chars of randomness so distinct incidents
                    # always get distinct ids; collision probability
                    # within a single (second, target) bucket is
                    # 1/65536 even before the (very rare) third
                    # outage in the same second.
                    ts_str = _dt.fromtimestamp(ev['ts']).strftime('%Y%m%d%H%M%S')
                    rand4  = os.urandom(2).hex().upper()
                    iid = f"INC-{ts_str}-{tid[:6].upper()}-{rand4}"
                    self._active_incidents[tid] = iid
                ev['incident_id'] = iid

            elif ns == 'green':
                if iid:
                    ev['incident_id'] = iid          # recovery closes incident
                    del self._active_incidents[tid]
                else:
                    ev['incident_id'] = None          # normal green→green

            elif ns == 'paused' or cat == 'paused':
                ev['incident_id'] = iid              # tag pause to prior incident
                if iid:
                    del self._active_incidents[tid]

            elif ns in ('gray',) or cat in ('resumed',):
                ev['incident_id'] = iid              # None if not in incident

            else:
                ev['incident_id'] = None

    def _rebuild_cum_stats_from_raw(self, conn) -> None:
        """Rebuild cum_stats from pings_raw via SQL aggregate (Bug91/93).

        One-shot correction for databases polluted before Bug90 (failed RTT
        counted in lat_avg/lat_n).  Uses GROUP BY in SQLite so v11 migration
        does not fetchall() millions of raw rows into Python on upgrade.
        """
        conn.execute("DELETE FROM cum_stats")
        now = int(time.time())
        conn.execute(
            f"INSERT INTO cum_stats "
            f"(target_id,total,success,lat_avg,lat_n,updated_at) "
            f"SELECT target_id, "
            f"COUNT(*), "
            f"SUM(CASE WHEN {_PROBE_OK_SQL} = 1 THEN 1 ELSE 0 END), "
            f"AVG(CASE WHEN {_PROBE_LAT_OK_SQL} THEN latency_ms END), "
            f"SUM(CASE WHEN {_PROBE_LAT_OK_SQL} THEN 1 ELSE 0 END), "
            f"? "
            f"FROM pings_raw GROUP BY target_id",
            (now,))

    def _update_cum_stats(self, conn, ping_buf: list):
        """用 CMA 公式更新累计统计——与 WebServer 端保持一致的数值稳定性。"""
        groups: dict[str, list] = defaultdict(list)
        for p in ping_buf:
            groups[p["target_id"]].append(p)

        for tid, pings in groups.items():
            row = conn.execute(
                "SELECT total, success, lat_avg, lat_n FROM cum_stats "
                "WHERE target_id=?", (tid,)).fetchone()
            total   = row[0] if row else 0
            success = row[1] if row else 0
            lat_avg = row[2] if row else 0.0
            lat_n   = row[3] if row else 0

            for p in pings:
                total += 1
                probe_ok = self.row_probe_success(
                    p.get("latency_ms"), p.get("status", "green"),
                    p.get("failure_reason"), p.get("probe_success"))
                if probe_ok:
                    success += 1
                if probe_ok and p["latency_ms"] is not None:
                    lat_n   += 1
                    lat_avg += (p["latency_ms"] - lat_avg) / lat_n

            conn.execute(
                "INSERT OR REPLACE INTO cum_stats "
                "(target_id,total,success,lat_avg,lat_n,updated_at) "
                "VALUES(?,?,?,?,?,?)",
                (tid, total, success, lat_avg, lat_n, int(time.time())))

    def _hourly_agg(self, conn, *, backfill_internal_gaps: bool = False):
        """Aggregate every hourly bucket we haven't already written.

        Previously hard-coded to "the immediately-prior full hour", which
        meant any hour the writer thread wasn't alive at minute :00 was
        skipped FOREVER -- once pings_raw cleanup ran (default 7 days
        later) and removed that hour's raw rows, the SLA report had a
        permanent hole in pings_hourly that no future agg could fill.
        Now, per target:
          - catch up every MISSING full hour from the per-target
            high-water mark up to (but not including) the current
            still-in-progress hour;
          - always recompute the most recent RECENT_RECOMPUTE_HOURS (3)
            completed hours (INSERT OR REPLACE) so pings that flush into
            pings_raw AFTER an hour bucket was first written are still
            folded in;
          - routine ticks never start earlier than the raw retention floor;
          - cleanup/backfill (``backfill_internal_gaps=True``) inserts only
            missing hourly buckets from the earliest raw hour still in the
            DB (Bug61) and re-touches the recent window for late flushes;
            it does not REPLACE existing old buckets with retention-edge
            partial raw (Bug62).

        TRADE-OFF (intentional): a raw row that flushes more than
        RECENT_RECOMPUTE_HOURS after its hour completed is NOT re-absorbed
        on the routine tick -- UNLESS that hour's bucket was missing
        entirely.  Internal gaps (raw present, hourly row missing, including
        hours before HWM) are rebuilt only when ``backfill_internal_gaps``
        is True -- daily cleanup / startup pre-cleanup -- via a DISTINCT+
        EXCEPT scan over raw retention.  Routine hourly ticks skip that
        scan so a caught-up 30-target DB does not re-walk ~7 days of raw
        rows every hour just to prove there are no gaps.

        Cost-idempotent: on a caught-up DB each routine tick processes at
        most RECENT_RECOMPUTE_HOURS buckets per target, and discovers
        targets only from raw rows in that recent window (not the full raw
        retention span).  Gap backfill and full target discovery run on the
        cleanup path (``backfill_internal_gaps=True``).
        """
        now          = int(time.time())
        current_hour = (now // 3600) * 3600   # exclude the in-progress hour
        # Routine floor: do not walk earlier than the retention window.
        # Cleanup/backfill uses per-target earliest raw hour instead (Bug61).
        recompute_floor = current_hour - int(self._raw_retention_days * 86400)
        # Routine recompute window: always re-touch the last few completed
        # hours so a late flush (lock contention / recovery / backfill) that
        # landed AFTER its hour was first aggregated still gets folded in.
        # This is the ONLY span re-scanned on a caught-up DB -- we do NOT
        # walk the whole retention window every tick (the old
        # `min(hour_start, recompute_floor)` did exactly that).  See the
        # method docstring for the late-flush trade-off this implies.
        RECENT_RECOMPUTE_HOURS = 3
        recent_floor = current_hour - RECENT_RECOMPUTE_HOURS * 3600

        # Targets from pings_raw (not targets_meta) so removed targets still
        # get trailing raw aggregated before cleanup.  Routine hourly only
        # scans the recent recompute window; cleanup/backfill scans all raw.
        if backfill_internal_gaps:
            target_ids = [r[0] for r in conn.execute(
                "SELECT DISTINCT target_id FROM pings_raw"
            ).fetchall()]
        else:
            target_ids = [r[0] for r in conn.execute(
                "SELECT DISTINCT target_id FROM pings_raw "
                "WHERE ts>=? AND ts<?",
                (recent_floor, current_hour),
            ).fetchall()]

        with conn:
            for tid in target_ids:
                # Per-target high-water mark.  Without this we would
                # rescan the entire raw retention window every hourly
                # tick to find missed hours; with it the work is
                # proportional to "hours since last successful agg"
                # (typically 1, occasionally a handful after a restart).
                hwm_row = conn.execute(
                    "SELECT MAX(hour_ts) FROM pings_hourly WHERE target_id=?",
                    (tid,)
                ).fetchone()
                hwm = hwm_row[0] if hwm_row else None

                if hwm is not None:
                    hour_start = hwm + 3600
                else:
                    # No prior aggregation: start from the earliest raw
                    # probe so a fresh target / first run after upgrade
                    # still gets its backlog aggregated.
                    first = conn.execute(
                        "SELECT MIN(ts) FROM pings_raw WHERE target_id=?",
                        (tid,)
                    ).fetchone()
                    if not first or first[0] is None:
                        continue
                    hour_start = (first[0] // 3600) * 3600

                first_raw = conn.execute(
                    "SELECT MIN(ts) FROM pings_raw WHERE target_id=?",
                    (tid,),
                ).fetchone()
                if (backfill_internal_gaps
                        and first_raw and first_raw[0] is not None):
                    agg_floor = (first_raw[0] // 3600) * 3600
                else:
                    agg_floor = recompute_floor

                hours_to_agg = set()
                if backfill_internal_gaps:
                    # Recent window: REPLACE to absorb true late flushes.
                    hour_start = max(recompute_floor,
                                     min(hour_start, recent_floor))
                    h = hour_start
                    while h < current_hour:
                        hours_to_agg.add(h)
                        h += 3600
                    # Rescue only MISSING buckets (raw present, hourly absent).
                    # Do not sweep agg_floor..current_hour — that would REPLACE
                    # complete old buckets with retention-edge partial raw.
                    gap_rows = conn.execute(
                        "SELECT hour_ts FROM ("
                        "  SELECT DISTINCT CAST(ts / 3600 AS INTEGER) * 3600 "
                        "         AS hour_ts "
                        "  FROM pings_raw "
                        "  WHERE target_id=? AND ts>=? AND ts<?"
                        ") EXCEPT "
                        "SELECT hour_ts FROM pings_hourly "
                        "WHERE target_id=? AND hour_ts>=? AND hour_ts<?",
                        (tid, agg_floor, current_hour,
                         tid, agg_floor, current_hour),
                    ).fetchall()
                    for (gap_h,) in gap_rows:
                        hours_to_agg.add(int(gap_h))
                else:
                    hour_start = max(agg_floor,
                                     min(hour_start, recent_floor))
                    h = hour_start
                    while h < current_hour:
                        hours_to_agg.add(h)
                        h += 3600

                for hour_ts in sorted(hours_to_agg):
                    if hour_ts >= current_hour:
                        continue
                    hour_end = hour_ts + 3600
                    rows = conn.execute(
                        "SELECT latency_ms, status, failure_reason, "
                        "probe_success FROM pings_raw "
                        "WHERE target_id=? AND ts >= ? AND ts < ?",
                        (tid, hour_ts, hour_end)
                    ).fetchall()
                    bucket = self._compute_hourly_bucket(rows)
                    if bucket:
                        # Old rescue buckets: insert-only.  Recent window:
                        # REPLACE for late-flush absorption.
                        if (backfill_internal_gaps
                                and hour_ts < recent_floor):
                            conn.execute(
                                "INSERT OR IGNORE INTO pings_hourly "
                                "(target_id,hour_ts,sample_n,success_n,"
                                " lat_avg,lat_max,lat_min,lat_p95,loss_rate) "
                                "VALUES(?,?,?,?,?,?,?,?,?)",
                                (tid, hour_ts,
                                 bucket["sample_n"], bucket["success_n"],
                                 bucket["lat_avg"], bucket["lat_max"],
                                 bucket["lat_min"], bucket["lat_p95"],
                                 bucket["loss_rate"]))
                        else:
                            conn.execute(
                                "INSERT OR REPLACE INTO pings_hourly "
                                "(target_id,hour_ts,sample_n,success_n,"
                                " lat_avg,lat_max,lat_min,lat_p95,loss_rate) "
                                "VALUES(?,?,?,?,?,?,?,?,?)",
                                (tid, hour_ts,
                                 bucket["sample_n"], bucket["success_n"],
                                 bucket["lat_avg"], bucket["lat_max"],
                                 bucket["lat_min"], bucket["lat_p95"],
                                 bucket["loss_rate"]))

    @staticmethod
    def _compute_outage_stats(start_ts: float, end_ts: float,
                              status_at_start: str, events,
                              now: float) -> tuple:
        """Return (outage_secs, outage_count) for one SLA window.

        Only counts outages with positive overlap inside [start_ts, end_ts].
        Zero-length boundary touches (recovery at window start, fault at
        window end) do not increment outage_count.
        """
        outage_secs = 0.0
        outage_count = 0
        fault_start = start_ts if status_at_start == "red" else None

        for ev in events:
            ts = ev["ts"]
            old_s = ev["old_status"]
            new_s = ev["new_status"]
            if new_s == "red" and fault_start is None:
                fault_start = ts
            elif (old_s == "red" and new_s != "red"
                  and fault_start is not None):
                dur = ts - fault_start
                if dur > 0:
                    outage_secs += dur
                    outage_count += 1
                fault_start = None

        if fault_start is not None:
            case_b_end = min(end_ts, now)
            dur = max(0.0, case_b_end - fault_start)
            if dur > 0:
                outage_secs += dur
                outage_count += 1

        return outage_secs, outage_count

    def get_dashboard_stats_batch(self, target_ids: list,
                                start_ts, end_ts) -> dict:
        """
        Per-target latency stats for Web /api/stats (total/success/avg/max/p95).

        Uses _bounded_hourly_buckets_for_targets + _rollup_hourly_buckets
        instead of fetching every pings_raw row per target into Python.
        """
        if not target_ids:
            return {}
        start_ts = int(start_ts)
        end_ts   = int(end_ts)
        CHUNK = 500
        if len(target_ids) > CHUNK:
            out = {}
            for i in range(0, len(target_ids), CHUNK):
                out.update(self.get_dashboard_stats_batch(
                    target_ids[i:i + CHUNK], start_ts, end_ts))
            return out
        empty = {
            "total": 0, "success": 0,
            "lat_avg": None, "lat_max": None, "lat_p95": None,
        }
        try:
            conn = self._read_conn()
            buckets_by_tid = self._bounded_hourly_buckets_for_targets(
                conn, target_ids, start_ts, end_ts)
            result = {}
            for tid in target_ids:
                merged = self._rollup_latency_for_window(
                    conn, tid, buckets_by_tid.get(tid, []), start_ts, end_ts)
                if not merged.get("total"):
                    result[tid] = dict(empty)
                    continue
                result[tid] = {
                    "total":   merged["total"],
                    "success": merged["success"],
                    "lat_avg": (round(merged["lat_avg"], 1)
                                if merged["lat_avg"] is not None else None),
                    "lat_max": (round(merged["lat_max"], 1)
                                if merged["lat_max"] is not None else None),
                    "lat_p95": (round(merged["lat_p95"], 1)
                                if merged["lat_p95"] is not None else None),
                }
            return result
        except Exception as e:
            print(f"[dashboard-stats-batch] {e}")
            return {tid: dict(empty) for tid in target_ids}

    def get_sla_stats(self, target_id: str,
                      start_ts: float, end_ts: float) -> dict:
        """
        Compute SLA metrics for one target over [start_ts, end_ts].

        Latency stats  — from pings_hourly plus the in-progress current
        hour aggregated from pings_raw (not written to pings_hourly).
        Outage stats   — from alert_events, status-transition based:

        The query now selects by old/new_status rather than category, so
        recovery events (which are stored with category='ok') are correctly
        included.  The three boundary cases are handled:
          Case A: outage started before start_ts → count from start_ts
          Case B: outage not resolved by end_ts  → count to end_ts
          Case C: node was red the ENTIRE period → whole span is downtime

        Outage definition: service is DOWN while new_status == 'red'.
        Recovery = first transition FROM red TO green/orange/gray.
        """
        import sqlite3, json
        conn = None
        try:
            conn = self._read_conn()
            conn.row_factory = sqlite3.Row

            # ── Latency stats from pings_hourly ─────────────────────
            # NOTE: weight by success_n, not sample_n.  lat_avg/lat_p95 in
            # pings_hourly are computed from non-null latencies only (see
            # _hourly_agg), so sample_n inflates the weight of failure-heavy
            # hours and pulls SLA latency toward whatever the 2 lucky probes
            # in a 3600-probe outage happened to clock.  NULLIF(SUM(success_n),0)
            # also gives the right zero-guard: if no probes succeeded in the
            # range, the latency average is genuinely undefined (NULL),
            # rather than dividing by a sample_n that includes only failures.
            buckets = self._bounded_hourly_buckets(
                conn, target_id, start_ts, end_ts)
            merged = self._rollup_latency_for_window(
                conn, target_id, buckets, start_ts, end_ts)

            total   = merged["total"]   or 0
            success = merged["success"] or 0
            # `... or 0.0` would coerce NULL (which means "no successful
            # probe in the range" — see the NULLIF(...) zero-guard in
            # the SQL above) to 0.0, making a hour of 100% packet loss
            # surface as "0.0 ms average".  Preserve None so consumers
            # can render "—" instead of an absurd zero.  The error-path
            # return below already uses None, so the calling shape is
            # unchanged for non-data hours.
            lat_avg = round(merged["lat_avg"], 2) if merged["lat_avg"] is not None else None
            lat_max = round(merged["lat_max"], 2) if merged["lat_max"] is not None else None
            lat_p95 = round(merged["lat_p95"], 2) if merged["lat_p95"] is not None else None

            # ── Outage stats from alert_events ───────────────────────
            # Query by STATUS TRANSITION, not by category.
            # Recovery events are stored with category='ok' (not 'availability'),
            # so a category filter would silently drop them — causing fault_start
            # to never be cleared and wildly overstating outage duration.
            #
            # Split into TWO bounded queries instead of fetching the
            # entire history with ts <= end_ts:
            #   Q1: the single most recent red-touching event strictly
            #       before start_ts → determines status_at_start.
            #   Q2: every red-touching event INSIDE [start_ts, end_ts]
            #       → walked to count outages and durations.
            # The previous one-query "ts <= end_ts" pulled every
            # red-touching event from the dawn of time to find Q1's
            # one row, which on a 3-year-old DB with 50k events meant
            # iterating 49,990 rows in Python just to determine the
            # boundary state.  Multiply by 50 nodes in the batch path
            # and a dashboard refresh was loading hundreds of MB of
            # alert history into Flask's worker memory per request.
            boundary = conn.execute("""
                SELECT new_status FROM alert_events
                WHERE target_id = ?
                  AND (new_status = 'red' OR old_status = 'red')
                  AND ts < ?
                ORDER BY ts DESC, id DESC LIMIT 1
            """, (target_id, start_ts)).fetchone()
            status_at_start = boundary["new_status"] if boundary else "green"

            events = conn.execute("""
                SELECT ts, old_status, new_status FROM alert_events
                WHERE target_id = ?
                  AND (new_status = 'red' OR old_status = 'red')
                  AND ts >= ? AND ts <= ?
                ORDER BY ts ASC, id ASC
            """, (target_id, start_ts, end_ts)).fetchall()

            now = time.time()
            outage_secs, outage_count = self._compute_outage_stats(
                float(start_ts), float(end_ts), status_at_start, events, now)

            # Connection is thread-local-cached — do not close here either.
            # Was previously closing mid-method then again in finally; both
            # are now removed.

            # uptime% denominator: use the smaller of (query window, elapsed so far)
            # so that a 7-day query made on day 1 isn't comparing against a full week.
            elapsed = min(end_ts, now) - start_ts
            period_secs = max(elapsed, 1)
            uptime_pct  = round(
                max(0.0, (period_secs - outage_secs) / period_secs) * 100, 3)
            return {
                "uptime_pct":     uptime_pct,
                "lat_avg":        lat_avg,
                "lat_max":        lat_max,
                "lat_p95":        lat_p95,
                "p95_from_raw":   bool(merged.get("p95_from_raw")),
                "outage_count":   outage_count,
                "outage_minutes": round(outage_secs / 60, 1),
                "sample_n":       total,
            }
        except Exception as e:
            print(f"[SLA] {e}")
            return {"uptime_pct": None, "lat_avg": None, "lat_max": None,
                    "lat_p95": None, "p95_from_raw": False,
                    "outage_count": 0,
                    "outage_minutes": 0, "sample_n": 0}
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def get_sla_stats_batch(self, target_ids: list,
                            start_ts: float, end_ts: float) -> dict:
        """
        Bulk version of get_sla_stats() for the entire dashboard.

        Old /api/sla loop: 50 nodes × 2 SQL each = 100 round-trips per
        request.  Each round-trip is cheap in isolation but they're
        serial because Flask's request handler is single-threaded.

        Latency: one batch pings_hourly SELECT plus at most a few pings_raw
        queries per target for partial boundary hours (not per target-hour).
        Outage: two alert_events IN (...) queries, then the same per-target
        walk as get_sla_stats().

        Returns {target_id: {uptime_pct, lat_avg, lat_max, lat_p95,
        outage_count, outage_minutes, sample_n}}.  Targets present in
        `target_ids` but absent from the DB get the "no data" defaults
        (uptime_pct=100.0, outage_count=0, sample_n=0).
        """
        if not target_ids:
            return {}

        # Defensive chunking: SQLite 3.32+ allows 32766 parameters but
        # older builds (some Python embed) cap at 999.  500 is comfortably
        # under both while still cutting round-trips by ~50× vs the
        # one-at-a-time loop.
        CHUNK = 500
        if len(target_ids) > CHUNK:
            out = {}
            for i in range(0, len(target_ids), CHUNK):
                out.update(self.get_sla_stats_batch(
                    target_ids[i:i+CHUNK], start_ts, end_ts))
            return out

        try:
            conn = self._read_conn()
            # row_factory persists on the cached conn (see _read_conn doc);
            # we still set it locally to make this method's logic explicit.
            conn.row_factory = sqlite3.Row

            placeholders = ",".join("?" * len(target_ids))

            # ── Query 1: aggregated latency per target ─────────────────
            # Same success_n weighting as get_sla_stats above -- see that
            # function for the full rationale.  This per-target rollup must
            # agree with the single-target rollup, otherwise the Web SLA view
            # and the Excel SLA sheet would disagree on the same data.
            buckets_by_tid = self._bounded_hourly_buckets_for_targets(
                conn, target_ids, start_ts, end_ts)
            lat_by_tid = {}
            for tid in target_ids:
                merged = self._rollup_latency_for_window(
                    conn, tid, buckets_by_tid.get(tid, []), start_ts, end_ts)
                if merged.get("total"):
                    lat_by_tid[tid] = {"target_id": tid, **merged}

            # ── Query 2: per-target boundary state (single row each).
            # JOIN'ing on (target_id, MAX(ts)) lets SQLite resolve each
            # target's "latest pre-window red-touching event" via the
            # composite (target_id, ts) index in a single statement.
            # The previous "ts <= end_ts" version pulled every historical
            # red-touching event for every target just to find one
            # boundary row per target — for 50 nodes × 50k events that
            # was hundreds of MB into Flask's per-request memory.
            boundary_rows = conn.execute(f"""
                SELECT e.target_id, e.new_status
                FROM alert_events e
                INNER JOIN (
                    SELECT target_id, MAX(ts) AS max_ts
                    FROM alert_events
                    WHERE target_id IN ({placeholders})
                      AND (new_status = 'red' OR old_status = 'red')
                      AND ts < ?
                    GROUP BY target_id
                ) latest ON e.target_id = latest.target_id
                        AND e.ts        = latest.max_ts
                WHERE e.target_id IN ({placeholders})
                  AND (e.new_status = 'red' OR e.old_status = 'red')
            """, (*target_ids, start_ts, *target_ids)).fetchall()
            status_at_start_by_tid = {
                r["target_id"]: r["new_status"] for r in boundary_rows
            }

            # ── Query 3: in-window red-touching events, sorted per target.
            ev_rows = conn.execute(f"""
                SELECT target_id, ts, old_status, new_status
                FROM alert_events
                WHERE target_id IN ({placeholders})
                  AND (new_status = 'red' OR old_status = 'red')
                  AND ts >= ? AND ts <= ?
                ORDER BY target_id ASC, ts ASC, id ASC
            """, (*target_ids, start_ts, end_ts)).fetchall()
            events_by_tid: dict = {}
            for r in ev_rows:
                events_by_tid.setdefault(r["target_id"], []).append(r)

            # ── Per-target outage walk (identical to get_sla_stats) ────
            now = time.time()
            elapsed = min(end_ts, now) - start_ts
            period_secs = max(elapsed, 1)

            result = {}
            for tid in target_ids:
                lat = lat_by_tid.get(tid)
                total   = (lat["total"]   if lat else None) or 0
                success = (lat["success"] if lat else None) or 0
                # Preserve None for "no successful probes" (same NULL→None
                # rationale as get_sla_stats above).  `(... or 0.0)`
                # collapses both "no row at all" and "row exists but
                # success_n=0" to a misleading "0.0ms".
                lat_avg = (round(lat["lat_avg"], 2) if lat and lat["lat_avg"] is not None
                           else None)
                lat_max = (round(lat["lat_max"], 2) if lat and lat["lat_max"] is not None
                           else None)
                lat_p95 = (round(lat["lat_p95"], 2) if lat and lat["lat_p95"] is not None
                           else None)

                events = events_by_tid.get(tid, [])
                status_at_start = status_at_start_by_tid.get(tid, "green")

                outage_secs, outage_count = self._compute_outage_stats(
                    float(start_ts), float(end_ts), status_at_start, events, now)

                uptime_pct = round(
                    max(0.0, (period_secs - outage_secs) / period_secs) * 100, 3)
                result[tid] = {
                    "uptime_pct":     uptime_pct,
                    "lat_avg":        lat_avg,
                    "lat_max":        lat_max,
                    "lat_p95":        lat_p95,
                    "p95_from_raw":   bool(lat.get("p95_from_raw")) if lat else False,
                    "outage_count":   outage_count,
                    "outage_minutes": round(outage_secs / 60, 1),
                    "sample_n":       total,
                }
            return result
        except Exception as e:
            print(f"[SLA-batch] {e}")
            # Fall back to per-target loop so a query bug doesn't break the
            # whole dashboard for everyone — only the batch path silently
            # degrades to the legacy 50× round-trip path.
            return {tid: self.get_sla_stats(tid, start_ts, end_ts)
                    for tid in target_ids}
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def get_outage_intervals(self, target_id: str,
                             start_ts: float, end_ts: float) -> dict:
        """
        Return merged red / orange intervals + raw transition events for a
        target within [start_ts, end_ts].

        Designed for the waveform-chart "alert overlay" feature: callers
        get back ready-to-render segments (no client-side event walking)
        plus the original transition list so the UI can also annotate
        precise red→/→green moments on top of the segment bands.

        Output:
          {
            "red":    [{"start": ts, "end": ts, "ongoing": bool}, ...],
            "orange": [{"start": ts, "end": ts, "ongoing": bool}, ...],
            "events": [{"ts": ts, "old": "x", "new": "y", "category": "z"}, ...]
          }

        Boundary handling (mirrors get_sla_stats Case A/B/C):
          - Case A: outage active before start_ts → segment clamped to start
          - Case B: outage active at end_ts      → segment clamped to
                    min(end_ts, now()), marked ongoing=True
          - Case C: status was red the entire span → one segment covering
                    the whole window
        red and orange are tracked independently; a node moving red→orange
        ends the red segment and opens an orange one, etc.
        """
        try:
            conn = self._read_conn()
            conn.row_factory = sqlite3.Row

            # Split into TWO bounded queries (same rationale as
            # get_sla_stats): one row for the pre-window boundary
            # state, then only the events that actually fall inside
            # [start_ts, end_ts].  Note we cannot filter by red-touching
            # here -- orange segments need full transition history too.
            boundary = conn.execute(
                "SELECT new_status FROM alert_events "
                "WHERE target_id=? AND ts<? "
                "ORDER BY ts DESC LIMIT 1",
                (target_id, start_ts)).fetchone()
            status_at_start = boundary["new_status"] if boundary else "green"

            rows = conn.execute(
                "SELECT ts, old_status, new_status, category "
                "FROM alert_events "
                "WHERE target_id=? AND ts>=? AND ts<=? "
                "ORDER BY ts ASC",
                (target_id, start_ts, end_ts)).fetchall()

            now_ts = time.time()

            red_segs:    list[dict] = []
            orange_segs: list[dict] = []

            # Initialise open segments based on pre-window status (Case A/C)
            red_open    = start_ts if status_at_start == "red"    else None
            orange_open = start_ts if status_at_start == "orange" else None

            for r in rows:
                ts_ev = r["ts"]
                old_s = r["old_status"]
                new_s = r["new_status"]

                # Close any RED segment when leaving red
                if old_s == "red" and new_s != "red" and red_open is not None:
                    red_segs.append({"start": red_open,
                                     "end":   ts_ev,
                                     "ongoing": False})
                    red_open = None

                # Close any ORANGE segment when leaving orange
                if (old_s == "orange" and new_s != "orange"
                        and orange_open is not None):
                    orange_segs.append({"start": orange_open,
                                        "end":   ts_ev,
                                        "ongoing": False})
                    orange_open = None

                # Open new segment when entering the corresponding state
                if new_s == "red" and red_open is None:
                    red_open = ts_ev
                if new_s == "orange" and orange_open is None:
                    orange_open = ts_ev

            # Case B: clamp ongoing segments to min(end_ts, now)
            # — same reasoning as get_sla_stats: avoids overstating an
            #   outage when end_ts is a future end-of-day timestamp.
            cap = min(end_ts, now_ts)
            if red_open is not None:
                red_segs.append({"start": red_open,
                                 "end":   cap,
                                 "ongoing": True})
            if orange_open is not None:
                orange_segs.append({"start": orange_open,
                                    "end":   cap,
                                    "ongoing": True})

            # Raw transition events within the window — used by the UI to
            # draw vertical ▼/▲ markers at precise transition points.
            # We include both directions touching red OR orange so the
            # markers are meaningful for both severity levels.
            events = [{"ts": r["ts"],
                       "old": r["old_status"],
                       "new": r["new_status"],
                       "category": r["category"]}
                      for r in rows
                      if start_ts <= r["ts"] <= end_ts
                      and (r["new_status"] in ("red", "orange")
                           or r["old_status"] in ("red", "orange"))]

            return {"red": red_segs, "orange": orange_segs, "events": events}
        except Exception as e:
            print(f"[outage] {e}")
            return {"red": [], "orange": [], "events": []}
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def record_traceroute(self, target_id: str, label: str,
                          ip: str, ts: float, hops: list,
                          route_changed: bool = False,
                          captured_generation: Optional[int] = None) -> None:
        """Queue an auto-traceroute result for async DB write.

        captured_generation: optional snapshot of
        current_target_generation(target_id) taken when the
        traceroute task captured its target identity in
        _TracerouteScheduler._run_one.  Travels with the queued
        message so the writer thread can re-check right before the
        INSERT -- if an IP edit / removal / history wipe lands
        between the scheduler's cache commit and this INSERT, the
        writer drops the row instead of repopulating traceroute_logs
        with stale-identity data.  Same shape as record_diag_result.
        """
        import json
        # Encode route_changed flag into the hops JSON metadata so we don't
        # need a schema migration — the flag travels with the data.
        if (captured_generation is not None and target_id is not None
            and self.current_target_generation(target_id) != captured_generation):
            return
        payload = {"hops": hops, "route_changed": bool(route_changed)}
        self._queue.put(("traceroute", ((
            target_id, label, ip, ts,
            json.dumps(payload, ensure_ascii=False),
        ), captured_generation)))

    def set_wipe_complete_callback(self, callback) -> None:
        """Register fn(target_id) called on the write thread after wipe commits."""
        self._wipe_complete_callback = callback

    def wipe_target_history(self, target_id: str) -> None:
        """Erase every historical row belonging to ``target_id``.

        Used by the "更换监控对象（清空历史）" path in MainWindow._on_edit,
        when the operator explicitly indicates that an IP/probe-type change
        means this slot is being repurposed for a different physical node
        rather than tracking the same logical entity across a renumbering.

        Without this, the new node's data would be silently appended to the
        old node's time series under the same target_id, producing nonsense
        SLA averages and a "Frankenstein timeline" in the waveform chart.

        Queued (not synchronous): the actual DELETE runs on the write
        thread, preserving serialisation with any concurrent record_ping /
        record_alert / record_traceroute / record_diag_result still in
        flight from probes that fired just before the wipe.  Tables
        covered: pings_raw, pings_hourly, alert_events, traceroute_logs,
        icmp_diagnostic_history, dns_diagnostic_history, cum_stats.
        targets_meta is intentionally NOT touched -- it will be overwritten
        by the next probe with the new (label, ip, ping_type), and clearing
        it here would just race with that write.

        Also pops target_id from in-memory _active_incidents on the same
        thread that owns that dict, so a stale open-incident pointer can't
        leak into the next red transition.
        """
        # Bump generation SYNCHRONOUSLY before queuing the wipe so any
        # ICMP / DNS diag task whose persist call lands AFTER the
        # wipe job runs sees a generation mismatch and skips the
        # write -- otherwise the stale diag result would re-insert
        # the very rows the wipe just deleted.  Mirrors what the
        # web_server._TracerouteScheduler.bump_generation hook does
        # for tracert.
        self.bump_target_generation(target_id)
        self._queue.put(("wipe_target", {"target_id": target_id}))

    def get_traceroute_history(self, target_id: str,
                               limit: int = 5) -> list:
        """Return the last N auto-traceroute snapshots for a target."""
        import json
        conn = None
        try:
            conn = self._read_conn()
            rows = conn.execute(
                "SELECT ts, hops FROM traceroute_logs"
                " WHERE target_id=? ORDER BY ts DESC LIMIT ?",
                (target_id, limit)).fetchall()
            results = []
            for r in rows:
                raw = json.loads(r[1]) if r[1] else {}
                # Support both old format (plain list) and new format (dict with hops+flag)
                if isinstance(raw, list):
                    results.append({"ts": r[0], "hops": raw, "route_changed": False})
                else:
                    results.append({"ts": r[0],
                                    "hops": raw.get("hops", []),
                                    "route_changed": raw.get("route_changed", False)})
            return results
        except Exception:
            return []
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def get_incident_traceroute(self, target_id: str, red_ts: float,
                                window_before: int = 10,
                                window_after: int = 180,
                                incident_id: Optional[str] = None) -> Optional[dict]:
        """Return the auto-traceroute snapshot captured for one incident.

        Used by export_excel to attach forensic "where did the path break"
        evidence to each fault incident, even after the link has recovered
        and the live tracert would show a healthy path.

        Correlation strategy (incremental — see HANDOFF design notes):
          • Level 2 (preferred, exact): if `incident_id` is given AND the
            traceroute_logs table carries an incident_id column, match on it
            directly.  No time heuristic.
          • Level 1 (fallback, time-window): match the earliest tracert whose
            ts falls in [red_ts - window_before, red_ts + window_after].
            The auto-tracert is request_now()'d the instant a node turns red
            and runs within ~5s, so the first row after red_ts is it.
            When Level 2 is active and an ``incident_id`` was supplied but
            exact match missed, fallback only considers rows whose
            ``incident_id IS NULL`` — legacy rows from before the column
            existed, or the rare brief-outage race where tracert INSERT ran
            after the incident was already popped.  Rows already bound to
            another incident_id are never borrowed across incidents.
            When no ``incident_id`` is supplied (Level-1-only schema), the
            time window is unrestricted as before.

        `red_ts` MUST be the timestamp of the *first red transition* of the
        incident — auto-tracert only fires on red, never on orange-only
        performance warnings.

        Returns a dict {ts, hops, route_changed} or None if no snapshot found.
        """
        import json
        conn = None
        try:
            conn = self._read_conn()
            row = None

            # ── Level 2: exact incident_id match (only if column present) ──
            if incident_id and self._traceroute_has_incident_col(conn):
                row = conn.execute(
                    "SELECT ts, hops FROM traceroute_logs "
                    "WHERE target_id=? AND incident_id=? "
                    "ORDER BY ts ASC LIMIT 1",
                    (target_id, incident_id)).fetchone()

            # ── Level 1: time-window fallback ──────────────────────────────
            if row is None:
                lo = red_ts - window_before
                hi = red_ts + window_after
                if incident_id and self._traceroute_has_incident_col(conn):
                    # Never borrow another incident's bound row (Bug 153).
                    row = conn.execute(
                        "SELECT ts, hops FROM traceroute_logs "
                        "WHERE target_id=? AND ts BETWEEN ? AND ? "
                        "AND (incident_id IS NULL OR incident_id = '') "
                        "ORDER BY ts ASC LIMIT 1",
                        (target_id, lo, hi)).fetchone()
                else:
                    row = conn.execute(
                        "SELECT ts, hops FROM traceroute_logs "
                        "WHERE target_id=? AND ts BETWEEN ? AND ? "
                        "ORDER BY ts ASC LIMIT 1",
                        (target_id, lo, hi)).fetchone()

            if row is None:
                return None

            raw = json.loads(row[1]) if row[1] else {}
            if isinstance(raw, list):       # legacy plain-list format
                hops = raw
                route_changed = False
            else:
                hops = raw.get("hops", [])
                route_changed = raw.get("route_changed", False)
            return {"ts": row[0], "hops": hops, "route_changed": route_changed}
        except Exception:
            return None
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def _traceroute_has_incident_col(self, conn) -> bool:
        """Whether traceroute_logs has the incident_id column (Level 2 schema).

        Cached after first probe.  Returns False on the Level-1-only schema
        so get_incident_traceroute degrades cleanly to time-window matching.
        """
        cached = getattr(self, "_tracert_incident_col", None)
        if cached is not None:
            return cached
        try:
            cols = {r[1] for r in
                    conn.execute("PRAGMA table_info(traceroute_logs)").fetchall()}
            has = "incident_id" in cols
        except Exception:
            has = False
        self._tracert_incident_col = has
        return has

    @staticmethod
    def summarize_break(hops: list) -> dict:
        """Delegate to shared traceroute_summary (Excel / webhook / UI)."""
        from src.traceroute_summary import summarize_break as _summarize_break
        return _summarize_break(hops)

    # ── ICMP diagnostic history (Phase 3) ─────────────────────────────

    def record_diag_result(self,
                           target_id: Optional[str],
                           host: str,
                           mode: str,
                           payload_size: int,
                           df: bool,
                           count: int,
                           success_count: int,
                           loss_count: int,
                           avg_rtt_ms: Optional[float],
                           min_rtt_ms: Optional[float],
                           max_rtt_ms: Optional[float],
                           estimated_path_mtu: Optional[int],
                           max_df_payload: Optional[int],
                           error_summary: Optional[str],
                           conclusion: str,
                           details: Optional[dict],
                           created_ts: Optional[float] = None,
                           captured_generation: Optional[int] = None) -> None:
        """Queue an ICMP diagnostic result for async insertion.

        target_id may be None for ad-hoc diagnostics not tied to a monitored
        node (future Phase: allow diagnosing arbitrary IPs).  details should
        be a small JSON-serialisable dict containing per-probe samples; it
        is stored as a TEXT blob and only inflated when the history detail
        panel is opened, keeping the history-list query lightweight.

        captured_generation: optional snapshot of
        current_target_generation(target_id) taken when the diag task
        STARTED.  Travels with the queued message so the writer
        thread can re-check it RIGHT BEFORE the INSERT.  The producer-
        side check below is just a fast path -- the queue is FIFO but
        the check + put are NOT atomic, so a wipe_target_history that
        runs between this generation read and the put() lands its
        DELETE first and our stale INSERT second.  The writer-side
        check is the authoritative guard.
        """
        if (captured_generation is not None and target_id is not None
            and self.current_target_generation(target_id) != captured_generation):
            return
        import json
        details_json = json.dumps(details, ensure_ascii=False, default=str) \
            if details is not None else None
        ts = created_ts if created_ts is not None else time.time()
        self._queue.put(("diag", ((
            target_id, host, mode, int(payload_size), 1 if df else 0,
            int(count), int(success_count), int(loss_count),
            avg_rtt_ms, min_rtt_ms, max_rtt_ms,
            estimated_path_mtu, max_df_payload,
            error_summary, conclusion, details_json, float(ts),
        ), captured_generation)))

    def get_diag_history(self, target_id: Optional[str] = None,
                         host: Optional[str] = None,
                         limit: int = 50) -> list[dict]:
        """Return recent diagnostic runs, optionally filtered by target.

        If target_id is given, returns rows matching that target; otherwise
        if host is given, returns rows matching the raw host string (useful
        for cross-node diagnoses of the same IP); otherwise returns the
        global recent-history (all diagnoses, newest first).

        Each row is a dict with all summary fields plus the parsed details
        dict (so callers can show per-probe samples without a second query).
        """
        import json
        clauses, args = [], []
        if target_id:
            clauses.append("target_id = ?")
            args.append(target_id)
        elif host:
            clauses.append("host = ?")
            args.append(host)
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        args.append(int(limit))

        try:
            conn = self._read_conn()
            rows = conn.execute(
                f"SELECT id, target_id, host, mode, payload_size, df, count, "
                f"       success_count, loss_count, avg_rtt_ms, min_rtt_ms, "
                f"       max_rtt_ms, estimated_path_mtu, max_df_payload, "
                f"       error_summary, conclusion, details_json, created_ts "
                f"FROM icmp_diagnostic_history {where} "
                f"ORDER BY created_ts DESC LIMIT ?",
                args).fetchall()
            results = []
            for r in rows:
                details = None
                if r[16]:
                    try: details = json.loads(r[16])
                    except Exception: details = None
                results.append({
                    "id":                 r[0],
                    "target_id":          r[1],
                    "host":               r[2],
                    "mode":               r[3],
                    "payload_size":       r[4],
                    "df":                 bool(r[5]),
                    "count":              r[6],
                    "success_count":      r[7],
                    "loss_count":         r[8],
                    "avg_rtt_ms":         r[9],
                    "min_rtt_ms":         r[10],
                    "max_rtt_ms":         r[11],
                    "estimated_path_mtu": r[12],
                    "max_df_payload":     r[13],
                    "error_summary":      r[14],
                    "conclusion":         r[15],
                    "details":            details,
                    "created_ts":         r[17],
                })
            return results
        except Exception:
            return []
        # _read_conn() returns a thread-local cached connection — do NOT close.

    # ── DNS diagnostic history (Phase 3A) ─────────────────────────────

    def record_dns_diag_result(self,
                               target_id: Optional[str],
                               domain: str,
                               qtype: str,
                               trace_chain: bool,
                               resolver: Optional[str],
                               success: bool,
                               rcode: Optional[str],
                               elapsed_ms: Optional[float],
                               answer_count: int,
                               error_summary: Optional[str],
                               message: Optional[str],
                               details: Optional[dict],
                               created_ts: Optional[float] = None,
                               captured_generation: Optional[int] = None) -> None:
        """Queue a DNS diagnostic result for async insertion.

        Mirrors record_diag_result() but for the DNS diagnostic engine
        (src.dns_diag).  target_id may be None for ad-hoc diagnoses not
        tied to a monitored node.  details should be a small
        JSON-serialisable dict produced by dns_result_to_dict() —
        stored as a TEXT blob and only inflated when the history detail
        panel is opened, so the list query stays cheap.

        Cancelled runs (rcode == ERROR_CANCELLED) MUST NOT reach this
        method — the caller (dns_diag._persist_dns_result) gates that.

        captured_generation: see record_diag_result.  Travels with
        the queued message so the writer thread can re-check it
        right before the INSERT -- the producer-side check below is
        a fast path; the writer-side check is what closes the
        check-vs-put race.
        """
        if (captured_generation is not None and target_id is not None
            and self.current_target_generation(target_id) != captured_generation):
            return
        import json
        details_json = json.dumps(details, ensure_ascii=False, default=str) \
            if details is not None else None
        ts = created_ts if created_ts is not None else time.time()
        self._queue.put(("dns_diag", ((
            target_id, domain, qtype, 1 if trace_chain else 0,
            resolver, 1 if success else 0, rcode,
            elapsed_ms, int(answer_count),
            error_summary, message, details_json, float(ts),
        ), captured_generation)))

    def get_dns_diag_history(self, target_id: Optional[str] = None,
                             domain: Optional[str] = None,
                             limit: int = 50) -> list[dict]:
        """Return recent DNS diagnostic runs, optionally filtered.

        Filter precedence (matches get_diag_history):
          target_id, then domain, then unfiltered (global recent).

        Each row dict includes both summary columns and the parsed
        details dict (cname_chain / answers / trace_steps as produced
        by dns_result_to_dict()), so callers can render a replay
        without a second query.
        """
        import json
        clauses, args = [], []
        if target_id:
            clauses.append("target_id = ?")
            args.append(target_id)
        elif domain:
            clauses.append("domain = ?")
            args.append(domain)
        where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
        args.append(int(limit))

        try:
            conn = self._read_conn()
            rows = conn.execute(
                f"SELECT id, target_id, domain, qtype, trace_chain, "
                f"       resolver, success, rcode, elapsed_ms, "
                f"       answer_count, error_summary, message, "
                f"       details_json, created_ts "
                f"FROM dns_diagnostic_history {where} "
                f"ORDER BY created_ts DESC LIMIT ?",
                args).fetchall()
            results = []
            for r in rows:
                details = None
                if r[12]:
                    try: details = json.loads(r[12])
                    except Exception: details = None
                results.append({
                    "id":            r[0],
                    "target_id":     r[1],
                    "domain":        r[2],
                    "qtype":         r[3],
                    "trace_chain":   bool(r[4]),
                    "resolver":      r[5],
                    "success":       bool(r[6]),
                    "rcode":         r[7],
                    "elapsed_ms":    r[8],
                    "answer_count":  r[9],
                    "error_summary": r[10],
                    "message":       r[11],
                    "details":       details,
                    "created_ts":    r[13],
                })
            return results
        except Exception:
            return []
        # _read_conn() returns a thread-local cached connection — do NOT close.

    def _cleanup(self, conn):
        """删除超出保留期的数据。"""
        now = int(time.time())
        with conn:
            conn.execute("DELETE FROM pings_raw WHERE ts<?",
                         (now - self._raw_retention_days * 86400,))
            conn.execute("DELETE FROM pings_hourly WHERE hour_ts<?",
                         (now - self._hourly_retention_days * 86400,))
            # Keep per-target anchors across retention cutoff (Bug94):
            #   * latest event overall — ongoing state / recovery tail;
            #   * latest event with ts < cutoff — left boundary for SLA
            #     status_at_start when a fault started before cutoff and
            #     recovered inside the retention window.
            alert_cutoff = now - self._alert_retention_days * 86400
            conn.execute(
                "DELETE FROM alert_events WHERE ts<? "
                "AND id NOT IN ("
                "  SELECT MAX(id) FROM alert_events GROUP BY target_id "
                "  UNION "
                "  SELECT MAX(id) FROM alert_events WHERE ts<? "
                "  GROUP BY target_id"
                ")",
                (alert_cutoff, alert_cutoff))
            conn.execute("DELETE FROM traceroute_logs WHERE ts<?",
                         (now - self._traceroute_retention_days * 86400,))
            conn.execute("DELETE FROM icmp_diagnostic_history WHERE created_ts<?",
                         (now - self._diag_retention_days * 86400,))
            conn.execute("DELETE FROM dns_diagnostic_history WHERE created_ts<?",
                         (now - self._diag_retention_days * 86400,))

    def _read_conn(self) -> sqlite3.Connection:
        """
        Return this thread's cached read-only SQLite connection, opening one
        on first use.  WAL mode is configured once at open time so subsequent
        get_*() calls go straight to queries with no PRAGMA overhead.

        IMPORTANT: callers MUST NOT call conn.close() on the returned
        connection.  Closing it would evict the cached instance but other
        already-running code paths on the same thread may still hold a
        reference and try to use it, raising sqlite3.ProgrammingError.
        The connection's lifetime is tied to the thread; daemon worker
        threads get cleaned up automatically at process exit, and SQLite's
        own finalisation closes the underlying file handle there.

        Side-effect cleanup: callers that set conn.row_factory (e.g. for
        sqlite3.Row access) DO need to be aware the setting persists to
        the next caller on the same thread.  This is acceptable today
        because the only caller setting row_factory (get_sla_stats) sets
        it to sqlite3.Row, which is harmless if inherited by tuple-index
        callers (Row supports both numeric and name indexing).
        """
        conn = getattr(self._tls, "conn", None)
        if conn is not None:
            return conn
        try:
            conn = sqlite3.connect(
                f"file:{self._db_path}?mode=ro", uri=True,
                check_same_thread=False)
        except Exception:
            # DB 文件可能还不存在（刚启动时 schema 还没建），回退到普通连接
            conn = sqlite3.connect(self._db_path, check_same_thread=False)
        # WAL is set once per connection's lifetime; subsequent set on the
        # same conn is a no-op but pointless overhead.
        conn.execute("PRAGMA journal_mode=WAL")
        self._tls.conn = conn
        return conn

    def release_read_conn(self) -> None:
        """Close this thread's cached read connection (Flask teardown hook)."""
        conn = getattr(self._tls, "conn", None)
        if conn is None:
            return
        try:
            conn.close()
        except Exception:
            pass
        try:
            del self._tls.conn
        except AttributeError:
            pass
